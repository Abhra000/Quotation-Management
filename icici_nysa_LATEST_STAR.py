"""
╔══════════════════════════════════════════════════════════════╗
║   ICICI NYSA — Motor Insurance Automation                    ║
║   Connects to YOUR existing Chrome — no new tab/login        ║
╚══════════════════════════════════════════════════════════════╝

STEP 1 — Start Chrome with remote debugging (run this ONCE):
  Press  Win + R  on your keyboard, paste this and press Enter:

  "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\\chrome-nysa"

  (A new Chrome window will open)

STEP 2 — In that Chrome window:
  Go to:  https://nysa.icicilombard.com
  Login with your ID, Password, Captcha, OTP — just like normal.
  Make sure you are on the HOME page.

STEP 3 — Fill details.xlsx row 5 with your vehicle data.

STEP 4 — Run this script:
  python icici_nysa.py

The script attaches to your already-logged-in Chrome.
No new tab. No new login. Starts directly from home page.
"""

import asyncio
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright

# ── Paths ──────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "details.xlsx"
LOG_PATH   = BASE_DIR / "proposal_log.txt"

# ── Must match the port you used to start Chrome ──────────────────
CHROME_PORT = 9222


# ═══════════════════════════════════════════════════════════════════
#  READ EXCEL
# ═══════════════════════════════════════════════════════════════════

def read_details(start_row=4):
    if not EXCEL_PATH.exists():
        print(f"\n  Cannot find: {EXCEL_PATH}")
        print("  Make sure details.xlsx is in the same folder.")
        input("\nPress ENTER to exit..."); raise SystemExit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "Details" not in wb.sheetnames:
        print("  Sheet 'Details' not found in details.xlsx")
        input("\nPress ENTER to exit..."); raise SystemExit(1)

    ws = wb["Details"]

    # ── Label map: handles any variation of header text ──────────
    # Strip special chars, lowercase, then match
    label_map = {
        # Motor Category & Type
        "motor category"                              : "motor_category",
        "motor category *"                            : "motor_category",
        "motor type"                                  : "motor_type",
        "motor type *"                                : "motor_type",
        # PID Number
        "pid number"                                  : "pid_number",
        "pid no"                                      : "pid_number",
        # Vehicle & Contact
        "vehicle reg no"                              : "vehicle_reg_number",
        "vehicle reg number"                          : "vehicle_reg_number",
        "registration number"                         : "vehicle_reg_number",
        "reg no"                                      : "vehicle_reg_number",
        "mobile number"                               : "mobile_number",
        "mobile"                                      : "mobile_number",
        "phone"                                       : "mobile_number",
        "email"                                       : "email",
        "policy type"                                 : "policy_type",
        "policy type *"                               : "policy_type",
        # Ownership
        "ownership"                                   : "ownership",
        # SAOD
        "saod"                                        : "saod",
        # Rollover / Used registration number
        "rollover registration number"                : "rollover_reg_number",
        "used registration number"                    : "rollover_reg_number",
        "registration number for rollover"            : "rollover_reg_number",
        # Registration Type — matches portal exact header
        "current registration type"                   : "current_reg_type",
        "current reg type"                            : "current_reg_type",
        "reg type"                                    : "current_reg_type",
        "registration type"                           : "current_reg_type",
        # RTO — matches portal exact header
        "city where vehicle is registered rto"        : "rto",
        "city where vehicle is registered"            : "rto",
        "rto"                                         : "rto",
        "rto code"                                    : "rto",
        "city rto"                                    : "rto",
        # Manufacturer/Model
        "manufacturer - model"                        : "manufacturer_model",
        "manufacturer-model"                          : "manufacturer_model",
        "manufacturer model"                          : "manufacturer_model",
        "make model"                                  : "manufacturer_model",
        "make - model"                                : "manufacturer_model",
        # Year
        "mfg year"                                    : "manufacturing_year",
        "tenure"                                      : "tenure",
        "tenure *"                                    : "tenure",
        "manufacturing year"                          : "manufacturing_year",
        "year"                                        : "manufacturing_year",
        # IDV — matches portal exact header
        "insured depreciation value idv"              : "idv",
        "insured declared value idv"                  : "idv",
        "insured depreciation value"                  : "idv",
        "insured declared value"                      : "idv",
        "idv"                                         : "idv",
        # KYC
        "pan number"                                  : "pan_number",
        "pan"                                         : "pan_number",
        "dob dd/mm/yyyy"                              : "dob",
        "dob dd mm yyyy"                              : "dob",
        "dob"                                         : "dob",
        "date of birth"                               : "dob",
        "aadhaar no"                                  : "aadhaar_number",
        "aadhaar"                                     : "aadhaar_number",
        # Engine / Chassis
        "engine number"                               : "engine_number",
        "engine no"                                   : "engine_number",
        "chassis number"                              : "chassis_number",
        "chassis no"                                  : "chassis_number",
        # GST — no bare "gst" (collides with "gst state")
        "gst number"                                  : "gst_number",
        "gstin"                                       : "gst_number",
        "gst state"                                   : "gst_state",
        # Pincode
        "pin code"                                    : "pin_code",
        "pincode"                                     : "pin_code",
        "pin code on gst certificate"                 : "pin_code",
        # Proposal form
        "insured name"                                : "insured_name",
        "address line 1"                              : "address_line",
        "address line"                                : "address_line",
        # Add-ons
        "add-ons comma separated"                     : "addons",
        "add ons comma separated"                     : "addons",
        "add-ons (comma separated)"                   : "addons",
        "addons"                                      : "addons",
        "add-ons"                                     : "addons",
        "add ons"                                     : "addons",
        "advanced pid"                                : "advanced_pid",
        "pid"                                         : "advanced_pid",
        # Payment
        "auto pay true/false"                         : "auto_pay",
        "auto pay (true/false)"                       : "auto_pay",
        "auto pay true false"                         : "auto_pay",
        "auto pay"                                    : "auto_pay",
        "autopay"                                     : "auto_pay",
        # Loan / Hypothecation
        "loan / lease / hypothecation details"         : "loan_hypothecation",
        "loan lease hypothecation details"             : "loan_hypothecation",
        "loan/lease/hypothecation details"             : "loan_hypothecation",
        "loan hypothecation"                           : "loan_hypothecation",
        "hypothecation"                                : "loan_hypothecation",
        "financier name"                               : "financier_name",
        "financier branch"                             : "financier_branch",
        "kyc pincode"                                  : "kyc_pincode",
        "kyc pin code"                                 : "kyc_pincode",
        "kyc pin"                                      : "kyc_pincode",
    }

    positional_map = {
        1:"motor_category",      2:"motor_type",
        3:"email",               4:"policy_type",         5:"ownership",
        6:"saod",                7:"rollover_reg_number", 8:"current_reg_type",
        9:"rto",                 10:"manufacturer_model", 11:"manufacturing_year",
        12:"tenure",             13:"idv",                14:"pan_number",
        15:"aadhaar_number",     16:"engine_number",      17:"chassis_number",
        18:"kyc_pincode",        19:"insured_name",       20:"dob",
        21:"address_line",       22:"mobile_number",      23:"email",
        24:"pin_code",           25:"gst_number",         26:"gst_state",
        27:"loan_hypothecation", 28:"financier_name",     29:"financier_branch",
        30:"addons",             31:"auto_pay",
        32:"pid_number",         33:"proposal_number",
    }

    import re as _re
    headers = {}
    header_row_found = False
    for col in range(1, 35):  # cols 1-34: includes new fields Ownership(7), SAOD(8), Rollover Reg No(9); KYC Pincode moved to col 20; PID Number(34), Proposal Number(35), Policy Number(36)
        val = ws.cell(row=3, column=col).value
        if val:
            header_row_found = True
            clean = _re.sub(r"[*()\[\]₹]", " ", str(val))
            clean = _re.sub(r"\s+", " ", clean).strip().lower()
            field = label_map.get(clean)
            if not field:
                c2 = clean.replace(" ","")
                for k,v in label_map.items():
                    if c2 == k.replace(" ",""): field = v; break
            if not field:
                for k,v in label_map.items():
                    if clean.startswith(k) and (len(clean)==len(k) or clean[len(k)]==" "):
                        field = v; break
            if not field:
                field = _re.sub(r"[^a-z0-9]+","_",clean).strip("_")
            headers[col] = field
            print(f"  [HEADER] col {col}: {repr(str(val))} → {field}")

    if not header_row_found or not headers:
        print("  [INFO] Using positional column map")
        headers = positional_map.copy()

    data = {}
    max_col = max(headers.keys()) + 1 if headers else 22
    for row in range(start_row, ws.max_row + 1):
        row_vals = [ws.cell(row=row, column=c).value for c in range(1, max_col)]
        if any(v and str(v).strip() for v in row_vals):
            for col, field in headers.items():
                raw = ws.cell(row=row, column=col).value
                if hasattr(raw, 'strftime'):
                    data[field] = raw.strftime("%d/%m/%Y")
                else:
                    data[field] = str(raw).strip() if raw is not None else ""
            print(f"  [INFO] Reading data from row {row}")
            data["_excel_row"] = row
            break

    if not data:
        print("  No data found in details.xlsx")
        input("\nPress ENTER to exit..."); raise SystemExit(1)

    # Store the Excel row number so we can write proposal number back later

    data["vehicle_reg_number"] = data.get("vehicle_reg_number","").upper().replace(" ","")
    data["pan_number"]         = data.get("pan_number","").upper()
    data["gst_number"]         = data.get("gst_number","").upper()
    data["pin_code"]           = str(data.get("pin_code","")).split(".")[0].strip()
    data["engine_number"]      = data.get("engine_number","").upper().replace(" ","")
    data["chassis_number"]     = data.get("chassis_number","").upper().replace(" ","")
    data["rto"]                = data.get("rto","").upper().replace(" ","")

    data["manufacturer_model"] = data.get("manufacturer_model","")
    data["manufacturing_year"] = str(data.get("manufacturing_year","")).split(".")[0].strip()
    data["idv"]                = (data.get("idv","")
                                  .replace(",","").replace("₹","").replace(" ","").strip())
    data["current_reg_type"]   = data.get("current_reg_type","")
    data["tenure"]             = data.get("tenure","1 year").strip()
    raw_addons                 = data.get("addons","")
    data["addons"]             = [a.strip() for a in raw_addons.split(",") if a.strip()]
    data["auto_pay"]           = str(data.get("auto_pay","false")).strip().lower() == "true"
    data["policy_type"]        = data.get("policy_type","New").strip()
    data["ownership"]          = data.get("ownership","").strip()
    data["saod"]               = str(data.get("saod","")).strip().lower()
    data["rollover_reg_number"]= data.get("rollover_reg_number","").upper().replace(" ","")

    # ── Debug: show what was read ─────────────────────────────────
    print("  [DEBUG] Fields read from Excel:")
    for k, v in data.items():
        if k != "addons":
            print(f"           {k:<22} = {repr(v)}")
    print(f"           {'addons':<22} = {data['addons']}")

    return data


# ═══════════════════════════════════════════════════════════════════
#  SAVE RESULT
# ═══════════════════════════════════════════════════════════════════

def save_proposal_number(D, proposal_no):
    """Write proposal number back to col 31 of the Details sheet."""
    try:
        excel_row = D.get("_excel_row", 4)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Details"]
        ws.cell(row=excel_row, column=33, value=proposal_no)
        wb.save(EXCEL_PATH)
        print(f"  Proposal number '{proposal_no}' saved to Details row {excel_row}, col 33")
    except Exception as e:
        print(f"  Could not save proposal number to Excel: {e}")


def save_policy_number(D, policy_no):
    """Write policy number back to col 32 of the Details sheet."""
    try:
        excel_row = D.get("_excel_row", 4)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Details"]
        ws.cell(row=excel_row, column=34, value=policy_no)
        wb.save(EXCEL_PATH)
        print(f"  Policy number '{policy_no}' saved to Details row {excel_row}, col 34")
    except Exception as e:
        print(f"  Could not save policy number to Excel: {e}")


def save_result(D, status, proposal_no, steps, error=""):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Results Log"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        s  = Side(style="thin", color="BFBFBF")
        b  = Border(left=s, right=s, top=s, bottom=s)
        bg = "E2EFDA" if status == "SUCCESS" else "FCE4D6"
        r  = max(ws.max_row + 1, 3)
        for col, val in enumerate([
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            D["vehicle_reg_number"], status, proposal_no,
            " -> ".join(steps), error,
            "YES" if D.get("auto_pay") else "NO",
        ], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = b
        wb.save(EXCEL_PATH)
        print("  Result saved to details.xlsx (Results Log tab)")
    except Exception as e:
        print(f"  Could not write to Excel: {e}")


# ═══════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════

def log(kind, msg):
    icons = {"step":"\n[STEP]","ok":"  [OK] ","warn":"  [WARN]","fill":"  [FILL]","info":"  [INFO]"}
    print(f"{icons.get(kind,'      ')} {msg}")


async def pause(msg="Done? Press ENTER to continue"):
    print(f"\n  [PAUSE] {msg}")
    await asyncio.get_event_loop().run_in_executor(
        None, input, "          Press ENTER: ")


async def smart_click(page, keywords, label, timeout=12000):
    """
    Tries every possible strategy to click an element on an Angular portal.
    If nothing works: prints what IS on the page, then asks user to click manually.
    """
    kws = [keywords] if isinstance(keywords, str) else keywords

    for kw in kws:
        kw_lower = kw.lower()

        # 1. Playwright text locator — exact then partial
        for exact in [True, False]:
            try:
                loc = page.get_by_text(kw, exact=exact).first
                await loc.wait_for(state="visible", timeout=3000)
                await loc.scroll_into_view_if_needed()
                await loc.click()
                log("info", f"Clicked '{label}'")
                return True
            except Exception:
                pass

        # 2. aria-label / title / alt attributes
        for attr in [f"[aria-label*='{kw}' i]", f"[title*='{kw}' i]",
                     f"[alt*='{kw}' i]", f"[data-label*='{kw}' i]"]:
            try:
                await page.wait_for_selector(attr, timeout=2000, state="visible")
                await page.click(attr)
                log("info", f"Clicked '{label}' via attribute")
                return True
            except Exception:
                pass

        # 3. Angular / Material Design selectors
        for sel in [
            f"mat-card:has-text('{kw}')",
            f"mat-list-item:has-text('{kw}')",
            f"mat-menu-item:has-text('{kw}')",
            f"mat-tab:has-text('{kw}')",
            f"button:has-text('{kw}')",
            f"a:has-text('{kw}')",
            f"li:has-text('{kw}')",
            f"[class*='card']:has-text('{kw}')",
            f"[class*='tile']:has-text('{kw}')",
            f"[class*='item']:has-text('{kw}')",
            f"[class*='product']:has-text('{kw}')",
            f"[class*='menu']:has-text('{kw}')",
            f"[class*='nav']:has-text('{kw}')",
            f"[class*='btn']:has-text('{kw}')",
        ]:
            try:
                await page.wait_for_selector(sel, timeout=2000, state="visible")
                await page.click(sel)
                log("info", f"Clicked '{label}' via Angular selector")
                return True
            except Exception:
                pass

        # 4. JavaScript innerText scan — catches anything the above missed
        try:
            result = await page.evaluate(f"""() => {{
                const kw = '{kw_lower}';
                const els = document.querySelectorAll(
                    'a, button, mat-card, mat-list-item, li, span, ' +
                    '[role="button"], [role="tab"], [role="menuitem"], ' +
                    '[class*="card"], [class*="tile"], [class*="item"], ' +
                    '[class*="menu"], [class*="nav"], [class*="btn"]'
                );
                for (const el of els) {{
                    const txt = (el.innerText || el.textContent || '').trim().toLowerCase();
                    if (txt === kw || txt.startsWith(kw)) {{
                        const rect = el.getBoundingClientRect();
                        if (rect.width > 0 && rect.height > 0) {{
                            el.click();
                            return txt.substring(0, 40);
                        }}
                    }}
                }}
                return null;
            }}""")
            if result:
                log("info", f"Clicked '{label}' via JS scan")
                return True
        except Exception:
            pass

    # ── All strategies failed — show what's on the page and ask user ──
    log("warn", f"Could not auto-click '{label}'. Visible elements on page:")
    try:
        items = await page.evaluate("""() => {
            const els = document.querySelectorAll(
                'a, button, mat-card, mat-list-item, li, [role="button"],
                 [class*="card"], [class*="tile"], [class*="nav"]'
            );
            return [...new Set(
                Array.from(els)
                    .map(e => (e.innerText || e.textContent || '').trim())
                    .filter(t => t.length > 1 && t.length < 60)
            )].slice(0, 25);
        }""")
        for item in items:
            print(f"         . {item}")
    except Exception:
        pass

    await pause(f"Click '{label}' manually in the browser, then press ENTER")
    return False


async def fill_field(page, selectors, value, label):
    """Try each selector until one works, then fill it."""
    for sel in selectors:
        try:
            el = await page.wait_for_selector(sel, timeout=5000, state="visible")
            await el.scroll_into_view_if_needed()
            await el.click()
            # Use Ctrl+A + Delete instead of triple_click (works on ElementHandle)
            await page.keyboard.press("Control+a")
            await page.keyboard.press("Delete")
            await el.type(value, delay=60)
            log("fill", f"{label}: {value}")
            return True
        except Exception:
            continue
    log("warn", f"Could not fill '{label}' — fill it manually")
    return False


async def fill_dropdown(page, value, label):
    """
    For Angular autocomplete/dropdown fields on NYSA portal.
    Strategy:
    1. Find ANY visible input that is currently active / empty
    2. Type the value slowly
    3. Wait for dropdown list to appear
    4. Click the best matching option
    5. If no dropdown appears, try pressing Enter
    """
    try:
        # Type into the currently focused/active input on the page
        # First click somewhere neutral to reset focus
        await page.wait_for_timeout(300)

        # Find all visible inputs and try each one that looks empty
        inputs = await page.query_selector_all("input:visible, mat-select:visible")

        # Try typing into the page directly after finding the right field
        # Use JS to find input by placeholder or nearby label text
        el = await page.evaluate_handle(f"""() => {{
            const allInputs = document.querySelectorAll('input[type="text"], input:not([type])');
            for (const inp of allInputs) {{
                const rect = inp.getBoundingClientRect();
                if (rect.width === 0 || rect.height === 0) continue;
                const ph = (inp.placeholder || '').toLowerCase();
                const label = (inp.getAttribute('aria-label') || '').toLowerCase();
                const formCtrl = (inp.getAttribute('formcontrolname') || '').toLowerCase();
                const val = '{value.lower()}';
                const labelLow = '{label.lower()}';
                // Match by placeholder, aria-label, or formcontrolname
                if (ph.includes(labelLow) || label.includes(labelLow) ||
                    formCtrl.includes(labelLow) || ph.includes(val.substring(0,4))) {{
                    return inp;
                }}
            }}
            return null;
        }}""")

        if el:
            await el.click()
            await el.fill("")
            await el.type(value, delay=80)
            log("fill", f"{label}: {value}")
        else:
            # Fallback: use keyboard focus — tab to next field and type
            await page.keyboard.type(value, delay=80)
            log("fill", f"{label} (keyboard): {value}")

        # Wait for dropdown/autocomplete to appear
        await page.wait_for_timeout(1000)

        # Try clicking matching option in any dropdown that appeared
        option_clicked = False
        for option_sel in [
            f"mat-option:has-text('{value}')",
            f"li:has-text('{value}')",
            f".mat-option:has-text('{value}')",
            f"[role='option']:has-text('{value}')",
            f".dropdown-item:has-text('{value}')",
            f".autocomplete-option:has-text('{value}')",
            f"span:has-text('{value}')",
        ]:
            try:
                await page.wait_for_selector(option_sel, timeout=2000, state="visible")
                await page.click(option_sel)
                log("info", f"Selected dropdown option: {value}")
                option_clicked = True
                break
            except Exception:
                continue

        if not option_clicked:
            # Try partial match on first word
            first_word = value.split()[0] if value else value
            for option_sel in [
                f"mat-option:has-text('{first_word}')",
                f"[role='option']:has-text('{first_word}')",
                f"li:has-text('{first_word}')",
            ]:
                try:
                    await page.wait_for_selector(option_sel, timeout=1500, state="visible")
                    await page.click(option_sel)
                    log("info", f"Selected dropdown option (partial): {first_word}")
                    option_clicked = True
                    break
                except Exception:
                    continue

        if not option_clicked:
            # Press Enter/Tab to confirm whatever is shown
            await page.keyboard.press("Enter")

        await page.wait_for_timeout(500)
        return True

    except Exception as ex:
        log("warn", f"Could not fill dropdown '{label}': {ex}")
        return False


# ═══════════════════════════════════════════════════════════════════
#  CONNECT TO YOUR EXISTING CHROME
# ═══════════════════════════════════════════════════════════════════

async def connect_to_chrome(playwright):
    """
    Attach to Chrome that was started with --remote-debugging-port=9222.
    Finds the NYSA tab automatically. No new tab is opened.
    """
    try:
        browser = await playwright.chromium.connect_over_cdp(
            f"http://localhost:{CHROME_PORT}"
        )
    except Exception as e:
        print(f"""
  ERROR: Cannot connect to Chrome on port {CHROME_PORT}

  Make sure you started Chrome with this command (Win+R → paste):
  "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\\chrome-nysa"

  Then login to nysa.icicilombard.com and press ENTER here.
  Details: {e}
        """)
        input("Press ENTER to exit..."); raise SystemExit(1)

    # Find the NYSA tab — look for icicilombard.com in any open tab
    nysa_page = None
    all_pages = [p for ctx in browser.contexts for p in ctx.pages]

    for p in all_pages:
        if "icicilombard.com" in p.url:
            nysa_page = p
            break

    # If not found, use the most recently active tab
    if not nysa_page:
        if all_pages:
            nysa_page = all_pages[-1]
            log("warn", f"NYSA tab not found — using: {nysa_page.url}")
            print("  Make sure the NYSA portal is open in Chrome!")
            await pause("Navigate to nysa.icicilombard.com home page, then press ENTER")
        else:
            print("  No open tabs found in Chrome.")
            input("Press ENTER to exit..."); raise SystemExit(1)

    await nysa_page.bring_to_front()
    log("ok", f"Connected! Tab URL: {nysa_page.url}")
    return browser, nysa_page


# ═══════════════════════════════════════════════════════════════════
#  MAIN AUTOMATION FLOW
# ═══════════════════════════════════════════════════════════════════

async def run(D, skip_to_step4=False):
    steps       = []
    proposal_no = ""

    make_model = D.get("manufacturer_model","(not set)")
    print(f"""
+--------------------------------------------------------------+
|  ICICI NYSA — New Car Insurance Automation                   |
|  Make/Model : {make_model:<48}|
|  RTO        : {D.get('rto',''):<48}|
|  Mobile     : {D['mobile_number']:<48}|
|  Auto Pay   : {'YES - will click PAY' if D['auto_pay'] else 'NO  - stops before payment':<48}|
+--------------------------------------------------------------+""")

    async with async_playwright() as p:
        browser, page = await connect_to_chrome(p)

        try:
            # Let page fully settle
            try:
                await page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass
            await page.wait_for_timeout(1500)
            log("ok", f"Automation starting from: {page.url}")
            steps.append("connected")

            if not skip_to_step4:
                # ── STEP 1: Motor Insurance ────────────────────────────
                log("step", "STEP 1 — Motor Insurance")
                await smart_click(page, ["Motor", "Motor Insurance"], "Motor")
                # Wait for Motor page to fully load — networkidle exits early when ready
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                # Also wait up to 30 attempts for the RTO/Quote form to actually render
                for _mot_wait in range(30):
                    page_loaded = await page.evaluate("""() => {
                        const txt = (document.body.innerText||'').toLowerCase();
                        return txt.includes('registration number') ||
                               txt.includes('fetch vehicle details') ||
                               txt.includes('city where vehicle') ||
                               txt.includes('four wheeler') ||
                               txt.includes('two wheeler');
                    }""")
                    if page_loaded:
                        log("info", f"  Motor page ready (attempt {_mot_wait+1})")
                        break
                    await page.wait_for_timeout(400)

                # ── STEP 1b: Click left-side slider (opens vehicle category modal) ─
                log("step", "STEP 1b — Click left slider to open category modal")
                # networkidle already fired above — no extra wait needed here
                # Click slider and verify modal opened (Four Wheeler/Two Wheeler visible)
                # If not opened, click again — nysa_openbtn toggles, so may need 2 clicks
                modal_open = False
                for _sb in range(25):
                    btn_coords = await page.evaluate("""() => {
                        const el = document.querySelector('.nysa_openbtn');
                        if (!el) return null;
                        const r = el.getBoundingClientRect();
                        if (r.width > 0 && r.height > 0)
                            return {x: Math.round(r.left + r.width/2), y: Math.round(r.top + r.height/2)};
                        return null;
                    }""")
                    if btn_coords:
                        log("info", f"  nysa_openbtn at x={btn_coords['x']} y={btn_coords['y']} (attempt {_sb+1})")
                        await page.mouse.click(btn_coords['x'], btn_coords['y'])
                        await page.wait_for_timeout(1000)
                        # Verify modal opened — Four Wheeler or Two Wheeler must be visible
                        modal_open = await page.evaluate("""() => {
                            for (const el of document.querySelectorAll('*')) {
                                const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                                if ((t === 'four wheeler' || t === 'two wheeler') &&
                                    el.getBoundingClientRect().width > 0) return true;
                            }
                            return false;
                        }""")
                        if modal_open:
                            log("ok", "  -> Slider clicked and modal opened")
                            break
                        log("info", f"  Modal not open yet — retrying slider click...")
                    else:
                        await page.wait_for_timeout(500)
                if not modal_open:
                    log("warn", "  Modal did not open after all attempts")

                # ── STEP 1c: Select Motor Category + Type from modal ──────
                # motor_category: "4w" → Four Wheeler, "2w" → Two Wheeler
                # motor_type: "Package" → Package tab, "TP" → TP tab
                motor_cat  = str(D.get("motor_category", "4w")).strip().lower()
                motor_type = str(D.get("motor_type", "Package")).strip()

                # Map category to modal label text
                cat_label_map = {
                    "4w": "Four Wheeler", "4 w": "Four Wheeler",
                    "four wheeler": "Four Wheeler", "fourwheeler": "Four Wheeler",
                    "2w": "Two Wheeler",  "2 w": "Two Wheeler",
                    "two wheeler": "Two Wheeler",   "twowheeler": "Two Wheeler",
                    "electric bike": "Electric Bike", "eb": "Electric Bike",
                    "gcv": "GCV", "pcv": "PCV", "misc d": "Misc D",
                }
                cat_text = cat_label_map.get(motor_cat, "Four Wheeler")

                # Click category using Playwright locator (most reliable — finds by visible text)
                log("info", f"  Selecting category: {cat_text}")
                cat_clicked = False
                for _ct in range(5):
                    try:
                        loc = page.get_by_text(cat_text, exact=True).first
                        await loc.scroll_into_view_if_needed()
                        await loc.click(force=True)
                        cat_clicked = True
                        log("ok", f"  -> Category '{cat_text}' clicked")
                        break
                    except Exception:
                        pass
                    # Fallback: XPath on exact text node
                    try:
                        loc2 = page.locator(f"xpath=//*[normalize-space(text())='{cat_text}']").first
                        await loc2.click(force=True)
                        cat_clicked = True
                        log("ok", f"  -> Category '{cat_text}' clicked (xpath)")
                        break
                    except Exception:
                        pass
                    await page.wait_for_timeout(800)
                if not cat_clicked:
                    log("warn", f"  Category '{cat_text}' not found in modal")
                await page.wait_for_timeout(1000)

                # Click Package or TP badge nearest to the category label (by X position)
                # After clicking "Two Wheeler", find the Package/TP badge closest to its center X
                log("info", f"  Selecting type: {motor_type}")
                type_clicked = await page.evaluate("""([catText, typeText]) => {
                    // Step 1: find the category label and get its center X
                    let catCenterX = null, catCenterY = null;
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim();
                        if (t.toLowerCase() === catText.toLowerCase()) {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0 && r.width < 300) {
                                catCenterX = r.left + r.width / 2;
                                catCenterY = r.top + r.height / 2;
                                break;
                            }
                        }
                    }
                    if (catCenterX === null) return false;

                    // Step 2: find ALL elements whose exact text is typeText
                    const candidates = [];
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim();
                        if (t.toLowerCase() !== typeText.toLowerCase()) continue;
                        const r = el.getBoundingClientRect();
                        if (r.width === 0 || r.height === 0) continue;
                        const cx = r.left + r.width / 2;
                        const cy = r.top + r.height / 2;
                        // Must be below the category label (badges are below the icon+label)
                        if (cy > catCenterY - 10) {
                            candidates.push({el, cx, cy, dist: Math.abs(cx - catCenterX)});
                        }
                    }
                    if (candidates.length === 0) return false;

                    // Step 3: click the one with smallest horizontal distance to category center
                    candidates.sort((a, b) => a.dist - b.dist);
                    candidates[0].el.scrollIntoView({block: 'center'});
                    candidates[0].el.click();
                    return true;
                }""", [cat_text, motor_type])
                if type_clicked:
                    log("ok", f"  -> Type '{motor_type}' clicked (nearest to '{cat_text}')")
                else:
                    log("warn", f"  Type '{motor_type}' not found near '{cat_text}' card")

                await page.wait_for_timeout(2000)
                steps.append("motor")

                # ── STEP 1d: SAOD — Standalone OD Policy ─────────────
                # If saod == "yes" in Excel, click the button beside "Standalone OD Policy"
                if D.get("saod", "").strip().lower() == "yes":
                    log("step", "STEP 1d — Standalone OD Policy (SAOD)")
                    saod_clicked = False
                    # Wait up to 10 attempts for Standalone OD Policy to appear
                    for _saod_wait in range(20):
                        saod_visible = await page.evaluate("""() => {
                            for (const el of document.querySelectorAll('*')) {
                                const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                                if (t.includes('standalone od') || t.includes('standalone od policy') || t.includes('saod')) {
                                    const r = el.getBoundingClientRect();
                                    if (r.width > 0) return true;
                                }
                            }
                            return false;
                        }""")
                        if saod_visible:
                            break
                        await page.wait_for_timeout(500)
                    # Find the clickable element (button/toggle) beside "Standalone OD Policy" text
                    saod_coords = await page.evaluate("""() => {
                        // Find label element for Standalone OD
                        let labelEl = null;
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if ((t === 'standalone od policy' || t === 'standalone od' || t === 'saod') &&
                                el.getBoundingClientRect().width > 0) {
                                labelEl = el;
                                break;
                            }
                        }
                        if (!labelEl) return null;
                        // Look for nearest button/toggle/checkbox to the label
                        let node = labelEl.parentElement;
                        for (let i = 0; i < 6; i++) {
                            if (!node) break;
                            const btn = node.querySelector(
                                'button, input[type="checkbox"], input[type="radio"], ' +
                                '[class*="toggle"], [class*="switch"], [class*="btn"]'
                            );
                            if (btn) {
                                btn.scrollIntoView({block: 'center'});
                                const r = btn.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                            node = node.parentElement;
                        }
                        // Fallback: click the label itself
                        labelEl.scrollIntoView({block: 'center'});
                        const r = labelEl.getBoundingClientRect();
                        return {x: r.left+r.width/2, y: r.top+r.height/2};
                    }""")
                    if saod_coords:
                        await page.mouse.click(saod_coords['x'], saod_coords['y'])
                        log("ok", f"  -> Standalone OD Policy (SAOD) clicked at ({saod_coords['x']:.0f},{saod_coords['y']:.0f})")
                        saod_clicked = True
                    if not saod_clicked:
                        log("warn", "  SAOD button not found — click it manually")
                        await pause("Click the Standalone OD Policy button manually, then press ENTER")
                    await page.wait_for_timeout(1000)
                    steps.append("saod")

                # ── STEP 2: Select Policy Type (New / Rollover / Used) ───
                policy_type_val = D.get("policy_type", "New").strip().lower()
                log("step", f"STEP 2 — Policy Type: {D.get('policy_type','New')}")

                if policy_type_val == "new":
                    await smart_click(page,
                        ["New", "New Policy", "New Business", "New Proposal"],
                        "New Policy")
                elif policy_type_val == "rollover":
                    await smart_click(page,
                        ["Rollover", "Roll Over"],
                        "Rollover")
                elif policy_type_val in ("used", "used car", "used vehicle"):
                    await smart_click(page,
                        ["Used", "Used Car", "Used Vehicle"],
                        "Used")
                else:
                    # fallback to New
                    log("warn", f"  Unknown policy_type '{D.get('policy_type','')}' — defaulting to New")
                    await smart_click(page,
                        ["New", "New Policy", "New Business", "New Proposal"],
                        "New Policy")

                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)
                steps.append("new")

                # ── STEP 3: Registration Number & Fetch ───────────────
                # New     → type "NEW"
                # Rollover/Used → type the actual reg number from Excel
                log("step", "STEP 3 — Enter Registration Number + Fetch Vehicle Details")
                if policy_type_val == "new":
                    reg_to_enter = "NEW"
                else:
                    reg_to_enter = D.get("rollover_reg_number", "").strip()
                    if not reg_to_enter:
                        # Fallback: use vehicle_reg_number column (header "Registration Number")
                        reg_to_enter = D.get("vehicle_reg_number", "").strip()
                    if not reg_to_enter:
                        log("warn", "  rollover_reg_number is empty in Excel — please fill it manually")
                        await pause("Enter Registration Number in browser manually, then press ENTER")
                        reg_to_enter = ""

                if reg_to_enter:
                    await fill_field(page, [
                        "input[placeholder*='Registration No' i]",
                        "input[placeholder*='Enter Registration' i]",
                        "input[formcontrolname*='registrationNumber' i]",
                        "input[formcontrolname*='regNo' i]",
                    ], reg_to_enter, "Registration Number")
                    await page.wait_for_timeout(500)

                await smart_click(page,
                    ["Fetch Vehicle Details", "Fetch Details"],
                    "Fetch Vehicle Details")
                await page.wait_for_timeout(2500)
                steps.append("fetch_details")

            else:
                # Skipping STEP 1-3: already on details page after Another Policy
                log("step", "Skipping STEP 1-3 (resuming from details page)")
                steps.append("motor"); steps.append("new"); steps.append("fetch_details")
                try:
                    await page.wait_for_load_state("networkidle", timeout=10000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)


            # ══════════════════════════════════════════════════════
            # STEP 4 — Fill form fields from Excel
            # From video the form shows:
            #   - City where vehicle is registered (RTO) [autocomplete]
            #   - Manufacturer - Model [single combined autocomplete field]
            #   - Insured Depreciation Value (IDV) [text + slider]
            #   - Current Registration Type [radio buttons]
            # ══════════════════════════════════════════════════════
            log("step", "STEP 4 — Filling Details from details.xlsx")

            async def type_and_pick(placeholder_hint, value, label):
                """
                Smart autocomplete filler:
                - Extracts the SEARCH TERM (text before the first '-')
                - Types only that into the field to trigger the dropdown
                - Then clicks the option that EXACTLY matches the full value from Excel

                Example:
                  Excel value  : KARNATAKA-BANGALORE
                  Types        : KARNATAKA   (before the hyphen)
                  Dropdown shows: KARNATAKA-BANGALORE, KARNATAKA-BANGALORE B ...
                  Clicks       : KARNATAKA-BANGALORE  (exact match)

                  Excel value  : MARUTI - SWIFT (1197 CC)
                  Types        : MARUTI          (before the hyphen)
                  Dropdown shows: MARUTI - 800 AC ..., MARUTI - SWIFT ...
                  Clicks       : MARUTI - SWIFT (1197 CC)  (exact match)
                """
                log("info", f"Filling [{label}]: {value}")

                # ── Extract search term (part before first hyphen) ────
                # This is what we type to open the dropdown
                search_term = value.split("-")[0].strip()
                log("info", f"  Search term: '{search_term}' (from '{value}')")

                # ── Find the input field ──────────────────────────────
                el = None
                for sel in [
                    f"input[placeholder*='{placeholder_hint}' i]",
                    f"input[placeholder*='{label}' i]",
                ]:
                    try:
                        found = await page.query_selector(sel)
                        if found:
                            el = found
                            break
                    except Exception:
                        pass

                # Fallback: find by nearby label text using JS
                if not el:
                    try:
                        el = await page.evaluate_handle("""(hint) => {
                            const inputs = document.querySelectorAll('input');
                            for (const inp of inputs) {
                                const r = inp.getBoundingClientRect();
                                if (r.width === 0 || r.height === 0) continue;
                                const container = inp.closest(
                                    'mat-form-field, .form-field, div'
                                );
                                if (container) {
                                    const lbl = container.querySelector(
                                        'label, mat-label, .label, span'
                                    );
                                    if (lbl && lbl.innerText.toLowerCase()
                                            .includes(hint.toLowerCase()))
                                        return inp;
                                }
                            }
                            return null;
                        }""", placeholder_hint)
                    except Exception:
                        pass

                if not el:
                    log("warn", f"  Input not found for '{label}' — please fill manually")
                    return False

                # ── Clear field and type ONLY the search term ─────────
                try:
                    await el.scroll_into_view_if_needed()
                    await el.click()
                    await page.wait_for_timeout(300)
                    await el.click()
                    await page.keyboard.press("Control+a")
                    await page.keyboard.press("Delete")
                    await page.wait_for_timeout(200)

                    # Type the search term slowly — triggers autocomplete
                    await el.type(search_term, delay=150)
                    log("fill", f"  Typed search term: '{search_term}'")

                except Exception as ex:
                    log("warn", f"  Could not type into '{label}': {ex}")
                    return False

                # ── Wait for dropdown list to appear ──────────────────
                await page.wait_for_timeout(1500)

                # ── Click the exact matching option ───────────────────
                # Use JS so we can do case-insensitive comparison
                # The full Excel value is what we match against
                full_value_upper = value.strip().upper()

                try:
                    result = await page.evaluate("""(full_val) => {
                        const v = full_val.trim().toUpperCase();

                        const opts = document.querySelectorAll(
                            'mat-option, [role="option"], li.mat-option, ' +
                            '.mat-autocomplete-panel mat-option, ' +
                            '.cdk-overlay-container mat-option'
                        );

                        if (opts.length === 0) return 'NO_OPTIONS';

                        // Pass 1: exact match (case-insensitive)
                        for (const o of opts) {
                            const t = (o.innerText || o.textContent || '').trim().toUpperCase();
                            if (t === v) {
                                o.click();
                                return 'EXACT||' + o.innerText.trim();
                            }
                        }

                        // Pass 2: option text starts with full value
                        for (const o of opts) {
                            const t = (o.innerText || o.textContent || '').trim().toUpperCase();
                            if (t.startsWith(v)) {
                                o.click();
                                return 'STARTS||' + o.innerText.trim();
                            }
                        }

                        // Pass 3: full value contains option text
                        for (const o of opts) {
                            const t = (o.innerText || o.textContent || '').trim().toUpperCase();
                            if (t.length > 3 && v.includes(t)) {
                                o.click();
                                return 'INCLUDED||' + o.innerText.trim();
                            }
                        }

                        // Pass 4: option text contains full value
                        for (const o of opts) {
                            const t = (o.innerText || o.textContent || '').trim().toUpperCase();
                            if (t.includes(v)) {
                                o.click();
                                return 'WITHIN||' + o.innerText.trim();
                            }
                        }

                        // Show all available options in console for debugging
                        const available = Array.from(opts).map(
                            o => (o.innerText || '').trim()
                        ).join(' | ');
                        return 'NO_MATCH||Available: ' + available;

                    }""", full_value_upper)

                    if result and "||" in result:
                        match_type, matched_text = result.split("||", 1)
                        if match_type == "NO_MATCH":
                            log("warn", f"  No match for '{value}'")
                            log("warn", f"  {matched_text[:120]}")
                            log("warn", f"  Update your Excel to match one of the above options exactly")
                            return False
                        elif match_type == "NO_OPTIONS":
                            log("warn", f"  Dropdown did not appear — portal may need more time")
                            return False
                        else:
                            log("ok", f"  Selected [{match_type}]: {matched_text}")
                            await page.wait_for_timeout(700)
                            return True
                    else:
                        log("warn", f"  Unexpected dropdown result: {result}")
                        return False

                except Exception as ex:
                    log("warn", f"  Dropdown error for '{label}': {ex}")
                    return False

            # ── 4a. RTO ─────────────────────────────────────────────
            # Types state name (before -), waits for dropdown, clicks exact match
            # Retry up to 5x — on slow networks dropdown may not load in time
            if D.get("rto"):
                for _rto_try in range(5):
                    rto_ok = await type_and_pick("RTO city", D["rto"], "RTO")
                    if rto_ok:
                        break
                    log("info", f"  RTO not matched (attempt {_rto_try+1}/5), retrying...")
                    await page.wait_for_timeout(1000)
                await page.wait_for_timeout(3000)

            # ── 4b. Manufacturer - Model ─────────────────────────────
            # Single combined field — types make name, clicks exact match
            if D.get("manufacturer_model"):
                await type_and_pick("Manufacturer", D["manufacturer_model"], "Manufacturer - Model")
                # Wait for form to reload after model selection (IDV slider appears after this)
                await page.wait_for_timeout(5000)
                # Scroll back to top so IDV field is in viewport with correct coordinates
                await page.evaluate("window.scrollTo(0, 0)")
                await page.wait_for_timeout(500)

            # ── 4d. Current Registration Type ────────────────────────
            # Radio buttons: Individual | Corporate | Partnership
            if D.get("current_reg_type"):
                reg_type = D["current_reg_type"].strip()
                log("info", f"Setting Reg Type: {reg_type}")
                try:
                    clicked = await page.evaluate("""(val) => {
                        const els = document.querySelectorAll(
                            'label, mat-radio-button, .mat-radio-label, ' +
                            'mat-radio-button .mat-radio-label-content'
                        );
                        const v = val.trim().toLowerCase();
                        for (const el of els) {
                            if ((el.innerText || '').trim().toLowerCase() === v) {
                                el.click(); return true;
                            }
                        }
                        return false;
                    }""", reg_type)
                    if clicked:
                        log("ok", f"  → Reg Type: {reg_type}")
                    else:
                        log("warn", f"Could not select reg type '{reg_type}'")
                except Exception as e:
                    log("warn", f"Reg Type error: {e}")
                await page.wait_for_timeout(2000)

            # ── 4e. Tenure ───────────────────────────────────────────
            # Radio buttons: 1 year (1yr OD + 3yrs TP) | 3 Years (3yrs OD + 3yrs TP)
            # Excel values: "1 year" or "3 Years"
            if D.get("tenure"):
                tenure_val = D["tenure"].strip()
                log("info", f"Setting Tenure: {tenure_val}")
                try:
                    clicked = await page.evaluate("""(val) => {
                        const v = val.trim().toLowerCase();
                        const els = document.querySelectorAll(
                            'label, mat-radio-button, .mat-radio-label, ' +
                            'mat-radio-button .mat-radio-label-content'
                        );
                        for (const el of els) {
                            const txt = (el.innerText || '').trim().toLowerCase();
                            // Match "1 year" or "3 years" at start of label text
                            if (txt === v || txt.startsWith(v)) {
                                el.click(); return true;
                            }
                        }
                        return false;
                    }""", tenure_val)
                    if clicked:
                        log("ok", f"  → Tenure: {tenure_val}")
                    else:
                        log("warn", f"Could not select tenure '{tenure_val}' — 1 year is default")
                except Exception as e:
                    log("warn", f"Tenure error: {e}")
                await page.wait_for_timeout(2000)

            # ── 4f. Year of Manufacture ──────────────────────────────
            # From video: dropdown (select element), shows years like 2026, 2025...
            if D.get("manufacturing_year"):
                mfg_year = str(D["manufacturing_year"]).split(".")[0].strip()
                log("info", f"Setting Year of Manufacture: {mfg_year}")
                # Try native select element first
                selected = False
                for sel in [
                    "select[formcontrolname*='year' i]",
                    "select[formcontrolname*='manufacture' i]",
                    "select[formcontrolname*='mfg' i]",
                    "mat-select[formcontrolname*='year' i]",
                    "mat-select[formcontrolname*='manufacture' i]",
                ]:
                    try:
                        el_yr = await page.query_selector(sel)
                        if el_yr:
                            tag = await el_yr.evaluate("el => el.tagName.toLowerCase()")
                            if tag == "select":
                                await el_yr.select_option(value=mfg_year)
                                log("ok", f"  → Year: {mfg_year} (select)")
                                selected = True
                                break
                            else:
                                # mat-select — click to open, then pick option
                                await el_yr.click()
                                await page.wait_for_timeout(800)
                                for opt_sel in [
                                    f"mat-option:has-text('{mfg_year}')",
                                    f"[role='option']:has-text('{mfg_year}')",
                                ]:
                                    try:
                                        await page.wait_for_selector(opt_sel, timeout=2000)
                                        await page.click(opt_sel)
                                        log("ok", f"  → Year: {mfg_year} (mat-select)")
                                        selected = True
                                        break
                                    except Exception:
                                        continue
                                if selected:
                                    break
                    except Exception:
                        continue

                if not selected:
                    # JS fallback — try both select and mat-select
                    try:
                        done = await page.evaluate("""(yr) => {
                            // Try native select
                            const selects = document.querySelectorAll('select');
                            for (const s of selects) {
                                const opts = Array.from(s.options).map(o => o.value || o.text);
                                if (opts.includes(yr)) {
                                    s.value = yr;
                                    s.dispatchEvent(new Event('change', {bubbles:true}));
                                    return 'select:' + yr;
                                }
                            }
                            // Try clicking mat-select then option
                            const matSels = document.querySelectorAll('mat-select');
                            for (const ms of matSels) {
                                const txt = (ms.innerText || '').trim();
                                if (txt === yr || ms.getAttribute('formcontrolname')
                                        ?.toLowerCase().includes('year')) {
                                    ms.click();
                                    return 'mat-select-opened';
                                }
                            }
                            return null;
                        }""", mfg_year)
                        if done:
                            log("ok", f"  → Year (JS): {done}")
                            await page.wait_for_timeout(800)
                            # Click the option if mat-select was opened
                            if done == "mat-select-opened":
                                for opt_sel in [
                                    f"mat-option:has-text('{mfg_year}')",
                                    f"[role='option']:has-text('{mfg_year}')",
                                ]:
                                    try:
                                        await page.wait_for_selector(opt_sel, timeout=2000)
                                        await page.click(opt_sel)
                                        log("ok", f"  → Year option clicked: {mfg_year}")
                                        break
                                    except Exception:
                                        continue
                    except Exception as e:
                        log("warn", f"Year of Manufacture error: {e}")

                await page.wait_for_timeout(2000)

            # ── 4c. IDV (Insured Depreciation Value) ─────────────────
            # IDV text input: id='sirangeinput', immediately BEFORE range slider.
            # KEY FIX: after scrollIntoView, ALWAYS re-read fresh coordinates before clicking.
            # Old coords are stale if the page scrolled between evaluate() and mouse.click().
            if D.get("idv"):
                idv_val = str(D["idv"]).strip().split(".")[0]
                log("info", f"Setting IDV: {idv_val}")

                # Step 1: wait for IDV input to appear (up to 20 x 500ms)
                idv_found = False
                for _attempt in range(20):
                    idv_found = await page.evaluate("""() => {
                        let inp = document.getElementById('sirangeinput');
                        if (!inp) inp = document.querySelector('input[name="amountInput"]');
                        if (!inp) {
                            const slider = document.querySelector(
                                'input[type="range"], input.custom-range');
                            if (!slider) return false;
                            const all = Array.from(document.querySelectorAll('input'));
                            const idx = all.indexOf(slider);
                            for (let i = idx - 1; i >= 0; i--) {
                                const c = all[i];
                                if (['range','checkbox','radio','hidden','submit','button']
                                    .includes(c.type)) continue;
                                if (c.getBoundingClientRect().width === 0) continue;
                                inp = c; break;
                            }
                        }
                        if (!inp) return false;
                        inp.scrollIntoView({block:'center'});
                        return true;
                    }""")
                    if idv_found:
                        log("fill", f"IDV input found (attempt {_attempt+1})")
                        break
                    await page.wait_for_timeout(500)

                # ── IDV field found but default value not yet populated ──
                # Wait up to 10s for the default numeric value to appear before writing Excel value
                if idv_found:
                    idv_default_ok = False
                    for _dv in range(20):
                        idv_current = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) {
                                const slider = document.querySelector(
                                    'input[type="range"], input.custom-range');
                                if (!slider) return '';
                                const all = Array.from(document.querySelectorAll('input'));
                                const idx = all.indexOf(slider);
                                for (let i = idx - 1; i >= 0; i--) {
                                    const c = all[i];
                                    if (['range','checkbox','radio','hidden','submit','button']
                                        .includes(c.type)) continue;
                                    if (c.getBoundingClientRect().width === 0) continue;
                                    return c.value;
                                }
                            }
                            return inp ? inp.value : '';
                        }""") or ""
                        if idv_current.strip() and idv_current.strip() not in ("0", ""):
                            log("fill", f"IDV default populated: {idv_current.strip()}")
                            idv_default_ok = True
                            break
                        await page.wait_for_timeout(500)
                    if not idv_default_ok:
                        log("warn", "IDV default never populated — will retry model entry")
                        idv_found = False  # force model re-type below

                # ── IDV not auto-populated: clear model and re-type to trigger IDV ──
                if not idv_found and D.get("manufacturer_model"):
                    log("warn", "IDV not populated — clearing model and re-typing to trigger IDV...")
                    # Save current RTO value before clearing model (model re-type resets RTO)
                    rto_saved = await page.evaluate("""() => {
                        const inputs = document.querySelectorAll('input');
                        for (const inp of inputs) {
                            const ph = (inp.placeholder||'').toLowerCase();
                            if (ph.includes('rto') || ph.includes('city where') ||
                                ph.includes('registered')) {
                                return inp.value || '';
                            }
                        }
                        // Also check by value pattern (e.g. "KARNATAKA-BANGALORE")
                        for (const inp of inputs) {
                            const v = (inp.value||'');
                            if (v.includes('-') && v.length > 5 && /^[A-Z]/.test(v))
                                return v;
                        }
                        return '';
                    }""") or D.get("rto", "")
                    log("info", f"  Saved RTO before model retry: '{rto_saved}'")

                    # Clear the Manufacturer-Model field and retype
                    cleared = await page.evaluate("""() => {
                        const inputs = document.querySelectorAll('input');
                        for (const inp of inputs) {
                            const ph = (inp.placeholder||'').toLowerCase();
                            const val = (inp.value||'').toLowerCase();
                            if (ph.includes('manufacturer') || ph.includes('model') ||
                                val.includes('pvt') || val.includes('ltd') || val.length > 10) {
                                inp.focus();
                                inp.value = '';
                                inp.dispatchEvent(new Event('input', {bubbles:true}));
                                return true;
                            }
                        }
                        return false;
                    }""")
                    if cleared:
                        await page.wait_for_timeout(500)
                        await type_and_pick("Manufacturer", D["manufacturer_model"], "Manufacturer - Model (retry)")
                        await page.wait_for_timeout(5000)
                        await page.evaluate("window.scrollTo(0, 0)")
                        await page.wait_for_timeout(500)

                        # Re-fill RTO after model retry (model re-type resets it)
                        rto_to_fill = rto_saved or D.get("rto", "")
                        if rto_to_fill:
                            log("info", f"  Re-filling RTO after model retry: '{rto_to_fill}'")
                            await type_and_pick("RTO city", rto_to_fill, "RTO (after model retry)")
                            await page.wait_for_timeout(2000)

                        # Retry IDV detection after re-typing model
                        for _attempt2 in range(20):
                            idv_found = await page.evaluate("""() => {
                                let inp = document.getElementById('sirangeinput');
                                if (!inp) inp = document.querySelector('input[name="amountInput"]');
                                if (!inp) {
                                    const slider = document.querySelector(
                                        'input[type="range"], input.custom-range');
                                    if (!slider) return false;
                                    const all = Array.from(document.querySelectorAll('input'));
                                    const idx = all.indexOf(slider);
                                    for (let i = idx - 1; i >= 0; i--) {
                                        const c = all[i];
                                        if (['range','checkbox','radio','hidden','submit','button']
                                            .includes(c.type)) continue;
                                        if (c.getBoundingClientRect().width === 0) continue;
                                        inp = c; break;
                                    }
                                }
                                if (!inp) return false;
                                inp.scrollIntoView({block:'center'});
                                return true;
                            }""")
                            if idv_found:
                                log("fill", f"IDV input found after model retry (attempt {_attempt2+1})")
                                break
                            await page.wait_for_timeout(500)

                if not idv_found:
                    log("warn", "IDV input not found — slider never appeared")
                else:
                    # Step 2: wait for scroll to settle, then re-read FRESH coords each attempt
                    await page.wait_for_timeout(400)

                    for _idv_try in range(3):
                        # Re-read coordinates fresh — never use stale coords from before scroll
                        idv_coords = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) {
                                const slider = document.querySelector(
                                    'input[type="range"], input.custom-range');
                                if (!slider) return null;
                                const all = Array.from(document.querySelectorAll('input'));
                                const idx = all.indexOf(slider);
                                for (let i = idx - 1; i >= 0; i--) {
                                    const c = all[i];
                                    if (['range','checkbox','radio','hidden','submit','button']
                                        .includes(c.type)) continue;
                                    if (c.getBoundingClientRect().width === 0) continue;
                                    inp = c; break;
                                }
                            }
                            if (!inp) return null;
                            inp.scrollIntoView({block:'center'});
                            const r = inp.getBoundingClientRect();
                            return {x: r.left + r.width/2, y: r.top + r.height/2};
                        }""")
                        if not idv_coords:
                            break
                        await page.wait_for_timeout(200)

                        # Click with fresh coords, select all, type value
                        await page.mouse.click(idv_coords['x'], idv_coords['y'],
                                               click_count=3)
                        await page.wait_for_timeout(150)
                        await page.keyboard.press("Delete")
                        await page.wait_for_timeout(80)
                        await page.keyboard.press("Backspace")
                        await page.wait_for_timeout(80)
                        await page.keyboard.type(idv_val, delay=120)
                        await page.wait_for_timeout(400)

                        # JS native setter to lock Angular's model binding
                        # NOTE: do NOT dispatch 'blur' — it triggers Angular to recalculate and reset IDV
                        await page.evaluate("""(val) => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) {
                                const slider = document.querySelector(
                                    'input[type="range"], input.custom-range');
                                if (!slider) return;
                                const all = Array.from(document.querySelectorAll('input'));
                                const idx = all.indexOf(slider);
                                for (let i = idx - 1; i >= 0; i--) {
                                    const c = all[i];
                                    if (['range','checkbox','radio','hidden','submit','button']
                                        .includes(c.type)) continue;
                                    if (c.getBoundingClientRect().width === 0) continue;
                                    inp = c; break;
                                }
                            }
                            if (!inp) return;
                            const setter = Object.getOwnPropertyDescriptor(
                                window.HTMLInputElement.prototype, 'value').set;
                            setter.call(inp, '');
                            setter.call(inp, val);
                            inp.dispatchEvent(new Event('input',  {bubbles: true}));
                            inp.dispatchEvent(new Event('change', {bubbles: true}));
                            // keep focus on inp — do NOT blur here
                            inp.focus();
                        }""", idv_val)
                        await page.wait_for_timeout(300)

                        # Verify — re-read value to confirm Angular didn't reset it
                        actual = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) {
                                const slider = document.querySelector(
                                    'input[type="range"], input.custom-range');
                                if (!slider) return 'no-slider';
                                const all = Array.from(document.querySelectorAll('input'));
                                const idx = all.indexOf(slider);
                                for (let i = idx - 1; i >= 0; i--) {
                                    const c = all[i];
                                    if (['range','checkbox','radio','hidden','submit','button']
                                        .includes(c.type)) continue;
                                    if (c.getBoundingClientRect().width === 0) continue;
                                    return c.value;
                                }
                                return 'not-found';
                            }
                            return inp.value;
                        }""")
                        log("fill", f"IDV try {_idv_try+1}: set={idv_val} actual={actual}")
                        if str(actual).strip() == idv_val:
                            log("ok", f"IDV confirmed: {actual}")
                            break
                        await page.wait_for_timeout(300)

                    # Click the IDV label text (non-interactive) to commit without triggering blur/reset
                    await page.evaluate("""() => {
                        // Click the IDV label element — it's a <label> or <span> with IDV text
                        // This moves mouse focus away without triggering Angular recalculation
                        for (const el of document.querySelectorAll('label, span, p, div')) {
                            const t = (el.innerText||'').trim().toLowerCase();
                            if (t.includes('insured depreciation') || t === 'idv') {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0 && el.querySelectorAll('input').length === 0) {
                                    el.click(); return;
                                }
                            }
                        }
                    }""")
                    await page.wait_for_timeout(500)
                    # Final verify
                    final_idv = await page.evaluate("""() => {
                        let inp = document.getElementById('sirangeinput');
                        if (!inp) inp = document.querySelector('input[name="amountInput"]');
                        if (!inp) {
                            const slider = document.querySelector('input[type="range"], input.custom-range');
                            if (!slider) return '';
                            const all = Array.from(document.querySelectorAll('input'));
                            const idx = all.indexOf(slider);
                            for (let i = idx-1; i >= 0; i--) {
                                const c = all[i];
                                if (['range','checkbox','radio','hidden','submit','button'].includes(c.type)) continue;
                                if (c.getBoundingClientRect().width === 0) continue;
                                return c.value;
                            }
                        }
                        return inp ? inp.value : '';
                    }""") or ""
                    log("ok" if str(final_idv).strip() == idv_val else "warn",
                        f"IDV final check: expected={idv_val} actual={final_idv}")
                    await page.wait_for_timeout(500)
                await page.wait_for_timeout(1000)


            # ── NOTE: Mobile and Email NOT filled here ───────────────
            # From video: no mobile/email fields visible on this form
            # They appear later in the proposal/KYC section

            # ── 4g. Customer State ───────────────────────────────────────
            try:
                rto_on_page = await page.evaluate("""() => {
                    for (const inp of document.querySelectorAll('input'))
                        if (/^[A-Z]{2,}-[A-Z]/i.test(inp.value||'')) return inp.value;
                    return null;
                }""")
                state_name = (rto_on_page or D.get('rto','')).split('-')[0].strip().title()
                if state_name:
                    log('info', f"Customer State: '{state_name}'")
                    state_set = await page.evaluate("""(sn) => {
                        const lower = sn.toLowerCase();
                        for (const sel of document.querySelectorAll('select')) {
                            let n = sel.parentElement;
                            for (let i=0;i<5;i++) {
                                if (!n) break;
                                if ((n.innerText||'').toLowerCase().includes('customer state')) {
                                    for (const o of sel.options) {
                                        if (o.text.toLowerCase().includes(lower)||lower.includes(o.text.toLowerCase().trim())){
                                            sel.value=o.value; sel.dispatchEvent(new Event('change',{bubbles:true})); return o.text;
                                        }
                                    }
                                }
                                n = n.parentElement;
                            }
                        }
                        return null;
                    }""", state_name)
                    if state_set: log('ok', f'  -> Customer State: {state_set}')
                    else: log('warn', f"  Customer State '{state_name}' not matched")
            except Exception as e:
                log('warn', f'  Customer State error: {e}')
            await page.wait_for_timeout(1000)

            steps.append("fill_details")
            await page.wait_for_timeout(1000)

            # ── IDV Guard: re-apply Excel IDV if it reverted to default ──
            # 25 retries + double-confirm: after a match, waits 800ms and re-reads
            # to catch Angular resetting it after the guard passes.
            if D.get("idv"):
                idv_val = str(D["idv"]).strip().split(".")[0]
                for _guard in range(25):
                    current_idv = await page.evaluate("""() => {
                        let inp = document.getElementById('sirangeinput');
                        if (!inp) inp = document.querySelector('input[name="amountInput"]');
                        if (!inp) {
                            const slider = document.querySelector('input[type="range"], input.custom-range');
                            if (!slider) return '';
                            const all = Array.from(document.querySelectorAll('input'));
                            const idx = all.indexOf(slider);
                            for (let i = idx-1; i >= 0; i--) {
                                const c = all[i];
                                if (['range','checkbox','radio','hidden','submit','button'].includes(c.type)) continue;
                                if (c.getBoundingClientRect().width === 0) continue;
                                return c.value;
                            }
                        }
                        return inp ? inp.value : '';
                    }""") or ""
                    if str(current_idv).strip() == idv_val:
                        # Double-confirm: wait 800ms and re-read to catch delayed Angular reset
                        await page.wait_for_timeout(800)
                        recheck_idv = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) {
                                const slider = document.querySelector('input[type="range"], input.custom-range');
                                if (!slider) return '';
                                const all = Array.from(document.querySelectorAll('input'));
                                const idx = all.indexOf(slider);
                                for (let i = idx-1; i >= 0; i--) {
                                    const c = all[i];
                                    if (['range','checkbox','radio','hidden','submit','button'].includes(c.type)) continue;
                                    if (c.getBoundingClientRect().width === 0) continue;
                                    return c.value;
                                }
                            }
                            return inp ? inp.value : '';
                        }""") or ""
                        if str(recheck_idv).strip() == idv_val:
                            log("ok", f"IDV guard: confirmed stable ({recheck_idv})")
                            break
                        else:
                            log("warn", f"IDV guard: reset after confirm ({recheck_idv}) — re-applying {idv_val}...")
                            current_idv = recheck_idv  # fall through to re-apply below
                    log("warn", f"IDV reverted to {current_idv} — re-applying {idv_val}...")
                    # Re-read coords and retype
                    idv_gc = await page.evaluate("""() => {
                        let inp = document.getElementById('sirangeinput');
                        if (!inp) inp = document.querySelector('input[name="amountInput"]');
                        if (!inp) {
                            const slider = document.querySelector('input[type="range"], input.custom-range');
                            if (!slider) return null;
                            const all = Array.from(document.querySelectorAll('input'));
                            const idx = all.indexOf(slider);
                            for (let i = idx-1; i >= 0; i--) {
                                const c = all[i];
                                if (['range','checkbox','radio','hidden','submit','button'].includes(c.type)) continue;
                                if (c.getBoundingClientRect().width === 0) continue;
                                inp = c; break;
                            }
                        }
                        if (!inp) return null;
                        inp.scrollIntoView({block:'center'});
                        const r = inp.getBoundingClientRect();
                        return {x: r.left+r.width/2, y: r.top+r.height/2};
                    }""")
                    if idv_gc:
                        await page.wait_for_timeout(200)
                        await page.mouse.click(idv_gc['x'], idv_gc['y'], click_count=3)
                        await page.wait_for_timeout(150)
                        await page.keyboard.press("Delete")
                        await page.wait_for_timeout(80)
                        await page.keyboard.press("Backspace")
                        await page.wait_for_timeout(80)
                        await page.keyboard.type(idv_val, delay=120)
                        await page.wait_for_timeout(400)
                        await page.evaluate("""(val) => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) {
                                const slider = document.querySelector('input[type="range"], input.custom-range');
                                if (!slider) return;
                                const all = Array.from(document.querySelectorAll('input'));
                                const idx = all.indexOf(slider);
                                for (let i = idx-1; i >= 0; i--) {
                                    const c = all[i];
                                    if (['range','checkbox','radio','hidden','submit','button'].includes(c.type)) continue;
                                    if (c.getBoundingClientRect().width === 0) continue;
                                    inp = c; break;
                                }
                            }
                            if (!inp) return;
                            const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
                            setter.call(inp,''); setter.call(inp,val);
                            inp.dispatchEvent(new Event('input',{bubbles:true}));
                            inp.dispatchEvent(new Event('change',{bubbles:true}));
                            inp.focus();
                        }""", idv_val)
                        await page.wait_for_timeout(500)

            # ── STEP 5: Click GET QUOTE ──────────────────────────────
            # ── STEP 5: Click GET QUOTE ──────────────────────────
            log("step", "STEP 5 — GET QUOTE")
            # Wait for page to fully settle after all field fills
            log("info", "Waiting for page to settle before GET QUOTE...")
            try:
                await page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(1000)

            async def click_by_text(texts):
                """Click any visible element matching text — works for any tag."""
                if isinstance(texts, str): texts = [texts]
                for text in texts:
                    for tag in ["button", "input", "a", "span", "div"]:
                        try:
                            loc = page.locator(f"{tag}:has-text('{text}')").first
                            await loc.wait_for(state="visible", timeout=2000)
                            await loc.scroll_into_view_if_needed()
                            await loc.click()
                            log("ok", f"  -> Clicked '{text}' via <{tag}>")
                            return True
                        except Exception:
                            continue
                    try:
                        clicked = await page.evaluate(f"""() => {{
                            for (const el of document.querySelectorAll('*')) {{
                                const rect = el.getBoundingClientRect();
                                if (rect.width === 0 || rect.height === 0) continue;
                                const txt = (el.innerText || el.value || '').trim();
                                if (txt === '{text}') {{ el.click(); return true; }}
                            }}
                            return false;
                        }}""")
                        if clicked:
                            log("ok", f"  -> Clicked '{text}' via JS scan")
                            return True
                    except Exception:
                        pass
                return False

            if not await click_by_text(["GET QUOTE", "Get Quote", "Get Quotation"]):
                log("warn", "GET QUOTE not found — check browser")
                await pause("Click 'GET QUOTE' manually, then press ENTER")
            try:
                await page.wait_for_load_state("networkidle", timeout=30000)
            except Exception:
                pass
            await page.wait_for_timeout(5000)
            steps.append("get_quote")

            # ── STEP 6: Customize Coverages ──────────────────────────
            log("step", "STEP 6 — Open Customize Coverages")
            # Wait for page to fully load (networkidle exits early when ready)
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            # Retry up to 25x (500ms each) — breaks immediately when link appears
            for _cust in range(25):
                if await page.evaluate("""() => { for (const el of document.querySelectorAll('a')) if ((el.parentElement?.innerText||'').toLowerCase().includes('customize')) return true; return false; }"""):
                    log("info", f"  Customize link found (attempt {_cust+1})")
                    break
                await page.wait_for_timeout(500)
            await page.evaluate("""() => { for (const el of document.querySelectorAll('a')) { if ((el.parentElement?.innerText||'').toLowerCase().includes('customize')) { el.scrollIntoView({block:'center'}); el.click(); return; } } }""")
            log("ok", "  -> Customize Coverages clicked")
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            for _ in range(40):
                if await page.evaluate("() => document.body.innerText.includes('Zero Depreciation')"):
                    log("ok", "  -> Add-ons page loaded"); break
                await page.wait_for_timeout(500)
            else:
                log("warn", "  Add-ons page timed out — continuing")
            await page.wait_for_timeout(1000)
            steps.append("customize_opened")

            # ── STEP 7: Tick Zero Depreciation Cover + Select ZD ─────
            log("step", "STEP 7 — Zero Depreciation Cover + ZD selection")

            # ── 7a: Click Zero Depreciation Cover checkbox ────────────
            # Strategy: find the element with text "Zero Depreciation Cover"
            # then find the nearest checkbox to it in the DOM
            log("info", "Ticking Zero Depreciation Cover")
            zd_clicked = False

            # Same approach as "Click here to customize the coverages":
            # Find the element by text, get its exact screen coordinates,
            # use page.mouse.click(x, y) — real physical click Angular responds to.
            # We target the CHECKBOX input near "Zero Depreciation Cover" text,
            # not the label, so Angular's change detection fires correctly.
            zd_info = await page.evaluate("""() => {
                // Find every checkbox — get the one whose nearby text = "Zero Depreciation Cover"
                const checkboxes = document.querySelectorAll('input[type="checkbox"]');
                for (const cb of checkboxes) {
                    let node = cb.parentElement;
                    for (let i = 0; i < 6; i++) {
                        if (!node) break;
                        const txt = (node.innerText || node.textContent || '').toLowerCase();
                        if (txt.includes('zero depreciation cover')) {
                            const r = cb.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) break;
                            return { x: r.left + r.width/2, y: r.top + r.height/2, tag: 'INPUT-checkbox' };
                        }
                        node = node.parentElement;
                    }
                }
                // Fallback: get coordinates of the text label itself
                let best = null, bestArea = Infinity;
                for (const el of document.querySelectorAll('*')) {
                    const rect = el.getBoundingClientRect();
                    if (rect.width === 0 || rect.height === 0) continue;
                    const txt = (el.innerText || el.textContent || '').trim();
                    if (txt.toLowerCase().includes('zero depreciation cover') && txt.length < 60) {
                        const area = rect.width * rect.height;
                        if (area < bestArea) { bestArea = area; best = el; }
                    }
                }
                if (best) {
                    const r = best.getBoundingClientRect();
                    return { x: r.left + r.width/2, y: r.top + r.height/2, tag: best.tagName + '-label' };
                }
                return null;
            }""")

            if zd_info:
                await page.mouse.click(zd_info['x'], zd_info['y'])
                log("ok", f"  -> Zero Depreciation clicked at ({zd_info['x']:.0f},{zd_info['y']:.0f}) via <{zd_info['tag']}>")
                zd_clicked = True
            else:
                log("warn", "Zero Depreciation Cover not found on page")
                await pause("Tick 'Zero Depreciation Cover' manually, then press ENTER")

            await page.wait_for_timeout(1500)

            # ── 7b: Select ZD (4w) or Silver (2w) from dropdown ─────
            # Wait for ZD dropdown to appear (networkidle + 25 attempts)
            try:
                await page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass
            motor_cat = str(D.get("motor_category", "2w")).strip().lower()
            zd_option = "Silver" if motor_cat in ("2w", "2 w", "two wheeler", "twowheeler") else "ZD"
            log("info", f"Selecting '{zd_option}' from ZD dropdown (motor_cat={motor_cat})")
            # Wait up to 25 attempts for the ZD dropdown to have options
            for _zd_wait in range(25):
                zd_ready = await page.evaluate("""(opt) => {
                    for (const sel of document.querySelectorAll('select')) {
                        if (Array.from(sel.options).some(o => o.text.trim().toLowerCase() === opt.toLowerCase()))
                            return true;
                    }
                    return false;
                }""", zd_option)
                if zd_ready:
                    log("info", f"  ZD dropdown ready (attempt {_zd_wait+1})")
                    break
                await page.wait_for_timeout(500)
            zd_selected = False

            # Pass 1: find <select> near "zero depreciation", select the right option
            all_selects = await page.query_selector_all('select')
            for sel_el in all_selects:
                try:
                    opts = await page.evaluate("""(s) =>
                        Array.from(s.options).map(o => o.text.trim())
                    """, sel_el)
                    if not any(o.lower() == zd_option.lower() for o in opts):
                        continue
                    await sel_el.scroll_into_view_if_needed()
                    await sel_el.select_option(label=zd_option)
                    await page.wait_for_timeout(600)
                    log("ok", f"  -> '{zd_option}' selected from ZD dropdown")
                    zd_selected = True
                    break
                except Exception:
                    continue

            # Pass 2: JS direct value set + change event (Angular fallback)
            if not zd_selected:
                all_selects = await page.query_selector_all('select')
                for sel_el in all_selects:
                    try:
                        matched = await page.evaluate("""([s, opt]) => {
                            const o = Array.from(s.options).find(
                                o => o.text.trim().toLowerCase() === opt.toLowerCase()
                                  || o.value.trim().toLowerCase() === opt.toLowerCase()
                            );
                            if (!o) return null;
                            s.value = o.value;
                            s.dispatchEvent(new Event('change', {bubbles: true}));
                            return o.text;
                        }""", [sel_el, zd_option])
                        if matched:
                            log("ok", f"  -> '{matched}' selected (JS fallback)")
                            zd_selected = True
                            break
                    except Exception:
                        continue

            if not zd_selected:
                log("warn", f"'{zd_option}' not selected automatically — select it manually")
                await pause(f"Select '{zd_option}' from the dropdown manually, then press ENTER")

            await page.wait_for_timeout(800)
            steps.append("zero_dep")

            # ── STEP 8: Click RECALCULATE button ─────────────────────
            # From video: the button has tooltip "Submit final quote" meaning
            # it could be any element type — use JS innerText scan on ALL
            # visible elements which reliably finds it regardless of tag.
            log("step", "STEP 8 — Recalculate Premium")
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(800)

            async def click_by_text(texts):
                if isinstance(texts, str):
                    texts = [texts]
                for text in texts:
                    # Playwright locator across all tags
                    for tag in ["button", "input", "a", "span", "div"]:
                        try:
                            loc = page.locator(f"{tag}:has-text('{text}')").first
                            await loc.wait_for(state="visible", timeout=2000)
                            await loc.scroll_into_view_if_needed()
                            await loc.click()
                            log("ok", f"  → Clicked '{text}' via <{tag}>")
                            return True
                        except Exception:
                            continue
                    # JS scan — catches input[value=...] and any element
                    try:
                        clicked = await page.evaluate(f"""() => {{
                            for (const el of document.querySelectorAll('*')) {{
                                const rect = el.getBoundingClientRect();
                                if (rect.width === 0 || rect.height === 0) continue;
                                const txt = (el.innerText || el.value || '').trim();
                                if (txt === '{text}') {{ el.click(); return true; }}
                            }}
                            return false;
                        }}""")
                        if clicked:
                            log("ok", f"  → Clicked '{text}' via JS scan")
                            return True
                    except Exception:
                        pass
                return False

            # Find smallest element with exact text "RECALCULATE" and mouse.click it
            recalc_info = await page.evaluate("""() => {
                let best = null, bestArea = Infinity;
                for (const el of document.querySelectorAll('*')) {
                    const rect = el.getBoundingClientRect();
                    if (rect.width === 0 || rect.height === 0) continue;
                    const txt = (el.innerText || el.value || '').trim();
                    if (txt === 'RECALCULATE' || txt === 'Recalculate') {
                        const area = rect.width * rect.height;
                        if (area < bestArea) { bestArea = area; best = el; }
                    }
                }
                if (best) {
                    const r = best.getBoundingClientRect();
                    return { x: r.left + r.width/2, y: r.top + r.height/2, tag: best.tagName };
                }
                return null;
            }""")
            if recalc_info:
                await page.mouse.click(recalc_info['x'], recalc_info['y'])
                log("ok", f"  -> RECALCULATE clicked at ({recalc_info['x']:.0f},{recalc_info['y']:.0f}) via <{recalc_info['tag']}>")
            else:
                log("warn", "RECALCULATE not found — check browser")
                await pause("Click 'RECALCULATE' manually, then press ENTER")

            # Wait for premium to recalculate and SUBMIT to appear.
            # Also detect if portal bounced back to RTO page (malfunction) —
            # if so, pause and let user fix it manually before continuing.
            log("info", "Waiting for premium to recalculate...")
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            await page.wait_for_timeout(4000)

            # Check if portal went back to RTO/details page (malfunction)
            on_rto_page = await page.evaluate("""() => {
                const txt = document.body.innerText.toLowerCase();
                return (txt.includes('city where vehicle is registered') ||
                        txt.includes('rto') && txt.includes('manufacturer')) &&
                       !txt.includes('recalculate') && !txt.includes('submit');
            }""")
            if on_rto_page:
                log("warn", "  MALFUNCTION: Portal returned to RTO/details page after RECALCULATE")
                log("warn", "  Please navigate back to the coverage/premium page manually")
                await pause("Fix the page manually (go back to premium page), then press ENTER")

            # Poll until SUBMIT button is visible (up to 30s)
            for _ in range(60):
                found = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('*')) {
                        const rect = el.getBoundingClientRect();
                        if (rect.width === 0 || rect.height === 0) continue;
                        const txt = (el.innerText || el.value || '').trim();
                        if (txt === 'SUBMIT' || txt === 'Submit') return true;
                    }
                    return false;
                }""")
                if found:
                    log("ok", "  -> SUBMIT button appeared")
                    break
                # Re-check for RTO malfunction during wait
                on_rto = await page.evaluate("""() => {
                    const txt = document.body.innerText.toLowerCase();
                    return txt.includes('city where vehicle is registered') &&
                           !txt.includes('submit');
                }""")
                if on_rto:
                    log("warn", "  Portal drifted to RTO page — pausing for manual fix")
                    await pause("Navigate back to premium/coverage page, then press ENTER")
                    break
                await page.wait_for_timeout(500)
            else:
                log("warn", "  SUBMIT button not detected — trying anyway")

            steps.append("recalculated")

            # ── STEP 9: Click SUBMIT button ───────────────────────────
            # After RECALCULATE the button label changes to SUBMIT
            log("step", "STEP 9 — SUBMIT")
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(800)

            submit_info = await page.evaluate("""() => {
                let best = null, bestArea = Infinity;
                for (const el of document.querySelectorAll('*')) {
                    const rect = el.getBoundingClientRect();
                    if (rect.width === 0 || rect.height === 0) continue;
                    const txt = (el.innerText || el.value || '').trim();
                    if (txt === 'SUBMIT' || txt === 'Submit') {
                        const area = rect.width * rect.height;
                        if (area < bestArea) { bestArea = area; best = el; }
                    }
                }
                if (best) {
                    const r = best.getBoundingClientRect();
                    return { x: r.left + r.width/2, y: r.top + r.height/2, tag: best.tagName };
                }
                return null;
            }""")
            if submit_info:
                await page.mouse.click(submit_info['x'], submit_info['y'])
                log("ok", f"  -> SUBMIT clicked at ({submit_info['x']:.0f},{submit_info['y']:.0f}) via <{submit_info['tag']}>")
            else:
                log("warn", "SUBMIT not found — check browser")
                await pause("Click 'SUBMIT' manually, then press ENTER")

            try:
                await page.wait_for_load_state("networkidle", timeout=30000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)

            # Verify SUBMIT navigated forward (not back to RTO page)
            on_rto_after_submit = await page.evaluate("""() => {
                const txt = document.body.innerText.toLowerCase();
                return txt.includes('city where vehicle is registered') &&
                       !txt.includes('engine number');
            }""")
            if on_rto_after_submit:
                log("warn", "  MALFUNCTION after SUBMIT: portal returned to RTO page")
                await pause("Manually navigate back to proposal/coverage page, then press ENTER")

            steps.append("submit_coverages")

            # ── STEP 11: Proposal Form (Engine/Chassis + KYC popup) ──
            log("step", "STEP 11 — Proposal Form")
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            # Wait until Engine Number AND Chassis fields are actually present in DOM
            # before attempting to fill them — prevents empty fills on slow pages.
            log("info", "Waiting for Engine/Chassis fields to appear...")
            for _ef in range(60):
                fields_ready = await page.evaluate("""() => {
                    const inputs = Array.from(document.querySelectorAll('input'));
                    const hasEngine  = inputs.some(i => (i.placeholder||'').toLowerCase().includes('engine'));
                    const hasChassis = inputs.some(i => (i.placeholder||'').toLowerCase().includes('chasis'));
                    return hasEngine && hasChassis;
                }""")
                if fields_ready:
                    log("info", f"  Engine/Chassis fields ready (attempt {_ef+1})")
                    break
                await page.wait_for_timeout(500)

            async def mouse_fill(find_js, value, field_name):
                await page.evaluate(f"() => {{ const el={find_js}; if(el) el.scrollIntoView({{block:'center'}}); }}")
                await page.wait_for_timeout(200)
                c = await page.evaluate(f"() => {{ const el={find_js}; if(!el) return null; const r=el.getBoundingClientRect(); return r.width>0?{{x:r.left+r.width/2,y:r.top+r.height/2}}:null; }}")
                if c:
                    await page.mouse.click(c['x'], c['y'])
                    await page.wait_for_timeout(120)
                    await page.keyboard.press("Control+a")
                    await page.keyboard.press("Delete")
                    await page.wait_for_timeout(80)
                    await page.keyboard.type(str(value), delay=40)
                    await page.wait_for_timeout(200)
                    log("fill", f"  -> {field_name}: {value}")
                    return True
                log("warn", f"  {field_name} not found")
                return False

            async def click_last_button(text_upper, label):
                log("info", f"Clicking {label}")
                await page.evaluate(f"""() => {{
                    const all = Array.from(document.querySelectorAll('button,input[type="submit"]'))
                        .filter(el => (el.innerText||el.value||'').trim().toUpperCase() === '{text_upper}');
                    if (all.length) all[all.length-1].scrollIntoView({{block:'center'}});
                }}""")
                await page.wait_for_timeout(400)
                c = await page.evaluate(f"""() => {{
                    const all = Array.from(document.querySelectorAll('button,input[type="submit"]'))
                        .filter(el => (el.innerText||el.value||'').trim().toUpperCase() === '{text_upper}')
                        .filter(el => {{ const r=el.getBoundingClientRect(); return r.width>0&&r.height>0; }});
                    if (!all.length) return null;
                    for (const el of [...all].reverse()) {{
                        let node=el.parentElement;
                        for (let i=0;i<12;i++) {{
                            if(!node) break;
                            const s=window.getComputedStyle(node);
                            if (s.position==='fixed'||s.position==='absolute'||
                                ['modal','popup','overlay','dialog','cdk-overlay'].some(c=>node.classList.contains(c))) {{
                                const r=el.getBoundingClientRect();
                                return {{x:r.left+r.width/2,y:r.top+r.height/2}};
                            }}
                            node=node.parentElement;
                        }}
                    }}
                    const el=all[all.length-1];
                    const r=el.getBoundingClientRect();
                    return {{x:r.left+r.width/2,y:r.top+r.height/2}};
                }}""")
                if c:
                    await page.mouse.click(c['x'], c['y'])
                    log("ok", f"  -> {label} clicked at ({c['x']:.0f},{c['y']:.0f})")
                    return True
                log("warn", f"  {label} not found")
                return False

            # ── 11a: Engine Number ────────────────────────────────────
            if D.get("engine_number"):
                await mouse_fill("Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('engine'))", D["engine_number"], "Engine Number")

            # ── 11b: Chassis Number (portal typo "Chasis") ───────────
            if D.get("chassis_number"):
                await mouse_fill("Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('chasis'))", D["chassis_number"], "Chassis Number")

            # ── 11c: Check if KYC already done (loop case) ──────────
            pin_val = D.get("pin_code","").strip()   # Applicant Details pincode — always available
            gst_val = D.get("gst_number","").strip()       # defined here so always available
            kyc_already_done = await page.evaluate("""() =>
                document.body.innerText.toLowerCase().includes('your kyc is done')
            """)

            if kyc_already_done:
                log("ok", "  -> KYC already done — skipping KYC popup steps")
                steps.append("kyc")
            else:
                # ── 11c: Complete KYC radio ───────────────────────────────
                log("info", "Clicking Complete KYC radio")
                await page.evaluate("""() => {
                    for (const r of document.querySelectorAll('input[type="radio"]')) {
                        const lbl=r.closest('label')||document.querySelector('label[for="'+r.id+'"]')||r.parentElement;
                        if ((lbl?.innerText||lbl?.textContent||'').toLowerCase().includes('complete kyc'))
                            { r.scrollIntoView({block:'center'}); return; }
                    }
                }""")
                await page.wait_for_timeout(300)
                kyc_c = await page.evaluate("""() => {
                    for (const r of document.querySelectorAll('input[type="radio"]')) {
                        const lbl=r.closest('label')||document.querySelector('label[for="'+r.id+'"]')||r.parentElement;
                        if ((lbl?.innerText||lbl?.textContent||'').toLowerCase().includes('complete kyc')) {
                            const rect=r.getBoundingClientRect();
                            if (rect.width>0) return {x:rect.left+rect.width/2,y:rect.top+rect.height/2};
                        }
                    }
                    for (const el of document.querySelectorAll('*')) {
                        if ((el.innerText||el.textContent||'').trim().toLowerCase()==='complete kyc') {
                            const rect=el.getBoundingClientRect();
                            if (rect.width>0) return {x:rect.left+rect.width/2,y:rect.top+rect.height/2};
                        }
                    }
                    return null;
                }""")
                if kyc_c:
                    await page.mouse.click(kyc_c['x'], kyc_c['y'])
                    log("ok", "  -> Complete KYC radio clicked")
                log("info", "Waiting for KYC popup...")
                for _ in range(20):
                    if await page.evaluate("""() =>
                        document.body.innerText.toLowerCase().includes('select entity type')||
                        document.body.innerText.toLowerCase().includes('customer onboarding')
                    """):
                        log("ok", "  -> KYC popup opened"); break
                    await page.wait_for_timeout(500)
                await page.wait_for_timeout(800)

                # ── 11d: Entity Type = Corporate ─────────────────────────
                log("info", "Selecting Entity Type: Corporate")
                for sel_el in await page.query_selector_all('select'):
                    try:
                        near = await page.evaluate("""(s)=>{ let n=s.parentElement; for(let i=0;i<8;i++){if(!n)break;if((n.innerText||'').toLowerCase().includes('entity type'))return true;n=n.parentElement;} return false;}""", sel_el)
                        if not near: continue
                        await sel_el.scroll_into_view_if_needed()
                        await sel_el.select_option(label="Corporate")
                        await page.wait_for_timeout(1200)
                        log("ok", "  -> Entity Type: Corporate"); break
                    except Exception as e:
                        log("warn", f"  Entity Type error: {e}")

                # ── 11e: GST radio ────────────────────────────────────────
                log("info", "Clicking GST radio")
                gst_radio_c = await page.evaluate("""() => {
                    for (const r of document.querySelectorAll('input[type="radio"]')) {
                        const lbl=r.closest('label')||document.querySelector('label[for="'+r.id+'"]')||r.parentElement;
                        if ((lbl?.innerText||lbl?.textContent||'').trim()==='GST') {
                            r.scrollIntoView({block:'center'});
                            const rect=r.getBoundingClientRect();
                            return rect.width>0?{x:rect.left+rect.width/2,y:rect.top+rect.height/2}:null;
                        }
                    }
                    for (const el of document.querySelectorAll('label,span,div')) {
                        if ((el.innerText||el.textContent||'').trim()==='GST') {
                            const r=el.getBoundingClientRect(); if(r.width>0) return {x:r.left+r.width/2,y:r.top+r.height/2};
                        }
                    }
                    return null;
                }""")
                if gst_radio_c:
                    await page.mouse.click(gst_radio_c['x'], gst_radio_c['y'])
                    log("ok", "  -> GST radio clicked")
                for _ in range(15):
                    if await page.evaluate("""() =>
                        document.body.innerText.toLowerCase().includes('gstin number')||
                        Array.from(document.querySelectorAll('input')).some(i=>(i.placeholder||'').toLowerCase().includes('e.g.')||(i.placeholder||'').toLowerCase().includes('aaach'))
                    """):
                        break
                    await page.wait_for_timeout(500)
                await page.wait_for_timeout(400)

                # ── 11f: GSTIN (KYC popup) ────────────────────────────────
                gst_val = D.get("gst_number","").strip()
                if gst_val:
                    for js in [
                        "Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('e.g.'))",
                        "Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('aaach'))",
                        "Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('gstin'))",
                    ]:
                        if await mouse_fill(js, gst_val, "GSTIN (KYC popup)"): break

                # ── 11g: PIN Code (KYC popup) ─────────────────────────────
                # Use separate variable — must NOT overwrite pin_val (used later for Applicant Details)
                kyc_pin_val = D.get("kyc_pincode", D.get("pin_code","")).strip()
                if kyc_pin_val:
                    filled = await mouse_fill(r"Array.from(document.querySelectorAll('input')).find(i=>/^\d{6}$/.test((i.placeholder||'').trim()))", kyc_pin_val, "PIN Code (KYC)")
                    if not filled:
                        await mouse_fill("Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('pin'))", kyc_pin_val, "PIN Code (KYC)")

                # ── 11h: Consent checkbox ─────────────────────────────────
                log("info", "Ticking consent checkbox")
                await page.wait_for_timeout(300)
                cb = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        if (el.checked) continue;
                        let node=el.parentElement;
                        for (let i=0;i<6;i++) {
                            if (!node) break;
                            const t=(node.innerText||node.textContent||'').toLowerCase();
                            if (t.includes('consent')||t.includes('hereby')) {
                                el.scrollIntoView({block:'center'});
                                const r=el.getBoundingClientRect();
                                return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2}:null;
                            }
                            node=node.parentElement;
                        }
                    }
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        if (el.checked) continue;
                        const r=el.getBoundingClientRect();
                        if (r.width>0) { el.scrollIntoView({block:'center'}); return {x:r.left+r.width/2,y:r.top+r.height/2}; }
                    }
                    return null;
                }""")
                if cb:
                    await page.mouse.click(cb['x'], cb['y'])
                    await page.wait_for_timeout(400)
                    log("ok", "  -> Consent checkbox ticked")
                else:
                    log("warn", "  Consent checkbox not found")

                await page.wait_for_timeout(300)
                await click_last_button("SUBMIT", "SUBMIT")
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)
                steps.append("kyc")

            # ── STEP 12: Declaration → PROCEED ────────────────────────
            # Skip PROCEED if KYC already done — Applicant Details already visible
            if kyc_already_done:
                log("step", "STEP 12 — Declaration PROCEED (skipped — KYC already done)")
                steps.append("proceed")
            else:
                log("step", "STEP 12 — Declaration PROCEED")
                log("info", "Waiting for page to fully load before PROCEED...")
                try:
                    await page.wait_for_load_state("networkidle", timeout=20000)
                except Exception:
                    pass
                decl_cb = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        if (el.checked) continue;
                        let node=el.parentElement;
                        for (let i=0;i<6;i++) {
                            if (!node) break;
                            const t=(node.innerText||node.textContent||'').toLowerCase();
                            if (t.includes('hereby declare')||t.includes('premium')||t.includes('declaration')) {
                                el.scrollIntoView({block:'center'});
                                const r=el.getBoundingClientRect();
                                return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2}:null;
                            }
                            node=node.parentElement;
                        }
                    }
                    return null;
                }""")
                if decl_cb:
                    await page.mouse.click(decl_cb['x'], decl_cb['y'])
                    await page.wait_for_timeout(400)
                    log("ok", "  -> Declaration checkbox ticked")
                proceed_clicked = False
                for _proc in range(25):
                    try:
                        for txt in ["PROCEED", "Proceed", "proceed"]:
                            loc = page.get_by_role("button", name=txt)
                            if await loc.count() > 0:
                                await loc.last.scroll_into_view_if_needed()
                                await page.wait_for_timeout(300)
                                await loc.last.click(force=True)
                                log("ok", f"  -> PROCEED clicked (locator)")
                                proceed_clicked = True
                                break
                        if proceed_clicked:
                            break
                    except Exception:
                        pass
                    coords = await page.evaluate("""() => {
                        const all = Array.from(document.querySelectorAll('button, input[type="submit"]'))
                            .filter(el => (el.innerText||el.value||'').trim().toLowerCase().includes('proceed'));
                        if (!all.length) return null;
                        const el = all[all.length - 1];
                        el.scrollIntoView({block:'center'});
                        const r = el.getBoundingClientRect();
                        return r.width > 0 ? {x: r.left+r.width/2, y: r.top+r.height/2} : null;
                    }""")
                    if coords:
                        await page.wait_for_timeout(300)
                        await page.mouse.click(coords['x'], coords['y'])
                        log("ok", f"  -> PROCEED clicked at ({coords['x']:.0f},{coords['y']:.0f})")
                        proceed_clicked = True
                        break
                    log("info", f"  PROCEED not found yet (attempt {_proc+1}/25), waiting...")
                    await page.wait_for_timeout(2000)
                if not proceed_clicked:
                    log("warn", "  PROCEED not found after all retries")
                try:
                    await page.wait_for_load_state("networkidle", timeout=20000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)
                steps.append("proceed")

            # ── STEP 12.5: Applicant Details Form ─────────────────────
            log("step", "STEP 12.5 — Applicant Details → Create Proposal")
            for _ in range(20):
                if await page.evaluate("""() => document.body.innerText.includes('Applicant Details')||document.body.innerText.includes('Create Proposal')"""):
                    log("ok", "  -> Applicant Details loaded"); break
                await page.wait_for_timeout(500)
            await page.wait_for_timeout(600)

            # ── Insured Name ──────────────────────────────────────────
            insured_name = D.get("insured_name","").strip()
            if insured_name:
                await mouse_fill(
                    "Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('insured name')||(i.placeholder||'').toLowerCase()==='insured name')",
                    insured_name, "Insured Name")

            # ── DOB — click, Escape to close calendar, type DD/MM/YYYY ──────
            dob_str = D.get("dob","").strip()
            if dob_str:
                log("info", f"Filling DOB: {dob_str}")
                try:
                    parts = dob_str.replace("-","/").split("/")
                    dob_day  = int(parts[0])
                    dob_mon  = int(parts[1])
                    dob_year = int(parts[2])
                    dob_formatted = f"{dob_day:02d}/{dob_mon:02d}/{dob_year}"
                    log("info", f"  Typing: {dob_formatted}")

                    dob_coords = await page.evaluate("""() => {
                        const inp = Array.from(document.querySelectorAll('input'))
                            .find(i => (i.placeholder||'').toLowerCase().includes('date of birth') ||
                                       (i.placeholder||'').toLowerCase() === 'dob');
                        if (!inp) return null;
                        inp.scrollIntoView({block:'center'});
                        const r = inp.getBoundingClientRect();
                        return r.width > 0 ? {x: r.left+r.width/2, y: r.top+r.height/2} : null;
                    }""")
                    if not dob_coords:
                        raise ValueError("DOB input not found")

                    # Click → calendar opens → Escape closes it, input stays focused
                    await page.mouse.click(dob_coords['x'], dob_coords['y'])
                    await page.wait_for_timeout(400)
                    await page.keyboard.press("Escape")
                    await page.wait_for_timeout(200)
                    # Click again to re-focus
                    await page.mouse.click(dob_coords['x'], dob_coords['y'])
                    await page.wait_for_timeout(300)
                    await page.keyboard.press("Escape")
                    await page.wait_for_timeout(150)
                    # Clear and type
                    await page.keyboard.press("Control+A")
                    await page.wait_for_timeout(80)
                    await page.keyboard.press("Delete")
                    await page.wait_for_timeout(80)
                    for ch in dob_formatted:
                        await page.keyboard.type(ch, delay=80)
                        await page.wait_for_timeout(50)
                    await page.wait_for_timeout(300)

                    # Lock via native setter + Angular events
                    await page.evaluate("""(val) => {
                        const inp = Array.from(document.querySelectorAll('input'))
                            .find(i => (i.placeholder||'').toLowerCase().includes('date of birth') ||
                                       (i.placeholder||'').toLowerCase() === 'dob');
                        if (!inp) return;
                        const setter = Object.getOwnPropertyDescriptor(
                            window.HTMLInputElement.prototype, 'value').set;
                        setter.call(inp, val);
                        inp.dispatchEvent(new Event('input',  {bubbles:true}));
                        inp.dispatchEvent(new Event('change', {bubbles:true}));
                        inp.dispatchEvent(new Event('blur',   {bubbles:true}));
                    }""", dob_formatted)
                    await page.wait_for_timeout(200)

                    actual_dob = await page.evaluate("""() => {
                        const inp = Array.from(document.querySelectorAll('input'))
                            .find(i => (i.placeholder||'').toLowerCase().includes('date of birth') ||
                                       (i.placeholder||'').toLowerCase() === 'dob');
                        return inp ? inp.value : null;
                    }""")
                    log("ok" if actual_dob else "warn",
                        f"  -> DOB set: {actual_dob}" if actual_dob
                        else "  DOB field empty after typing")

                except Exception as e:
                    log("warn", f"  DOB error: {e}")

            # ── Address Line ──────────────────────────────────────────
            addr_val = D.get("address_line","").strip()
            if addr_val:
                addr_coords = await page.evaluate("""() => {
                    const inp = Array.from(document.querySelectorAll('input,textarea'))
                        .find(i => (i.placeholder||'').toLowerCase().includes('address') ||
                                   (i.id||'').toLowerCase().includes('address') ||
                                   (i.name||'').toLowerCase().includes('address'));
                    if (!inp) return null;
                    inp.scrollIntoView({block:'center'});
                    const r = inp.getBoundingClientRect();
                    return r.width > 0 ? {x: r.left+r.width/2, y: r.top+r.height/2} : null;
                }""")
                if addr_coords:
                    await page.mouse.click(addr_coords['x'], addr_coords['y'], click_count=3)
                    await page.wait_for_timeout(150)
                    await page.keyboard.press("Delete")
                    await page.wait_for_timeout(80)
                    await page.keyboard.type(addr_val, delay=40)
                    await page.wait_for_timeout(200)
                    await page.evaluate("""(val) => {
                        const inp = Array.from(document.querySelectorAll('input,textarea'))
                            .find(i => (i.placeholder||'').toLowerCase().includes('address') ||
                                       (i.id||'').toLowerCase().includes('address') ||
                                       (i.name||'').toLowerCase().includes('address'));
                        if (!inp) return;
                        const setter = Object.getOwnPropertyDescriptor(
                            window.HTMLInputElement.prototype, 'value').set;
                        setter.call(inp, val);
                        inp.dispatchEvent(new Event('input',  {bubbles:true}));
                        inp.dispatchEvent(new Event('change', {bubbles:true}));
                        inp.dispatchEvent(new Event('blur',   {bubbles:true}));
                    }""", addr_val)
                    log("fill", f"  -> Address Line: {addr_val}")
                else:
                    log("warn", "  Address Line field not found")

            # ── Mobile ────────────────────────────────────────────────
            mob_val = D.get("mobile_number","").strip()
            if mob_val:
                await mouse_fill("Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('mobile'))", mob_val, "Mobile")

            # ── Email ─────────────────────────────────────────────────
            email_val = D.get("email","").strip()
            if email_val:
                await mouse_fill("Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase()==='email'||((i.placeholder||'').toLowerCase().includes('email')&&!(i.placeholder||'').toLowerCase().includes('mobile')))", email_val, "Email")

            # ── Pincode — retry until State field populates ──────────
            # After typing pincode the portal auto-fills State.
            # If State is empty it means the pincode wasn't accepted —
            # clear and retype until State appears (up to 10 attempts).
            if pin_val:
                for _pin_try in range(10):
                    filled = await mouse_fill("Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('pincode')||(i.placeholder||'').toLowerCase()==='pincode')", pin_val, "Pincode")
                    if not filled:
                        await mouse_fill("Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('pin')&&(i.placeholder||'').toLowerCase().includes('code'))", pin_val, "Pincode")
                    # Wait for State to auto-populate
                    await page.wait_for_timeout(1000)
                    state_populated = await page.evaluate("""() => {
                        for (const inp of document.querySelectorAll('input')) {
                            const ph = (inp.placeholder||'').toLowerCase();
                            const val = (inp.value||'').trim();
                            if ((ph.includes('state') || ph === 'state') && val.length > 1)
                                return true;
                        }
                        // Also check read-only state fields (disabled/readonly inputs)
                        for (const inp of document.querySelectorAll('input[readonly],input[disabled]')) {
                            const val = (inp.value||'').trim();
                            if (val.length > 1) return true;
                        }
                        // Check for state text in nearby spans/divs
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText||'').trim().toLowerCase();
                            if (t === 'state' || t.includes('state')) {
                                const sib = el.nextElementSibling;
                                if (sib && (sib.innerText||sib.value||'').trim().length > 1)
                                    return true;
                            }
                        }
                        return false;
                    }""")
                    if state_populated:
                        log("ok", f"  State populated after pincode (attempt {_pin_try+1})")
                        break
                    log("warn", f"  State not populated — retyping pincode (attempt {_pin_try+1}/10)...")
                    await page.wait_for_timeout(500)

            # ── City/District: custom Angular dropdown ─────────────────
            # Flow: click the dropdown trigger → search input + option list appears
            #       → click the first/BANGALORE option
            log("info", "Selecting City/District")
            # Wait for state to populate City/District options (networkidle + 25 attempts)
            try:
                await page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass
            for _city_wait in range(25):
                city_ready = await page.evaluate("""() => {
                    // Check if the City/District select has options beyond placeholder
                    for (const sel of document.querySelectorAll('select')) {
                        let n = sel.parentElement;
                        for (let i=0;i<6;i++) {
                            if (!n) break;
                            const t = (n.innerText||n.textContent||'').toLowerCase();
                            if (t.includes('city') || t.includes('district')) {
                                // Has real options (more than just placeholder)
                                const real = Array.from(sel.options).filter(
                                    o => o.value && o.text.trim() &&
                                    !['','enter city / district','select'].includes(o.text.trim().toLowerCase())
                                );
                                return real.length > 0;
                            }
                            n = n.parentElement;
                        }
                    }
                    // Also check ng-select/custom dropdown has a visible value
                    for (const el of document.querySelectorAll('[class*="ng-select"],[class*="dropdown"]')) {
                        const t = (el.innerText||'').toLowerCase();
                        if (t.includes('city') || t.includes('district') || t.includes('bangalore')) {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0) return true;
                        }
                    }
                    return false;
                }""")
                if city_ready:
                    log("info", f"  City/District ready (attempt {_city_wait+1})")
                    break
                await page.wait_for_timeout(500)
            await page.wait_for_timeout(400)
            city_done = False

            # Try clicking the dropdown trigger (the arrow/select-box for City/District)
            city_trigger = await page.evaluate("""() => {
                // Find label containing "city" or "district"
                for (const lbl of document.querySelectorAll('label,span,div')) {
                    const t=(lbl.innerText||lbl.textContent||'').toLowerCase();
                    if ((t.includes('city')||t.includes('district'))&&t.length<30) {
                        // Look for the dropdown trigger near this label
                        let node=lbl.parentElement;
                        for (let i=0;i<5;i++) {
                            if(!node) break;
                            // Find a select-like element or dropdown arrow button
                            const trigger = node.querySelector('select,[role="combobox"],[role="listbox"],'+
                                '[class*="dropdown"],[class*="select"],[class*="ng-select"]');
                            if (trigger) {
                                trigger.scrollIntoView({block:'center'});
                                const r=trigger.getBoundingClientRect();
                                return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2}:null;
                            }
                            node=node.parentElement;
                        }
                    }
                }
                // Fallback: find any element showing "BANGALORE" or "Enter City"
                for (const el of document.querySelectorAll('*')) {
                    const t=(el.innerText||'').trim();
                    if (t==='BANGALORE'||t==='Enter City / District') {
                        const r=el.getBoundingClientRect();
                        if (r.width>0&&r.height>0) return {x:r.left+r.width/2,y:r.top+r.height/2};
                    }
                }
                return null;
            }""")
            if city_trigger:
                await page.mouse.click(city_trigger['x'], city_trigger['y'])
                await page.wait_for_timeout(500)
                log("info", "  -> City dropdown trigger clicked")

                # Now pick first visible option that is not a placeholder
                city_opt = await page.evaluate("""() => {
                    // Options appear as li elements or divs in a dropdown list
                    for (const el of document.querySelectorAll('li,[role="option"],div.option,div.item')) {
                        const t=(el.innerText||el.textContent||'').trim();
                        if (t&&t!=='Enter City / District'&&t.length>1&&t.length<40) {
                            const r=el.getBoundingClientRect();
                            if (r.width>0&&r.height>0) return {x:r.left+r.width/2,y:r.top+r.height/2,txt:t};
                        }
                    }
                    return null;
                }""")
                if city_opt:
                    await page.mouse.click(city_opt['x'], city_opt['y'])
                    await page.wait_for_timeout(400)
                    log("ok", f"  -> City selected: {city_opt['txt']}")
                    city_done = True

            if not city_done:
                # Fallback: try native <select>
                for sel_el in await page.query_selector_all('select'):
                    try:
                        near = await page.evaluate("""(s)=>{ let n=s.parentElement; for(let i=0;i<5;i++){if(!n)break;const t=(n.innerText||n.textContent||'').toLowerCase();if(t.includes('city')||t.includes('district'))return true;n=n.parentElement;}return false;}""", sel_el)
                        if not near: continue
                        await sel_el.scroll_into_view_if_needed()
                        opts = await page.evaluate("(s)=>Array.from(s.options).map(o=>({val:o.value,txt:o.text.trim()}))", sel_el)
                        best = next((o for o in opts if o['val'] and o['txt'] and o['txt'].upper() not in ('','ENTER CITY / DISTRICT','SELECT')), None)
                        if best:
                            await sel_el.select_option(value=best['val'])
                            await page.wait_for_timeout(400)
                            log("ok", f"  -> City (select): {best['txt']}")
                            city_done = True; break
                    except Exception as e:
                        log("warn", f"  City select error: {e}")

            # ── Add GST Details checkbox ──────────────────────────────
            if gst_val:
                log("info", "Ticking Add GST Details checkbox")
                await page.evaluate("window.scrollBy(0, 400)")
                await page.wait_for_timeout(400)
                gst_cb = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        let node = el.parentElement;
                        for (let i = 0; i < 6; i++) {
                            if (!node) break;
                            const t = (node.innerText||node.textContent||'').toLowerCase();
                            if (t.includes('add gst') || t.includes('gst details')) {
                                el.scrollIntoView({block:'center'});
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                            node = node.parentElement;
                        }
                    }
                    for (const lbl of document.querySelectorAll('label')) {
                        const t = (lbl.innerText||lbl.textContent||'').toLowerCase();
                        if (t.includes('add gst') || t.includes('gst details')) {
                            lbl.scrollIntoView({block:'center'});
                            const r = lbl.getBoundingClientRect();
                            if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                        }
                    }
                    return null;
                }""")
                if gst_cb:
                    await page.mouse.click(gst_cb['x'], gst_cb['y'])
                    await page.wait_for_timeout(800)
                    log("ok", "  -> Add GST Details checkbox ticked")
                else:
                    log("warn", "  Add GST Details checkbox not found")

                # ── GSTIN field (now visible after checkbox) ──────────
                await mouse_fill("Array.from(document.querySelectorAll('input')).find(i=>(i.placeholder||'').toLowerCase().includes('gstin')||(i.placeholder||'').toLowerCase().includes('enter gstin'))", gst_val, "GSTIN (proposal form)")

            # ── Loan / Lease / Hypothecation Details ─────────────────
            loan_val       = str(D.get("loan_hypothecation", "")).strip().lower()
            financier_name = str(D.get("financier_name",   "")).strip()
            financier_branch = str(D.get("financier_branch", "")).strip()

            if loan_val == "yes":
                log("info", "Ticking Loan / Lease / Hypothecation Details checkbox")
                await page.evaluate("window.scrollBy(0, 300)")
                await page.wait_for_timeout(400)

                # Tick the checkbox
                loan_cb = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        let node = el.parentElement;
                        for (let i = 0; i < 6; i++) {
                            if (!node) break;
                            const t = (node.innerText||node.textContent||'').toLowerCase();
                            if (t.includes('loan') || t.includes('hypothecation') || t.includes('lease')) {
                                el.scrollIntoView({block:'center'});
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                            node = node.parentElement;
                        }
                    }
                    for (const lbl of document.querySelectorAll('label')) {
                        const t = (lbl.innerText||lbl.textContent||'').toLowerCase();
                        if (t.includes('loan') || t.includes('hypothecation') || t.includes('lease')) {
                            lbl.scrollIntoView({block:'center'});
                            const r = lbl.getBoundingClientRect();
                            if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                        }
                    }
                    return null;
                }""")
                if loan_cb:
                    await page.mouse.click(loan_cb['x'], loan_cb['y'])
                    await page.wait_for_timeout(800)
                    log("ok", "  -> Loan/Hypothecation checkbox ticked")
                else:
                    log("warn", "  Loan/Hypothecation checkbox not found")

                # Select "Loan/Hypothecation" radio button
                await page.wait_for_timeout(500)
                loan_radio = await page.evaluate("""() => {
                    for (const r of document.querySelectorAll('input[type="radio"]')) {
                        const lbl = r.closest('label') ||
                                    document.querySelector('label[for="'+r.id+'"]') ||
                                    r.parentElement;
                        const t = (lbl?.innerText||lbl?.textContent||'').toLowerCase();
                        if (t.includes('loan') || t.includes('hypothecation')) {
                            r.scrollIntoView({block:'center'});
                            const rect = r.getBoundingClientRect();
                            if (rect.width > 0) return {x: rect.left+rect.width/2, y: rect.top+rect.height/2};
                        }
                    }
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                        if (t === 'loan/hypothecation' || t === 'loan / hypothecation') {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                        }
                    }
                    return null;
                }""")
                if loan_radio:
                    await page.mouse.click(loan_radio['x'], loan_radio['y'])
                    await page.wait_for_timeout(600)
                    log("ok", "  -> Loan/Hypothecation radio selected")
                else:
                    log("warn", "  Loan/Hypothecation radio not found")

                # Fill Financier Name
                if financier_name:
                    await mouse_fill(
                        "Array.from(document.querySelectorAll('input')).find(i=>"
                        "(i.placeholder||'').toLowerCase().includes('financier name')||"
                        "(i.placeholder||'').toLowerCase().includes('financier'))",
                        financier_name, "Financier Name"
                    )

                # Fill Financier Branch
                if financier_branch:
                    await mouse_fill(
                        "Array.from(document.querySelectorAll('input')).find(i=>"
                        "(i.placeholder||'').toLowerCase().includes('branch'))",
                        financier_branch, "Financier Branch"
                    )

            # ── Create Proposal ───────────────────────────────────────
            log("info", "Clicking Create Proposal")
            await page.wait_for_timeout(400)
            cp_c = await page.evaluate("""() => {
                for (const el of document.querySelectorAll('button,input[type="submit"]')) {
                    if ((el.innerText||el.value||'').trim().toLowerCase().includes('create proposal')) {
                        el.scrollIntoView({block:'center'});
                        const r=el.getBoundingClientRect();
                        return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2}:null;
                    }
                }
                return null;
            }""")
            if cp_c:
                await page.mouse.click(cp_c['x'], cp_c['y'])
                log("ok", f"  -> Create Proposal clicked at ({cp_c['x']:.0f},{cp_c['y']:.0f})")
            else:
                log("warn", "  Create Proposal not found")
                await pause("Click 'Create Proposal' manually, then press ENTER")
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            await page.wait_for_timeout(2000)
            steps.append("proposal_created")

            # ── STEP 14: Payment ─────────────────────────────────────
            log("step", "STEP 14 — Payment")

            # Scroll to bottom so Pay Now button is in viewport
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(1000)

            # ── STEP 13: Proposal Number ──────────────────────────────
            # Read proposal number — retry up to 10x (page may take a few seconds to load)
            log("step", "STEP 13 — Saving Proposal Number")

            proposal_no = ""
            for _pn_try in range(25):
                proposal_no = await page.evaluate("""() => {
                    const isPN = s => /^[0-9]{8,15}$/.test((s||'').trim());
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                        if (t !== 'proposal number' && t !== 'proposal no') continue;
                        let card = el.parentElement;
                        for (let i = 0; i < 5; i++) {
                            if (!card) break;
                            const walker = document.createTreeWalker(
                                card, NodeFilter.SHOW_TEXT);
                            let node;
                            while ((node = walker.nextNode())) {
                                const v = node.nodeValue.trim();
                                if (isPN(v)) return v;
                            }
                            for (const child of card.querySelectorAll('*')) {
                                if (child === el) continue;
                                const v = (child.innerText||'').trim();
                                if (isPN(v)) return v;
                            }
                            card = card.parentElement;
                        }
                    }
                    return null;
                }""") or ""
                if proposal_no:
                    break
                log("info", f"  Proposal number not visible yet, waiting... (attempt {_pn_try+1}/25)")
                await page.wait_for_timeout(1000)

            if proposal_no:
                log("ok", f"Proposal Number: {proposal_no}")
            else:
                log("warn", "Could not auto-read proposal number")
                proposal_no = input("      Type the Proposal Number from browser, press ENTER: ").strip()

            with open(LOG_PATH, "a") as f:
                f.write(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
                        f"{D['manufacturer_model']}  |  {D.get('engine_number','')}  |  {proposal_no}\n")
            log("ok", "Saved to proposal_log.txt")
            save_proposal_number(D, proposal_no)
            steps.append("proposal_saved")

            # Now click Pay Now — scroll to bottom, get fresh coords after scroll settles
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(800)
            pay_now_c = await page.evaluate("""() => {
                // Find Pay Now — scroll it into view, wait, then return fresh coords
                for (const el of document.querySelectorAll('*')) {
                    const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                    if (t !== 'pay now') continue;
                    el.scrollIntoView({block:'center'});
                    return true;
                }
                return false;
            }""")
            await page.wait_for_timeout(600)  # wait for scroll to settle
            pay_now_coords = await page.evaluate("""() => {
                for (const el of document.querySelectorAll('*')) {
                    const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                    if (t !== 'pay now') continue;
                    const r = el.getBoundingClientRect();
                    if (r.width > 0 && r.height > 0) {
                        return {x: r.left+r.width/2, y: r.top+r.height/2};
                    }
                }
                return null;
            }""")
            if pay_now_coords:
                await page.mouse.click(pay_now_coords['x'], pay_now_coords['y'])
                log("ok", "  -> Pay Now clicked")
                # Wait for payment page to fully load before attempting Advance PID
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(5000)

                # ── STEP 13 retry: read proposal number from payment page ──
                # The Premium Details card on the payment page shows proposal number
                # as bold plain text — try to capture it here if STEP 13 failed earlier
                if not proposal_no:
                    log("info", "  Retrying proposal number read from payment page...")
                    proposal_no_retry = await page.evaluate("""() => {
                        const isPN = s => /^[0-9]{8,15}$/.test((s||'').trim());
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if (t !== 'proposal number' && t !== 'proposal no') continue;
                            let card = el.parentElement;
                            for (let i = 0; i < 5; i++) {
                                if (!card) break;
                                const walker = document.createTreeWalker(
                                    card, NodeFilter.SHOW_TEXT);
                                let node;
                                while ((node = walker.nextNode())) {
                                    const v = node.nodeValue.trim();
                                    if (isPN(v)) return v;
                                }
                                for (const child of card.querySelectorAll('*')) {
                                    if (child === el) continue;
                                    const v = (child.innerText||'').trim();
                                    if (isPN(v)) return v;
                                }
                                card = card.parentElement;
                            }
                        }
                        return null;
                    }""") or ""
                    if proposal_no_retry:
                        proposal_no = proposal_no_retry
                        log("ok", f"  Proposal Number (payment page): {proposal_no}")
                        with open(LOG_PATH, "a") as f:
                            f.write(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
                                    f"{D['manufacturer_model']}  |  {D.get('engine_number','')}  |  {proposal_no}\n")
                        save_proposal_number(D, proposal_no)
                    else:
                        log("warn", "  Proposal number still not found on payment page")

            else:
                log("warn", "  Pay Now button not found")

            if D["auto_pay"]:
                # Wait for payment page to fully load before searching for Advance PID
                log("info", "Waiting for payment page to load before Advance PID...")
                try:
                    await page.wait_for_load_state("networkidle", timeout=25000)
                except Exception:
                    pass
                # Auto-complete payment — click Advance PID
                # Try on current page AND any newly opened page/tab
                adv_clicked = False

                async def try_click_advance_pid(pg):
                    # Strategy 1: Playwright locator by text
                    try:
                        loc = pg.get_by_text("Advance PID", exact=True)
                        if await loc.count() > 0:
                            await loc.first.scroll_into_view_if_needed()
                            await pg.wait_for_timeout(300)
                            await loc.first.click(force=True)
                            return True
                    except Exception:
                        pass
                    # Strategy 2: XPath — find element whose text is exactly "Advance PID"
                    try:
                        loc2 = pg.locator("xpath=//*[normalize-space(text())='Advance PID']")
                        if await loc2.count() > 0:
                            await loc2.first.scroll_into_view_if_needed()
                            await pg.wait_for_timeout(300)
                            await loc2.first.click(force=True)
                            return True
                    except Exception:
                        pass
                    # Strategy 3: JS — click parent of text node "Advance PID"
                    try:
                        done = await pg.evaluate("""() => {
                            const walker = document.createTreeWalker(
                                document.body, NodeFilter.SHOW_TEXT);
                            let node;
                            while ((node = walker.nextNode())) {
                                if (node.nodeValue.trim().toLowerCase() === 'advance pid') {
                                    const el = node.parentElement;
                                    el.scrollIntoView({block:'center'});
                                    el.click();
                                    return true;
                                }
                            }
                            return false;
                        }""")
                        if done:
                            return True
                    except Exception:
                        pass
                    return False

                # Poll until "Advance PID" text appears on the page —
                # the payment page can take several minutes to fully load.
                # Wait up to 10 minutes (120 × 5s) before giving up.
                log("info", "  Waiting for Advance PID to appear on page...")
                for _adv_wait in range(120):
                    adv_visible = await page.evaluate("""() => {
                        const walker = document.createTreeWalker(
                            document.body, NodeFilter.SHOW_TEXT);
                        let node;
                        while ((node = walker.nextNode())) {
                            if (node.nodeValue.trim().toLowerCase() === 'advance pid')
                                return true;
                        }
                        return false;
                    }""")
                    if adv_visible:
                        log("info", f"  Advance PID found (attempt {_adv_wait+1})")
                        break
                    # Also check other open tabs
                    try:
                        for pg in page.context.pages:
                            if pg == page:
                                continue
                            adv_visible = await pg.evaluate("""() => {
                                const walker = document.createTreeWalker(
                                    document.body, NodeFilter.SHOW_TEXT);
                                let node;
                                while ((node = walker.nextNode())) {
                                    if (node.nodeValue.trim().toLowerCase() === 'advance pid')
                                        return true;
                                }
                                return false;
                            }""")
                            if adv_visible:
                                break
                    except Exception:
                        pass
                    if adv_visible:
                        break
                    if _adv_wait % 6 == 5:   # log every 30s so user knows it's working
                        log("info", f"  Still waiting for Advance PID... ({(_adv_wait+1)*5}s)")
                    await page.wait_for_timeout(5000)

                # Now click it
                for _adv_try in range(5):
                    adv_clicked = await try_click_advance_pid(page)
                    if adv_clicked:
                        break
                    try:
                        for pg in page.context.pages:
                            if pg == page:
                                continue
                            adv_clicked = await try_click_advance_pid(pg)
                            if adv_clicked:
                                break
                    except Exception:
                        pass
                    if adv_clicked:
                        break
                    await page.wait_for_timeout(1000)

                if adv_clicked:
                    log("ok", "  -> Advance PID clicked")
                    # Wait for Advance PID page to fully load
                    try:
                        await page.wait_for_load_state("networkidle", timeout=15000)
                    except Exception:
                        pass
                    await page.wait_for_timeout(3000)

                    # ── Tick the PID Number checkbox matching Excel pid_number ──
                    pid_val = str(D.get("pid_number", "")).strip()
                    if pid_val:
                        log("info", f"  Selecting PID Number: {pid_val}")
                        pid_ticked = False

                        # ── DEBUG: dump all rows and cells to find PID structure ──
                        debug_rows = await page.evaluate("""() => {
                            const out = [];
                            document.querySelectorAll('tr').forEach((row, ri) => {
                                const cells = Array.from(row.querySelectorAll('td,th'))
                                    .map(c => (c.innerText||'').trim().slice(0,40));
                                const hasCb = !!row.querySelector('input[type="checkbox"]');
                                if (cells.length > 0)
                                    out.push({ri, cells, hasCb});
                            });
                            return out.slice(0,20);
                        }""")
                        for dr in debug_rows:
                            log("info", f"  [ROW {dr['ri']}] cb={dr['hasCb']} cells={dr['cells']}")

                        # ── Also dump all inputs on page ──
                        debug_inputs = await page.evaluate("""() => {
                            return Array.from(document.querySelectorAll('input')).slice(0,10).map(i => ({
                                type: i.type, ph: i.placeholder, val: (i.value||'').slice(0,30),
                                cls: (i.className||'').slice(0,40)
                            }));
                        }""")
                        for di in debug_inputs:
                            log("info", f"  [INP] type={di['type']} ph='{di['ph']}' val='{di['val']}' cls='{di['cls']}'")

                        async def tick_pid_on_page(pg):
                            """Type PID in search box to filter table, then click its checkbox."""
                            # Step 1: type PID number into the Search box to filter the table
                            try:
                                search_inp = await pg.query_selector('input[placeholder*="Search" i], input[type="search"]')
                                if search_inp:
                                    await search_inp.triple_click()
                                    await search_inp.type(pid_val, delay=80)
                                    await pg.wait_for_timeout(1500)
                                    log("info", f"  Typed '{pid_val}' into search box")
                                else:
                                    log("info", "  No search box found")
                            except Exception as ex:
                                log("info", f"  Search box error: {ex}")

                            for _s in range(5):
                                # Step 2: scroll checkbox into view
                                scrolled = await pg.evaluate("""(pidVal) => {
                                    const rows = document.querySelectorAll('tr');
                                    for (const row of rows) {
                                        for (const cell of row.querySelectorAll('td, th')) {
                                            const t = (cell.innerText||'').trim();
                                            if (t === pidVal) {
                                                const cb = row.querySelector('input[type="checkbox"]');
                                                if (cb) { cb.scrollIntoView({block:'center'}); return 'cb-found'; }
                                                return 'row-found-no-cb';
                                            }
                                        }
                                    }
                                    return 'not-found';
                                }""", pid_val)
                                log("info", f"  PID scroll attempt {_s+1}: {scrolled}")
                                if scrolled == 'not-found':
                                    await pg.wait_for_timeout(1000)
                                    continue
                                if scrolled == 'row-found-no-cb':
                                    # No checkbox in tr — try clicking the row itself
                                    coords = await pg.evaluate("""(pidVal) => {
                                        const rows = document.querySelectorAll('tr');
                                        for (const row of rows) {
                                            for (const cell of row.querySelectorAll('td,th')) {
                                                if ((cell.innerText||'').trim() === pidVal) {
                                                    const r = row.getBoundingClientRect();
                                                    if (r.width > 0) return {x: r.left+20, y: r.top+r.height/2};
                                                }
                                            }
                                        }
                                        return null;
                                    }""", pid_val)
                                    if coords:
                                        await pg.wait_for_timeout(400)
                                        await pg.mouse.click(coords['x'], coords['y'])
                                        return True
                                # Step 3: re-read fresh coords after scroll settles
                                await pg.wait_for_timeout(500)
                                coords = await pg.evaluate("""(pidVal) => {
                                    const rows = document.querySelectorAll('tr');
                                    for (const row of rows) {
                                        for (const cell of row.querySelectorAll('td, th')) {
                                            if ((cell.innerText||'').trim() === pidVal) {
                                                const cb = row.querySelector('input[type="checkbox"]');
                                                if (cb) {
                                                    const cr = cb.getBoundingClientRect();
                                                    if (cr.width > 0)
                                                        return {x: cr.left+cr.width/2, y: cr.top+cr.height/2};
                                                }
                                            }
                                        }
                                    }
                                    return null;
                                }""", pid_val)
                                if coords:
                                    await pg.mouse.click(coords['x'], coords['y'])
                                    await pg.wait_for_timeout(300)
                                    return True
                                await pg.wait_for_timeout(1000)
                            return False

                        # Try on current page first
                        pid_ticked = await tick_pid_on_page(page)
                        # If not found, try all other open tabs
                        if not pid_ticked:
                            try:
                                for pg in page.context.pages:
                                    if pg == page:
                                        continue
                                    log("info", f"  Trying PID on tab: {pg.url}")
                                    pid_ticked = await tick_pid_on_page(pg)
                                    if pid_ticked:
                                        break
                            except Exception:
                                pass

                        if pid_ticked:
                            log("ok", f"  -> PID {pid_val} checkbox ticked")
                        else:
                            log("warn", f"  PID {pid_val} not found — tick manually")
                    else:
                        log("info", "  No PID Number in Excel — skipping PID checkbox")

                    await page.wait_for_timeout(1000)

                    # ── Click Pay button after PID checkbox ticked ──
                    log("info", "Clicking Pay button...")
                    pay_clicked = False
                    for _pay in range(5):
                        pay_coords = await page.evaluate("""() => {
                            for (const el of document.querySelectorAll('button, input[type="submit"], a')) {
                                const t = (el.innerText||el.textContent||el.value||'').trim().toLowerCase();
                                if (t === 'pay') {
                                    el.scrollIntoView({block:'center'});
                                    const r = el.getBoundingClientRect();
                                    if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                }
                            }
                            return null;
                        }""")
                        if pay_coords:
                            await page.wait_for_timeout(400)
                            # Re-read coords fresh after scroll
                            pay_coords2 = await page.evaluate("""() => {
                                for (const el of document.querySelectorAll('button, input[type="submit"], a')) {
                                    const t = (el.innerText||el.textContent||el.value||'').trim().toLowerCase();
                                    if (t === 'pay') {
                                        const r = el.getBoundingClientRect();
                                        if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                    }
                                }
                                return null;
                            }""")
                            if pay_coords2:
                                await page.mouse.click(pay_coords2['x'], pay_coords2['y'])
                                log("ok", "  -> Pay clicked")
                                pay_clicked = True
                                break
                        await page.wait_for_timeout(1000)
                    if not pay_clicked:
                        log("warn", "  Pay button not found — click manually")
                        await pause("Click 'Pay' manually, then press ENTER")

                    log("ok", "Payment initiated!")
                    steps.append("payment_done")
                else:
                    log("warn", "Advance PID button not found — complete payment manually")
                    await pause("Complete payment manually in the browser, then press ENTER")
                    steps.append("payment_manual")
            else:
                log("info", "auto_pay = FALSE — complete payment manually")
                await pause("Complete payment manually in the browser, then press ENTER")
                steps.append("payment_manual")

            save_result(D, "SUCCESS", proposal_no, steps)

            print(f"""
+--------------------------------------------------------------+
|  ALL DONE!                                                   |
|  Vehicle  : {D['manufacturer_model']:<51}|
|  Proposal : {proposal_no:<51}|
|  Saved    : proposal_log.txt + details.xlsx (Results tab)    |
+--------------------------------------------------------------+""")

            # ── Capture Policy Number from success page ──────────────
            log("info", "Waiting for success page and Policy Number...")
            try:
                await page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)

            # Retry up to 60x × 2s = 2 minutes — success page can take time.
            # Uses 4 strategies so no layout variation is missed.
            policy_no = ""
            for _pol in range(60):
                policy_no = await page.evaluate("""() => {
                    // Policy number pattern: 4digits/9digits/2digits/3digits
                    function isPolicyNo(v) {
                        return v && v.length > 6 &&
                               /[0-9]/.test(v) &&
                               v.includes('/') &&
                               !/policy number/i.test(v);
                    }

                    // ── Strategy 1: find label "Policy Number", read next sibling ──
                    let labelEl = null, labelSize = Infinity;
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim();
                        if (t.toLowerCase() === 'policy number') {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0) {
                                const sz = r.width * r.height;
                                if (sz < labelSize) { labelEl = el; labelSize = sz; }
                            }
                        }
                    }
                    if (labelEl) {
                        // Check next siblings
                        let sib = labelEl.nextElementSibling;
                        for (let i = 0; i < 5; i++) {
                            if (!sib) break;
                            const v = (sib.innerText||sib.textContent||'').trim();
                            if (isPolicyNo(v)) return v;
                            sib = sib.nextElementSibling;
                        }
                        // Walk up parent chain, collect all text nodes, find value after label
                        let row = labelEl.parentElement;
                        for (let i = 0; i < 6; i++) {
                            if (!row) break;
                            const texts = [];
                            const walker = document.createTreeWalker(row, NodeFilter.SHOW_TEXT);
                            let node;
                            while ((node = walker.nextNode())) {
                                const v = node.nodeValue.trim();
                                if (v) texts.push(v);
                            }
                            const idx = texts.findIndex(t => t.toLowerCase() === 'policy number');
                            if (idx !== -1) {
                                for (let j = idx+1; j < Math.min(idx+5, texts.length); j++) {
                                    if (isPolicyNo(texts[j])) return texts[j];
                                }
                            }
                            row = row.parentElement;
                        }
                    }

                    // ── Strategy 2: scan ALL text nodes for policy number pattern ──
                    const walker2 = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
                    let node2;
                    while ((node2 = walker2.nextNode())) {
                        const v = node2.nodeValue.trim();
                        // Match format like 3005/433241549/00/000
                        if (v.match(/^[0-9]{4}[/][0-9]{9}[/][0-9]{2}[/][0-9]{3}$/)) return v;
                    }

                    // ── Strategy 3: scan visible elements for policy-like value ──
                    for (const el of document.querySelectorAll('span,p,div,td,h1,h2,h3,h4,strong,b')) {
                        const v = (el.innerText||'').trim();
                        if (v.match(/^[0-9]{4}[/][0-9]+[/][0-9]{2}[/][0-9]{3}$/)) return v;
                    }

                    return null;
                }""") or ""
                if policy_no:
                    log("ok", f"  Policy Number found (attempt {_pol+1}): {policy_no}")
                    break
                if _pol % 5 == 4:
                    log("info", f"  Still waiting for Policy Number... ({(_pol+1)*2}s)")
                await page.wait_for_timeout(2000)

            if policy_no:
                log("ok", f"  Policy Number: {policy_no}")
                save_policy_number(D, policy_no)
            else:
                log("warn", "  Policy Number not found — saving empty")

            # ── Click "Another Policy" to go back for next row ──────

            another_clicked = False
            for _ap in range(15):
                # Strategy 1: Playwright get_by_text (most reliable)
                try:
                    loc = page.get_by_text("Another Policy", exact=True)
                    if await loc.count() > 0:
                        await loc.first.scroll_into_view_if_needed()
                        await page.wait_for_timeout(400)
                        await loc.first.click(force=True)
                        log("ok", "  -> Another Policy clicked (get_by_text)")
                        another_clicked = True
                        break
                except Exception:
                    pass
                # Strategy 2: XPath exact text
                try:
                    loc2 = page.locator("xpath=//*[normalize-space(text())='Another Policy']")
                    if await loc2.count() > 0:
                        await loc2.first.scroll_into_view_if_needed()
                        await page.wait_for_timeout(400)
                        await loc2.first.click(force=True)
                        log("ok", "  -> Another Policy clicked (xpath)")
                        another_clicked = True
                        break
                except Exception:
                    pass
                # Strategy 3: JS scan all elements including div/span
                coords = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                        if (t === 'another policy') {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0) {
                                el.scrollIntoView({block:'center'});
                                return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                        }
                    }
                    return null;
                }""")
                if coords:
                    await page.wait_for_timeout(400)
                    await page.mouse.click(coords['x'], coords['y'])
                    log("ok", "  -> Another Policy clicked (JS scan)")
                    another_clicked = True
                    break
                await page.wait_for_timeout(1000)
            if not another_clicked:
                log("warn", "  Another Policy button not found — waiting extra 30s and retrying once")
                await page.wait_for_timeout(30000)
                try:
                    loc = page.get_by_text("Another Policy")
                    if await loc.count() > 0:
                        await loc.first.click(force=True)
                        log("ok", "  -> Another Policy clicked (late retry)")
                        another_clicked = True
                except Exception:
                    pass
                if not another_clicked:
                    log("warn", "  Another Policy not found after all retries — skipping")

            # Wait for the details page to reload
            try:
                await page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)
            return D["_excel_row"] + 1  # signal to process next row

        except Exception as e:
            log("warn", f"Error: {e}")
            save_result(D, "FAILED", proposal_no, steps, str(e))
            await pause("Check the browser. Press ENTER to continue to next row.")
            return D.get("_excel_row", 4) + 1  # skip failed row, try next


# ═══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # ── Read first row to preview and confirm before starting ────
    D_first = read_details(start_row=4)

    print("\n  First record from details.xlsx:")
    print(f"     Vehicle      : {D_first['vehicle_reg_number']}")
    print(f"     Mobile       : {D_first['mobile_number']}")
    print(f"     Policy Type  : {D_first.get('policy_type','')}")
    print(f"     Ownership    : {D_first.get('ownership','') or '(not set)'}")
    print(f"     SAOD         : {D_first.get('saod','') or 'no'}")
    print(f"     Rollover Reg : {D_first.get('rollover_reg_number','') or '(N/A for New)'}")
    print(f"     Reg Type     : {D_first.get('current_reg_type','')}")
    print(f"     RTO          : {D_first.get('rto','')}")
    print(f"     Make/Model   : {D_first.get('manufacturer_model','')}")
    print(f"     Mfg Year     : {D_first.get('manufacturing_year','')}")
    print(f"     Tenure       : {D_first.get('tenure','')}")
    print(f"     IDV (Rs)     : {D_first.get('idv','')}")
    print(f"     PAN          : {D_first['pan_number']}")
    print(f"     DOB          : {D_first['dob']}")
    print(f"     GST          : {D_first['gst_number'] or '(not provided)'}")
    print(f"     Add-ons      : {', '.join(D_first['addons']) or '(none)'}")
    print(f"     Auto Pay     : {D_first['auto_pay']}")

    print("""
  ──────────────────────────────────────────────────────────────
  BEFORE pressing ENTER, make sure:

  1. Chrome is open with remote debugging. Start it with Win+R:
     "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
     --remote-debugging-port=9222 --user-data-dir="C:\\chrome-nysa"

  2. You are logged in to nysa.icicilombard.com (home page visible)
  ──────────────────────────────────────────────────────────────
    """)

    input("  Press ENTER to connect to your Chrome and start...")

    # ── Loop through all data rows ───────────────────────────────
    next_row = 4
    row_count = 0
    while True:
        try:
            D = read_details(start_row=next_row)
        except SystemExit:
            print("\n  No more data rows found. All done!")
            break

        row_count += 1
        print(f"\n{'='*60}")
        print(f"  Processing row {D['_excel_row']} (record #{row_count})")
        print(f"  Vehicle: {D['vehicle_reg_number']}  |  Model: {D['manufacturer_model'][:40]}")
        print(f"{'='*60}")

        next_row = asyncio.run(run(D, skip_to_step4=(row_count > 1)))

        if next_row is None:
            # run() didn't return a next row — stop
            print("\n  Processing stopped.")
            break

        # Check if there are more rows
        try:
            read_details(start_row=next_row)  # just check — will raise SystemExit if no data
        except SystemExit:
            print(f"\n  All {row_count} record(s) processed successfully!")
            break

    print("\n  Script finished.")