"""
╔══════════════════════════════════════════════════════════════╗
║   ICICI NYSA — Motor Insurance Automation                    ║
║   Connects to YOUR existing Chrome — no new tab/login        ║
╚══════════════════════════════════════════════════════════════╝

STEP 1 — Start Chrome with remote debugging (run this ONCE):
  Press  Win + R  on your keyboard, paste this and press Enter:

  "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\\chrome-nysa"

  (A new Chrome window will open)

STEP 2 — In that Chrome window:
  Go to:  https://nysa.icicilombard.com
  Login with your ID, Password, Captcha, OTP — just like normal.
  Make sure you are on the HOME page.

STEP 3 — Fill details.xlsx row 5 with your vehicle data.

STEP 4 — Run this script:
  python icici_nysa.py

The script attaches to your already-logged-in Chrome.
No new tab. No new login. Starts directly from home page.
"""

import asyncio
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright

# ── Paths ──────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "details.xlsx"
LOG_PATH   = BASE_DIR / "proposal_log.txt"

# ── Must match the port you used to start Chrome ──────────────────
CHROME_PORT = 9222


# ═══════════════════════════════════════════════════════════════════
#  READ EXCEL
# ═══════════════════════════════════════════════════════════════════

def read_details(start_row=4):
    if not EXCEL_PATH.exists():
        print(f"\n  Cannot find: {EXCEL_PATH}")
        print("  Make sure details.xlsx is in the same folder.")
        input("\nPress ENTER to exit..."); raise SystemExit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "Details" not in wb.sheetnames:
        print("  Sheet 'Details' not found in details.xlsx")
        input("\nPress ENTER to exit..."); raise SystemExit(1)

    ws = wb["Details"]

    # ── Label map: handles any variation of header text ──────────
    # Strip special chars, lowercase, then match
    label_map = {
        # Motor Category & Type
        "motor category"                              : "motor_category",
        "motor category *"                            : "motor_category",
        "motor type"                                  : "motor_type",
        "motor type *"                                : "motor_type",
        # PID Number
        "pid number"                                  : "pid_number",
        "pid no"                                      : "pid_number",
        # Vehicle & Contact
        "vehicle reg no"                              : "vehicle_reg_number",
        "vehicle reg number"                          : "vehicle_reg_number",
        "registration number"                         : "vehicle_reg_number",
        "reg no"                                      : "vehicle_reg_number",
        "mobile number"                               : "mobile_number",
        "mobile"                                      : "mobile_number",
        "phone"                                       : "mobile_number",
        "email"                                       : "email",
        "policy type"                                 : "policy_type",
        "policy type *"                               : "policy_type",
        # Ownership
        "ownership"                                   : "ownership",
        # SAOD
        "saod"                                        : "saod",
        # Rollover / Used registration number
        "rollover registration number"                : "rollover_reg_number",
        "used registration number"                    : "rollover_reg_number",
        "registration number for rollover"            : "rollover_reg_number",
        # Registration Type — matches portal exact header
        "current registration type"                   : "current_reg_type",
        "current reg type"                            : "current_reg_type",
        "reg type"                                    : "current_reg_type",
        "registration type"                           : "current_reg_type",
        # RTO — matches portal exact header
        "city where vehicle is registered rto"        : "rto",
        "city where vehicle is registered"            : "rto",
        "rto"                                         : "rto",
        "rto code"                                    : "rto",
        "city rto"                                    : "rto",
        # Manufacturer/Model
        "manufacturer - model"                        : "manufacturer_model",
        "manufacturer-model"                          : "manufacturer_model",
        "manufacturer model"                          : "manufacturer_model",
        "make model"                                  : "manufacturer_model",
        "make - model"                                : "manufacturer_model",
        # Year
        "mfg year"                                    : "manufacturing_year",
        "tenure"                                      : "tenure",
        "tenure *"                                    : "tenure",
        "manufacturing year"                          : "manufacturing_year",
        "year"                                        : "manufacturing_year",
        # IDV — matches portal exact header
        "insured depreciation value idv"              : "idv",
        "insured declared value idv"                  : "idv",
        "insured depreciation value"                  : "idv",
        "insured declared value"                      : "idv",
        "idv"                                         : "idv",
        # KYC
        "pan number"                                  : "pan_number",
        "pan"                                         : "pan_number",
        "dob dd/mm/yyyy"                              : "dob",
        "dob dd mm yyyy"                              : "dob",
        "dob"                                         : "dob",
        "date of birth"                               : "dob",
        "aadhaar no"                                  : "aadhaar_number",
        "aadhaar"                                     : "aadhaar_number",
        # Engine / Chassis
        "engine number"                               : "engine_number",
        "engine no"                                   : "engine_number",
        "chassis number"                              : "chassis_number",
        "chassis no"                                  : "chassis_number",
        # GST — no bare "gst" (collides with "gst state")
        "gst number"                                  : "gst_number",
        "gstin"                                       : "gst_number",
        "gst state"                                   : "gst_state",
        # Pincode
        "pin code"                                    : "pin_code",
        "pincode"                                     : "pin_code",
        "pin code on gst certificate"                 : "pin_code",
        # City (proposal form)
        "city"                                        : "city",
        "city district"                               : "city",
        "city/district"                               : "city",
        "city / district"                             : "city",
        # Proposal form
        "insured name"                                : "insured_name",
        "address line 1"                              : "address_line",
        "address line"                                : "address_line",
        # Add-ons
        "add-ons comma separated"                     : "addons",
        "add ons comma separated"                     : "addons",
        "add-ons (comma separated)"                   : "addons",
        "addons"                                      : "addons",
        "add-ons"                                     : "addons",
        "add ons"                                     : "addons",
        "advanced pid"                                : "advanced_pid",
        "pid"                                         : "advanced_pid",
        # Payment
        "auto pay true/false"                         : "auto_pay",
        "auto pay (true/false)"                       : "auto_pay",
        "auto pay true false"                         : "auto_pay",
        "auto pay"                                    : "auto_pay",
        "autopay"                                     : "auto_pay",
        # Loan / Hypothecation
        "loan / lease / hypothecation details"         : "loan_hypothecation",
        "loan lease hypothecation details"             : "loan_hypothecation",
        "loan/lease/hypothecation details"             : "loan_hypothecation",
        "loan hypothecation"                           : "loan_hypothecation",
        "hypothecation"                                : "loan_hypothecation",
        "financier name"                               : "financier_name",
        "financier branch"                             : "financier_branch",
        "kyc pincode"                                  : "kyc_pincode",
        "kyc pin code"                                 : "kyc_pincode",
        "kyc pin"                                      : "kyc_pincode",
        # Additional Discount & Loading
        "additional discount yes/no"                   : "additional_discount",
        "additional discount yes no"                   : "additional_discount",
        "additional discount (yes/no)"                 : "additional_discount",
        "additional discount"                          : "additional_discount",
        "loading%"                                     : "loading_pct",
        "loading %"                                    : "loading_pct",
        "loading"                                      : "loading_pct",
    }

    positional_map = {
        1:"motor_category",      2:"motor_type",
        3:"email",               4:"policy_type",         5:"ownership",
        6:"saod",                7:"rollover_reg_number", 8:"current_reg_type",
        9:"rto",                 10:"manufacturer_model", 11:"manufacturing_year",
        12:"tenure",             13:"idv",                14:"pan_number",
        15:"aadhaar_number",     16:"engine_number",      17:"chassis_number",
        18:"kyc_pincode",        19:"insured_name",       20:"dob",
        21:"address_line",       22:"mobile_number",      23:"email",
        24:"pin_code",           25:"gst_number",         26:"gst_state",
        27:"loan_hypothecation", 28:"financier_name",     29:"financier_branch",
        30:"addons",             31:"additional_discount",
        32:"loading_pct",        33:"auto_pay",
        34:"pid_number",         35:"proposal_number",
        36:"policy_number",
    }

    import re as _re
    headers = {}
    header_row_found = False
    for col in range(1, 38):  # cols 1-37: includes Additional Discount(31), Loading%(32), Auto Pay(33), PID(34), Proposal(35), Policy(36)
        val = ws.cell(row=3, column=col).value
        if val:
            header_row_found = True
            clean = _re.sub(r"[*()\[\]₹]", " ", str(val))
            clean = _re.sub(r"\s+", " ", clean).strip().lower()
            field = label_map.get(clean)
            if not field:
                c2 = clean.replace(" ","")
                for k,v in label_map.items():
                    if c2 == k.replace(" ",""): field = v; break
            if not field:
                for k,v in label_map.items():
                    if clean.startswith(k) and (len(clean)==len(k) or clean[len(k)]==" "):
                        field = v; break
            if not field:
                field = _re.sub(r"[^a-z0-9]+","_",clean).strip("_")
            headers[col] = field
            print(f"  [HEADER] col {col}: {repr(str(val))} → {field}")

    if not header_row_found or not headers:
        print("  [INFO] Using positional column map")
        headers = positional_map.copy()

    data = {}
    max_col = max(headers.keys()) + 1 if headers else 22
    for row in range(start_row, ws.max_row + 1):
        row_vals = [ws.cell(row=row, column=c).value for c in range(1, max_col)]
        if any(v and str(v).strip() for v in row_vals):
            for col, field in headers.items():
                raw = ws.cell(row=row, column=col).value
                if hasattr(raw, 'strftime'):
                    data[field] = raw.strftime("%d/%m/%Y")
                else:
                    data[field] = str(raw).strip() if raw is not None else ""
            print(f"  [INFO] Reading data from row {row}")
            data["_excel_row"] = row
            break

    if not data:
        print("  No data found in details.xlsx")
        input("\nPress ENTER to exit..."); raise SystemExit(1)

    # Store the Excel row number so we can write proposal number back later

    data["vehicle_reg_number"] = data.get("vehicle_reg_number","").upper().replace(" ","")
    data["pan_number"]         = data.get("pan_number","").upper()
    data["gst_number"]         = data.get("gst_number","").upper()
    data["pin_code"]           = str(data.get("pin_code","")).split(".")[0].strip()
    data["engine_number"]      = data.get("engine_number","").upper().replace(" ","")
    data["chassis_number"]     = data.get("chassis_number","").upper().replace(" ","")
    data["rto"]                = data.get("rto","").upper().replace(" ","")

    data["manufacturer_model"] = data.get("manufacturer_model","")
    data["manufacturing_year"] = str(data.get("manufacturing_year","")).split(".")[0].strip()
    data["idv"]                = (data.get("idv","")
                                  .replace(",","").replace("₹","").replace(" ","").strip())
    data["current_reg_type"]   = data.get("current_reg_type","")
    data["tenure"]             = data.get("tenure","1 year").strip()
    raw_addons                 = data.get("addons","")
    data["addons"]             = [a.strip() for a in raw_addons.split(",") if a.strip()]
    data["auto_pay"]           = str(data.get("auto_pay","false")).strip().lower() == "true"
    data["policy_type"]        = data.get("policy_type","New").strip()
    data["ownership"]          = data.get("ownership","").strip()
    data["saod"]               = str(data.get("saod","")).strip().lower()
    data["rollover_reg_number"]= data.get("rollover_reg_number","").upper().replace(" ","")
    data["additional_discount"]= str(data.get("additional_discount","no")).strip().lower() == "yes"
    data["loading_pct"]        = str(data.get("loading_pct","")).replace("%","").strip()

    # ── Debug: show what was read ─────────────────────────────────
    print("  [DEBUG] Fields read from Excel:")
    for k, v in data.items():
        if k != "addons":
            print(f"           {k:<22} = {repr(v)}")
    print(f"           {'addons':<22} = {data['addons']}")

    return data


# ═══════════════════════════════════════════════════════════════════
#  SAVE RESULT
# ═══════════════════════════════════════════════════════════════════

def save_proposal_number(D, proposal_no):
    """Write proposal number back to col 31 of the Details sheet."""
    try:
        excel_row = D.get("_excel_row", 4)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Details"]
        ws.cell(row=excel_row, column=35, value=proposal_no)
        wb.save(EXCEL_PATH)
        print(f"  Proposal number '{proposal_no}' saved to Details row {excel_row}, col 35")
    except Exception as e:
        print(f"  Could not save proposal number to Excel: {e}")


def save_policy_number(D, policy_no):
    """Write policy number back to col 32 of the Details sheet."""
    try:
        excel_row = D.get("_excel_row", 4)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Details"]
        ws.cell(row=excel_row, column=36, value=policy_no)
        wb.save(EXCEL_PATH)
        print(f"  Policy number '{policy_no}' saved to Details row {excel_row}, col 36")
    except Exception as e:
        print(f"  Could not save policy number to Excel: {e}")


def save_result(D, status, proposal_no, steps, error=""):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Results Log"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        s  = Side(style="thin", color="BFBFBF")
        b  = Border(left=s, right=s, top=s, bottom=s)
        bg = "E2EFDA" if status == "SUCCESS" else "FCE4D6"
        r  = max(ws.max_row + 1, 3)
        for col, val in enumerate([
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            D["vehicle_reg_number"], status, proposal_no,
            " -> ".join(steps), error,
            "YES" if D.get("auto_pay") else "NO",
        ], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = b
        wb.save(EXCEL_PATH)
        print("  Result saved to details.xlsx (Results Log tab)")
    except Exception as e:
        print(f"  Could not write to Excel: {e}")


def mark_engine_chassis_green(D):
    """Colour the Engine Number and Chassis Number cells green after successful entry."""
    try:
        from openpyxl.styles import PatternFill
        excel_row = D.get("_excel_row", 4)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Details"]
        green = PatternFill("solid", fgColor="92D050")   # bright green
        # col 16 = engine_number, col 17 = chassis_number (from positional_map)
        ws.cell(row=excel_row, column=16).fill = green
        ws.cell(row=excel_row, column=17).fill = green
        wb.save(EXCEL_PATH)
    except Exception as e:
        print(f"  Could not mark engine/chassis green in Excel: {e}")


# ═══════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════

def log(kind, msg):
    icons = {"step":"\n[STEP]","ok":"  [OK] ","warn":"  [WARN]","fill":"  [FILL]","info":"  [INFO]"}
    print(f"{icons.get(kind,'      ')} {msg}")


async def pause(msg="Done? Press ENTER to continue"):
    print(f"\n  [PAUSE] {msg}")
    await asyncio.get_event_loop().run_in_executor(
        None, input, "          Press ENTER: ")


async def smart_click(page, keywords, label, timeout=12000):
    """
    Tries every possible strategy to click an element on an Angular portal.
    If nothing works: prints what IS on the page, then asks user to click manually.
    """
    kws = [keywords] if isinstance(keywords, str) else keywords

    for kw in kws:
        kw_lower = kw.lower()

        # 1. Playwright text locator — exact then partial
        for exact in [True, False]:
            try:
                loc = page.get_by_text(kw, exact=exact).first
                await loc.wait_for(state="visible", timeout=3000)
                await loc.scroll_into_view_if_needed()
                await loc.click()
                log("info", f"Clicked '{label}'")
                return True
            except Exception:
                pass

        # 2. aria-label / title / alt attributes
        for attr in [f"[aria-label*='{kw}' i]", f"[title*='{kw}' i]",
                     f"[alt*='{kw}' i]", f"[data-label*='{kw}' i]"]:
            try:
                await page.wait_for_selector(attr, timeout=2000, state="visible")
                await page.click(attr)
                log("info", f"Clicked '{label}' via attribute")
                return True
            except Exception:
                pass

        # 3. Angular / Material Design selectors
        for sel in [
            f"mat-card:has-text('{kw}')",
            f"mat-list-item:has-text('{kw}')",
            f"mat-menu-item:has-text('{kw}')",
            f"mat-tab:has-text('{kw}')",
            f"button:has-text('{kw}')",
            f"a:has-text('{kw}')",
            f"li:has-text('{kw}')",
            f"[class*='card']:has-text('{kw}')",
            f"[class*='tile']:has-text('{kw}')",
            f"[class*='item']:has-text('{kw}')",
            f"[class*='product']:has-text('{kw}')",
            f"[class*='menu']:has-text('{kw}')",
            f"[class*='nav']:has-text('{kw}')",
            f"[class*='btn']:has-text('{kw}')",
        ]:
            try:
                await page.wait_for_selector(sel, timeout=2000, state="visible")
                await page.click(sel)
                log("info", f"Clicked '{label}' via Angular selector")
                return True
            except Exception:
                pass

        # 4. JavaScript innerText scan — catches anything the above missed
        try:
            result = await page.evaluate(f"""() => {{
                const kw = '{kw_lower}';
                const els = document.querySelectorAll(
                    'a, button, mat-card, mat-list-item, li, span, ' +
                    '[role="button"], [role="tab"], [role="menuitem"], ' +
                    '[class*="card"], [class*="tile"], [class*="item"], ' +
                    '[class*="menu"], [class*="nav"], [class*="btn"]'
                );
                for (const el of els) {{
                    const txt = (el.innerText || el.textContent || '').trim().toLowerCase();
                    if (txt === kw || txt.startsWith(kw)) {{
                        const rect = el.getBoundingClientRect();
                        if (rect.width > 0 && rect.height > 0) {{
                            el.click();
                            return txt.substring(0, 40);
                        }}
                    }}
                }}
                return null;
            }}""")
            if result:
                log("info", f"Clicked '{label}' via JS scan")
                return True
        except Exception:
            pass

    # ── All strategies failed — show what's on the page and ask user ──
    log("warn", f"Could not auto-click '{label}'. Visible elements on page:")
    try:
        items = await page.evaluate("""() => {
            const els = document.querySelectorAll(
                'a, button, mat-card, mat-list-item, li, [role="button"],
                 [class*="card"], [class*="tile"], [class*="nav"]'
            );
            return [...new Set(
                Array.from(els)
                    .map(e => (e.innerText || e.textContent || '').trim())
                    .filter(t => t.length > 1 && t.length < 60)
            )].slice(0, 25);
        }""")
        for item in items:
            print(f"         . {item}")
    except Exception:
        pass

    await pause(f"Click '{label}' manually in the browser, then press ENTER")
    return False


async def fill_field(page, selectors, value, label):
    """Try each selector until one works, then fill it."""
    for sel in selectors:
        try:
            el = await page.wait_for_selector(sel, timeout=5000, state="visible")
            await el.scroll_into_view_if_needed()
            await el.click()
            # Use Ctrl+A + Delete instead of triple_click (works on ElementHandle)
            await page.keyboard.press("Control+a")
            await page.keyboard.press("Delete")
            await el.type(value, delay=60)
            log("fill", f"{label}: {value}")
            return True
        except Exception:
            continue
    log("warn", f"Could not fill '{label}' — fill it manually")
    return False


async def fill_dropdown(page, value, label):
    """
    For Angular autocomplete/dropdown fields on NYSA portal.
    Strategy:
    1. Find ANY visible input that is currently active / empty
    2. Type the value slowly
    3. Wait for dropdown list to appear
    4. Click the best matching option
    5. If no dropdown appears, try pressing Enter
    """
    try:
        # Type into the currently focused/active input on the page
        # First click somewhere neutral to reset focus
        await page.wait_for_timeout(300)

        # Find all visible inputs and try each one that looks empty
        inputs = await page.query_selector_all("input:visible, mat-select:visible")

        # Try typing into the page directly after finding the right field
        # Use JS to find input by placeholder or nearby label text
        el = await page.evaluate_handle(f"""() => {{
            const allInputs = document.querySelectorAll('input[type="text"], input:not([type])');
            for (const inp of allInputs) {{
                const rect = inp.getBoundingClientRect();
                if (rect.width === 0 || rect.height === 0) continue;
                const ph = (inp.placeholder || '').toLowerCase();
                const label = (inp.getAttribute('aria-label') || '').toLowerCase();
                const formCtrl = (inp.getAttribute('formcontrolname') || '').toLowerCase();
                const val = '{value.lower()}';
                const labelLow = '{label.lower()}';
                // Match by placeholder, aria-label, or formcontrolname
                if (ph.includes(labelLow) || label.includes(labelLow) ||
                    formCtrl.includes(labelLow) || ph.includes(val.substring(0,4))) {{
                    return inp;
                }}
            }}
            return null;
        }}""")

        if el:
            await el.click()
            await el.fill("")
            await el.type(value, delay=80)
            log("fill", f"{label}: {value}")
        else:
            # Fallback: use keyboard focus — tab to next field and type
            await page.keyboard.type(value, delay=80)
            log("fill", f"{label} (keyboard): {value}")

        # Wait for dropdown/autocomplete to appear
        await page.wait_for_timeout(1000)

        # Try clicking matching option in any dropdown that appeared
        option_clicked = False
        for option_sel in [
            f"mat-option:has-text('{value}')",
            f"li:has-text('{value}')",
            f".mat-option:has-text('{value}')",
            f"[role='option']:has-text('{value}')",
            f".dropdown-item:has-text('{value}')",
            f".autocomplete-option:has-text('{value}')",
            f"span:has-text('{value}')",
        ]:
            try:
                await page.wait_for_selector(option_sel, timeout=2000, state="visible")
                await page.click(option_sel)
                log("info", f"Selected dropdown option: {value}")
                option_clicked = True
                break
            except Exception:
                continue

        if not option_clicked:
            # Try partial match on first word
            first_word = value.split()[0] if value else value
            for option_sel in [
                f"mat-option:has-text('{first_word}')",
                f"[role='option']:has-text('{first_word}')",
                f"li:has-text('{first_word}')",
            ]:
                try:
                    await page.wait_for_selector(option_sel, timeout=1500, state="visible")
                    await page.click(option_sel)
                    log("info", f"Selected dropdown option (partial): {first_word}")
                    option_clicked = True
                    break
                except Exception:
                    continue

        if not option_clicked:
            # Press Enter/Tab to confirm whatever is shown
            await page.keyboard.press("Enter")

        await page.wait_for_timeout(500)
        return True

    except Exception as ex:
        log("warn", f"Could not fill dropdown '{label}': {ex}")
        return False


# ═══════════════════════════════════════════════════════════════════
#  CONNECT TO YOUR EXISTING CHROME
# ═══════════════════════════════════════════════════════════════════

async def connect_to_chrome(playwright):
    """
    Attach to Chrome that was started with --remote-debugging-port=9222.
    Finds the NYSA tab automatically. No new tab is opened.
    """
    try:
        browser = await playwright.chromium.connect_over_cdp(
            f"http://localhost:{CHROME_PORT}"
        )
    except Exception as e:
        print(f"""
  ERROR: Cannot connect to Chrome on port {CHROME_PORT}

  Make sure you started Chrome with this command (Win+R → paste):
  "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\\chrome-nysa"

  Then login to nysa.icicilombard.com and press ENTER here.
  Details: {e}
        """)
        input("Press ENTER to exit..."); raise SystemExit(1)

    # Find the NYSA tab — look for icicilombard.com in any open tab
    nysa_page = None
    all_pages = [p for ctx in browser.contexts for p in ctx.pages]

    for p in all_pages:
        if "icicilombard.com" in p.url:
            nysa_page = p
            break

    # If not found, use the most recently active tab
    if not nysa_page:
        if all_pages:
            nysa_page = all_pages[-1]
            log("warn", f"NYSA tab not found — using: {nysa_page.url}")
            print("  Make sure the NYSA portal is open in Chrome!")
            await pause("Navigate to nysa.icicilombard.com home page, then press ENTER")
        else:
            print("  No open tabs found in Chrome.")
            input("Press ENTER to exit..."); raise SystemExit(1)

    await nysa_page.bring_to_front()
    log("ok", f"Connected! Tab URL: {nysa_page.url}")
    return browser, nysa_page


# ═══════════════════════════════════════════════════════════════════
#  MAIN AUTOMATION FLOW
# ═══════════════════════════════════════════════════════════════════

async def run(D, skip_to_step4=False):
    steps       = []
    proposal_no = ""

    make_model = D.get("manufacturer_model","(not set)")
    print(f"""
+--------------------------------------------------------------+
|  ICICI NYSA — New Car Insurance Automation                   |
|  Make/Model : {make_model:<48}|
|  RTO        : {D.get('rto',''):<48}|
|  Mobile     : {D['mobile_number']:<48}|
|  Auto Pay   : {'YES - will click PAY' if D['auto_pay'] else 'NO  - stops before payment':<48}|
+--------------------------------------------------------------+""")

    async with async_playwright() as p:
        browser, page = await connect_to_chrome(p)

        try:
            # Let page fully settle
            try:
                await page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass
            await page.wait_for_timeout(1500)
            log("ok", f"Automation starting from: {page.url}")
            steps.append("connected")

            if not skip_to_step4:
                # ── STEP 1: Motor Insurance ────────────────────────────
                log("step", "STEP 1 — Motor Insurance")
                await smart_click(page, ["Motor", "Motor Insurance"], "Motor")
                # Wait for Motor page to fully load — networkidle exits early when ready
                # The Motor click often triggers a navigation which destroys the
                # execution context. We need to wait for navigation to complete first.
                try:
                    await page.wait_for_load_state("domcontentloaded", timeout=15000)
                except Exception:
                    pass
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)

                # Also wait up to 30 attempts for the RTO/Quote form to actually render
                # Each evaluate is wrapped in try/except because the portal may still
                # be navigating (destroying the execution context)
                for _mot_wait in range(30):
                    try:
                        page_loaded = await page.evaluate("""() => {
                            const txt = (document.body.innerText||'').toLowerCase();
                            return txt.includes('registration number') ||
                                   txt.includes('fetch vehicle details') ||
                                   txt.includes('fetch vehicle data') ||
                                   txt.includes('city where vehicle') ||
                                   txt.includes('city of registration') ||
                                   txt.includes('four wheeler') ||
                                   txt.includes('two wheeler');
                        }""")
                        if page_loaded:
                            log("info", f"  Motor page ready (attempt {_mot_wait+1})")
                            break
                    except Exception:
                        # Context destroyed by navigation — wait and retry
                        log("info", f"  Motor page still loading (attempt {_mot_wait+1})...")
                        pass
                    await page.wait_for_timeout(500)

                # ── STEP 1b: Click left-side slider (opens vehicle category modal) ─
                log("step", "STEP 1b — Click left slider to open category modal")
                # networkidle already fired above — no extra wait needed here
                # Click slider and verify modal opened (Four Wheeler/Two Wheeler visible)
                # If not opened, click again — nysa_openbtn toggles, so may need 2 clicks
                # NOTE: Each page.evaluate is wrapped in try/except because the portal
                # sometimes triggers a navigation during the click, which destroys the
                # execution context. That's expected — we just retry instead of failing.
                modal_open = False
                for _sb in range(25):
                    # ── Pre-check: is the modal already open? ──
                    # nysa_openbtn is a TOGGLE. If a previous click opened the modal
                    # but the verify eval crashed on a navigation, clicking again
                    # would CLOSE it. So always check state before clicking.
                    try:
                        modal_open = await page.evaluate("""() => {
                            for (const el of document.querySelectorAll('*')) {
                                const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                                if ((t === 'four wheeler' || t === 'two wheeler') &&
                                    el.getBoundingClientRect().width > 0) return true;
                            }
                            return false;
                        }""")
                    except Exception as e:
                        log("info", f"  (pre-check interrupted by navigation — retrying)")
                        await page.wait_for_timeout(800)
                        continue

                    if modal_open:
                        log("ok", "  -> Modal is open")
                        break

                    # ── Find slider button coordinates ──
                    try:
                        btn_coords = await page.evaluate("""() => {
                            const el = document.querySelector('.nysa_openbtn');
                            if (!el) return null;
                            const r = el.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0)
                                return {x: Math.round(r.left + r.width/2), y: Math.round(r.top + r.height/2)};
                            return null;
                        }""")
                    except Exception as e:
                        log("info", f"  (coord lookup interrupted by navigation — retrying)")
                        await page.wait_for_timeout(800)
                        continue

                    if btn_coords:
                        log("info", f"  nysa_openbtn at x={btn_coords['x']} y={btn_coords['y']} (attempt {_sb+1})")
                        try:
                            await page.mouse.click(btn_coords['x'], btn_coords['y'])
                        except Exception:
                            # Click itself can fail if page navigates mid-click
                            pass
                        # Let any navigation triggered by the click settle
                        try:
                            await page.wait_for_load_state("networkidle", timeout=3000)
                        except Exception:
                            pass
                        await page.wait_for_timeout(800)
                        log("info", f"  Checking if modal opened...")
                    else:
                        await page.wait_for_timeout(500)
                if not modal_open:
                    log("warn", "  Modal did not open after all attempts")

                # ── STEP 1c: Select Motor Category + Type from modal ──────
                # motor_category: "4w" → Four Wheeler, "2w" → Two Wheeler
                # motor_type: "Package" → Package tab, "TP" → TP tab
                motor_cat  = str(D.get("motor_category", "4w")).strip().lower()
                motor_type = str(D.get("motor_type", "Package")).strip()

                # Map category to modal label text
                cat_label_map = {
                    "4w": "Four Wheeler", "4 w": "Four Wheeler",
                    "four wheeler": "Four Wheeler", "fourwheeler": "Four Wheeler",
                    "2w": "Two Wheeler",  "2 w": "Two Wheeler",
                    "two wheeler": "Two Wheeler",   "twowheeler": "Two Wheeler",
                    "electric bike": "Electric Bike", "eb": "Electric Bike",
                    "gcv": "GCV", "pcv": "PCV", "misc d": "Misc D",
                }
                cat_text = cat_label_map.get(motor_cat, "Four Wheeler")

                # Click category using multiple strategies
                log("info", f"  Selecting category: {cat_text}")
                cat_clicked = False
                for _ct in range(5):
                    # Strategy 1: Playwright text locator (non-exact to handle "Two Wheeler\nPackage\nTP")
                    try:
                        loc = page.get_by_text(cat_text, exact=True).first
                        await loc.scroll_into_view_if_needed()
                        await page.wait_for_timeout(300)
                        box = await loc.bounding_box()
                        if box:
                            await page.mouse.click(box['x'] + box['width']/2, box['y'] + box['height']/2)
                            cat_clicked = True
                            log("ok", f"  -> Category '{cat_text}' clicked (locator+mouse)")
                            break
                    except Exception:
                        pass

                    # Strategy 2: XPath on exact text node
                    try:
                        loc2 = page.locator(f"xpath=//*[normalize-space(text())='{cat_text}']").first
                        box2 = await loc2.bounding_box()
                        if box2:
                            await page.mouse.click(box2['x'] + box2['width']/2, box2['y'] + box2['height']/2)
                            cat_clicked = True
                            log("ok", f"  -> Category '{cat_text}' clicked (xpath+mouse)")
                            break
                    except Exception:
                        pass

                    # Strategy 3: JS — find the INNERMOST element whose own text
                    # matches "Two Wheeler" (not subtree text), then click its parent card.
                    # This handles: <div class="card"><img/><span>Two Wheeler</span><div>Package TP</div></div>
                    try:
                        cat_coords = await page.evaluate("""(catText) => {
                            const lower = catText.toLowerCase().trim();

                            // Walk ALL elements, find ones whose DIRECT childNodes text matches
                            let bestEl = null, bestArea = Infinity;
                            for (const el of document.querySelectorAll('*')) {
                                // Get only DIRECT text content (not children's text)
                                let directText = '';
                                for (const child of el.childNodes) {
                                    if (child.nodeType === 3) directText += child.nodeValue;
                                }
                                directText = directText.trim().toLowerCase();

                                // Also check full innerText for leaf elements (no child elements with text)
                                const inner = (el.innerText||'').trim().toLowerCase();
                                const childEls = el.querySelectorAll('*');
                                const isLeaf = childEls.length === 0 ||
                                    Array.from(childEls).every(c => !(c.innerText||'').trim());

                                if (directText === lower || (isLeaf && inner === lower)) {
                                    const r = el.getBoundingClientRect();
                                    if (r.width > 0 && r.height > 0) {
                                        const area = r.width * r.height;
                                        if (area < bestArea) { bestArea = area; bestEl = el; }
                                    }
                                }
                            }
                            if (!bestEl) return null;

                            // Walk UP to find the clickable card container (has border/img/card class)
                            let clickTarget = bestEl;
                            let parent = bestEl.parentElement;
                            for (let i = 0; i < 5; i++) {
                                if (!parent) break;
                                const cls = (parent.className||'').toLowerCase();
                                const tag = parent.tagName.toLowerCase();
                                const style = window.getComputedStyle(parent);
                                const hasBorder = style.borderWidth && style.borderWidth !== '0px';
                                if (cls.includes('card') || cls.includes('item') ||
                                    cls.includes('tile') || cls.includes('product') ||
                                    tag === 'mat-card' || tag === 'a' ||
                                    parent.querySelector('img') || hasBorder) {
                                    clickTarget = parent;
                                    break;
                                }
                                parent = parent.parentElement;
                            }
                            clickTarget.scrollIntoView({block:'center'});
                            const r = clickTarget.getBoundingClientRect();
                            return {x: r.left + r.width/2, y: r.top + r.height/2};
                        }""", cat_text)
                        if cat_coords:
                            await page.wait_for_timeout(300)
                            await page.mouse.click(cat_coords['x'], cat_coords['y'])
                            cat_clicked = True
                            log("ok", f"  -> Category '{cat_text}' clicked (JS card at {cat_coords['x']:.0f},{cat_coords['y']:.0f})")
                            break
                    except Exception:
                        pass
                    await page.wait_for_timeout(800)
                if not cat_clicked:
                    log("warn", f"  Category '{cat_text}' not found in modal — click it manually")
                    await pause(f"Click '{cat_text}' in the modal, then press ENTER")
                await page.wait_for_timeout(1000)

                # Click Package or TP badge nearest to the category label (by X position)
                # After clicking "Two Wheeler", find the Package/TP badge closest to its center X
                log("info", f"  Selecting type: {motor_type}")
                type_clicked = await page.evaluate("""([catText, typeText]) => {
                    // Step 1: find the category label and get its center X
                    let catCenterX = null, catCenterY = null;
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim();
                        if (t.toLowerCase() === catText.toLowerCase()) {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0 && r.width < 300) {
                                catCenterX = r.left + r.width / 2;
                                catCenterY = r.top + r.height / 2;
                                break;
                            }
                        }
                    }
                    if (catCenterX === null) return false;

                    // Step 2: find ALL elements whose exact text is typeText
                    const candidates = [];
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim();
                        if (t.toLowerCase() !== typeText.toLowerCase()) continue;
                        const r = el.getBoundingClientRect();
                        if (r.width === 0 || r.height === 0) continue;
                        const cx = r.left + r.width / 2;
                        const cy = r.top + r.height / 2;
                        // Must be below the category label (badges are below the icon+label)
                        if (cy > catCenterY - 10) {
                            candidates.push({el, cx, cy, dist: Math.abs(cx - catCenterX)});
                        }
                    }
                    if (candidates.length === 0) return false;

                    // Step 3: click the one with smallest horizontal distance to category center
                    candidates.sort((a, b) => a.dist - b.dist);
                    candidates[0].el.scrollIntoView({block: 'center'});
                    candidates[0].el.click();
                    return true;
                }""", [cat_text, motor_type])
                if type_clicked:
                    log("ok", f"  -> Type '{motor_type}' clicked (nearest to '{cat_text}')")
                else:
                    log("warn", f"  Type '{motor_type}' not found near '{cat_text}' card")

                await page.wait_for_timeout(2000)
                steps.append("motor")

                # ── STEP 1d: SAOD — Standalone OD Policy ─────────────
                # If saod == "yes" in Excel, click the button beside "Standalone OD Policy"
                if D.get("saod", "").strip().lower() == "yes":
                    log("step", "STEP 1d — Standalone OD Policy (SAOD)")
                    saod_clicked = False
                    # Wait up to 10 attempts for Standalone OD Policy to appear
                    for _saod_wait in range(20):
                        saod_visible = await page.evaluate("""() => {
                            for (const el of document.querySelectorAll('*')) {
                                const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                                if (t.includes('standalone od') || t.includes('standalone od policy') || t.includes('saod')) {
                                    const r = el.getBoundingClientRect();
                                    if (r.width > 0) return true;
                                }
                            }
                            return false;
                        }""")
                        if saod_visible:
                            break
                        await page.wait_for_timeout(500)
                    # Find the clickable element (button/toggle) beside "Standalone OD Policy" text
                    saod_coords = await page.evaluate("""() => {
                        // Find label element for Standalone OD
                        let labelEl = null;
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if ((t === 'standalone od policy' || t === 'standalone od' || t === 'saod') &&
                                el.getBoundingClientRect().width > 0) {
                                labelEl = el;
                                break;
                            }
                        }
                        if (!labelEl) return null;
                        // Look for nearest button/toggle/checkbox to the label
                        let node = labelEl.parentElement;
                        for (let i = 0; i < 6; i++) {
                            if (!node) break;
                            const btn = node.querySelector(
                                'button, input[type="checkbox"], input[type="radio"], ' +
                                '[class*="toggle"], [class*="switch"], [class*="btn"]'
                            );
                            if (btn) {
                                btn.scrollIntoView({block: 'center'});
                                const r = btn.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                            node = node.parentElement;
                        }
                        // Fallback: click the label itself
                        labelEl.scrollIntoView({block: 'center'});
                        const r = labelEl.getBoundingClientRect();
                        return {x: r.left+r.width/2, y: r.top+r.height/2};
                    }""")
                    if saod_coords:
                        await page.mouse.click(saod_coords['x'], saod_coords['y'])
                        log("ok", f"  -> Standalone OD Policy (SAOD) clicked at ({saod_coords['x']:.0f},{saod_coords['y']:.0f})")
                        saod_clicked = True
                    if not saod_clicked:
                        log("warn", "  SAOD button not found — click it manually")
                        await pause("Click the Standalone OD Policy button manually, then press ENTER")
                    await page.wait_for_timeout(1000)
                    steps.append("saod")

                # ── STEP 2: Select Policy Type (New / Rollover / Used) ───
                policy_type_val = D.get("policy_type", "New").strip().lower()
                log("step", f"STEP 2 — Policy Type: {D.get('policy_type','New')}")

                if policy_type_val == "new":
                    await smart_click(page,
                        ["New", "New Policy", "New Business", "New Proposal"],
                        "New Policy")
                elif policy_type_val == "rollover":
                    await smart_click(page,
                        ["Rollover", "Roll Over"],
                        "Rollover")
                elif policy_type_val in ("used", "used car", "used vehicle"):
                    await smart_click(page,
                        ["Used", "Used Car", "Used Vehicle"],
                        "Used")
                else:
                    # fallback to New
                    log("warn", f"  Unknown policy_type '{D.get('policy_type','')}' — defaulting to New")
                    await smart_click(page,
                        ["New", "New Policy", "New Business", "New Proposal"],
                        "New Policy")

                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)
                steps.append("new")

                # ── STEP 3: Registration Number & Fetch ───────────────
                # New     → skip reg number, go directly to vehicle details
                # Rollover/Used → type the actual reg number from Excel
                log("step", "STEP 3 — Enter Registration Number + Fetch Vehicle Details")

                if policy_type_val == "new":
                    # For New policy: portal already shows "NEW" in reg field
                    # and vehicle details form is already visible — skip reg + fetch
                    log("info", "  Policy Type = New → skipping Registration Number & Fetch")
                    # Just wait for the vehicle details form to be ready
                    try:
                        await page.wait_for_load_state("networkidle", timeout=10000)
                    except Exception:
                        pass
                    await page.wait_for_timeout(2000)
                    # Verify vehicle details form is visible
                    for _vd_wait in range(20):
                        try:
                            vd_ready = await page.evaluate("""() => {
                                const txt = (document.body.innerText||'').toLowerCase();
                                return txt.includes('city of registration') ||
                                       txt.includes('city where vehicle') ||
                                       txt.includes('select make') ||
                                       txt.includes('manufacturer') ||
                                       txt.includes('vehicle details');
                            }""")
                            if vd_ready:
                                log("ok", f"  Vehicle Details form ready (attempt {_vd_wait+1})")
                                break
                        except Exception:
                            pass
                        await page.wait_for_timeout(500)
                else:
                    # Rollover / Used — type the actual reg number
                    reg_to_enter = D.get("rollover_reg_number", "").strip()
                    if not reg_to_enter:
                        # Fallback: use vehicle_reg_number column (header "Registration Number")
                        reg_to_enter = D.get("vehicle_reg_number", "").strip()
                    if not reg_to_enter:
                        log("warn", "  rollover_reg_number is empty in Excel — please fill it manually")
                        await pause("Enter Registration Number in browser manually, then press ENTER")
                        reg_to_enter = ""

                    if reg_to_enter:
                        await fill_field(page, [
                            "input[placeholder*='Registration No' i]",
                            "input[placeholder*='Registration Number' i]",
                            "input[placeholder*='Enter Registration' i]",
                            "input[formcontrolname*='registrationNumber' i]",
                            "input[formcontrolname*='regNo' i]",
                            "input[formcontrolname*='regNumber' i]",
                        ], reg_to_enter, "Registration Number")
                        await page.wait_for_timeout(500)

                    await smart_click(page,
                        ["Fetch Vehicle Data", "Fetch Vehicle Details", "Fetch Details"],
                        "Fetch Vehicle Data")
                    await page.wait_for_timeout(2500)

                steps.append("fetch_details")

            else:
                # Skipping STEP 1-3: already on details page after Another Policy
                log("step", "Skipping STEP 1-3 (resuming from details page)")
                steps.append("motor"); steps.append("new"); steps.append("fetch_details")
                try:
                    await page.wait_for_load_state("networkidle", timeout=10000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)


            # ══════════════════════════════════════════════════════
            # STEP 4 — Fill Vehicle Details (6 fields only)
            #   1. City of registration     [autocomplete - 1st input]
            #   2. Select make & model      [autocomplete - 2nd input]
            #   3. Customer State           [auto-fills or select]
            #   4. Year of Manufacture      [select dropdown]
            #   5. CPA Tenure               [select dropdown]
            #   6. Current Registration Type [select dropdown]
            # ══════════════════════════════════════════════════════
            log("step", "STEP 4 — Filling Vehicle Details")

            # ── Helper: find the Nth visible text input in Vehicle Details section ──
            async def get_vehicle_input(index):
                """Get the Nth (0-based) visible text input INSIDE the Vehicle Details card only."""
                coords = await page.evaluate("""(idx) => {
                    // Find the "Vehicle Details" container first
                    let vehicleCard = null;
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim();
                        if (t.startsWith('Vehicle Details') && el.tagName !== 'BODY' && el.tagName !== 'HTML') {
                            const r = el.getBoundingClientRect();
                            // Must be a card-like container (not a tiny label)
                            if (r.width > 400 && r.height > 200) {
                                vehicleCard = el;
                                break;
                            }
                        }
                    }
                    // Fallback: search the whole page but skip inputs above Vehicle Details
                    const searchRoot = vehicleCard || document.body;
                    
                    const inputs = Array.from(searchRoot.querySelectorAll('input'))
                        .filter(inp => {
                            const r = inp.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) return false;
                            const t = (inp.type||'').toLowerCase();
                            if (['checkbox','radio','hidden','submit','button','date'].includes(t)) return false;
                            if (inp.readOnly || inp.disabled) return false;
                            // Skip the Registration Number input (it has value "NEW" or is near that label)
                            const val = (inp.value||'').trim().toUpperCase();
                            if (val === 'NEW') return false;
                            // Skip date inputs (value looks like DD/MM/YYYY)
                            if (/^[0-9]{2}[/][0-9]{2}[/][0-9]{4}$/.test(inp.value||'')) return false;
                            return true;
                        });
                    if (idx >= inputs.length) return null;
                    const inp = inputs[idx];
                    inp.scrollIntoView({block:'center'});
                    const r = inp.getBoundingClientRect();
                    return {x: r.left+r.width/2, y: r.top+r.height/2};
                }""", index)
                return coords

            # ── Helper: type into input at coords, wait for dropdown, scroll & pick match ──
            async def type_and_pick_at(coords, value, label):
                """Click input at coords, type search term, scroll dropdown, pick match."""
                search_term = value.split("-")[0].strip()
                log("info", f"  [{label}]: {value}")

                # Click, clear, type
                await page.mouse.click(coords['x'], coords['y'])
                await page.wait_for_timeout(200)
                await page.keyboard.press("Control+a")
                await page.keyboard.press("Delete")
                await page.wait_for_timeout(200)
                await page.keyboard.type(search_term, delay=120)
                log("fill", f"  Typed: '{search_term}'")
                await page.wait_for_timeout(1500)

                # Scroll dropdown and match — up to 10 attempts
                full_val = value.strip().upper()
                for _scroll in range(10):
                    result = await page.evaluate("""(v) => {
                        const val = v.trim().toUpperCase();
                        const opts = document.querySelectorAll(
                            'mat-option, [role="option"], .cdk-overlay-container mat-option'
                        );
                        if (opts.length === 0) return 'NO_OPTIONS';
                        for (const o of opts) {
                            const t = (o.innerText||o.textContent||'').trim().toUpperCase();
                            if (t === val) { o.scrollIntoView({block:'center'}); o.click(); return 'EXACT||'+o.innerText.trim(); }
                        }
                        for (const o of opts) {
                            const t = (o.innerText||o.textContent||'').trim().toUpperCase();
                            if (t.startsWith(val) || val.startsWith(t)) { o.scrollIntoView({block:'center'}); o.click(); return 'PARTIAL||'+o.innerText.trim(); }
                        }
                        return 'NO_MATCH||' + opts.length;
                    }""", full_val)

                    if result.startswith("EXACT") or result.startswith("PARTIAL"):
                        log("ok", f"  Selected: {result.split('||')[1]}")
                        await page.wait_for_timeout(500)
                        return True
                    elif result == "NO_OPTIONS":
                        if _scroll == 0:
                            log("info", f"  Dropdown not ready, waiting...")
                            await page.wait_for_timeout(1000)
                            continue
                        log("warn", f"  Dropdown did not appear for '{label}'")
                        return False
                    else:
                        # Scroll dropdown slowly
                        await page.evaluate("""() => {
                            const panel = document.querySelector(
                                '.mat-autocomplete-panel, .cdk-overlay-pane, [role="listbox"]'
                            );
                            if (panel) panel.scrollBy({top: 200, behavior: 'smooth'});
                            const opts = document.querySelectorAll('mat-option, [role="option"]');
                            if (opts.length > 0) opts[opts.length-1].scrollIntoView({block:'end', behavior:'smooth'});
                        }""")
                        log("info", f"  Scrolling dropdown... ({_scroll+1}/10)")
                        await page.wait_for_timeout(1000)

                log("warn", f"  No match for '{value}' after scrolling")
                return False

            # ── Helper: click a dropdown by label, pick value from overlay ──
            async def select_dropdown(label_text, value):
                """Find label text, click below it to open dropdown, pick option."""
                log("info", f"  [{label_text}]: {value}")

                # Dismiss any open overlay first
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(500)
                # Click on a neutral area (page title) to remove focus
                await page.mouse.click(400, 200)
                await page.wait_for_timeout(300)

                # Find label bounding box
                try:
                    label_loc = page.get_by_text(label_text, exact=False).first
                    await label_loc.wait_for(state="visible", timeout=3000)
                    box = await label_loc.bounding_box()
                    if not box:
                        log("warn", f"  Label '{label_text}' not visible")
                        return False
                except Exception as e:
                    log("warn", f"  Label '{label_text}' not found: {e}")
                    return False

                # Check if dropdown already shows the correct value
                # The dropdown value text is displayed below the label
                current_val = await page.evaluate("""(pos) => {
                    // Find element at the dropdown position (below label)
                    const el = document.elementFromPoint(pos.x, pos.y);
                    if (!el) return '';
                    // Walk up to find the mat-select or select and read its displayed value
                    let node = el;
                    for (let i = 0; i < 5; i++) {
                        if (!node) break;
                        // mat-select shows value in .mat-select-value-text
                        const valEl = node.querySelector('.mat-select-value-text, .mat-select-value');
                        if (valEl) return (valEl.innerText||'').trim();
                        // native select
                        if (node.tagName === 'SELECT') return node.options[node.selectedIndex]?.text?.trim() || '';
                        node = node.parentElement;
                    }
                    return (el.innerText||el.value||'').trim();
                }""", {"x": box['x'] + box['width']/2, "y": box['y'] + box['height'] + 25})

                if current_val.lower() == value.strip().lower():
                    log("ok", f"  Already set to '{current_val}' — skipping")
                    return True

                # Click below the label to open dropdown
                click_x = box['x'] + box['width'] / 2
                click_y = box['y'] + box['height'] + 25
                await page.mouse.click(click_x, click_y)
                log("info", f"  Opened dropdown (current: '{current_val}')")
                await page.wait_for_timeout(800)

                # Pick option from the overlay
                v = value.strip()
                selected = False
                for try_val in [v, v.title(), v.upper(), v.lower()]:
                    if selected:
                        break
                    try:
                        opt = page.get_by_text(try_val, exact=True).first
                        await opt.wait_for(state="visible", timeout=1000)
                        await opt.click()
                        log("ok", f"  -> {try_val}")
                        selected = True
                    except Exception:
                        continue

                if not selected:
                    # JS fallback
                    try:
                        result = await page.evaluate("""(val) => {
                            const v = val.toLowerCase();
                            const opts = document.querySelectorAll('mat-option, [role="option"]');
                            for (const o of opts) {
                                const t = (o.innerText||o.textContent||'').trim();
                                if (t.toLowerCase() === v) { o.click(); return 'OK||' + t; }
                            }
                            for (const o of opts) {
                                const t = (o.innerText||o.textContent||'').trim();
                                if (t.toLowerCase().includes(v) || v.includes(t.toLowerCase())) { o.click(); return 'OK||' + t; }
                            }
                            return 'FAIL||' + Array.from(opts).map(o=>(o.innerText||'').trim()).filter(t=>t).join(', ');
                        }""", v)
                        if result.startswith("OK"):
                            log("ok", f"  -> {result.split('||')[1]}")
                            selected = True
                        else:
                            log("warn", f"  '{v}' not in: {result.split('||')[1][:100] if '||' in result else ''}")
                    except Exception as e:
                        log("warn", f"  Error: {e}")

                # Close dropdown: Escape + click neutral area
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(300)
                await page.mouse.click(400, 200)
                await page.wait_for_timeout(500)
                return selected


            # ────────────────────────────────────────────────────────
            # 1. City of registration (1st text input — autocomplete)
            # ────────────────────────────────────────────────────────
            if D.get("rto"):
                c1 = await get_vehicle_input(0)
                if c1:
                    for _try in range(3):
                        if await type_and_pick_at(c1, D["rto"], "City of registration"):
                            break
                        c1 = await get_vehicle_input(0)  # re-read coords
                        if not c1: break
                        await page.wait_for_timeout(1000)
                else:
                    log("warn", "  City of registration input not found")
                await page.wait_for_timeout(1000)

            # ────────────────────────────────────────────────────────
            # 2. Select make & model (2nd text input — autocomplete)
            # ────────────────────────────────────────────────────────
            if D.get("manufacturer_model"):
                c2 = await get_vehicle_input(1)
                if c2:
                    for _try in range(3):
                        if await type_and_pick_at(c2, D["manufacturer_model"], "Select make & model"):
                            break
                        c2 = await get_vehicle_input(1)
                        if not c2: break
                        await page.wait_for_timeout(1000)
                else:
                    log("warn", "  Make & Model input not found")
                await page.wait_for_timeout(2000)

            # ────────────────────────────────────────────────────────
            # 3. Customer State (auto-fills from RTO)
            # ────────────────────────────────────────────────────────
            state_name = D.get("rto", "").split("-")[0].strip().title()
            if state_name:
                already = await page.evaluate("""(sn) => {
                    for (const inp of document.querySelectorAll('input')) {
                        if ((inp.value||'').trim().toUpperCase() === sn.toUpperCase()) return true;
                    }
                    return false;
                }""", state_name)
                if already:
                    log("ok", f"  Customer State: {state_name} (auto-filled)")
                else:
                    await select_dropdown("Customer State", state_name)
                await page.wait_for_timeout(500)

            # ────────────────────────────────────────────────────────
            # 4. Year of Manufacture (select dropdown)
            # ────────────────────────────────────────────────────────
            if D.get("manufacturing_year"):
                mfg_year = str(D["manufacturing_year"]).split(".")[0].strip()
                await select_dropdown("Year of Manufacture", mfg_year)
                await page.wait_for_timeout(500)

            # ────────────────────────────────────────────────────────
            # 5. CPA Tenure (select dropdown)
            # ────────────────────────────────────────────────────────
            if D.get("tenure"):
                await select_dropdown("CPA Tenure", D["tenure"].strip())
                await page.wait_for_timeout(500)

            # ────────────────────────────────────────────────────────
            # 6. Current Registration Type (select dropdown)
            # ────────────────────────────────────────────────────────
            if D.get("current_reg_type"):
                reg = D["current_reg_type"].strip()
                ok = await select_dropdown("Current Registration Type", reg)
                if not ok:
                    await select_dropdown("Current Registration Type", reg.title())
                await page.wait_for_timeout(500)

            steps.append("fill_details")
            log("ok", "══ SECTION DONE: RTO / Vehicle Details filled ══")
            await page.wait_for_timeout(1000)


            # ── STEP 5: Click GET QUOTE ──────────────────────────────
            log("step", "STEP 5 — GET QUOTE")
            # Wait for page to fully settle after all field fills
            log("info", "Waiting for page to settle before GET QUOTE...")
            try:
                await page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(1000)

            async def click_by_text(texts):
                """Click any visible element matching text — works for any tag."""
                if isinstance(texts, str): texts = [texts]
                for text in texts:
                    for tag in ["button", "input", "a", "span", "div"]:
                        try:
                            loc = page.locator(f"{tag}:has-text('{text}')").first
                            await loc.wait_for(state="visible", timeout=2000)
                            await loc.scroll_into_view_if_needed()
                            await loc.click()
                            log("ok", f"  -> Clicked '{text}' via <{tag}>")
                            return True
                        except Exception:
                            continue
                    # JS scan — find smallest matching element and mouse.click it
                    try:
                        coords = await page.evaluate(f"""() => {{
                            let best = null, bestArea = Infinity;
                            for (const el of document.querySelectorAll('*')) {{
                                const rect = el.getBoundingClientRect();
                                if (rect.width === 0 || rect.height === 0) continue;
                                const txt = (el.innerText || el.value || '').trim();
                                if (txt === '{text}') {{
                                    const area = rect.width * rect.height;
                                    if (area < bestArea) {{ bestArea = area; best = el; }}
                                }}
                            }}
                            if (best) {{
                                best.scrollIntoView({{block:'center'}});
                                const r = best.getBoundingClientRect();
                                return {{x: r.left+r.width/2, y: r.top+r.height/2}};
                            }}
                            return null;
                        }}""")
                        if coords:
                            await page.wait_for_timeout(300)
                            await page.mouse.click(coords['x'], coords['y'])
                            log("ok", f"  -> Clicked '{text}' via JS mouse.click at ({coords['x']:.0f},{coords['y']:.0f})")
                            return True
                    except Exception:
                        pass
                return False

            if not await click_by_text(["Get Quote", "GET QUOTE", "Get Quotation"]):
                log("warn", "GET QUOTE not found — check browser")
                await pause("Click 'GET QUOTE' manually, then press ENTER")
            try:
                await page.wait_for_load_state("networkidle", timeout=30000)
            except Exception:
                pass
            await page.wait_for_timeout(5000)
            steps.append("get_quote")

            # ── IDV fill — after Get Quote (portal shows IDV input on the quote/customize page) ──
            if D.get("idv"):
                idv_val = str(D["idv"]).strip().split(".")[0]
                log("info", f"Setting IDV (post-Get Quote): {idv_val}")

                # Wait for IDV input to appear (up to 20 x 500ms = 10s)
                idv_found = False
                for _attempt in range(20):
                    idv_found = await page.evaluate("""() => {
                        let inp = document.getElementById('sirangeinput');
                        if (!inp) inp = document.querySelector('input[name="amountInput"]');
                        if (!inp) {
                            const slider = document.querySelector(
                                'input[type="range"], input.custom-range');
                            if (!slider) return false;
                            const all = Array.from(document.querySelectorAll('input'));
                            const idx = all.indexOf(slider);
                            for (let i = idx - 1; i >= 0; i--) {
                                const c = all[i];
                                if (['range','checkbox','radio','hidden','submit','button']
                                    .includes(c.type)) continue;
                                if (c.getBoundingClientRect().width === 0) continue;
                                inp = c; break;
                            }
                        }
                        if (!inp) return false;
                        inp.scrollIntoView({block:'center'});
                        return true;
                    }""")
                    if idv_found:
                        log("fill", f"IDV input found (attempt {_attempt+1})")
                        break
                    await page.wait_for_timeout(500)

                # Wait for default value to populate before overwriting
                if idv_found:
                    for _dv in range(20):
                        idv_current = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) {
                                const slider = document.querySelector(
                                    'input[type="range"], input.custom-range');
                                if (!slider) return '';
                                const all = Array.from(document.querySelectorAll('input'));
                                const idx = all.indexOf(slider);
                                for (let i = idx - 1; i >= 0; i--) {
                                    const c = all[i];
                                    if (['range','checkbox','radio','hidden','submit','button']
                                        .includes(c.type)) continue;
                                    if (c.getBoundingClientRect().width === 0) continue;
                                    return c.value;
                                }
                            }
                            return inp ? inp.value : '';
                        }""") or ""
                        if idv_current.strip() and idv_current.strip() not in ("0", ""):
                            log("fill", f"IDV default populated: {idv_current.strip()}")
                            break
                        await page.wait_for_timeout(500)

                if not idv_found:
                    log("warn", "IDV input not found on quote page — skipping IDV fill")
                else:
                    await page.wait_for_timeout(400)
                    # Single fill attempt — the guard loop below re-applies only if Angular resets it
                    idv_coords = await page.evaluate("""() => {
                        let inp = document.getElementById('sirangeinput');
                        if (!inp) inp = document.querySelector('input[name="amountInput"]');
                        if (!inp) {
                            const slider = document.querySelector(
                                'input[type="range"], input.custom-range');
                            if (!slider) return null;
                            const all = Array.from(document.querySelectorAll('input'));
                            const idx = all.indexOf(slider);
                            for (let i = idx - 1; i >= 0; i--) {
                                const c = all[i];
                                if (['range','checkbox','radio','hidden','submit','button']
                                    .includes(c.type)) continue;
                                if (c.getBoundingClientRect().width === 0) continue;
                                inp = c; break;
                            }
                        }
                        if (!inp) return null;
                        inp.scrollIntoView({block:'center'});
                        const r = inp.getBoundingClientRect();
                        return {x: r.left + r.width/2, y: r.top + r.height/2};
                    }""")
                    if idv_coords:
                        await page.mouse.click(idv_coords['x'], idv_coords['y'], click_count=3)
                        await page.wait_for_timeout(150)
                        await page.keyboard.press("Control+a")
                        await page.keyboard.press("Delete")
                        await page.wait_for_timeout(80)
                        await page.keyboard.type(idv_val, delay=80)
                        await page.wait_for_timeout(400)
                        # Commit via Angular native input setter
                        await page.evaluate("""(val) => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) {
                                const slider = document.querySelector(
                                    'input[type="range"], input.custom-range');
                                if (!slider) return;
                                const all = Array.from(document.querySelectorAll('input'));
                                const idx = all.indexOf(slider);
                                for (let i = idx - 1; i >= 0; i--) {
                                    const c = all[i];
                                    if (['range','checkbox','radio','hidden','submit','button']
                                        .includes(c.type)) continue;
                                    if (c.getBoundingClientRect().width === 0) continue;
                                    inp = c; break;
                                }
                            }
                            if (!inp) return;
                            const setter = Object.getOwnPropertyDescriptor(
                                window.HTMLInputElement.prototype, 'value').set;
                            setter.call(inp, ''); setter.call(inp, val);
                            inp.dispatchEvent(new Event('input', {bubbles: true}));
                            inp.dispatchEvent(new Event('change', {bubbles: true}));
                        }""", idv_val)
                        await page.wait_for_timeout(600)
                        actual = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            return inp ? inp.value : '';
                        }""") or ""
                        if actual.strip() == idv_val:
                            log("ok", f"IDV confirmed: {actual.strip()}")
                        else:
                            log("warn", f"IDV got '{actual.strip()}', expected '{idv_val}' — guard will re-apply if Angular resets")

                    # IDV Guard: re-apply up to 10x if Angular resets it
                    for _guard in range(10):
                        current_idv = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            return inp ? inp.value : '';
                        }""") or ""
                        if str(current_idv).strip() == idv_val:
                            await page.wait_for_timeout(600)
                            recheck = await page.evaluate("""() => {
                                let inp = document.getElementById('sirangeinput');
                                if (!inp) inp = document.querySelector('input[name="amountInput"]');
                                return inp ? inp.value : '';
                            }""") or ""
                            if str(recheck).strip() == idv_val:
                                log("ok", f"IDV stable: {recheck}")
                                break
                        log("warn", f"IDV reverted to '{current_idv}' — re-applying {idv_val}")
                        idv_gc = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) return null;
                            inp.scrollIntoView({block:'center'});
                            const r = inp.getBoundingClientRect();
                            return {x: r.left+r.width/2, y: r.top+r.height/2};
                        }""")
                        if idv_gc:
                            await page.mouse.click(idv_gc['x'], idv_gc['y'], click_count=3)
                            await page.wait_for_timeout(150)
                            await page.keyboard.press("Delete")
                            await page.keyboard.type(idv_val, delay=80)
                            await page.wait_for_timeout(500)
                        else:
                            break

                    # Click "Update" button after IDV is set
                    log("info", "Clicking Update button after IDV...")
                    update_clicked = False
                    for _upd in range(5):
                        try:
                            upd = page.get_by_text("Update", exact=True).first
                            await upd.wait_for(state="visible", timeout=2000)
                            await upd.click()
                            log("ok", "  -> Update clicked")
                            update_clicked = True
                            break
                        except Exception:
                            pass
                        # JS fallback
                        try:
                            coords = await page.evaluate("""() => {
                                for (const el of document.querySelectorAll('button, input[type="submit"], a')) {
                                    const t = (el.innerText||el.value||'').trim();
                                    if (t === 'Update' || t === 'UPDATE') {
                                        el.scrollIntoView({block:'center'});
                                        const r = el.getBoundingClientRect();
                                        if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                    }
                                }
                                return null;
                            }""")
                            if coords:
                                await page.mouse.click(coords['x'], coords['y'])
                                log("ok", "  -> Update clicked (JS)")
                                update_clicked = True
                                break
                        except Exception:
                            pass
                        await page.wait_for_timeout(500)
                    if not update_clicked:
                        log("warn", "  Update button not found — click manually")
                    try:
                        await page.wait_for_load_state("networkidle", timeout=15000)
                    except Exception:
                        pass
                    await page.wait_for_timeout(2000)
            log("step", "STEP 6 — Open Add/Remove Cover from Basic Plan card")
            # Wait for Plan Selection page to fully load
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass

            # Wait up to 25x for Basic Plan card to appear
            for _cust in range(25):
                found = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText || el.textContent || '').trim();
                        if (t.toLowerCase().includes('basic plan') && t.length < 80) return true;
                    }
                    return false;
                }""")
                if found:
                    log("info", f"  Basic Plan card found (attempt {_cust+1})")
                    break
                await page.wait_for_timeout(500)

            # Click "Add/Remove Cover" link inside the Basic Plan card
            add_remove_clicked = await page.evaluate("""() => {
                // Find the Basic Plan card first, then locate Add/Remove Cover within it
                for (const card of document.querySelectorAll('*')) {
                    const cardText = (card.innerText || card.textContent || '').trim().toLowerCase();
                    if (!cardText.includes('basic plan')) continue;
                    // Search for Add/Remove Cover link inside this card
                    for (const el of card.querySelectorAll('a, button, span')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if (t.includes('add') && t.includes('remove') && t.includes('cover')) {
                            el.scrollIntoView({block: 'center'});
                            el.click();
                            return true;
                        }
                    }
                }
                // Fallback: click any "Add/Remove Cover" visible on page
                for (const el of document.querySelectorAll('a, button, span')) {
                    const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                    if (t.includes('add') && t.includes('remove') && t.includes('cover')) {
                        const r = el.getBoundingClientRect();
                        if (r.width > 0 && r.height > 0) {
                            el.scrollIntoView({block: 'center'});
                            el.click();
                            return true;
                        }
                    }
                }
                return false;
            }""")

            if add_remove_clicked:
                log("ok", "  -> Add/Remove Cover clicked on Basic Plan card")
            else:
                log("warn", "  Add/Remove Cover not found — click it manually on the Basic Plan card")
                await pause("Click 'Add/Remove Cover' on the Basic Plan card, then press ENTER")

            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            # Wait for Additional Covers / Zero Depreciation page to load
            for _ in range(40):
                if await page.evaluate("() => document.body.innerText.includes('Zero Depreciation')"):
                    log("ok", "  -> Additional Covers page loaded"); break
                await page.wait_for_timeout(500)
            else:
                log("warn", "  Additional Covers page timed out — continuing")
            await page.wait_for_timeout(1000)
            steps.append("customize_opened")

            # ── STEP 7: Tick only the add-ons listed in Excel ──────────
            desired_addons = [a.strip() for a in D.get("addons", []) if a.strip()]
            log("step", f"STEP 7 — Add-ons selection: {desired_addons or ['(none)']}")

            for addon_name in desired_addons:
                log("info", f"Looking for '{addon_name}' on page...")

                # Use Playwright get_by_text to find the EXACT label on the page
                addon_clicked = False

                # Strategy 1: Playwright get_by_text — most reliable
                for exact in [True, False]:
                    try:
                        loc = page.get_by_text(addon_name, exact=exact)
                        count = await loc.count()
                        if count > 0:
                            # Found the label — now click the row/checkbox area
                            el = loc.first
                            await el.scroll_into_view_if_needed()
                            await page.wait_for_timeout(300)

                            # The checkbox is typically to the LEFT of the label text.
                            # Get the label's bounding box, then click ~20px to the left
                            # of its left edge (where the checkbox square is).
                            box = await el.bounding_box()
                            if box:
                                # Click the checkbox square area (left of the label)
                                click_x = box['x'] - 20
                                click_y = box['y'] + box['height'] / 2
                                # Ensure x is not negative
                                if click_x < 5:
                                    click_x = box['x'] + 5
                                await page.mouse.click(click_x, click_y)
                                log("ok", f"  -> Clicked checkbox area for '{addon_name}' at ({click_x:.0f},{click_y:.0f})")
                                addon_clicked = True
                                await page.wait_for_timeout(800)
                                break
                    except Exception as e:
                        log("info", f"  get_by_text({'exact' if exact else 'partial'}) failed: {e}")

                # Strategy 2: If Strategy 1 didn't work, try JS to find text and click its row
                if not addon_clicked:
                    log("info", f"  Trying JS text search for '{addon_name}'...")
                    click_coords = await page.evaluate("""(addonName) => {
                        const lower = addonName.toLowerCase();
                        // Find all elements whose text matches
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText || el.textContent || '').trim();
                            if (t.toLowerCase().includes(lower) && t.length < 100) {
                                // Check this isn't a huge container
                                const childCount = el.querySelectorAll('*').length;
                                if (childCount > 20) continue;
                                const r = el.getBoundingClientRect();
                                if (r.width === 0 || r.height === 0) continue;
                                el.scrollIntoView({block: 'center'});
                                const r2 = el.getBoundingClientRect();
                                // Click left of the text where checkbox square is
                                return {
                                    x: Math.max(r2.left - 20, 5),
                                    y: r2.top + r2.height / 2
                                };
                            }
                        }
                        return null;
                    }""", addon_name)

                    if click_coords:
                        await page.wait_for_timeout(300)
                        await page.mouse.click(click_coords['x'], click_coords['y'])
                        log("ok", f"  -> Clicked via JS at ({click_coords['x']:.0f},{click_coords['y']:.0f})")
                        addon_clicked = True
                        await page.wait_for_timeout(800)

                # Strategy 3: Click the label text itself (some Angular checkboxes
                # toggle when you click anywhere on the row)
                if not addon_clicked:
                    log("info", f"  Trying direct label click for '{addon_name}'...")
                    try:
                        loc = page.get_by_text(addon_name, exact=False)
                        if await loc.count() > 0:
                            await loc.first.click()
                            log("ok", f"  -> Direct label click for '{addon_name}'")
                            addon_clicked = True
                            await page.wait_for_timeout(800)
                    except Exception:
                        pass

                if not addon_clicked:
                    log("warn", f"  '{addon_name}' not found — tick it manually")
                    await pause(f"Tick '{addon_name}' manually, then press ENTER")

                # ── After ticking, check if a "Select Plan" dropdown appeared ──
                # Some covers (e.g. Zero Depreciation) show a plan dropdown (Silver/Gold etc.)
                # Auto-select "Silver" if the dropdown appears.
                if addon_clicked:
                    await page.wait_for_timeout(1000)
                    log("info", f"  Checking for 'Select Plan' dropdown...")

                    select_plan_found = False
                    try:
                        # Look for a Select/dropdown near the addon row
                        sel_loc = page.get_by_text("Select Plan", exact=False)
                        if await sel_loc.count() > 0:
                            select_plan_found = True
                        else:
                            # Also try looking for a <select> or dropdown with "Select" text
                            sel_loc2 = page.locator("select, [class*='dropdown'], [class*='select']")
                            for i in range(await sel_loc2.count()):
                                el = sel_loc2.nth(i)
                                txt = await el.inner_text()
                                if 'select' in txt.lower():
                                    select_plan_found = True
                                    break
                    except Exception:
                        pass

                    if select_plan_found:
                        log("info", f"  'Select Plan' dropdown found — selecting Silver")
                        dd_clicked = False

                        # Method 1: Find a <select> element and use Playwright select_option
                        selects = page.locator("select")
                        for si in range(await selects.count()):
                            sel_el = selects.nth(si)
                            if await sel_el.is_visible():
                                try:
                                    await sel_el.select_option(label="Silver")
                                    log("ok", f"  -> Selected 'Silver' via <select>")
                                    dd_clicked = True
                                    break
                                except Exception:
                                    try:
                                        await sel_el.select_option(value="Silver")
                                        log("ok", f"  -> Selected 'Silver' via <select> value")
                                        dd_clicked = True
                                        break
                                    except Exception:
                                        pass

                        # Method 2: Click the dropdown trigger, then click "Silver" text
                        if not dd_clicked:
                            try:
                                dd_trigger = page.locator("text=Select").first
                                await dd_trigger.scroll_into_view_if_needed()
                                await dd_trigger.click()
                                await page.wait_for_timeout(500)
                                silver_loc = page.get_by_text("Silver", exact=True)
                                if await silver_loc.count() > 0:
                                    await silver_loc.first.click()
                                    log("ok", f"  -> Selected 'Silver' from dropdown")
                                    dd_clicked = True
                                await page.wait_for_timeout(500)
                            except Exception:
                                pass

                        # Method 3: JS — find dropdown and click Silver option
                        if not dd_clicked:
                            js_result = await page.evaluate("""() => {
                                for (const sel of document.querySelectorAll('select')) {
                                    for (const opt of sel.options) {
                                        if (opt.text.trim().toLowerCase() === 'silver') {
                                            sel.value = opt.value;
                                            sel.dispatchEvent(new Event('change', {bubbles:true}));
                                            return 'select';
                                        }
                                    }
                                }
                                for (const el of document.querySelectorAll('*')) {
                                    const t = (el.innerText || el.textContent || '').trim();
                                    if (t === 'Silver') {
                                        const r = el.getBoundingClientRect();
                                        if (r.width > 0 && r.height > 0) {
                                            el.click();
                                            return 'click';
                                        }
                                    }
                                }
                                return null;
                            }""")
                            if js_result:
                                log("ok", f"  -> Selected 'Silver' via JS ({js_result})")
                                dd_clicked = True

                        if not dd_clicked:
                            log("warn", f"  Could not select plan — select 'Silver' manually")
                            await pause("Select 'Silver' from the plan dropdown, then press ENTER")

                        await page.wait_for_timeout(800)
                    else:
                        log("info", f"  No 'Select Plan' dropdown — continuing")

                await page.wait_for_timeout(500)

            await page.wait_for_timeout(1000)
            log("info", "Add-on selection complete")
            await page.wait_for_timeout(800)
            steps.append("addons_selected")

            # ── STEP 7c: Click RECALCULATE (only if a distinct button exists) ──
            # Recalculate appears when a cover with price impact is added.
            # Guard: if its coordinates match Submit's coordinates it is NOT real.
            log("step", "STEP 7c — Recalculate premium after ZD selection")
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(800)

            recalc_info = await page.evaluate("""() => {
                let recalcEl = null, submitEl = null;
                let bestRC = Infinity, bestSB = Infinity;
                for (const el of document.querySelectorAll('*')) {
                    const r = el.getBoundingClientRect();
                    if (r.width === 0 || r.height === 0) continue;
                    const txt = (el.innerText || el.value || '').trim().toLowerCase();
                    const a = r.width * r.height;
                    if (txt === 'recalculate' && a < bestRC) { bestRC = a; recalcEl = el; }
                    if ((txt === 'submit' || txt === 'submit ') && a < bestSB) { bestSB = a; submitEl = el; }
                }
                if (!recalcEl) return null;
                recalcEl.scrollIntoView({block: 'center'});
                const rr = recalcEl.getBoundingClientRect();
                const rx = rr.left + rr.width / 2, ry = rr.top + rr.height / 2;
                // If Submit is at same screen position, skip — it's the same button
                if (submitEl) {
                    submitEl.scrollIntoView({block: 'center'});
                    const rs = submitEl.getBoundingClientRect();
                    const sx = rs.left + rs.width / 2, sy = rs.top + rs.height / 2;
                    if (Math.abs(rx - sx) < 10 && Math.abs(ry - sy) < 10) return null;
                }
                return {x: rx, y: ry, tag: recalcEl.tagName};
            }""")

            if recalc_info:
                await page.wait_for_timeout(400)
                await page.mouse.click(recalc_info['x'], recalc_info['y'])
                log("ok", f"  -> Recalculate clicked at ({recalc_info['x']:.0f},{recalc_info['y']:.0f}) via <{recalc_info['tag']}>")
                try:
                    await page.wait_for_load_state("networkidle", timeout=12000)
                except Exception:
                    pass
                await page.wait_for_timeout(1500)
            else:
                log("info", "  No separate Recalculate button — premium updates automatically")
                await page.wait_for_timeout(1000)

            steps.append("recalculated")

            # ── STEP 8: Click SUBMIT button on Additional Covers page ─
            log("step", "STEP 8 — Submit Additional Covers")
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(800)

            # Poll up to 15s for Submit button to be visible
            for _ in range(30):
                found = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('*')) {
                        const rect = el.getBoundingClientRect();
                        if (rect.width === 0 || rect.height === 0) continue;
                        const txt = (el.innerText || el.value || '').trim();
                        if (txt === 'SUBMIT' || txt === 'Submit') return true;
                    }
                    return false;
                }""")
                if found:
                    break
                await page.wait_for_timeout(500)

            submit_info = await page.evaluate("""() => {
                let best = null, bestArea = Infinity;
                for (const el of document.querySelectorAll('*')) {
                    const rect = el.getBoundingClientRect();
                    if (rect.width === 0 || rect.height === 0) continue;
                    const txt = (el.innerText || el.value || '').trim();
                    if (txt === 'SUBMIT' || txt === 'Submit') {
                        const area = rect.width * rect.height;
                        if (area < bestArea) { bestArea = area; best = el; }
                    }
                }
                if (best) {
                    best.scrollIntoView({block: 'center'});
                    const r = best.getBoundingClientRect();
                    return {x: r.left + r.width/2, y: r.top + r.height/2, tag: best.tagName};
                }
                return null;
            }""")
            if submit_info:
                await page.mouse.click(submit_info['x'], submit_info['y'])
                log("ok", f"  -> Submit clicked at ({submit_info['x']:.0f},{submit_info['y']:.0f}) via <{submit_info['tag']}>")
            else:
                log("warn", "Submit button not found — click it manually")
                await pause("Click 'Submit' manually, then press ENTER")

            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            await page.wait_for_timeout(4000)

            # Check if portal went back to RTO/details page (malfunction)
            on_rto_after_submit = await page.evaluate("""() => {
                const txt = document.body.innerText.toLowerCase();
                return txt.includes('city where vehicle is registered') &&
                       txt.includes('rto') && txt.includes('manufacturer');
            }""")
            if on_rto_after_submit:
                log("warn", "  MALFUNCTION: Portal returned to RTO page after Submit")
                await pause("Navigate back to the proposal/premium page, then press ENTER")

            steps.append("submit_coverages")

            # ── STEP 9: Additional Discounts (if enabled in Excel) ────
            if D.get("additional_discount"):
                log("step", "STEP 9 — Additional Discounts (Loading)")

                # Wait for Plan Selection page to load (it has "Buy Now" buttons)
                log("info", "  Waiting for Plan Selection page to load...")
                for _ps in range(40):
                    has_plan = await page.evaluate("""() => {
                        const txt = document.body.innerText.toLowerCase();
                        return txt.includes('buy now') || txt.includes('plan selection')
                            || txt.includes('additional discounts');
                    }""")
                    if has_plan:
                        log("ok", "  -> Plan Selection page loaded")
                        break
                    await page.wait_for_timeout(500)
                else:
                    log("warn", "  Plan Selection page timed out — continuing anyway")

                await page.wait_for_timeout(2000)

                # Scroll to bottom to find "+Add" button in Additional Discounts section
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await page.wait_for_timeout(1000)

                # Click "+Add" button in Additional Discounts section
                add_clicked = False

                # Strategy 1: Find the "Additional Discounts" section, then click "+ Add" near it
                try:
                    # The button shows as "+ Add" — try multiple text patterns
                    for btn_text in ["+ Add", "+Add", "Add"]:
                        loc = page.get_by_text(btn_text, exact=True)
                        cnt = await loc.count()
                        if cnt > 0:
                            # Click the last match (likely the one at the bottom near Additional Discounts)
                            el = loc.last
                            await el.scroll_into_view_if_needed()
                            await page.wait_for_timeout(300)
                            await el.click()
                            log("ok", f"  -> '{btn_text}' button clicked")
                            add_clicked = True
                            break
                except Exception as e:
                    log("info", f"  get_by_text failed: {e}")

                # Strategy 2: Use Playwright locator with role
                if not add_clicked:
                    try:
                        btn_loc = page.get_by_role("button", name="Add")
                        if await btn_loc.count() > 0:
                            await btn_loc.last.scroll_into_view_if_needed()
                            await page.wait_for_timeout(300)
                            await btn_loc.last.click()
                            log("ok", "  -> 'Add' button clicked via role")
                            add_clicked = True
                    except Exception:
                        pass

                # Strategy 3: JS — find element containing "Add" near "Additional Discounts"
                if not add_clicked:
                    coords = await page.evaluate("""() => {
                        // First find "Additional Discounts" section y-position
                        let adY = 0;
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText || '').trim();
                            if (t.includes('Additional Discounts') && t.length < 100) {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) { adY = r.top; break; }
                            }
                        }
                        // Now find a clickable element with "Add" text near that y-position
                        for (const el of document.querySelectorAll('button, a, div, span')) {
                            const t = (el.innerText || el.textContent || '').trim();
                            if (t.includes('Add') && t.length < 20) {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0 && r.height > 0) {
                                    // Must be near the Additional Discounts section (within 100px)
                                    if (adY > 0 && Math.abs(r.top - adY) > 100) continue;
                                    el.scrollIntoView({block: 'center'});
                                    const r2 = el.getBoundingClientRect();
                                    return {x: r2.left + r2.width/2, y: r2.top + r2.height/2};
                                }
                            }
                        }
                        return null;
                    }""")
                    if coords:
                        await page.wait_for_timeout(300)
                        await page.mouse.click(coords['x'], coords['y'])
                        log("ok", f"  -> '+Add' clicked via JS at ({coords['x']:.0f},{coords['y']:.0f})")
                        add_clicked = True

                if not add_clicked:
                    log("warn", "  '+Add' button not found")
                    await pause("Click '+Add' in Additional Discounts section, then press ENTER")

                await page.wait_for_timeout(1500)

                # ── Tick "Other Loading" checkbox ──
                log("info", "  Looking for 'Other Loading' checkbox...")
                ol_clicked = False

                try:
                    ol_loc = page.get_by_text("Other Loading", exact=False)
                    if await ol_loc.count() > 0:
                        el = ol_loc.first
                        await el.scroll_into_view_if_needed()
                        await page.wait_for_timeout(300)
                        box = await el.bounding_box()
                        if box:
                            # Click the checkbox area to the left of the text
                            click_x = box['x'] - 20
                            if click_x < 5:
                                click_x = box['x'] + 5
                            click_y = box['y'] + box['height'] / 2
                            await page.mouse.click(click_x, click_y)
                            log("ok", f"  -> 'Other Loading' clicked at ({click_x:.0f},{click_y:.0f})")
                            ol_clicked = True
                            await page.wait_for_timeout(800)
                except Exception as e:
                    log("info", f"  get_by_text for Other Loading failed: {e}")

                # Fallback: try clicking the text itself
                if not ol_clicked:
                    try:
                        ol_loc = page.get_by_text("Other Loading", exact=False)
                        if await ol_loc.count() > 0:
                            await ol_loc.first.click()
                            log("ok", "  -> 'Other Loading' clicked (direct text)")
                            ol_clicked = True
                            await page.wait_for_timeout(800)
                    except Exception:
                        pass

                if not ol_clicked:
                    log("warn", "  'Other Loading' not found")
                    await pause("Tick 'Other Loading' manually, then press ENTER")

                await page.wait_for_timeout(500)

                # ── Fill the Loading% value ──
                loading_val = D.get("loading_pct", "")
                if loading_val:
                    log("info", f"  Entering Loading% value: {loading_val}")
                    filled = False

                    # Strategy 1: Find input by placeholder "Enter loading %"
                    try:
                        inp = page.get_by_placeholder("Enter loading", exact=False)
                        if await inp.count() > 0:
                            await inp.first.scroll_into_view_if_needed()
                            await page.wait_for_timeout(300)
                            await inp.first.click()
                            await inp.first.fill(str(loading_val))
                            log("ok", f"  -> Loading% '{loading_val}' entered via placeholder match")
                            filled = True
                    except Exception as e:
                        log("info", f"  Placeholder search failed: {e}")

                    # Strategy 2: Find input inside the "Other Loading" expanded section
                    if not filled:
                        try:
                            js_filled = await page.evaluate("""(val) => {
                                // Find input with placeholder containing 'loading'
                                for (const inp of document.querySelectorAll('input')) {
                                    const ph = (inp.placeholder || '').toLowerCase();
                                    if (ph.includes('loading') || ph.includes('load')) {
                                        const r = inp.getBoundingClientRect();
                                        if (r.width > 0 && r.height > 0) {
                                            inp.focus();
                                            inp.value = '';
                                            inp.dispatchEvent(new Event('input', {bubbles:true}));
                                            inp.value = val;
                                            inp.dispatchEvent(new Event('input', {bubbles:true}));
                                            inp.dispatchEvent(new Event('change', {bubbles:true}));
                                            return 'placeholder';
                                        }
                                    }
                                }
                                // Find "Loading Percent" label, then get the next input
                                for (const el of document.querySelectorAll('*')) {
                                    const t = (el.innerText || '').trim().toLowerCase();
                                    if (t.includes('loading percent') && t.length < 30) {
                                        const r = el.getBoundingClientRect();
                                        if (r.width === 0) continue;
                                        // Look for input below this label
                                        const inputs = document.querySelectorAll('input');
                                        let closest = null, closestDist = Infinity;
                                        for (const inp of inputs) {
                                            const ir = inp.getBoundingClientRect();
                                            if (ir.width === 0 || ir.height === 0) continue;
                                            const dist = Math.abs(ir.top - r.bottom);
                                            if (dist < closestDist && ir.top >= r.top - 10) {
                                                closestDist = dist;
                                                closest = inp;
                                            }
                                        }
                                        if (closest && closestDist < 80) {
                                            closest.focus();
                                            closest.value = '';
                                            closest.dispatchEvent(new Event('input', {bubbles:true}));
                                            closest.value = val;
                                            closest.dispatchEvent(new Event('input', {bubbles:true}));
                                            closest.dispatchEvent(new Event('change', {bubbles:true}));
                                            return 'label';
                                        }
                                    }
                                }
                                return null;
                            }""", str(loading_val))
                            if js_filled:
                                log("ok", f"  -> Loading% '{loading_val}' entered via JS ({js_filled})")
                                filled = True
                        except Exception as e:
                            log("info", f"  JS fill failed: {e}")

                    if not filled:
                        log("warn", f"  Could not enter Loading% value")
                        await pause(f"Enter '{loading_val}' in the Loading% field, then press ENTER")

                await page.wait_for_timeout(800)

                # ── Click "Apply & Continue" button ──
                log("info", "  Clicking 'Apply & Continue'...")
                ac_clicked = False

                # Strategy 1: Exact text match for the button
                try:
                    ac_loc = page.get_by_text("Apply & Continue", exact=True)
                    if await ac_loc.count() > 0:
                        await ac_loc.first.scroll_into_view_if_needed()
                        await page.wait_for_timeout(300)
                        await ac_loc.first.click()
                        log("ok", "  -> 'Apply & Continue' clicked (exact match)")
                        ac_clicked = True
                except Exception:
                    pass

                # Strategy 2: Playwright get_by_role button
                if not ac_clicked:
                    try:
                        btn = page.get_by_role("button", name="Apply & Continue")
                        if await btn.count() > 0:
                            await btn.first.scroll_into_view_if_needed()
                            await page.wait_for_timeout(300)
                            await btn.first.click()
                            log("ok", "  -> 'Apply & Continue' clicked (role button)")
                            ac_clicked = True
                    except Exception:
                        pass

                # Strategy 3: JS — find button/clickable with exact "Apply & Continue" text
                if not ac_clicked:
                    coords = await page.evaluate("""() => {
                        for (const el of document.querySelectorAll('button, a, div, span')) {
                            const t = (el.innerText || el.textContent || '').trim();
                            if (t === 'Apply & Continue' || t === 'Apply &amp; Continue'
                                || t === 'Apply and Continue') {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0 && r.height > 0) {
                                    el.scrollIntoView({block: 'center'});
                                    const r2 = el.getBoundingClientRect();
                                    return {x: r2.left + r2.width/2, y: r2.top + r2.height/2};
                                }
                            }
                        }
                        return null;
                    }""")
                    if coords:
                        await page.wait_for_timeout(300)
                        await page.mouse.click(coords['x'], coords['y'])
                        log("ok", f"  -> 'Apply & Continue' clicked via JS at ({coords['x']:.0f},{coords['y']:.0f})")
                        ac_clicked = True

                if not ac_clicked:
                    log("warn", "  'Apply & Continue' not found")
                    await pause("Click 'Apply & Continue' manually, then press ENTER")

                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)
                steps.append("additional_discount")
            else:
                log("info", "STEP 9 — Additional Discounts: SKIPPED (not enabled in Excel)")

            # ── STEP 10: Click "Buy Now" on Basic Plan card ──────────
            log("step", "STEP 10 — Click Buy Now on Basic Plan")

            # Wait for Plan Selection page to be ready
            await page.wait_for_timeout(2000)
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(500)
            await page.evaluate("window.scrollTo(0, 0)")
            await page.wait_for_timeout(500)

            buy_clicked = False

            # Strategy 1: Find "Buy Now" button on Basic Plan (first/left card)
            try:
                buy_loc = page.get_by_text("Buy Now", exact=True)
                if await buy_loc.count() > 0:
                    # Click the FIRST "Buy Now" (Basic Plan is the left card)
                    await buy_loc.first.scroll_into_view_if_needed()
                    await page.wait_for_timeout(300)
                    await buy_loc.first.click()
                    log("ok", "  -> 'Buy Now' clicked (Basic Plan)")
                    buy_clicked = True
            except Exception as e:
                log("info", f"  get_by_text failed: {e}")

            # Strategy 2: Playwright role
            if not buy_clicked:
                try:
                    btn = page.get_by_role("button", name="Buy Now")
                    if await btn.count() > 0:
                        await btn.first.scroll_into_view_if_needed()
                        await page.wait_for_timeout(300)
                        await btn.first.click()
                        log("ok", "  -> 'Buy Now' clicked via role")
                        buy_clicked = True
                except Exception:
                    pass

            # Strategy 3: JS
            if not buy_clicked:
                coords = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('button, a, div, span')) {
                        const t = (el.innerText || el.textContent || '').trim();
                        if (t === 'Buy Now') {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0) {
                                el.scrollIntoView({block: 'center'});
                                const r2 = el.getBoundingClientRect();
                                return {x: r2.left + r2.width/2, y: r2.top + r2.height/2};
                            }
                        }
                    }
                    return null;
                }""")
                if coords:
                    await page.mouse.click(coords['x'], coords['y'])
                    log("ok", f"  -> 'Buy Now' clicked via JS at ({coords['x']:.0f},{coords['y']:.0f})")
                    buy_clicked = True

            if not buy_clicked:
                log("warn", "  'Buy Now' button not found")
                await pause("Click 'Buy Now' on the Basic Plan card, then press ENTER")

            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)
            steps.append("buy_now")

            # ── STEP 11: Proposal Form (Engine/Chassis + KYC popup) ──
            log("step", "STEP 11 — Proposal Form")
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            # Wait until Engine Number AND Chassis fields are actually present in DOM
            # before attempting to fill them — prevents empty fills on slow pages.
            log("info", "Waiting for Engine/Chassis fields to appear...")
            for _ef in range(60):
                fields_ready = await page.evaluate("""() => {
                    const txt = document.body.innerText.toLowerCase();
                    const hasEngine  = txt.includes('engine number');
                    const hasChassis = txt.includes('chassis number');
                    // Also check for inputs on the page
                    const inputs = document.querySelectorAll('input');
                    return (hasEngine || hasChassis) && inputs.length > 0;
                }""")
                if fields_ready:
                    log("info", f"  Engine/Chassis fields ready (attempt {_ef+1})")
                    break
                await page.wait_for_timeout(500)

            async def mouse_fill(find_js, value, field_name):
                await page.evaluate(f"() => {{ const el={find_js}; if(el) el.scrollIntoView({{block:'center'}}); }}")
                await page.wait_for_timeout(200)
                c = await page.evaluate(f"() => {{ const el={find_js}; if(!el) return null; const r=el.getBoundingClientRect(); return r.width>0?{{x:r.left+r.width/2,y:r.top+r.height/2}}:null; }}")
                if c:
                    await page.mouse.click(c['x'], c['y'])
                    await page.wait_for_timeout(120)
                    await page.keyboard.press("Control+a")
                    await page.keyboard.press("Delete")
                    await page.wait_for_timeout(80)
                    await page.keyboard.type(str(value), delay=40)
                    await page.wait_for_timeout(200)
                    log("fill", f"  -> {field_name}: {value}")
                    return True
                log("warn", f"  {field_name} not found")
                return False

            async def click_last_button(text_upper, label):
                log("info", f"Clicking {label}")
                await page.evaluate(f"""() => {{
                    const all = Array.from(document.querySelectorAll('button,input[type="submit"]'))
                        .filter(el => (el.innerText||el.value||'').trim().toUpperCase() === '{text_upper}');
                    if (all.length) all[all.length-1].scrollIntoView({{block:'center'}});
                }}""")
                await page.wait_for_timeout(400)
                c = await page.evaluate(f"""() => {{
                    const all = Array.from(document.querySelectorAll('button,input[type="submit"]'))
                        .filter(el => (el.innerText||el.value||'').trim().toUpperCase() === '{text_upper}')
                        .filter(el => {{ const r=el.getBoundingClientRect(); return r.width>0&&r.height>0; }});
                    if (!all.length) return null;
                    for (const el of [...all].reverse()) {{
                        let node=el.parentElement;
                        for (let i=0;i<12;i++) {{
                            if(!node) break;
                            const s=window.getComputedStyle(node);
                            if (s.position==='fixed'||s.position==='absolute'||
                                ['modal','popup','overlay','dialog','cdk-overlay'].some(c=>node.classList.contains(c))) {{
                                const r=el.getBoundingClientRect();
                                return {{x:r.left+r.width/2,y:r.top+r.height/2}};
                            }}
                            node=node.parentElement;
                        }}
                    }}
                    const el=all[all.length-1];
                    const r=el.getBoundingClientRect();
                    return {{x:r.left+r.width/2,y:r.top+r.height/2}};
                }}""")
                if c:
                    await page.mouse.click(c['x'], c['y'])
                    log("ok", f"  -> {label} clicked at ({c['x']:.0f},{c['y']:.0f})")
                    return True
                log("warn", f"  {label} not found")
                return False

            # ── 11a: Engine Number ────────────────────────────────────
            if D.get("engine_number"):
                # Find input by placeholder OR by nearby label text "Engine Number"
                engine_js = """(() => {
                    // Try placeholder first
                    const byPh = Array.from(document.querySelectorAll('input')).find(i =>
                        (i.placeholder||'').toLowerCase().includes('engine'));
                    if (byPh) return byPh;
                    // Try by label text: find "Engine Number" text, then get nearest input
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if (t.includes('engine number') && t.length < 30) {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0) continue;
                            // Find the closest input below or near this label
                            let best = null, bestDist = Infinity;
                            for (const inp of document.querySelectorAll('input')) {
                                const ir = inp.getBoundingClientRect();
                                if (ir.width === 0 || ir.height === 0) continue;
                                const dist = Math.abs(ir.left - r.left) + Math.abs(ir.top - r.bottom);
                                if (dist < bestDist) { bestDist = dist; best = inp; }
                            }
                            if (best && bestDist < 150) return best;
                        }
                    }
                    return null;
                })()"""
                await mouse_fill(engine_js, D["engine_number"], "Engine Number")

            # ── 11b: Chassis Number ───────────────────────────────────
            if D.get("chassis_number"):
                chassis_js = """(() => {
                    // Try placeholder first (portal sometimes uses "Chasis" typo)
                    const byPh = Array.from(document.querySelectorAll('input')).find(i => {
                        const ph = (i.placeholder||'').toLowerCase();
                        return ph.includes('chassis') || ph.includes('chasis');
                    });
                    if (byPh) return byPh;
                    // Try by label text
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if ((t.includes('chassis number') || t.includes('chasis number')) && t.length < 30) {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0) continue;
                            let best = null, bestDist = Infinity;
                            for (const inp of document.querySelectorAll('input')) {
                                const ir = inp.getBoundingClientRect();
                                if (ir.width === 0 || ir.height === 0) continue;
                                const dist = Math.abs(ir.left - r.left) + Math.abs(ir.top - r.bottom);
                                if (dist < bestDist) { bestDist = dist; best = inp; }
                            }
                            if (best && bestDist < 150) return best;
                        }
                    }
                    return null;
                })()"""
                await mouse_fill(chassis_js, D["chassis_number"], "Chassis Number")

            # Mark engine/chassis cells green in Excel (Req 2 — prevent duplicate entry)
            mark_engine_chassis_green(D)

            # ── 11c: Check if KYC already done (loop case) ──────────
            pin_val = D.get("pin_code","").strip()   # Applicant Details pincode — always available
            gst_val = D.get("gst_number","").strip()       # defined here so always available
            kyc_already_done = await page.evaluate("""() =>
                document.body.innerText.toLowerCase().includes('your kyc is done')
            """)

            if kyc_already_done:
                log("ok", "  -> KYC already done — skipping KYC popup steps")
                steps.append("kyc")
                log("ok", "══ SECTION DONE: KYC ══")
            else:
                # ── 11c: Complete KYC radio ───────────────────────────────
                log("info", "Clicking Complete KYC radio")
                await page.evaluate("""() => {
                    for (const r of document.querySelectorAll('input[type="radio"]')) {
                        const lbl=r.closest('label')||document.querySelector('label[for="'+r.id+'"]')||r.parentElement;
                        if ((lbl?.innerText||lbl?.textContent||'').toLowerCase().includes('complete kyc'))
                            { r.scrollIntoView({block:'center'}); return; }
                    }
                }""")
                await page.wait_for_timeout(300)
                kyc_c = await page.evaluate("""() => {
                    for (const r of document.querySelectorAll('input[type="radio"]')) {
                        const lbl=r.closest('label')||document.querySelector('label[for="'+r.id+'"]')||r.parentElement;
                        if ((lbl?.innerText||lbl?.textContent||'').toLowerCase().includes('complete kyc')) {
                            const rect=r.getBoundingClientRect();
                            if (rect.width>0) return {x:rect.left+rect.width/2,y:rect.top+rect.height/2};
                        }
                    }
                    for (const el of document.querySelectorAll('*')) {
                        if ((el.innerText||el.textContent||'').trim().toLowerCase()==='complete kyc') {
                            const rect=el.getBoundingClientRect();
                            if (rect.width>0) return {x:rect.left+rect.width/2,y:rect.top+rect.height/2};
                        }
                    }
                    return null;
                }""")
                if kyc_c:
                    await page.mouse.click(kyc_c['x'], kyc_c['y'])
                    log("ok", "  -> Complete KYC radio clicked")
                log("info", "Waiting for KYC popup...")
                for _ in range(20):
                    if await page.evaluate("""() =>
                        document.body.innerText.toLowerCase().includes('select entity type')||
                        document.body.innerText.toLowerCase().includes('customer onboarding')
                    """):
                        log("ok", "  -> KYC popup opened"); break
                    await page.wait_for_timeout(500)
                await page.wait_for_timeout(800)

                # ── 11d: Entity Type = Corporate ─────────────────────────
                log("info", "Selecting Entity Type: Corporate")
                for sel_el in await page.query_selector_all('select'):
                    try:
                        near = await page.evaluate("""(s)=>{ let n=s.parentElement; for(let i=0;i<8;i++){if(!n)break;if((n.innerText||'').toLowerCase().includes('entity type'))return true;n=n.parentElement;} return false;}""", sel_el)
                        if not near: continue
                        await sel_el.scroll_into_view_if_needed()
                        await sel_el.select_option(label="Corporate")
                        await page.wait_for_timeout(1200)
                        log("ok", "  -> Entity Type: Corporate"); break
                    except Exception as e:
                        log("warn", f"  Entity Type error: {e}")

                # ── 11e: Click GST document-type button ───────────────────
                # The portal renders document-type choices (PAN/CKYC/CIN/GST) as
                # styled button-like elements, NOT standard radio <input>s.
                # Strategy: find the element whose ONLY visible text is exactly "GST"
                # and whose bounding box is small (it's a button badge, not a container).
                # Retry up to 10 times in case the popup is still animating in.
                log("info", "Clicking GST document type button")
                gst_clicked = False
                for _gst_try in range(10):
                    gst_coords = await page.evaluate("""() => {
                        // Walk every element; pick the one whose trimmed innerText === 'GST'
                        // with the smallest bounding area (the button, not a container).
                        let best = null, bestArea = Infinity;
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText || el.textContent || '').trim();
                            if (t !== 'GST') continue;
                            const r = el.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) continue;
                            const area = r.width * r.height;
                            if (area < bestArea) { bestArea = area; best = el; }
                        }
                        if (!best) return null;
                        best.scrollIntoView({block: 'center'});
                        const r = best.getBoundingClientRect();
                        return {x: r.left + r.width / 2, y: r.top + r.height / 2};
                    }""")
                    if gst_coords:
                        await page.mouse.click(gst_coords['x'], gst_coords['y'])
                        log("ok", f"  -> GST button clicked at ({gst_coords['x']:.0f},{gst_coords['y']:.0f})")
                        gst_clicked = True
                        break
                    await page.wait_for_timeout(400)
                if not gst_clicked:
                    log("warn", "  GST button not found")

                # Wait until the label below the doc-type buttons changes to
                # something GST-related ("GSTIN Number", "GST", "GSTIN") —
                # meaning the portal has re-rendered the input fields for GST.
                log("info", "  Waiting for GSTIN field to appear...")
                for _ in range(25):
                    gstin_label_visible = await page.evaluate("""() => {
                        const body = (document.body.innerText || '').toLowerCase();
                        return body.includes('gstin') || body.includes('gst number');
                    }""")
                    if gstin_label_visible:
                        log("ok", "  -> GSTIN field visible")
                        break
                    await page.wait_for_timeout(400)
                await page.wait_for_timeout(500)

                # ── 11f: Fill GSTIN ───────────────────────────────────────
                # ── 11g: Fill PIN Code ────────────────────────────────────
                # After switching to GST tab the popup shows:
                #   Field label "GSTIN Number" → text input
                #   Field label "Pin Code"     → text input
                # Fill by finding the input closest to each label.

                async def fill_kyc_field_by_label(label_texts, value, field_name):
                    """Find visible input nearest to a label containing any of label_texts, fill it."""
                    coords = await page.evaluate("""(labels) => {
                        const lowerLabels = labels.map(l => l.toLowerCase());
                        // Find the label element
                        let labelEl = null;
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                            if (lowerLabels.some(l => t === l || t.startsWith(l)) && t.length < 40) {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0 && r.height > 0) { labelEl = el; break; }
                            }
                        }
                        if (!labelEl) return null;
                        const lr = labelEl.getBoundingClientRect();
                        // Find the closest visible, editable input below/near this label
                        let best = null, bestDist = Infinity;
                        for (const inp of document.querySelectorAll('input')) {
                            if (inp.readOnly || inp.disabled) continue;
                            const t = (inp.type || '').toLowerCase();
                            if (['checkbox','radio','hidden','submit','button'].includes(t)) continue;
                            const r = inp.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) continue;
                            // Prefer inputs below the label (positive dy), penalise ones above
                            const dy = r.top - lr.bottom;
                            const dx = Math.abs((r.left + r.width/2) - (lr.left + lr.width/2));
                            const dist = dx + Math.max(dy, 0) * 0.5 + Math.max(-dy, 0) * 3;
                            if (dist < bestDist) { bestDist = dist; best = inp; }
                        }
                        if (!best || bestDist > 300) return null;
                        best.scrollIntoView({block: 'center'});
                        const r = best.getBoundingClientRect();
                        return {x: r.left + r.width / 2, y: r.top + r.height / 2};
                    }""", label_texts)
                    if not coords:
                        log("warn", f"  {field_name} input not found by label")
                        return False
                    await page.mouse.click(coords['x'], coords['y'])
                    await page.wait_for_timeout(150)
                    await page.keyboard.press("Control+a")
                    await page.keyboard.press("Delete")
                    await page.wait_for_timeout(80)
                    await page.keyboard.type(str(value), delay=60)
                    await page.wait_for_timeout(300)
                    # Commit via Angular native setter
                    await page.evaluate("""({cx, cy, val}) => {
                        const el = document.elementFromPoint(cx, cy);
                        if (!el) return;
                        const inp = el.tagName === 'INPUT' ? el : el.closest('input');
                        if (!inp) return;
                        const setter = Object.getOwnPropertyDescriptor(
                            window.HTMLInputElement.prototype, 'value').set;
                        setter.call(inp, val);
                        inp.dispatchEvent(new Event('input',  {bubbles: true}));
                        inp.dispatchEvent(new Event('change', {bubbles: true}));
                        inp.dispatchEvent(new Event('blur',   {bubbles: true}));
                    }""", {"cx": coords['x'], "cy": coords['y'], "val": str(value)})
                    log("ok", f"  -> {field_name}: {value}")
                    return True

                gst_val = D.get("gst_number", "").strip()
                kyc_pin_val = D.get("kyc_pincode", D.get("pin_code", "")).strip()

                if gst_val:
                    filled = await fill_kyc_field_by_label(
                        ["gstin number", "gstin", "gst number", "gst no"],
                        gst_val, "GSTIN (KYC popup)")
                    if not filled:
                        # Positional fallback: first visible text input in popup
                        await mouse_fill(
                            "Array.from(document.querySelectorAll('input')).filter(i=>{ const r=i.getBoundingClientRect(); return r.width>0&&r.height>0&&!i.readOnly&&!i.disabled&&!['checkbox','radio','hidden','submit','button'].includes((i.type||'').toLowerCase()); })[0]",
                            gst_val, "GSTIN (KYC popup) [positional]")

                if kyc_pin_val:
                    filled = await fill_kyc_field_by_label(
                        ["pin code", "pincode", "pin"],
                        kyc_pin_val, "PIN Code (KYC)")
                    if not filled:
                        # Positional fallback: second visible text input in popup
                        await mouse_fill(
                            "Array.from(document.querySelectorAll('input')).filter(i=>{ const r=i.getBoundingClientRect(); return r.width>0&&r.height>0&&!i.readOnly&&!i.disabled&&!['checkbox','radio','hidden','submit','button'].includes((i.type||'').toLowerCase()); })[1]",
                            kyc_pin_val, "PIN Code (KYC) [positional]")

                # ── 11h: Consent checkbox ─────────────────────────────────
                # Ensure the consent checkbox is ticked. It may already be pre-checked.
                log("info", "Ensuring consent checkbox is ticked")
                await page.wait_for_timeout(300)
                cb = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        let node = el.parentElement;
                        for (let i = 0; i < 8; i++) {
                            if (!node) break;
                            const t = (node.innerText || node.textContent || '').toLowerCase();
                            if (t.includes('consent') || t.includes('hereby')) {
                                el.scrollIntoView({block: 'center'});
                                const r = el.getBoundingClientRect();
                                return r.width > 0
                                    ? {x: r.left+r.width/2, y: r.top+r.height/2, checked: el.checked}
                                    : null;
                            }
                            node = node.parentElement;
                        }
                    }
                    return null;
                }""")
                if cb:
                    if not cb.get('checked', False):
                        await page.mouse.click(cb['x'], cb['y'])
                        await page.wait_for_timeout(400)
                        log("ok", f"  -> Consent checkbox ticked at ({cb['x']:.0f},{cb['y']:.0f})")
                    else:
                        log("ok", f"  -> Consent checkbox already ticked at ({cb['x']:.0f},{cb['y']:.0f})")
                else:
                    log("warn", "  Consent checkbox not found")

                await page.wait_for_timeout(300)
                await click_last_button("SUBMIT", "SUBMIT")
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)
                steps.append("kyc")
                log("ok", "══ SECTION DONE: KYC ══")

            # ── STEP 12: Declaration → PROCEED ────────────────────────
            # Skip PROCEED if KYC already done — Applicant Details already visible
            if kyc_already_done:
                log("step", "STEP 12 — Declaration PROCEED (skipped — KYC already done)")
                steps.append("proceed")
            else:
                log("step", "STEP 12 — Declaration PROCEED")
                log("info", "Waiting for page to fully load before PROCEED...")
                try:
                    await page.wait_for_load_state("networkidle", timeout=20000)
                except Exception:
                    pass
                decl_cb = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        if (el.checked) continue;
                        let node=el.parentElement;
                        for (let i=0;i<6;i++) {
                            if (!node) break;
                            const t=(node.innerText||node.textContent||'').toLowerCase();
                            if (t.includes('hereby declare')||t.includes('premium')||t.includes('declaration')) {
                                el.scrollIntoView({block:'center'});
                                const r=el.getBoundingClientRect();
                                return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2}:null;
                            }
                            node=node.parentElement;
                        }
                    }
                    return null;
                }""")
                if decl_cb:
                    await page.mouse.click(decl_cb['x'], decl_cb['y'])
                    await page.wait_for_timeout(400)
                    log("ok", "  -> Declaration checkbox ticked")
                proceed_clicked = False
                for _proc in range(25):
                    try:
                        for txt in ["PROCEED", "Proceed", "proceed"]:
                            loc = page.get_by_role("button", name=txt)
                            if await loc.count() > 0:
                                await loc.last.scroll_into_view_if_needed()
                                await page.wait_for_timeout(300)
                                await loc.last.click(force=True)
                                log("ok", f"  -> PROCEED clicked (locator)")
                                proceed_clicked = True
                                break
                        if proceed_clicked:
                            break
                    except Exception:
                        pass
                    coords = await page.evaluate("""() => {
                        const all = Array.from(document.querySelectorAll('button, input[type="submit"]'))
                            .filter(el => (el.innerText||el.value||'').trim().toLowerCase().includes('proceed'));
                        if (!all.length) return null;
                        const el = all[all.length - 1];
                        el.scrollIntoView({block:'center'});
                        const r = el.getBoundingClientRect();
                        return r.width > 0 ? {x: r.left+r.width/2, y: r.top+r.height/2} : null;
                    }""")
                    if coords:
                        await page.wait_for_timeout(300)
                        await page.mouse.click(coords['x'], coords['y'])
                        log("ok", f"  -> PROCEED clicked at ({coords['x']:.0f},{coords['y']:.0f})")
                        proceed_clicked = True
                        break
                    log("info", f"  PROCEED not found yet (attempt {_proc+1}/25), waiting...")
                    await page.wait_for_timeout(2000)
                if not proceed_clicked:
                    log("warn", "  PROCEED not found after all retries")
                try:
                    await page.wait_for_load_state("networkidle", timeout=20000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)
                steps.append("proceed")

            # ── STEP 12.5: Applicant Details Form ─────────────────────
            log("step", "STEP 12.5 — Applicant Details → Create Proposal")
            for _ in range(20):
                if await page.evaluate("""() => document.body.innerText.includes('Applicant Details')||document.body.innerText.includes('Customer Details')||document.body.innerText.includes('Create Proposal')"""):
                    log("ok", "  -> Applicant Details loaded"); break
                await page.wait_for_timeout(500)
            await page.wait_for_timeout(600)

            # ── DIAGNOSTIC: dump all visible inputs in Customer Details ──
            # This tells us the real attributes (placeholder/formControlName/id/name)
            # so we know exactly how to target each field.
            input_map = await page.evaluate("""() => {
                // Find Customer Details section
                let section = null;
                for (const el of document.querySelectorAll('*')) {
                    const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                    if ((t === 'customer details' || t === 'kyc details') &&
                         el.querySelectorAll('input').length >= 2) {
                        section = el; break;
                    }
                }
                const root = section || document;
                return Array.from(root.querySelectorAll('input,textarea,mat-select,select'))
                    .filter(el => {
                        const r = el.getBoundingClientRect();
                        return r.width > 0 && r.height > 0;
                    })
                    .map((el, idx) => ({
                        idx,
                        tag:         el.tagName,
                        type:        el.type || '',
                        placeholder: el.placeholder || '',
                        fc:          el.getAttribute('formcontrolname') || '',
                        id:          el.id || '',
                        name:        el.name || '',
                        ariaLabel:   el.getAttribute('aria-label') || '',
                        value:       el.value || '',
                        x:           Math.round(el.getBoundingClientRect().left + el.getBoundingClientRect().width/2),
                        y:           Math.round(el.getBoundingClientRect().top  + el.getBoundingClientRect().height/2),
                    }));
            }""")
            log("info", f"  Customer Details inputs found: {len(input_map)}")
            for im in input_map:
                log("info", f"  [{im['idx']}] {im['tag']} type={im['type']!r:8} ph={im['placeholder']!r:20} fc={im['fc']!r:20} id={im['id']!r:20} name={im['name']!r:15} aria={im['ariaLabel']!r:20} val={im['value']!r}")

            # ── Helper: find input by any known attribute, then by position ──
            async def find_and_fill(field_name, value, *,
                                    ph_includes=(), ph_exact=(),
                                    fc_includes=(), id_includes=(),
                                    aria_includes=(), name_includes=(),
                                    position=None):
                """
                Find the right input using every available attribute, then fill it.
                Falls back to positional index within Customer Details as last resort.
                Always uses: click → Ctrl+A → Delete → type → JS native setter.
                """
                js = """([phInc, phExact, fcInc, idInc, ariaInc, nameInc, pos]) => {
                    let section = null;
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                        if ((t === 'customer details' || t === 'kyc details') &&
                             el.querySelectorAll('input').length >= 2) {
                            section = el; break;
                        }
                    }
                    const root = section || document;
                    const visibleInputs = Array.from(
                        root.querySelectorAll('input,textarea')
                    ).filter(el => {
                        if (el.readOnly || el.disabled) return false;
                        const t = (el.type||'').toLowerCase();
                        if (['checkbox','radio','hidden','submit','button'].includes(t)) return false;
                        const r = el.getBoundingClientRect();
                        return r.width > 0 && r.height > 0;
                    });

                    // Try attribute-based match first
                    for (const inp of visibleInputs) {
                        const ph   = (inp.placeholder||'').toLowerCase();
                        const fc   = (inp.getAttribute('formcontrolname')||'').toLowerCase();
                        const id   = (inp.id||'').toLowerCase();
                        const aria = (inp.getAttribute('aria-label')||'').toLowerCase();
                        const nm   = (inp.name||'').toLowerCase();
                        const matched =
                            phExact.some(v => ph === v.toLowerCase()) ||
                            phInc.some(v   => ph.includes(v.toLowerCase())) ||
                            fcInc.some(v   => fc.includes(v.toLowerCase())) ||
                            idInc.some(v   => id.includes(v.toLowerCase())) ||
                            ariaInc.some(v => aria.includes(v.toLowerCase())) ||
                            nameInc.some(v => nm.includes(v.toLowerCase()));
                        if (matched) {
                            inp.scrollIntoView({block:'center'});
                            const r = inp.getBoundingClientRect();
                            return {x: r.left+r.width/2, y: r.top+r.height/2, how:'attr'};
                        }
                    }
                    // Positional fallback
                    if (pos !== null && pos < visibleInputs.length) {
                        const inp = visibleInputs[pos];
                        inp.scrollIntoView({block:'center'});
                        const r = inp.getBoundingClientRect();
                        return {x: r.left+r.width/2, y: r.top+r.height/2, how:'pos:'+pos};
                    }
                    return null;
                }"""
                coords = await page.evaluate(js, [
                    list(ph_includes), list(ph_exact),
                    list(fc_includes), list(id_includes),
                    list(aria_includes), list(name_includes),
                    position
                ])
                if not coords:
                    log("warn", f"  {field_name} not found")
                    return False
                log("info", f"  Filling {field_name} via {coords['how']} at ({coords['x']},{coords['y']})")
                await page.mouse.click(coords['x'], coords['y'])
                await page.wait_for_timeout(150)
                await page.keyboard.press("Control+A")
                await page.wait_for_timeout(60)
                await page.keyboard.press("Delete")
                await page.wait_for_timeout(60)
                await page.keyboard.type(str(value), delay=40)
                await page.wait_for_timeout(250)
                # Angular native setter
                await page.evaluate("""({cx, cy, val}) => {
                    const el = document.elementFromPoint(cx, cy);
                    if (!el) return;
                    const inp = (el.tagName==='INPUT'||el.tagName==='TEXTAREA') ? el
                        : (el.closest('input') || el.closest('textarea'));
                    if (!inp) return;
                    const proto = inp.tagName==='TEXTAREA'
                        ? window.HTMLTextAreaElement.prototype
                        : window.HTMLInputElement.prototype;
                    const setter = Object.getOwnPropertyDescriptor(proto, 'value').set;
                    setter.call(inp, val);
                    inp.dispatchEvent(new Event('input',  {bubbles:true}));
                    inp.dispatchEvent(new Event('change', {bubbles:true}));
                    inp.dispatchEvent(new Event('blur',   {bubbles:true}));
                }""", {"cx": coords['x'], "cy": coords['y'], "val": str(value)})
                log("ok", f"  -> {field_name}: {value}")
                return True

            # ── Title dropdown ─────────────────────────────────────────
            # Diagnostic: MAT-SELECT fc='title' id='mat-select-6'
            # Must click the mat-select trigger to open the overlay, then click M/s mat-option.
            log("info", "Selecting Title: M/s")
            title_done = False

            # Find the mat-select trigger by formcontrolname='title'
            title_trigger = await page.evaluate("""() => {
                // Direct: mat-select[formcontrolname='title']
                const ms = document.querySelector('mat-select[formcontrolname="title"]');
                if (ms) {
                    ms.scrollIntoView({block:'center'});
                    const r = ms.getBoundingClientRect();
                    return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                }
                // Fallback: mat-form-field whose label = 'title'
                for (const ff of document.querySelectorAll('mat-form-field')) {
                    const lbl = ff.querySelector('mat-label,label');
                    if (!lbl) continue;
                    const t = (lbl.innerText||lbl.textContent||'').trim().toLowerCase();
                    if (t === 'title' || t === 'title *') {
                        const ms2 = ff.querySelector('mat-select');
                        if (ms2) {
                            ms2.scrollIntoView({block:'center'});
                            const r = ms2.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                    }
                }
                return null;
            }""")
            if title_trigger:
                await page.mouse.click(title_trigger['x'], title_trigger['y'])
                await page.wait_for_timeout(500)
                # Wait for mat-option overlay to render
                for _tw in range(10):
                    has_opts = await page.evaluate("() => document.querySelectorAll('mat-option').length > 0")
                    if has_opts: break
                    await page.wait_for_timeout(150)
                # Click the M/s option
                clicked = await page.evaluate("""() => {
                    for (const opt of document.querySelectorAll('mat-option')) {
                        const t = (opt.innerText||opt.textContent||'').trim();
                        if (t === 'M/s' || t.toLowerCase() === 'm/s') {
                            opt.click(); return t;
                        }
                    }
                    // If M/s not listed, click first available option
                    for (const opt of document.querySelectorAll('mat-option')) {
                        const t = (opt.innerText||opt.textContent||'').trim();
                        const r = opt.getBoundingClientRect();
                        if (t && r.width>0 && r.height>0) { opt.click(); return '(first) ' + t; }
                    }
                    return null;
                }""")
                if clicked:
                    log("ok", f"  -> Title: {clicked}")
                    title_done = True
                else:
                    await page.keyboard.press("Escape")
                    log("warn", "  Title overlay opened but no mat-option found")
            else:
                log("warn", "  Title mat-select not found — skipping")
            await page.wait_for_timeout(400)

            # ── Insured Name ──────────────────────────────────────────
            # Motor form: position 0 in Customer Details (after Title select)
            # fc candidates: 'insuredName','insured','name','applicantName'
            insured_name = D.get("insured_name","").strip()
            if insured_name:
                await find_and_fill("Insured Name", insured_name,
                    ph_includes=("insured name","insured","applicant name"),
                    fc_includes=("insuredname","insured","applicantname","name"),
                    id_includes=("insured","applicant"),
                    aria_includes=("insured name","insured"),
                    position=0)
            await page.wait_for_timeout(300)

            # ── DOB ───────────────────────────────────────────────────
            # Find by formcontrolname='dob' first — the only reliable attribute on this form.
            # Type DD/MM/YYYY with slashes so no digits bleed into the next field.
            # Pattern: click → Escape → click → Escape → Ctrl+A → Delete → type
            # Commit via Angular native setter + blur — do NOT press Tab.
            dob_str = D.get("dob","").strip()
            if dob_str:
                log("info", f"Filling DOB: {dob_str}")
                try:
                    parts = dob_str.replace("-","/").split("/")
                    dob_formatted = f"{int(parts[0]):02d}/{int(parts[1]):02d}/{int(parts[2])}"
                    log("info", f"  Typing: {dob_formatted}")

                    dob_coords = await page.evaluate("""() => {
                        // Strategy 1: exact formcontrolname='dob'
                        let dobInp = document.querySelector("input[formcontrolname='dob']");
                        // Strategy 2: broader attribute match
                        if (!dobInp) {
                            const inputs = Array.from(document.querySelectorAll('input')).filter(el => {
                                if (el.readOnly||el.disabled) return false;
                                const t=(el.type||'').toLowerCase();
                                if (['checkbox','radio','hidden','submit','button'].includes(t)) return false;
                                const r=el.getBoundingClientRect();
                                return r.width>0&&r.height>0;
                            });
                            dobInp = inputs.find(i => {
                                const ph   = (i.placeholder||'').toLowerCase();
                                const fc   = (i.getAttribute('formcontrolname')||'').toLowerCase();
                                const id   = (i.id||'').toLowerCase();
                                const aria = (i.getAttribute('aria-label')||'').toLowerCase();
                                return ph.includes('dob')||ph.includes('date')||ph.includes('birth')
                                    || fc.includes('dob')||fc.includes('date')||fc.includes('birth')
                                    || id.includes('dob')||id.includes('date')||id.includes('birth')
                                    || aria.includes('dob')||aria.includes('date')||aria.includes('birth');
                            }) || inputs[1];  // positional fallback: index 1
                        }
                        if (!dobInp) return null;
                        dobInp.scrollIntoView({block:'center'});
                        const r = dobInp.getBoundingClientRect();
                        return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                    }""")
                    if not dob_coords:
                        raise ValueError("DOB input not found")

                    # STAR-file Escape pattern: click → Escape → click → Escape → clear → type
                    await page.mouse.click(dob_coords['x'], dob_coords['y'])
                    await page.wait_for_timeout(300)
                    await page.keyboard.press("Escape")
                    await page.wait_for_timeout(200)
                    await page.mouse.click(dob_coords['x'], dob_coords['y'])
                    await page.wait_for_timeout(200)
                    await page.keyboard.press("Escape")
                    await page.wait_for_timeout(200)
                    await page.keyboard.press("Control+A")
                    await page.wait_for_timeout(60)
                    await page.keyboard.press("Delete")
                    await page.wait_for_timeout(60)
                    # Type with slashes DD/MM/YYYY — prevents digit bleeding to next field
                    for ch in dob_formatted:
                        await page.keyboard.type(ch, delay=80)
                        await page.wait_for_timeout(50)
                    await page.wait_for_timeout(300)
                    # Commit via Angular native setter + blur — do NOT press Tab
                    await page.evaluate("""({cx, cy, val}) => {
                        const el = document.elementFromPoint(cx,cy);
                        const inp = el&&(el.tagName==='INPUT'?el:el.closest('input'));
                        if (!inp) return;
                        const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
                        setter.call(inp, val);
                        inp.dispatchEvent(new Event('input',  {bubbles:true}));
                        inp.dispatchEvent(new Event('change', {bubbles:true}));
                        inp.dispatchEvent(new Event('blur',   {bubbles:true}));
                    }""", {"cx": dob_coords['x'], "cy": dob_coords['y'], "val": dob_formatted})
                    await page.wait_for_timeout(200)
                    actual_dob = await page.evaluate("""({cx,cy}) => {
                        const el=document.elementFromPoint(cx,cy);
                        const inp=el&&(el.tagName==='INPUT'?el:el.closest('input'));
                        return inp?inp.value:null;
                    }""", {"cx": dob_coords['x'], "cy": dob_coords['y']})
                    log("ok" if actual_dob else "warn",
                        f"  -> DOB set: {actual_dob}" if actual_dob else "  DOB field empty after typing")
                except Exception as e:
                    log("warn", f"  DOB error: {e}")

            # ── Address Line ──────────────────────────────────────────
            # Motor form: position 2
            addr_val = D.get("address_line","").strip()
            if addr_val:
                await find_and_fill("Address Line", addr_val,
                    ph_includes=("address",),
                    fc_includes=("address","addressline","addr"),
                    id_includes=("address","addr"),
                    aria_includes=("address",),
                    name_includes=("address",),
                    position=2)
            await page.wait_for_timeout(300)

            # ── Mobile ────────────────────────────────────────────────
            # Motor form: position 3
            mob_val = D.get("mobile_number","").strip()
            if mob_val:
                await find_and_fill("Mobile", mob_val,
                    ph_includes=("mobile","phone"),
                    fc_includes=("mobile","phone","mobileNo","contactNo","contact"),
                    id_includes=("mobile","phone"),
                    aria_includes=("mobile","phone"),
                    name_includes=("mobile","phone"),
                    position=3)
            await page.wait_for_timeout(300)

            # ── Email ─────────────────────────────────────────────────
            # Motor form: position 4
            email_val = D.get("email","").strip()
            if email_val:
                await find_and_fill("Email", email_val,
                    ph_includes=("email",),
                    fc_includes=("email","emailId","emailAddress"),
                    id_includes=("email",),
                    aria_includes=("email",),
                    name_includes=("email",),
                    position=4)
            await page.wait_for_timeout(300)

            # ── Pincode — retry until State field populates ──────────
            # Motor form: position 5
            pin_val = D.get("pin_code","").strip()
            if pin_val:
                for _pin_try in range(10):
                    filled = await find_and_fill("Pincode", pin_val,
                        ph_includes=("pincode","pin code","pin"),
                        fc_includes=("pincode","pin","zipcode","zip","postalcode"),
                        id_includes=("pin","zip"),
                        aria_includes=("pincode","pin code"),
                        name_includes=("pin","zip"),
                        position=5)
                    await page.wait_for_timeout(1000)
                    state_populated = await page.evaluate("""() => {
                        for (const inp of document.querySelectorAll('input[readonly],input[disabled]')) {
                            if ((inp.value||'').trim().length > 1) return true;
                        }
                        const body = (document.body.innerText||'').toLowerCase();
                        return ['rajasthan','maharashtra','gujarat','karnataka','delhi',
                                'west bengal','uttar pradesh','tamil nadu','punjab',
                                'haryana','telangana','kerala'].some(s => body.includes(s));
                    }""")
                    if state_populated:
                        log("ok", f"  State populated after pincode (attempt {_pin_try+1})")
                        break
                    log("warn", f"  State not populated — retyping pincode (attempt {_pin_try+1}/10)...")
                    await page.wait_for_timeout(500)

            # ── City/District ─────────────────────────────────────────
            # Strategy 1: mat-select[formcontrolname="city"] — most precise selector.
            # After clicking, loop up to 15×200ms (3s) for mat-option elements to appear.
            # Strategy 2 fallback: label-proximity. Strategy 3 fallback: native <select>.
            log("info", "Selecting City/District")
            try:
                await page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass
            city_done = False
            # Wait up to 12.5s for the mat-select[formcontrolname="city"] to appear
            for _city_wait in range(25):
                city_ready = await page.evaluate("""() => {
                    if (document.querySelector('mat-select[formcontrolname="city"]')) return true;
                    for (const sel of document.querySelectorAll('select')) {
                        let n = sel.parentElement;
                        for (let i=0;i<6;i++) {
                            if (!n) break;
                            const t=(n.innerText||n.textContent||'').toLowerCase();
                            if (t.includes('city')||t.includes('district')) {
                                const real = Array.from(sel.options).filter(
                                    o=>o.value&&o.text.trim()&&
                                    !['','enter city / district','select'].includes(o.text.trim().toLowerCase())
                                );
                                return real.length > 0;
                            }
                            n = n.parentElement;
                        }
                    }
                    for (const ff of document.querySelectorAll('mat-form-field')) {
                        const lbl=ff.querySelector('mat-label,label');
                        if(!lbl) continue;
                        const t=(lbl.innerText||lbl.textContent||'').trim().toLowerCase();
                        if (t.includes('city')||t.includes('district')) {
                            const ms=ff.querySelector('mat-select');
                            if(ms) return true;
                        }
                    }
                    return false;
                }""")
                if city_ready:
                    log("info", f"  City/District ready (attempt {_city_wait+1})")
                    break
                await page.wait_for_timeout(500)
            await page.wait_for_timeout(400)

            # Find the trigger coords — formcontrolname="city" first
            city_trigger = await page.evaluate("""() => {
                const ms = document.querySelector('mat-select[formcontrolname="city"]');
                if (ms) {
                    ms.scrollIntoView({block:'center'});
                    const r = ms.getBoundingClientRect();
                    return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2, via:'fc'} : null;
                }
                // Fallback: label-proximity via mat-form-field
                for (const ff of document.querySelectorAll('mat-form-field')) {
                    const lbl=ff.querySelector('mat-label,label');
                    if(!lbl) continue;
                    const t=(lbl.innerText||lbl.textContent||'').trim().toLowerCase();
                    if (t.includes('city')||t.includes('district')) {
                        const ms2=ff.querySelector('mat-select');
                        if(ms2){ms2.scrollIntoView({block:'center'});const r=ms2.getBoundingClientRect();return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2,via:'lbl'}:null;}
                    }
                }
                for (const lbl of document.querySelectorAll('label,span,div')) {
                    const t=(lbl.innerText||lbl.textContent||'').toLowerCase();
                    if((t.includes('city')||t.includes('district'))&&t.length<30){
                        let node=lbl.parentElement;
                        for(let i=0;i<5;i++){
                            if(!node) break;
                            const trigger=node.querySelector('select,[role="combobox"],[role="listbox"],[class*="dropdown"],[class*="select"],[class*="ng-select"]');
                            if(trigger){trigger.scrollIntoView({block:'center'});const r=trigger.getBoundingClientRect();return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2,via:'prox'}:null;}
                            node=node.parentElement;
                        }
                    }
                }
                for (const el of document.querySelectorAll('*')) {
                    const t=(el.innerText||'').trim();
                    if(t==='Enter City / District'){const r=el.getBoundingClientRect();if(r.width>0&&r.height>0)return{x:r.left+r.width/2,y:r.top+r.height/2,via:'ph'};}
                }
                return null;
            }""")
            if city_trigger:
                await page.mouse.click(city_trigger['x'], city_trigger['y'])
                log("info", f"  -> City dropdown trigger clicked (via={city_trigger.get('via','')})")
                # Loop up to 15×200ms = 3s waiting for mat-option elements to appear in DOM
                city_opt = None
                for _opt_wait in range(15):
                    await page.wait_for_timeout(200)
                    city_opt = await page.evaluate("""() => {
                        for (const el of document.querySelectorAll('mat-option,li,[role="option"],div.option,div.item')) {
                            const t=(el.innerText||el.textContent||'').trim();
                            if(t&&t!=='Enter City / District'&&t.length>1&&t.length<40){
                                const r=el.getBoundingClientRect();
                                if(r.width>0&&r.height>0)return{x:r.left+r.width/2,y:r.top+r.height/2,txt:t};
                            }
                        }
                        return null;
                    }""")
                    if city_opt:
                        log("info", f"  -> City options appeared after {(_opt_wait+1)*200}ms")
                        break
                if city_opt:
                    await page.mouse.click(city_opt['x'], city_opt['y'])
                    await page.wait_for_timeout(400)
                    log("ok", f"  -> City selected: {city_opt['txt']}")
                    city_done = True
                else:
                    log("warn", "  City options did not appear after 3s — trying native select fallback")

            if not city_done:
                for sel_el in await page.query_selector_all('select'):
                    try:
                        near = await page.evaluate("""(s)=>{let n=s.parentElement;for(let i=0;i<5;i++){if(!n)break;const t=(n.innerText||n.textContent||'').toLowerCase();if(t.includes('city')||t.includes('district'))return true;n=n.parentElement;}return false;}""", sel_el)
                        if not near: continue
                        opts = await page.evaluate("(s)=>Array.from(s.options).map(o=>({val:o.value,txt:o.text.trim()}))", sel_el)
                        best = next((o for o in opts if o['val'] and o['txt'] and o['txt'].upper() not in ('','ENTER CITY / DISTRICT','SELECT')), None)
                        if best:
                            await sel_el.select_option(value=best['val'])
                            await page.wait_for_timeout(400)
                            log("ok", f"  -> City (select): {best['txt']}")
                            city_done = True; break
                    except Exception as e:
                        log("warn", f"  City select error: {e}")

            # ── Add GST Details checkbox ──────────────────────────────
            if gst_val:
                log("info", "Ticking Add GST Details checkbox")
                await page.evaluate("window.scrollBy(0, 400)")
                await page.wait_for_timeout(400)
                gst_cb = await page.evaluate("""() => {
                    // Match the checkbox whose OWN label/sibling says "Add GST Details"
                    // — NOT the whole container which includes all three checkbox labels.
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        // Check: <label> wrapping this checkbox
                        const wrap = el.closest('label');
                        if (wrap) {
                            const t = (wrap.innerText||wrap.textContent||'').toLowerCase();
                            if (t.includes('add gst') || (t.includes('gst') && t.includes('detail'))) {
                                el.scrollIntoView({block:'center'});
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                            continue;
                        }
                        // Check: <label for=id> associated with this checkbox
                        if (el.id) {
                            const lbl = document.querySelector('label[for="'+el.id+'"]');
                            if (lbl) {
                                const t = (lbl.innerText||lbl.textContent||'').toLowerCase();
                                if (t.includes('add gst') || (t.includes('gst') && t.includes('detail'))) {
                                    el.scrollIntoView({block:'center'});
                                    const r = el.getBoundingClientRect();
                                    if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                }
                            }
                        }
                        // Check: immediately adjacent sibling text node / span
                        let sib = el.nextElementSibling || el.parentElement;
                        if (sib) {
                            const t = (sib.innerText||sib.textContent||'').toLowerCase().trim();
                            if (t.length < 40 && (t.includes('add gst') || (t.includes('gst') && t.includes('detail')))) {
                                el.scrollIntoView({block:'center'});
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                        }
                    }
                    return null;
                }""")
                if gst_cb:
                    await page.mouse.click(gst_cb['x'], gst_cb['y'])
                    await page.wait_for_timeout(800)
                    log("ok", "  -> Add GST Details checkbox ticked")
                else:
                    log("warn", "  Add GST Details checkbox not found")

                # ── GSTIN field ───────────────────────────────────────
                # Wait up to 3s for the GSTIN input to appear after checkbox tick
                gstin_coords = None
                for _gst_wait in range(15):
                    await page.wait_for_timeout(200)
                    gstin_coords = await page.evaluate("""() => {
                        // 1. formcontrolname contains 'gstin' or 'gst'
                        const byFc = Array.from(document.querySelectorAll('input')).find(i => {
                            const fc = (i.getAttribute('formcontrolname')||'').toLowerCase();
                            return fc.includes('gstin') || fc === 'gstno' || fc === 'gstnumber';
                        });
                        if (byFc && !byFc.readOnly && !byFc.disabled) {
                            byFc.scrollIntoView({block:'center'});
                            const r = byFc.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // 2. placeholder includes gstin
                        const byPh = Array.from(document.querySelectorAll('input')).find(i =>
                            (i.placeholder||'').toLowerCase().includes('gstin') ||
                            (i.placeholder||'').toLowerCase().includes('enter gstin')
                        );
                        if (byPh && !byPh.readOnly && !byPh.disabled) {
                            byPh.scrollIntoView({block:'center'});
                            const r = byPh.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // 3. nearest input to a label that says "GSTIN Number" or "GSTIN"
                        for (const el of document.querySelectorAll('label,mat-label,span,div')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if (t.includes('gstin') || t === 'gst number') {
                                const r = el.getBoundingClientRect();
                                if (r.width === 0) continue;
                                let best = null, bestDist = Infinity;
                                for (const inp of document.querySelectorAll('input')) {
                                    if (inp.readOnly || inp.disabled) continue;
                                    const ir = inp.getBoundingClientRect();
                                    if (ir.width === 0 || ir.height === 0) continue;
                                    const dist = Math.abs(ir.left - r.left) + Math.abs(ir.top - r.bottom);
                                    if (dist < bestDist) { bestDist = dist; best = inp; }
                                }
                                if (best && bestDist < 200) {
                                    best.scrollIntoView({block:'center'});
                                    const br = best.getBoundingClientRect();
                                    return {x:br.left+br.width/2, y:br.top+br.height/2};
                                }
                            }
                        }
                        return null;
                    }""")
                    if gstin_coords:
                        log("info", f"  GSTIN field appeared after {(_gst_wait+1)*200}ms")
                        break
                if gstin_coords:
                    # Fill with confirm loop: up to 3 attempts in case Angular re-renders
                    gstin_confirmed = False
                    for _gst_fill in range(3):
                        await page.mouse.click(gstin_coords['x'], gstin_coords['y'])
                        await page.wait_for_timeout(150)
                        await page.keyboard.press("Control+A")
                        await page.wait_for_timeout(60)
                        await page.keyboard.press("Delete")
                        await page.wait_for_timeout(60)
                        await page.keyboard.type(gst_val, delay=40)
                        await page.wait_for_timeout(200)
                        await page.evaluate("""({cx,cy,val}) => {
                            const el = document.elementFromPoint(cx,cy);
                            const inp = el&&(el.tagName==='INPUT'?el:el.closest('input'));
                            if (!inp) return;
                            const s = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
                            s.call(inp, val);
                            inp.dispatchEvent(new Event('input',  {bubbles:true}));
                            inp.dispatchEvent(new Event('change', {bubbles:true}));
                            inp.dispatchEvent(new Event('blur',   {bubbles:true}));
                        }""", {"cx": gstin_coords['x'], "cy": gstin_coords['y'], "val": gst_val})
                        # Poll up to 10×200ms to confirm field holds the value
                        for _chk in range(10):
                            await page.wait_for_timeout(200)
                            held = await page.evaluate("""({cx,cy,val}) => {
                                const el = document.elementFromPoint(cx,cy);
                                const inp = el&&(el.tagName==='INPUT'?el:el.closest('input'));
                                return inp ? inp.value.trim().toUpperCase() === val.trim().toUpperCase() : false;
                            }""", {"cx": gstin_coords['x'], "cy": gstin_coords['y'], "val": gst_val})
                            if held:
                                gstin_confirmed = True
                                break
                        if gstin_confirmed:
                            log("ok", f"  -> GSTIN filled and confirmed: {gst_val} (attempt {_gst_fill+1})")
                            break
                        log("warn", f"  GSTIN value lost — retrying (attempt {_gst_fill+1}/3)...")
                    if not gstin_confirmed:
                        log("warn", "  GSTIN could not be confirmed after 3 attempts")
                    # Let Angular fully settle before the next section touches the form
                    await page.wait_for_timeout(1200)
                else:
                    log("warn", "  GSTIN input not found — skipping")

            # ── Loan / Lease / Hypothecation Details ─────────────────
            loan_val         = str(D.get("loan_hypothecation","")).strip().lower()
            financier_name   = str(D.get("financier_name",   "")).strip()
            financier_branch = str(D.get("financier_branch", "")).strip()

            if loan_val == "yes":
                log("info", "Ticking Loan / Lease / Hypothecation Details checkbox")
                await page.evaluate("window.scrollBy(0, 300)")
                await page.wait_for_timeout(400)
                loan_cb = await page.evaluate("""() => {
                    // Match the checkbox whose OWN label/sibling mentions loan/hypothecation
                    // — NOT the whole container which includes all three checkbox labels.
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        const wrap = el.closest('label');
                        if (wrap) {
                            const t = (wrap.innerText||wrap.textContent||'').toLowerCase();
                            if (t.includes('loan') || t.includes('hypothecation') || t.includes('lease')) {
                                el.scrollIntoView({block:'center'});
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                            continue;
                        }
                        if (el.id) {
                            const lbl = document.querySelector('label[for="'+el.id+'"]');
                            if (lbl) {
                                const t = (lbl.innerText||lbl.textContent||'').toLowerCase();
                                if (t.includes('loan') || t.includes('hypothecation') || t.includes('lease')) {
                                    el.scrollIntoView({block:'center'});
                                    const r = el.getBoundingClientRect();
                                    if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                }
                            }
                        }
                        let sib = el.nextElementSibling || el.parentElement;
                        if (sib) {
                            const t = (sib.innerText||sib.textContent||'').toLowerCase().trim();
                            if (t.length < 60 && (t.includes('loan') || t.includes('hypothecation') || t.includes('lease'))) {
                                el.scrollIntoView({block:'center'});
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                        }
                    }
                    return null;
                }""")
                if loan_cb:
                    await page.mouse.click(loan_cb['x'], loan_cb['y'])
                    await page.wait_for_timeout(800)
                    log("ok", "  -> Loan/Hypothecation checkbox ticked")
                else:
                    log("warn", "  Loan/Hypothecation checkbox not found")

                # Wait for the Loan section to expand, then click Loan/Hypothecation radio
                # The portal uses mat-radio-button. Retry up to 8× waiting for it to appear.
                await page.wait_for_timeout(600)
                loan_radio = None
                for _lr in range(8):
                    loan_radio = await page.evaluate("""() => {
                        // Strategy 1: mat-radio-button whose text includes loan/hypothecation
                        for (const rb of document.querySelectorAll('mat-radio-button')) {
                            const t = (rb.innerText||rb.textContent||'').trim().toLowerCase();
                            if (t.includes('loan') || t.includes('hypothecation')) {
                                rb.scrollIntoView({block:'center'});
                                const r = rb.getBoundingClientRect();
                                if (r.width > 0) return {x:r.left+r.width/2, y:r.top+r.height/2};
                            }
                        }
                        // Strategy 2: input[type=radio] near loan/hypothecation label
                        for (const r of document.querySelectorAll('input[type="radio"]')) {
                            const lbl = r.closest('label') ||
                                        document.querySelector('label[for="'+r.id+'"]') ||
                                        r.parentElement;
                            const t = (lbl?.innerText||lbl?.textContent||'').toLowerCase();
                            if (t.includes('loan') || t.includes('hypothecation')) {
                                r.scrollIntoView({block:'center'});
                                const rect = r.getBoundingClientRect();
                                if (rect.width > 0) return {x:rect.left+rect.width/2, y:rect.top+rect.height/2};
                            }
                        }
                        // Strategy 3: any element whose exact text is 'Loan/Hypothecation'
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText||el.textContent||'').trim();
                            if (t === 'Loan/Hypothecation' || t === 'Loan / Hypothecation') {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x:r.left+r.width/2, y:r.top+r.height/2};
                            }
                        }
                        return null;
                    }""")
                    if loan_radio: break
                    await page.wait_for_timeout(400)
                if loan_radio:
                    await page.mouse.click(loan_radio['x'], loan_radio['y'])
                    await page.wait_for_timeout(600)
                    log("ok", "  -> Loan/Hypothecation radio selected")
                else:
                    log("warn", "  Loan/Hypothecation radio not found")

                # Wait for Financier Name/Branch fields to appear
                await page.wait_for_timeout(500)

                # Fill Financier Name — find by formcontrolname OR label proximity
                if financier_name:
                    fn_coords = await page.evaluate("""() => {
                        // formcontrolname
                        const byFc = Array.from(document.querySelectorAll('input')).find(i => {
                            const fc = (i.getAttribute('formcontrolname')||'').toLowerCase();
                            return fc.includes('financier') && (fc.includes('name') || fc === 'financiername' || fc === 'financier');
                        });
                        if (byFc && !byFc.readOnly && !byFc.disabled) {
                            byFc.scrollIntoView({block:'center'});
                            const r = byFc.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // placeholder
                        const byPh = Array.from(document.querySelectorAll('input')).find(i =>
                            (i.placeholder||'').toLowerCase().includes('financier name') ||
                            (i.placeholder||'').toLowerCase().includes('financier')
                        );
                        if (byPh && !byPh.readOnly && !byPh.disabled) {
                            byPh.scrollIntoView({block:'center'});
                            const r = byPh.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // label proximity
                        for (const el of document.querySelectorAll('label,mat-label,span,div')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if (t === 'financier name' || t === 'financier name *') {
                                const r = el.getBoundingClientRect();
                                if (r.width === 0) continue;
                                let best = null, bd = Infinity;
                                for (const inp of document.querySelectorAll('input')) {
                                    if (inp.readOnly||inp.disabled) continue;
                                    const ir = inp.getBoundingClientRect();
                                    if (!ir.width||!ir.height) continue;
                                    const d = Math.abs(ir.left-r.left)+Math.abs(ir.top-r.bottom);
                                    if (d < bd) { bd=d; best=inp; }
                                }
                                if (best && bd < 200) { best.scrollIntoView({block:'center'}); const br=best.getBoundingClientRect(); return {x:br.left+br.width/2,y:br.top+br.height/2}; }
                            }
                        }
                        return null;
                    }""")
                    if fn_coords:
                        await page.mouse.click(fn_coords['x'], fn_coords['y'])
                        await page.wait_for_timeout(120)
                        await page.keyboard.press("Control+A")
                        await page.keyboard.press("Delete")
                        await page.wait_for_timeout(60)
                        await page.keyboard.type(financier_name, delay=40)
                        await page.wait_for_timeout(200)
                        await page.evaluate("""({cx,cy,val})=>{const el=document.elementFromPoint(cx,cy);const inp=el&&(el.tagName==='INPUT'?el:el.closest('input'));if(!inp)return;const s=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;s.call(inp,val);inp.dispatchEvent(new Event('input',{bubbles:true}));inp.dispatchEvent(new Event('change',{bubbles:true}));inp.dispatchEvent(new Event('blur',{bubbles:true}));}""",
                                            {"cx": fn_coords['x'], "cy": fn_coords['y'], "val": financier_name})
                        log("ok", f"  -> Financier Name: {financier_name}")
                    else:
                        log("warn", "  Financier Name not found")

                # Fill Financier Branch — find by formcontrolname OR label proximity
                if financier_branch:
                    fb_coords = await page.evaluate("""() => {
                        // formcontrolname
                        const byFc = Array.from(document.querySelectorAll('input')).find(i => {
                            const fc = (i.getAttribute('formcontrolname')||'').toLowerCase();
                            return fc.includes('branch');
                        });
                        if (byFc && !byFc.readOnly && !byFc.disabled) {
                            byFc.scrollIntoView({block:'center'});
                            const r = byFc.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // placeholder
                        const byPh = Array.from(document.querySelectorAll('input')).find(i =>
                            (i.placeholder||'').toLowerCase().includes('branch')
                        );
                        if (byPh && !byPh.readOnly && !byPh.disabled) {
                            byPh.scrollIntoView({block:'center'});
                            const r = byPh.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // label proximity
                        for (const el of document.querySelectorAll('label,mat-label,span,div')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if (t === 'financier branch' || t === 'financier branch *') {
                                const r = el.getBoundingClientRect();
                                if (r.width === 0) continue;
                                let best = null, bd = Infinity;
                                for (const inp of document.querySelectorAll('input')) {
                                    if (inp.readOnly||inp.disabled) continue;
                                    const ir = inp.getBoundingClientRect();
                                    if (!ir.width||!ir.height) continue;
                                    const d = Math.abs(ir.left-r.left)+Math.abs(ir.top-r.bottom);
                                    if (d < bd) { bd=d; best=inp; }
                                }
                                if (best && bd < 200) { best.scrollIntoView({block:'center'}); const br=best.getBoundingClientRect(); return {x:br.left+br.width/2,y:br.top+br.height/2}; }
                            }
                        }
                        return null;
                    }""")
                    if fb_coords:
                        await page.mouse.click(fb_coords['x'], fb_coords['y'])
                        await page.wait_for_timeout(120)
                        await page.keyboard.press("Control+A")
                        await page.keyboard.press("Delete")
                        await page.wait_for_timeout(60)
                        await page.keyboard.type(financier_branch, delay=40)
                        await page.wait_for_timeout(200)
                        await page.evaluate("""({cx,cy,val})=>{const el=document.elementFromPoint(cx,cy);const inp=el&&(el.tagName==='INPUT'?el:el.closest('input'));if(!inp)return;const s=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;s.call(inp,val);inp.dispatchEvent(new Event('input',{bubbles:true}));inp.dispatchEvent(new Event('change',{bubbles:true}));inp.dispatchEvent(new Event('blur',{bubbles:true}));}""",
                                            {"cx": fb_coords['x'], "cy": fb_coords['y'], "val": financier_branch})
                        log("ok", f"  -> Financier Branch: {financier_branch}")
                    else:
                        log("warn", "  Financier Branch not found")

            # ── Create Proposal ───────────────────────────────────────
            log("info", "Clicking Create Proposal")
            await page.wait_for_timeout(400)
            cp_c = await page.evaluate("""() => {
                for (const el of document.querySelectorAll('button,input[type="submit"]')) {
                    if ((el.innerText||el.value||'').trim().toLowerCase().includes('create proposal')) {
                        el.scrollIntoView({block:'center'});
                        const r=el.getBoundingClientRect();
                        return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2}:null;
                    }
                }
                return null;
            }""")
            if cp_c:
                await page.mouse.click(cp_c['x'], cp_c['y'])
                log("ok", f"  -> Create Proposal clicked at ({cp_c['x']:.0f},{cp_c['y']:.0f})")
            else:
                log("warn", "  Create Proposal not found")
                await pause("Click 'Create Proposal' manually, then press ENTER")
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            await page.wait_for_timeout(2000)
            steps.append("proposal_created")
            log("ok", "══ SECTION DONE: Applicant Details / Proposal Created ══")

            # ── STEP 14: Payment ─────────────────────────────────────
            log("step", "STEP 14 — Payment")

            # Scroll to bottom so Pay Now button is in viewport
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(1000)

            # ── STEP 13: Proposal Number ──────────────────────────────
            # Read proposal number — retry up to 10x (page may take a few seconds to load)
            log("step", "STEP 13 — Saving Proposal Number")

            proposal_no = ""
            for _pn_try in range(25):
                proposal_no = await page.evaluate("""() => {
                    // Portal shows "Proposal Number: 4500823477" as a single element text.
                    // Strategy 1: find element containing "proposal number" and extract value after colon.
                    for (const el of document.querySelectorAll('*')) {
                        if ((el.children || []).length > 8) continue;
                        const t = (el.innerText || el.textContent || '').trim();
                        if (t.toLowerCase().includes('proposal number') ||
                            t.toLowerCase().includes('proposal no')) {
                            const m = t.match(/proposal\s*(?:number|no\.?)\s*[:\-]?\s*([A-Z0-9]{6,20})/i);
                            if (m && m[1]) return m[1].trim();
                        }
                    }
                    // Strategy 2: full body text scan
                    const allText = document.body.innerText || '';
                    const m = allText.match(/Proposal\s*(?:Number|No\.?)\s*[:\-]?\s*([A-Z0-9]{6,20})/i);
                    if (m && m[1]) return m[1].trim();
                    return null;
                }""") or ""
                if proposal_no:
                    break
                log("info", f"  Proposal number not visible yet, waiting... (attempt {_pn_try+1}/25)")
                await page.wait_for_timeout(1000)

            if proposal_no:
                log("ok", f"Proposal Number: {proposal_no}")
            else:
                log("warn", "Could not auto-read proposal number")
                proposal_no = input("      Type the Proposal Number from browser, press ENTER: ").strip()

            with open(LOG_PATH, "a") as f:
                f.write(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
                        f"{D['manufacturer_model']}  |  {D.get('engine_number','')}  |  {proposal_no}\n")
            log("ok", "Saved to proposal_log.txt")
            save_proposal_number(D, proposal_no)
            steps.append("proposal_saved")

            # Now click Pay Now — scroll to bottom, get fresh coords after scroll settles
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(800)
            pay_now_c = await page.evaluate("""() => {
                // Find Pay Now — scroll it into view, wait, then return fresh coords
                for (const el of document.querySelectorAll('*')) {
                    const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                    if (t !== 'pay now') continue;
                    el.scrollIntoView({block:'center'});
                    return true;
                }
                return false;
            }""")
            await page.wait_for_timeout(600)  # wait for scroll to settle
            pay_now_coords = await page.evaluate("""() => {
                for (const el of document.querySelectorAll('*')) {
                    const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                    if (t !== 'pay now') continue;
                    const r = el.getBoundingClientRect();
                    if (r.width > 0 && r.height > 0) {
                        return {x: r.left+r.width/2, y: r.top+r.height/2};
                    }
                }
                return null;
            }""")
            if pay_now_coords:
                await page.mouse.click(pay_now_coords['x'], pay_now_coords['y'])
                log("ok", "  -> Pay Now clicked")
                # Wait for payment page to fully load before attempting Advance PID
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(5000)

                # ── STEP 13 retry: read proposal number from payment page ──
                # The Premium Details card on the payment page shows proposal number
                # as bold plain text — try to capture it here if STEP 13 failed earlier
                if not proposal_no:
                    log("info", "  Retrying proposal number read from payment page...")
                    proposal_no_retry = await page.evaluate("""() => {
                        const isPN = s => /^[0-9]{8,15}$/.test((s||'').trim());
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if (t !== 'proposal number' && t !== 'proposal no') continue;
                            let card = el.parentElement;
                            for (let i = 0; i < 5; i++) {
                                if (!card) break;
                                const walker = document.createTreeWalker(
                                    card, NodeFilter.SHOW_TEXT);
                                let node;
                                while ((node = walker.nextNode())) {
                                    const v = node.nodeValue.trim();
                                    if (isPN(v)) return v;
                                }
                                for (const child of card.querySelectorAll('*')) {
                                    if (child === el) continue;
                                    const v = (child.innerText||'').trim();
                                    if (isPN(v)) return v;
                                }
                                card = card.parentElement;
                            }
                        }
                        return null;
                    }""") or ""
                    if proposal_no_retry:
                        proposal_no = proposal_no_retry
                        log("ok", f"  Proposal Number (payment page): {proposal_no}")
                        with open(LOG_PATH, "a") as f:
                            f.write(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
                                    f"{D['manufacturer_model']}  |  {D.get('engine_number','')}  |  {proposal_no}\n")
                        save_proposal_number(D, proposal_no)
                    else:
                        log("warn", "  Proposal number still not found on payment page")

            else:
                log("warn", "  Pay Now button not found")

            if D["auto_pay"]:
                # Wait for payment page to fully load before searching for Advance PID
                log("info", "Waiting for payment page to load before Advance PID...")
                try:
                    await page.wait_for_load_state("networkidle", timeout=25000)
                except Exception:
                    pass
                # Auto-complete payment — click Advance PID
                # Try on current page AND any newly opened page/tab
                adv_clicked = False

                async def try_click_advance_pid(pg):
                    # Strategy 1: Playwright locator by text
                    try:
                        loc = pg.get_by_text("Advance PID", exact=True)
                        if await loc.count() > 0:
                            await loc.first.scroll_into_view_if_needed()
                            await pg.wait_for_timeout(300)
                            await loc.first.click(force=True)
                            return True
                    except Exception:
                        pass
                    # Strategy 2: XPath — find element whose text is exactly "Advance PID"
                    try:
                        loc2 = pg.locator("xpath=//*[normalize-space(text())='Advance PID']")
                        if await loc2.count() > 0:
                            await loc2.first.scroll_into_view_if_needed()
                            await pg.wait_for_timeout(300)
                            await loc2.first.click(force=True)
                            return True
                    except Exception:
                        pass
                    # Strategy 3: JS — click parent of text node "Advance PID"
                    try:
                        done = await pg.evaluate("""() => {
                            const walker = document.createTreeWalker(
                                document.body, NodeFilter.SHOW_TEXT);
                            let node;
                            while ((node = walker.nextNode())) {
                                if (node.nodeValue.trim().toLowerCase() === 'advance pid') {
                                    const el = node.parentElement;
                                    el.scrollIntoView({block:'center'});
                                    el.click();
                                    return true;
                                }
                            }
                            return false;
                        }""")
                        if done:
                            return True
                    except Exception:
                        pass
                    return False

                # Poll until "Advance PID" text appears on the page —
                # the payment page can take several minutes to fully load.
                # Wait up to 10 minutes (120 × 5s) before giving up.
                log("info", "  Waiting for Advance PID to appear on page...")
                for _adv_wait in range(120):
                    adv_visible = await page.evaluate("""() => {
                        const walker = document.createTreeWalker(
                            document.body, NodeFilter.SHOW_TEXT);
                        let node;
                        while ((node = walker.nextNode())) {
                            if (node.nodeValue.trim().toLowerCase() === 'advance pid')
                                return true;
                        }
                        return false;
                    }""")
                    if adv_visible:
                        log("info", f"  Advance PID found (attempt {_adv_wait+1})")
                        break
                    # Also check other open tabs
                    try:
                        for pg in page.context.pages:
                            if pg == page:
                                continue
                            adv_visible = await pg.evaluate("""() => {
                                const walker = document.createTreeWalker(
                                    document.body, NodeFilter.SHOW_TEXT);
                                let node;
                                while ((node = walker.nextNode())) {
                                    if (node.nodeValue.trim().toLowerCase() === 'advance pid')
                                        return true;
                                }
                                return false;
                            }""")
                            if adv_visible:
                                break
                    except Exception:
                        pass
                    if adv_visible:
                        break
                    if _adv_wait % 6 == 5:   # log every 30s so user knows it's working
                        log("info", f"  Still waiting for Advance PID... ({(_adv_wait+1)*5}s)")
                    await page.wait_for_timeout(5000)

                # Now click it
                for _adv_try in range(5):
                    adv_clicked = await try_click_advance_pid(page)
                    if adv_clicked:
                        break
                    try:
                        for pg in page.context.pages:
                            if pg == page:
                                continue
                            adv_clicked = await try_click_advance_pid(pg)
                            if adv_clicked:
                                break
                    except Exception:
                        pass
                    if adv_clicked:
                        break
                    await page.wait_for_timeout(1000)

                if adv_clicked:
                    log("ok", "  -> Advance PID clicked")
                    # Wait for Advance PID page to fully load
                    try:
                        await page.wait_for_load_state("networkidle", timeout=15000)
                    except Exception:
                        pass
                    await page.wait_for_timeout(3000)

                    # ── Tick the PID Number checkbox matching Excel pid_number ──
                    pid_val = str(D.get("pid_number", "")).strip()
                    if pid_val:
                        log("info", f"  Selecting PID Number: {pid_val}")
                        pid_ticked = False

                        # ── DEBUG: dump all rows and cells to find PID structure ──
                        debug_rows = await page.evaluate("""() => {
                            const out = [];
                            document.querySelectorAll('tr').forEach((row, ri) => {
                                const cells = Array.from(row.querySelectorAll('td,th'))
                                    .map(c => (c.innerText||'').trim().slice(0,40));
                                const hasCb = !!row.querySelector('input[type="checkbox"]');
                                if (cells.length > 0)
                                    out.push({ri, cells, hasCb});
                            });
                            return out.slice(0,20);
                        }""")
                        for dr in debug_rows:
                            log("info", f"  [ROW {dr['ri']}] cb={dr['hasCb']} cells={dr['cells']}")

                        # ── Also dump all inputs on page ──
                        debug_inputs = await page.evaluate("""() => {
                            return Array.from(document.querySelectorAll('input')).slice(0,10).map(i => ({
                                type: i.type, ph: i.placeholder, val: (i.value||'').slice(0,30),
                                cls: (i.className||'').slice(0,40)
                            }));
                        }""")
                        for di in debug_inputs:
                            log("info", f"  [INP] type={di['type']} ph='{di['ph']}' val='{di['val']}' cls='{di['cls']}'")

                        async def tick_pid_on_page(pg):
                            """Type PID in search box to filter table, then click its checkbox."""
                            # Step 1: type PID number into the Search box to filter the table
                            try:
                                search_inp = await pg.query_selector('input[placeholder*="Search" i], input[type="search"]')
                                if search_inp:
                                    await search_inp.triple_click()
                                    await search_inp.type(pid_val, delay=80)
                                    await pg.wait_for_timeout(1500)
                                    log("info", f"  Typed '{pid_val}' into search box")
                                else:
                                    log("info", "  No search box found")
                            except Exception as ex:
                                log("info", f"  Search box error: {ex}")

                            for _s in range(5):
                                # Step 2: scroll checkbox into view
                                scrolled = await pg.evaluate("""(pidVal) => {
                                    const rows = document.querySelectorAll('tr');
                                    for (const row of rows) {
                                        for (const cell of row.querySelectorAll('td, th')) {
                                            const t = (cell.innerText||'').trim();
                                            if (t === pidVal) {
                                                const cb = row.querySelector('input[type="checkbox"]');
                                                if (cb) { cb.scrollIntoView({block:'center'}); return 'cb-found'; }
                                                return 'row-found-no-cb';
                                            }
                                        }
                                    }
                                    return 'not-found';
                                }""", pid_val)
                                log("info", f"  PID scroll attempt {_s+1}: {scrolled}")
                                if scrolled == 'not-found':
                                    await pg.wait_for_timeout(1000)
                                    continue
                                if scrolled == 'row-found-no-cb':
                                    # No checkbox in tr — try clicking the row itself
                                    coords = await pg.evaluate("""(pidVal) => {
                                        const rows = document.querySelectorAll('tr');
                                        for (const row of rows) {
                                            for (const cell of row.querySelectorAll('td,th')) {
                                                if ((cell.innerText||'').trim() === pidVal) {
                                                    const r = row.getBoundingClientRect();
                                                    if (r.width > 0) return {x: r.left+20, y: r.top+r.height/2};
                                                }
                                            }
                                        }
                                        return null;
                                    }""", pid_val)
                                    if coords:
                                        await pg.wait_for_timeout(400)
                                        await pg.mouse.click(coords['x'], coords['y'])
                                        return True
                                # Step 3: re-read fresh coords after scroll settles
                                await pg.wait_for_timeout(500)
                                coords = await pg.evaluate("""(pidVal) => {
                                    const rows = document.querySelectorAll('tr');
                                    for (const row of rows) {
                                        for (const cell of row.querySelectorAll('td, th')) {
                                            if ((cell.innerText||'').trim() === pidVal) {
                                                const cb = row.querySelector('input[type="checkbox"]');
                                                if (cb) {
                                                    const cr = cb.getBoundingClientRect();
                                                    if (cr.width > 0)
                                                        return {x: cr.left+cr.width/2, y: cr.top+cr.height/2};
                                                }
                                            }
                                        }
                                    }
                                    return null;
                                }""", pid_val)
                                if coords:
                                    await pg.mouse.click(coords['x'], coords['y'])
                                    await pg.wait_for_timeout(300)
                                    return True
                                await pg.wait_for_timeout(1000)
                            return False

                        # Try on current page first
                        pid_ticked = await tick_pid_on_page(page)
                        # If not found, try all other open tabs
                        if not pid_ticked:
                            try:
                                for pg in page.context.pages:
                                    if pg == page:
                                        continue
                                    log("info", f"  Trying PID on tab: {pg.url}")
                                    pid_ticked = await tick_pid_on_page(pg)
                                    if pid_ticked:
                                        break
                            except Exception:
                                pass

                        if pid_ticked:
                            log("ok", f"  -> PID {pid_val} checkbox ticked")
                        else:
                            log("warn", f"  PID {pid_val} not found — tick manually")
                    else:
                        log("info", "  No PID Number in Excel — skipping PID checkbox")

                    await page.wait_for_timeout(1000)

                    # ── Click Pay button after PID checkbox ticked ──
                    log("info", "Clicking Pay button...")
                    pay_clicked = False
                    for _pay in range(5):
                        pay_coords = await page.evaluate("""() => {
                            for (const el of document.querySelectorAll('button, input[type="submit"], a')) {
                                const t = (el.innerText||el.textContent||el.value||'').trim().toLowerCase();
                                if (t === 'pay') {
                                    el.scrollIntoView({block:'center'});
                                    const r = el.getBoundingClientRect();
                                    if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                }
                            }
                            return null;
                        }""")
                        if pay_coords:
                            await page.wait_for_timeout(400)
                            # Re-read coords fresh after scroll
                            pay_coords2 = await page.evaluate("""() => {
                                for (const el of document.querySelectorAll('button, input[type="submit"], a')) {
                                    const t = (el.innerText||el.textContent||el.value||'').trim().toLowerCase();
                                    if (t === 'pay') {
                                        const r = el.getBoundingClientRect();
                                        if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                    }
                                }
                                return null;
                            }""")
                            if pay_coords2:
                                await page.mouse.click(pay_coords2['x'], pay_coords2['y'])
                                log("ok", "  -> Pay clicked")
                                pay_clicked = True
                                break
                        await page.wait_for_timeout(1000)
                    if not pay_clicked:
                        log("warn", "  Pay button not found — click manually")
                        await pause("Click 'Pay' manually, then press ENTER")

                    log("ok", "Payment initiated!")
                    steps.append("payment_done")
                else:
                    log("warn", "Advance PID button not found — complete payment manually")
                    await pause("Complete payment manually in the browser, then press ENTER")
                    steps.append("payment_manual")
            else:
                log("info", "auto_pay = FALSE — complete payment manually")
                await pause("Complete payment manually in the browser, then press ENTER")
                steps.append("payment_manual")

            save_result(D, "SUCCESS", proposal_no, steps)

            print(f"""
+--------------------------------------------------------------+
|  ALL DONE!                                                   |
|  Vehicle  : {D['manufacturer_model']:<51}|
|  Proposal : {proposal_no:<51}|
|  Saved    : proposal_log.txt + details.xlsx (Results tab)    |
+--------------------------------------------------------------+""")

            # ── Capture Policy Number from success page ──────────────
            log("info", "Waiting for success page and Policy Number...")
            try:
                await page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)

            # Retry up to 60x × 2s = 2 minutes — success page can take time.
            # Uses 4 strategies so no layout variation is missed.
            policy_no = ""
            for _pol in range(60):
                policy_no = await page.evaluate("""() => {
                    // Policy number pattern: 4digits/9digits/2digits/3digits
                    function isPolicyNo(v) {
                        return v && v.length > 6 &&
                               /[0-9]/.test(v) &&
                               v.includes('/') &&
                               !/policy number/i.test(v);
                    }

                    // ── Strategy 1: find label "Policy Number", read next sibling ──
                    let labelEl = null, labelSize = Infinity;
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim();
                        if (t.toLowerCase() === 'policy number') {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0) {
                                const sz = r.width * r.height;
                                if (sz < labelSize) { labelEl = el; labelSize = sz; }
                            }
                        }
                    }
                    if (labelEl) {
                        // Check next siblings
                        let sib = labelEl.nextElementSibling;
                        for (let i = 0; i < 5; i++) {
                            if (!sib) break;
                            const v = (sib.innerText||sib.textContent||'').trim();
                            if (isPolicyNo(v)) return v;
                            sib = sib.nextElementSibling;
                        }
                        // Walk up parent chain, collect all text nodes, find value after label
                        let row = labelEl.parentElement;
                        for (let i = 0; i < 6; i++) {
                            if (!row) break;
                            const texts = [];
                            const walker = document.createTreeWalker(row, NodeFilter.SHOW_TEXT);
                            let node;
                            while ((node = walker.nextNode())) {
                                const v = node.nodeValue.trim();
                                if (v) texts.push(v);
                            }
                            const idx = texts.findIndex(t => t.toLowerCase() === 'policy number');
                            if (idx !== -1) {
                                for (let j = idx+1; j < Math.min(idx+5, texts.length); j++) {
                                    if (isPolicyNo(texts[j])) return texts[j];
                                }
                            }
                            row = row.parentElement;
                        }
                    }

                    // ── Strategy 2: scan ALL text nodes for policy number pattern ──
                    const walker2 = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
                    let node2;
                    while ((node2 = walker2.nextNode())) {
                        const v = node2.nodeValue.trim();
                        // Match format like 3005/433241549/00/000
                        if (v.match(/^[0-9]{4}[/][0-9]{9}[/][0-9]{2}[/][0-9]{3}$/)) return v;
                    }

                    // ── Strategy 3: scan visible elements for policy-like value ──
                    for (const el of document.querySelectorAll('span,p,div,td,h1,h2,h3,h4,strong,b')) {
                        const v = (el.innerText||'').trim();
                        if (v.match(/^[0-9]{4}[/][0-9]+[/][0-9]{2}[/][0-9]{3}$/)) return v;
                    }

                    return null;
                }""") or ""
                if policy_no:
                    log("ok", f"  Policy Number found (attempt {_pol+1}): {policy_no}")
                    break
                if _pol % 5 == 4:
                    log("info", f"  Still waiting for Policy Number... ({(_pol+1)*2}s)")
                await page.wait_for_timeout(2000)

            if policy_no:
                log("ok", f"  Policy Number: {policy_no}")
                save_policy_number(D, policy_no)
            else:
                log("warn", "  Policy Number not found — saving empty")

            # ── Click "Another Policy" to go back for next row ──────

            another_clicked = False
            for _ap in range(15):
                # Strategy 1: Playwright get_by_text (most reliable)
                try:
                    loc = page.get_by_text("Another Policy", exact=True)
                    if await loc.count() > 0:
                        await loc.first.scroll_into_view_if_needed()
                        await page.wait_for_timeout(400)
                        await loc.first.click(force=True)
                        log("ok", "  -> Another Policy clicked (get_by_text)")
                        another_clicked = True
                        break
                except Exception:
                    pass
                # Strategy 2: XPath exact text
                try:
                    loc2 = page.locator("xpath=//*[normalize-space(text())='Another Policy']")
                    if await loc2.count() > 0:
                        await loc2.first.scroll_into_view_if_needed()
                        await page.wait_for_timeout(400)
                        await loc2.first.click(force=True)
                        log("ok", "  -> Another Policy clicked (xpath)")
                        another_clicked = True
                        break
                except Exception:
                    pass
                # Strategy 3: JS scan all elements including div/span
                coords = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                        if (t === 'another policy') {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0) {
                                el.scrollIntoView({block:'center'});
                                return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                        }
                    }
                    return null;
                }""")
                if coords:
                    await page.wait_for_timeout(400)
                    await page.mouse.click(coords['x'], coords['y'])
                    log("ok", "  -> Another Policy clicked (JS scan)")
                    another_clicked = True
                    break
                await page.wait_for_timeout(1000)
            if not another_clicked:
                log("warn", "  Another Policy button not found — waiting extra 30s and retrying once")
                await page.wait_for_timeout(30000)
                try:
                    loc = page.get_by_text("Another Policy")
                    if await loc.count() > 0:
                        await loc.first.click(force=True)
                        log("ok", "  -> Another Policy clicked (late retry)")
                        another_clicked = True
                except Exception:
                    pass
                if not another_clicked:
                    log("warn", "  Another Policy not found after all retries — skipping")

            # Wait for the details page to reload
            try:
                await page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)
            return D["_excel_row"] + 1  # signal to process next row

        except Exception as e:
            log("warn", f"Error: {e}")
            save_result(D, "FAILED", proposal_no, steps, str(e))
            await pause("Check the browser. Press ENTER to continue to next row.")
            return D.get("_excel_row", 4) + 1  # skip failed row, try next


# ═══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # ── Read first row to preview and confirm before starting ────
    D_first = read_details(start_row=4)

    print("\n  First record from details.xlsx:")
    print(f"     Vehicle      : {D_first['vehicle_reg_number']}")
    print(f"     Mobile       : {D_first['mobile_number']}")
    print(f"     Policy Type  : {D_first.get('policy_type','')}")
    print(f"     Ownership    : {D_first.get('ownership','') or '(not set)'}")
    print(f"     SAOD         : {D_first.get('saod','') or 'no'}")
    print(f"     Rollover Reg : {D_first.get('rollover_reg_number','') or '(N/A for New)'}")
    print(f"     Reg Type     : {D_first.get('current_reg_type','')}")
    print(f"     RTO          : {D_first.get('rto','')}")
    print(f"     Make/Model   : {D_first.get('manufacturer_model','')}")
    print(f"     Mfg Year     : {D_first.get('manufacturing_year','')}")
    print(f"     Tenure       : {D_first.get('tenure','')}")
    print(f"     IDV (Rs)     : {D_first.get('idv','')}")
    print(f"     PAN          : {D_first['pan_number']}")
    print(f"     DOB          : {D_first['dob']}")
    print(f"     GST          : {D_first['gst_number'] or '(not provided)'}")
    print(f"     Add-ons      : {', '.join(D_first['addons']) or '(none)'}")
    print(f"     Auto Pay     : {D_first['auto_pay']}")

    print("""
  ──────────────────────────────────────────────────────────────
  BEFORE pressing ENTER, make sure:

  1. Chrome is open with remote debugging. Start it with Win+R:
     "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
     --remote-debugging-port=9222 --user-data-dir="C:\\chrome-nysa"

  2. You are logged in to nysa.icicilombard.com (home page visible)
  ──────────────────────────────────────────────────────────────
    """)

    input("  Press ENTER to connect to your Chrome and start...")

    # ── Loop through all data rows ───────────────────────────────
    next_row = 4
    row_count = 0
    while True:
        try:
            D = read_details(start_row=next_row)
        except SystemExit:
            print("\n  No more data rows found. All done!")
            break

        row_count += 1

        # ── Skip if Proposal Number already filled ───────────────
        existing_proposal = str(D.get("proposal_number", "") or "").strip()
        if existing_proposal:
            print(f"\n[SKIP] Row {D['_excel_row']} already has Proposal Number: {existing_proposal} — skipping")
            next_row = D["_excel_row"] + 1
            continue

        # ── Skip if Engine/Chassis already marked green (full flow done) ──
        if is_engine_chassis_green(D):
            print(f"\n[SKIP] Row {D['_excel_row']} engine/chassis already green (policy issued) — skipping")
            next_row = D["_excel_row"] + 1
            continue

        print(f"\n{'='*60}")
        print(f"  Processing row {D['_excel_row']} (record #{row_count})")
        print(f"  Vehicle: {D['vehicle_reg_number']}  |  Model: {D['manufacturer_model'][:40]}")
        print(f"{'='*60}")

        next_row = asyncio.run(run(D, skip_to_step4=(row_count > 1)))

        if next_row is None:
            # run() didn't return a next row — stop
            print("\n  Processing stopped.")
            break

        # Check if there are more rows
        try:
            read_details(start_row=next_row)  # just check — will raise SystemExit if no data
        except SystemExit:
            print(f"\n  All {row_count} record(s) processed successfully!")
            break

    print("\n  Script finished.")