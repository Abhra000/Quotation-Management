"""
╔══════════════════════════════════════════════════════════════╗
║   ICICI NYSA — EV (Electric Bike) Insurance Automation       ║
║   Reads ONLY from OnlyEV.xlsx                                ║
║   Selects "Electric Bike" on Motor screen                    ║
╚══════════════════════════════════════════════════════════════╝

STEP 1 — Start Chrome with remote debugging (run this ONCE):
  Press  Win + R  on your keyboard, paste this and press Enter:

  "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\\chrome-nysa"

  (A new Chrome window will open)

STEP 2 — In that Chrome window:
  Go to:  https://nysa.icicilombard.com
  Login with your ID, Password, Captcha, OTP — just like normal.
  Make sure you are on the HOME page.

STEP 3 — Fill OnlyEV.xlsx row 4 (or later) with your EV details.

STEP 4 — Run this script:
  python OnlyEV.py

The script attaches to your already-logged-in Chrome.
No new tab. No new login. Starts directly from home page.
"""

import asyncio
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright

# ── Paths ──────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "OnlyEV.xlsx"
LOG_PATH   = BASE_DIR / "proposal_log.txt"

# ── Must match the port you used to start Chrome ──────────────────
CHROME_PORT = 9222


# ═══════════════════════════════════════════════════════════════════
#  READ EXCEL
# ═══════════════════════════════════════════════════════════════════

def read_details(start_row=4):
    if not EXCEL_PATH.exists():
        print(f"\n  Cannot find: {EXCEL_PATH}")
        print("  Make sure OnlyEV.xlsx is in the same folder.")
        input("\nPress ENTER to exit..."); raise SystemExit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "Details" not in wb.sheetnames:
        print("  Sheet 'Details' not found in OnlyEV.xlsx")
        input("\nPress ENTER to exit..."); raise SystemExit(1)

    ws = wb["Details"]

    # ── Label map: handles any variation of header text ──────────
    # Strip special chars, lowercase, then match
    label_map = {
        # ─── EV Vehicle Details fields (Step 1d page) ───
        "transaction type"                            : "transaction_type",
        "transaction"                                 : "transaction_type",
        "battery number"                              : "battery_number",
        "battery no"                                  : "battery_number",
        "battery"                                     : "battery_number",
        "purchase date"                               : "purchase_date",
        "customer type"                               : "customer_type",
        "customer category"                           : "customer_type",
        "manufacturer"                                : "manufacturer",
        "model"                                       : "model",
        "watt"                                        : "watt",
        "wattage"                                     : "watt",
        "year of manufacture"                         : "year_of_manufacture",
        "year of mfg"                                 : "year_of_manufacture",
        "manufacturing year"                          : "year_of_manufacture",
        "mfg year"                                    : "year_of_manufacture",
        "policy start date"                           : "policy_start_date",
        "policy start"                                : "policy_start_date",
        "policy end date"                             : "policy_end_date",
        "policy end"                                  : "policy_end_date",
        "unique identification number of asset"       : "uin_asset",
        "unique identification number"                : "uin_asset",
        "uin"                                         : "uin_asset",
        "uin of asset"                                : "uin_asset",
        "invoice number"                              : "invoice_number",
        "invoice no"                                  : "invoice_number",
        "invoice"                                     : "invoice_number",
        "customer state"                              : "customer_state",
        # ─── Quote / Plan ───
        "idv"                                         : "idv",
        "insured declared value idv"                  : "idv",
        "insured declared value"                      : "idv",
        "insured depreciation value"                  : "idv",
        "insured depreciation value idv"              : "idv",
        "mandatory covers"                            : "mandatory_covers",
        "mandatory covers yes/no"                     : "mandatory_covers",
        "mandatory covers (yes/no)"                   : "mandatory_covers",
        "mandatory covers yes no"                     : "mandatory_covers",
        # ─── Add-ons / Discount ───
        "additional discount"                         : "additional_discount",
        "additional discount yes/no"                  : "additional_discount",
        "additional discount (yes/no)"                : "additional_discount",
        "additional discount yes no"                  : "additional_discount",
        "additional discounts dropdown"               : "additional_discounts_dropdown",
        "additional discount dropdown"                : "additional_discounts_dropdown",
        "additional discounts"                        : "additional_discounts_dropdown",
        "discount type"                               : "additional_discounts_dropdown",
        "loading%"                                    : "loading_pct",
        "loading %"                                   : "loading_pct",
        "loading"                                     : "loading_pct",
        # ─── KYC ───
        "gst number"                                  : "gst_number",
        "gstin"                                       : "gst_number",
        "gst state"                                   : "gst_state",
        "kyc pincode"                                 : "kyc_pincode",
        "kyc pin code"                                : "kyc_pincode",
        "kyc pin"                                     : "kyc_pincode",
        "pan number"                                  : "pan_number",
        "pan"                                         : "pan_number",
        "aadhaar no"                                  : "aadhaar_number",
        "aadhaar"                                     : "aadhaar_number",
        # ─── Insured details ───
        "insured name"                                : "insured_name",
        "date of birth"                               : "dob",
        "dob"                                         : "dob",
        "dob dd/mm/yyyy"                              : "dob",
        "dob dd mm yyyy"                              : "dob",
        "address line"                                : "address_line",
        "address line 1"                              : "address_line",
        "mobile"                                      : "mobile_number",
        "mobile number"                               : "mobile_number",
        "phone"                                       : "mobile_number",
        "email"                                       : "email",
        "pincode"                                     : "pin_code",
        "pin code"                                    : "pin_code",
        "city"                                        : "city",
        "city/district"                               : "city",
        "city / district"                             : "city",
        "city district"                               : "city",
        "state"                                       : "state",
        # ─── Loan / Financier ───
        "loan / lease / hypothecation details"        : "loan_hypothecation",
        "loan lease hypothecation details"            : "loan_hypothecation",
        "loan/lease/hypothecation details"            : "loan_hypothecation",
        "loan hypothecation"                          : "loan_hypothecation",
        "hypothecation"                               : "loan_hypothecation",
        "financier name"                              : "financier_name",
        "financier branch"                            : "financier_branch",
        # ─── Payment ───
        "auto pay"                                    : "auto_pay",
        "auto pay true/false"                         : "auto_pay",
        "auto pay (true/false)"                       : "auto_pay",
        "auto pay true false"                         : "auto_pay",
        "autopay"                                     : "auto_pay",
        # ─── Output (filled by automation) ───
        "pid number"                                  : "pid_number",
        "pid no"                                      : "pid_number",
        "proposal number"                             : "proposal_number",
        "policy number"                               : "policy_number",
        # ─── Add-ons (kept for compatibility, not used in EV flow by default) ───
        "add-ons comma separated"                     : "addons",
        "add ons comma separated"                     : "addons",
        "add-ons (comma separated)"                   : "addons",
        "addons"                                      : "addons",
        "add-ons"                                     : "addons",
        "add ons"                                     : "addons",
    }

    positional_map = {
        # EV Vehicle Details (col 1-13)
        1:"transaction_type",    2:"battery_number",       3:"purchase_date",
        4:"customer_type",       5:"manufacturer",         6:"model",
        7:"watt",                8:"year_of_manufacture",  9:"policy_start_date",
        10:"policy_end_date",    11:"uin_asset",           12:"invoice_number",
        13:"customer_state",
        # Quote (col 14-15)
        14:"idv",                15:"mandatory_covers",
        # Add-ons / Discount (col 16-18)
        16:"additional_discount", 17:"additional_discounts_dropdown", 18:"loading_pct",
        # KYC (col 19-20)
        19:"gst_number",         20:"kyc_pincode",
        # Insured details (col 21-28)
        21:"insured_name",       22:"dob",                 23:"address_line",
        24:"mobile_number",      25:"email",               26:"pin_code",
        27:"city",               28:"state",
        # Loan / Financier (col 29-31)
        29:"loan_hypothecation", 30:"financier_name",      31:"financier_branch",
        # Payment (col 32)
        32:"auto_pay",
        # Output (col 33-35)
        33:"pid_number",         34:"proposal_number",     35:"policy_number",
    }

    import re as _re
    headers = {}
    header_row_found = False
    for col in range(1, 37):  # cols 1-36: EV layout — PID(33), Proposal(34), Policy(35)
        val = ws.cell(row=3, column=col).value
        if val:
            header_row_found = True
            clean = _re.sub(r"[*()\[\]₹]", " ", str(val))
            clean = _re.sub(r"\s+", " ", clean).strip().lower()
            field = label_map.get(clean)
            if not field:
                c2 = clean.replace(" ","")
                for k,v in label_map.items():
                    if c2 == k.replace(" ",""): field = v; break
            if not field:
                for k,v in label_map.items():
                    if clean.startswith(k) and (len(clean)==len(k) or clean[len(k)]==" "):
                        field = v; break
            if not field:
                field = _re.sub(r"[^a-z0-9]+","_",clean).strip("_")
            headers[col] = field
            print(f"  [HEADER] col {col}: {repr(str(val))} → {field}")

    if not header_row_found or not headers:
        print("  [INFO] Using positional column map")
        headers = positional_map.copy()

    data = {}
    max_col = max(headers.keys()) + 1 if headers else 22
    for row in range(start_row, ws.max_row + 1):
        row_vals = [ws.cell(row=row, column=c).value for c in range(1, max_col)]
        if any(v and str(v).strip() for v in row_vals):
            for col, field in headers.items():
                raw = ws.cell(row=row, column=col).value
                if hasattr(raw, 'strftime'):
                    data[field] = raw.strftime("%d/%m/%Y")
                else:
                    data[field] = str(raw).strip() if raw is not None else ""
            print(f"  [INFO] Reading data from row {row}")
            data["_excel_row"] = row
            break

    if not data:
        print("  No data found in OnlyEV.xlsx")
        input("\nPress ENTER to exit..."); raise SystemExit(1)

    # Store the Excel row number so we can write proposal number back later

    # ── Pre-set EV-specific defaults so existing Step 1 logic works ──
    # The original try.py used motor_category + motor_type. For EV we
    # hard-set these so Step 1 picks "Electric Bike". Policy type is
    # always "New" for EV transactions on this portal.
    data["motor_category"] = "Electric Bike"
    data["motor_type"]     = ""           # Electric Bike has no Package/TP toggle
    data["policy_type"]    = data.get("transaction_type", "New").strip() or "New"
    if data["policy_type"].lower() in ("renewal", "rollover"):
        data["policy_type"] = "Rollover"
    else:
        data["policy_type"] = "New"

    # ── Clean / normalize EV fields ──
    data["transaction_type"]    = str(data.get("transaction_type", "New") or "New").strip()
    data["battery_number"]      = str(data.get("battery_number", "") or "").strip()
    data["purchase_date"]       = str(data.get("purchase_date", "") or "").strip()
    data["customer_type"]       = str(data.get("customer_type", "Individual") or "Individual").strip()
    data["manufacturer"]        = str(data.get("manufacturer", "") or "").strip()
    data["model"]               = str(data.get("model", "") or "").strip()
    data["watt"]                = str(data.get("watt", "") or "").split(".")[0].strip()
    data["year_of_manufacture"] = str(data.get("year_of_manufacture", "") or "").split(".")[0].strip()
    data["policy_start_date"]   = str(data.get("policy_start_date", "") or "").strip()
    data["policy_end_date"]     = str(data.get("policy_end_date", "") or "").strip()
    data["uin_asset"]           = str(data.get("uin_asset", "") or "").strip()
    data["invoice_number"]      = str(data.get("invoice_number", "") or "").strip()
    data["customer_state"]      = str(data.get("customer_state", "") or "").strip()
    data["state"]               = str(data.get("state", "") or data.get("customer_state", "") or "").strip()

    # ── Common fields shared with try.py downstream logic ──
    data["pan_number"]          = str(data.get("pan_number", "") or "").upper()
    data["gst_number"]          = str(data.get("gst_number", "") or "").upper().strip()
    data["gst_state"]           = str(data.get("gst_state", "") or data.get("customer_state", "") or "").strip()
    data["pin_code"]            = str(data.get("pin_code", "") or "").split(".")[0].strip()
    data["kyc_pincode"]         = str(data.get("kyc_pincode", "") or "").split(".")[0].strip()
    data["mobile_number"]       = str(data.get("mobile_number", "") or "").split(".")[0].strip()
    data["email"]               = str(data.get("email", "") or "").strip()
    data["insured_name"]        = str(data.get("insured_name", "") or "").strip()
    data["dob"]                 = str(data.get("dob", "") or "").strip()
    data["address_line"]        = str(data.get("address_line", "") or "").strip()
    data["city"]                = str(data.get("city", "") or "").strip()
    data["loan_hypothecation"]  = str(data.get("loan_hypothecation", "") or "").strip()
    data["financier_name"]      = str(data.get("financier_name", "") or "").strip()
    data["financier_branch"]    = str(data.get("financier_branch", "") or "").strip()
    data["idv"]                 = (str(data.get("idv", "") or "")
                                   .replace(",", "").replace("₹", "").replace(" ", "")
                                   .split(".")[0].strip())

    # ── Mandatory Covers + Add-ons (none expected by default in EV) ──
    data["mandatory_covers"]    = str(data.get("mandatory_covers", "no") or "no").strip().lower() == "yes"
    raw_addons                  = data.get("addons", "") or ""
    data["addons"]              = [a.strip() for a in str(raw_addons).split(",") if a.strip()]

    # ── Discount + payment ──
    data["additional_discount"] = str(data.get("additional_discount", "no") or "no").strip().lower() == "yes"
    data["additional_discounts_dropdown"] = str(data.get("additional_discounts_dropdown", "") or "").strip()
    data["loading_pct"]         = str(data.get("loading_pct", "") or "").replace("%", "").strip()
    data["auto_pay"]            = str(data.get("auto_pay", "false") or "false").strip().lower() == "true"

    # ── Compatibility: stub fields not present in EV Excel but referenced
    # by downstream try.py logic so .get() lookups never KeyError ──
    data.setdefault("vehicle_reg_number", "NEW")
    data.setdefault("engine_number", "")
    data.setdefault("chassis_number", "")
    data.setdefault("rto", "")
    data.setdefault("manufacturer_model",
                    f"{data['manufacturer']} - {data['model']}".strip(" -"))
    data.setdefault("manufacturing_year", data["year_of_manufacture"])
    data.setdefault("current_reg_type",
                    "Corporate" if data["customer_type"].lower().startswith("corp") else "Individual")
    data.setdefault("tenure", "1 year")
    data.setdefault("ownership", "")
    data.setdefault("saod", "")
    data.setdefault("rollover_reg_number", "")
    data.setdefault("aadhaar_number", "")
    data.setdefault("advanced_pid", "")

    # ── Debug: show what was read ─────────────────────────────────
    print("  [DEBUG] Fields read from OnlyEV.xlsx:")
    for k, v in data.items():
        if k != "addons":
            print(f"           {k:<22} = {repr(v)}")
    print(f"           {'addons':<22} = {data['addons']}")

    return data


# ═══════════════════════════════════════════════════════════════════
#  SAVE RESULT
# ═══════════════════════════════════════════════════════════════════

def save_proposal_number(D, proposal_no):
    """Write proposal number back to col 34 of the Details sheet (EV layout)."""
    try:
        excel_row = D.get("_excel_row", 4)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Details"]
        ws.cell(row=excel_row, column=34, value=proposal_no)
        wb.save(EXCEL_PATH)
        print(f"  Proposal number '{proposal_no}' saved to Details row {excel_row}, col 34")
    except Exception as e:
        print(f"  Could not save proposal number to Excel: {e}")


def save_policy_number(D, policy_no):
    """Write policy number back to col 35 of the Details sheet (EV layout)."""
    try:
        excel_row = D.get("_excel_row", 4)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Details"]
        ws.cell(row=excel_row, column=35, value=policy_no)
        wb.save(EXCEL_PATH)
        print(f"  Policy number '{policy_no}' saved to Details row {excel_row}, col 35")
    except Exception as e:
        print(f"  Could not save policy number to Excel: {e}")


def save_result(D, status, proposal_no, steps, error=""):
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Results Log"]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        s  = Side(style="thin", color="BFBFBF")
        b  = Border(left=s, right=s, top=s, bottom=s)
        bg = "E2EFDA" if status == "SUCCESS" else "FCE4D6"
        r  = max(ws.max_row + 1, 3)
        # For EV, the unique identifier is Battery Number, not Vehicle Reg
        ev_id = D.get("battery_number") or D.get("uin_asset") or "-"
        for col, val in enumerate([
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            ev_id, status, proposal_no,
            " -> ".join(steps), error,
            "YES" if D.get("auto_pay") else "NO",
        ], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border    = b
        wb.save(EXCEL_PATH)
        print("  Result saved to OnlyEV.xlsx (Results Log tab)")
    except Exception as e:
        print(f"  Could not write to Excel: {e}")


def mark_engine_chassis_green(D):
    """For EV, no engine/chassis columns exist — mark the Battery Number cell green
    after a successful policy issuance, so the row can be skipped on re-runs."""
    try:
        from openpyxl.styles import PatternFill
        excel_row = D.get("_excel_row", 4)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Details"]
        green = PatternFill("solid", fgColor="92D050")   # bright green
        # col 2 = battery_number (EV layout)
        ws.cell(row=excel_row, column=2).fill = green
        wb.save(EXCEL_PATH)
    except Exception as e:
        print(f"  Could not mark battery number green in Excel: {e}")


def is_engine_chassis_green(D):
    """For EV: returns True if the Battery Number cell is green (=> policy already issued)."""
    try:
        excel_row = D.get("_excel_row", 4)
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb["Details"]
        cell = ws.cell(row=excel_row, column=2)   # battery_number cell
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            rgb = str(cell.fill.fgColor.rgb).upper()
            return "92D050" in rgb
    except Exception:
        pass
    return False


# ═══════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════

def log(kind, msg):
    icons = {"step":"\n[STEP]","ok":"  [OK] ","warn":"  [WARN]","fill":"  [FILL]","info":"  [INFO]"}
    print(f"{icons.get(kind,'      ')} {msg}")


async def pause(msg="Done? Press ENTER to continue"):
    print(f"\n  [PAUSE] {msg}")
    await asyncio.get_event_loop().run_in_executor(
        None, input, "          Press ENTER: ")


async def smart_click(page, keywords, label, timeout=12000):
    """
    Tries every possible strategy to click an element on an Angular portal.
    If nothing works: prints what IS on the page, then asks user to click manually.
    """
    kws = [keywords] if isinstance(keywords, str) else keywords

    for kw in kws:
        kw_lower = kw.lower()

        # 1. Playwright text locator — exact then partial
        for exact in [True, False]:
            try:
                loc = page.get_by_text(kw, exact=exact).first
                await loc.wait_for(state="visible", timeout=3000)
                await loc.scroll_into_view_if_needed()
                await loc.click()
                log("info", f"Clicked '{label}'")
                return True
            except Exception:
                pass

        # 2. aria-label / title / alt attributes
        for attr in [f"[aria-label*='{kw}' i]", f"[title*='{kw}' i]",
                     f"[alt*='{kw}' i]", f"[data-label*='{kw}' i]"]:
            try:
                await page.wait_for_selector(attr, timeout=2000, state="visible")
                await page.click(attr)
                log("info", f"Clicked '{label}' via attribute")
                return True
            except Exception:
                pass

        # 3. Angular / Material Design selectors
        for sel in [
            f"mat-card:has-text('{kw}')",
            f"mat-list-item:has-text('{kw}')",
            f"mat-menu-item:has-text('{kw}')",
            f"mat-tab:has-text('{kw}')",
            f"button:has-text('{kw}')",
            f"a:has-text('{kw}')",
            f"li:has-text('{kw}')",
            f"[class*='card']:has-text('{kw}')",
            f"[class*='tile']:has-text('{kw}')",
            f"[class*='item']:has-text('{kw}')",
            f"[class*='product']:has-text('{kw}')",
            f"[class*='menu']:has-text('{kw}')",
            f"[class*='nav']:has-text('{kw}')",
            f"[class*='btn']:has-text('{kw}')",
        ]:
            try:
                await page.wait_for_selector(sel, timeout=2000, state="visible")
                await page.click(sel)
                log("info", f"Clicked '{label}' via Angular selector")
                return True
            except Exception:
                pass

        # 4. JavaScript innerText scan — catches anything the above missed
        try:
            result = await page.evaluate(f"""() => {{
                const kw = '{kw_lower}';
                const els = document.querySelectorAll(
                    'a, button, mat-card, mat-list-item, li, span, ' +
                    '[role="button"], [role="tab"], [role="menuitem"], ' +
                    '[class*="card"], [class*="tile"], [class*="item"], ' +
                    '[class*="menu"], [class*="nav"], [class*="btn"]'
                );
                for (const el of els) {{
                    const txt = (el.innerText || el.textContent || '').trim().toLowerCase();
                    if (txt === kw || txt.startsWith(kw)) {{
                        const rect = el.getBoundingClientRect();
                        if (rect.width > 0 && rect.height > 0) {{
                            el.click();
                            return txt.substring(0, 40);
                        }}
                    }}
                }}
                return null;
            }}""")
            if result:
                log("info", f"Clicked '{label}' via JS scan")
                return True
        except Exception:
            pass

    # ── All strategies failed — show what's on the page and ask user ──
    log("warn", f"Could not auto-click '{label}'. Visible elements on page:")
    try:
        items = await page.evaluate("""() => {
            const els = document.querySelectorAll(
                'a, button, mat-card, mat-list-item, li, [role="button"],
                 [class*="card"], [class*="tile"], [class*="nav"]'
            );
            return [...new Set(
                Array.from(els)
                    .map(e => (e.innerText || e.textContent || '').trim())
                    .filter(t => t.length > 1 && t.length < 60)
            )].slice(0, 25);
        }""")
        for item in items:
            print(f"         . {item}")
    except Exception:
        pass

    await pause(f"Click '{label}' manually in the browser, then press ENTER")
    return False


async def fill_field(page, selectors, value, label):
    """Try each selector until one works, then fill it."""
    for sel in selectors:
        try:
            el = await page.wait_for_selector(sel, timeout=5000, state="visible")
            await el.scroll_into_view_if_needed()
            await el.click()
            # Use Ctrl+A + Delete instead of triple_click (works on ElementHandle)
            await page.keyboard.press("Control+a")
            await page.keyboard.press("Delete")
            await el.type(value, delay=60)
            log("fill", f"{label}: {value}")
            return True
        except Exception:
            continue
    log("warn", f"Could not fill '{label}' — fill it manually")
    return False


async def fill_dropdown(page, value, label):
    """
    For Angular autocomplete/dropdown fields on NYSA portal.
    Strategy:
    1. Find ANY visible input that is currently active / empty
    2. Type the value slowly
    3. Wait for dropdown list to appear
    4. Click the best matching option
    5. If no dropdown appears, try pressing Enter
    """
    try:
        # Type into the currently focused/active input on the page
        # First click somewhere neutral to reset focus
        await page.wait_for_timeout(300)

        # Find all visible inputs and try each one that looks empty
        inputs = await page.query_selector_all("input:visible, mat-select:visible")

        # Try typing into the page directly after finding the right field
        # Use JS to find input by placeholder or nearby label text
        el = await page.evaluate_handle(f"""() => {{
            const allInputs = document.querySelectorAll('input[type="text"], input:not([type])');
            for (const inp of allInputs) {{
                const rect = inp.getBoundingClientRect();
                if (rect.width === 0 || rect.height === 0) continue;
                const ph = (inp.placeholder || '').toLowerCase();
                const label = (inp.getAttribute('aria-label') || '').toLowerCase();
                const formCtrl = (inp.getAttribute('formcontrolname') || '').toLowerCase();
                const val = '{value.lower()}';
                const labelLow = '{label.lower()}';
                // Match by placeholder, aria-label, or formcontrolname
                if (ph.includes(labelLow) || label.includes(labelLow) ||
                    formCtrl.includes(labelLow) || ph.includes(val.substring(0,4))) {{
                    return inp;
                }}
            }}
            return null;
        }}""")

        if el:
            await el.click()
            await el.fill("")
            await el.type(value, delay=80)
            log("fill", f"{label}: {value}")
        else:
            # Fallback: use keyboard focus — tab to next field and type
            await page.keyboard.type(value, delay=80)
            log("fill", f"{label} (keyboard): {value}")

        # Wait for dropdown/autocomplete to appear
        await page.wait_for_timeout(1000)

        # Try clicking matching option in any dropdown that appeared
        option_clicked = False
        for option_sel in [
            f"mat-option:has-text('{value}')",
            f"li:has-text('{value}')",
            f".mat-option:has-text('{value}')",
            f"[role='option']:has-text('{value}')",
            f".dropdown-item:has-text('{value}')",
            f".autocomplete-option:has-text('{value}')",
            f"span:has-text('{value}')",
        ]:
            try:
                await page.wait_for_selector(option_sel, timeout=2000, state="visible")
                await page.click(option_sel)
                log("info", f"Selected dropdown option: {value}")
                option_clicked = True
                break
            except Exception:
                continue

        if not option_clicked:
            # Try partial match on first word
            first_word = value.split()[0] if value else value
            for option_sel in [
                f"mat-option:has-text('{first_word}')",
                f"[role='option']:has-text('{first_word}')",
                f"li:has-text('{first_word}')",
            ]:
                try:
                    await page.wait_for_selector(option_sel, timeout=1500, state="visible")
                    await page.click(option_sel)
                    log("info", f"Selected dropdown option (partial): {first_word}")
                    option_clicked = True
                    break
                except Exception:
                    continue

        if not option_clicked:
            # Press Enter/Tab to confirm whatever is shown
            await page.keyboard.press("Enter")

        await page.wait_for_timeout(500)
        return True

    except Exception as ex:
        log("warn", f"Could not fill dropdown '{label}': {ex}")
        return False


# ═══════════════════════════════════════════════════════════════════
#  CONNECT TO YOUR EXISTING CHROME
# ═══════════════════════════════════════════════════════════════════

async def connect_to_chrome(playwright):
    """
    Attach to Chrome that was started with --remote-debugging-port=9222.
    Finds the NYSA tab automatically. No new tab is opened.
    """
    try:
        browser = await playwright.chromium.connect_over_cdp(
            f"http://localhost:{CHROME_PORT}"
        )
    except Exception as e:
        print(f"""
  ERROR: Cannot connect to Chrome on port {CHROME_PORT}

  Make sure you started Chrome with this command (Win+R → paste):
  "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\\chrome-nysa"

  Then login to nysa.icicilombard.com and press ENTER here.
  Details: {e}
        """)
        input("Press ENTER to exit..."); raise SystemExit(1)

    # Find the NYSA tab — look for icicilombard.com in any open tab
    nysa_page = None
    all_pages = [p for ctx in browser.contexts for p in ctx.pages]

    for p in all_pages:
        if "icicilombard.com" in p.url:
            nysa_page = p
            break

    # If not found, use the most recently active tab
    if not nysa_page:
        if all_pages:
            nysa_page = all_pages[-1]
            log("warn", f"NYSA tab not found — using: {nysa_page.url}")
            print("  Make sure the NYSA portal is open in Chrome!")
            await pause("Navigate to nysa.icicilombard.com home page, then press ENTER")
        else:
            print("  No open tabs found in Chrome.")
            input("Press ENTER to exit..."); raise SystemExit(1)

    await nysa_page.bring_to_front()
    log("ok", f"Connected! Tab URL: {nysa_page.url}")
    return browser, nysa_page


# ═══════════════════════════════════════════════════════════════════
#  MAIN AUTOMATION FLOW
# ═══════════════════════════════════════════════════════════════════

async def run(D, skip_to_step4=False):
    steps       = []
    proposal_no = ""

    make_model = f"{D.get('manufacturer','')} - {D.get('model','')}".strip(" -") or "(not set)"
    print(f"""
+--------------------------------------------------------------+
|  ICICI NYSA — EV (Electric Bike) Insurance Automation        |
|  Make/Model : {make_model[:48]:<48}|
|  Battery    : {D.get('battery_number',''):<48}|
|  Mobile     : {D.get('mobile_number',''):<48}|
|  Auto Pay   : {'YES - will click PAY' if D.get('auto_pay') else 'NO  - stops before payment':<48}|
+--------------------------------------------------------------+""")

    async with async_playwright() as p:
        browser, page = await connect_to_chrome(p)

        try:
            # Let page fully settle
            try:
                await page.wait_for_load_state("networkidle", timeout=8000)
            except Exception:
                pass
            await page.wait_for_timeout(1500)
            log("ok", f"Automation starting from: {page.url}")
            steps.append("connected")

            if not skip_to_step4:
                # ── STEP 1: Motor Insurance ────────────────────────────
                log("step", "STEP 1 — Motor Insurance")
                await smart_click(page, ["Motor", "Motor Insurance"], "Motor")
                # Wait for Motor page to fully load — networkidle exits early when ready
                # The Motor click often triggers a navigation which destroys the
                # execution context. We need to wait for navigation to complete first.
                try:
                    await page.wait_for_load_state("domcontentloaded", timeout=15000)
                except Exception:
                    pass
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)

                # Also wait up to 30 attempts for the RTO/Quote form to actually render
                # Each evaluate is wrapped in try/except because the portal may still
                # be navigating (destroying the execution context)
                for _mot_wait in range(30):
                    try:
                        page_loaded = await page.evaluate("""() => {
                            const txt = (document.body.innerText||'').toLowerCase();
                            return txt.includes('registration number') ||
                                   txt.includes('fetch vehicle details') ||
                                   txt.includes('fetch vehicle data') ||
                                   txt.includes('city where vehicle') ||
                                   txt.includes('city of registration') ||
                                   txt.includes('four wheeler') ||
                                   txt.includes('two wheeler');
                        }""")
                        if page_loaded:
                            log("info", f"  Motor page ready (attempt {_mot_wait+1})")
                            break
                    except Exception:
                        # Context destroyed by navigation — wait and retry
                        log("info", f"  Motor page still loading (attempt {_mot_wait+1})...")
                        pass
                    await page.wait_for_timeout(500)

                # ── STEP 1b: Click left-side slider (opens vehicle category modal) ─
                log("step", "STEP 1b — Click left slider to open category modal")
                # networkidle already fired above — no extra wait needed here
                # Click slider and verify modal opened (Four Wheeler/Two Wheeler visible)
                # If not opened, click again — nysa_openbtn toggles, so may need 2 clicks
                # NOTE: Each page.evaluate is wrapped in try/except because the portal
                # sometimes triggers a navigation during the click, which destroys the
                # execution context. That's expected — we just retry instead of failing.
                modal_open = False
                for _sb in range(25):
                    # ── Pre-check: is the modal already open? ──
                    # nysa_openbtn is a TOGGLE. If a previous click opened the modal
                    # but the verify eval crashed on a navigation, clicking again
                    # would CLOSE it. So always check state before clicking.
                    # CRITICAL FIX for EV flow: require the cards to be ON-SCREEN
                    # at positive coordinates. The portal has the card text in a
                    # hidden left side-menu (negative X coords) which would
                    # falsely indicate modal-open.
                    try:
                        modal_open = await page.evaluate("""() => {
                            const vw = window.innerWidth;
                            for (const el of document.querySelectorAll('*')) {
                                const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                                if (t === 'four wheeler' || t === 'two wheeler' || t === 'electric bike') {
                                    const r = el.getBoundingClientRect();
                                    // Require ON-SCREEN: positive X within viewport
                                    if (r.width > 0 && r.left >= 0 && r.left < vw) return true;
                                }
                            }
                            return false;
                        }""")
                    except Exception as e:
                        log("info", f"  (pre-check interrupted by navigation — retrying)")
                        await page.wait_for_timeout(800)
                        continue

                    if modal_open:
                        log("ok", "  -> Modal is open (ON-SCREEN)")
                        break

                    # ── Find slider button coordinates ──
                    try:
                        btn_coords = await page.evaluate("""() => {
                            const el = document.querySelector('.nysa_openbtn');
                            if (!el) return null;
                            const r = el.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0)
                                return {x: Math.round(r.left + r.width/2), y: Math.round(r.top + r.height/2)};
                            return null;
                        }""")
                    except Exception as e:
                        log("info", f"  (coord lookup interrupted by navigation — retrying)")
                        await page.wait_for_timeout(800)
                        continue

                    if btn_coords:
                        log("info", f"  nysa_openbtn at x={btn_coords['x']} y={btn_coords['y']} (attempt {_sb+1})")
                        try:
                            await page.mouse.click(btn_coords['x'], btn_coords['y'])
                        except Exception:
                            # Click itself can fail if page navigates mid-click
                            pass
                        # Let any navigation triggered by the click settle
                        try:
                            await page.wait_for_load_state("networkidle", timeout=3000)
                        except Exception:
                            pass
                        await page.wait_for_timeout(800)
                        log("info", f"  Checking if modal opened...")
                    else:
                        await page.wait_for_timeout(500)
                if not modal_open:
                    log("warn", "  Modal did not open after all attempts")

                # ── STEP 1c: Select "Electric Bike" card from modal ───────
                # The modal has 7 cards in the top row: Four Wheeler / Two Wheeler /
                # Electric Bike / GCV / PCV / MISCD / FMS. Each card is an image +
                # label. Electric Bike is the ONLY one without a Package/TP badge.
                # We need to click the CARD CONTAINER (not just the text label),
                # because clicking the wrong child element triggers the regular
                # 2W Package flow instead of the EV flow.
                cat_text = "Electric Bike"
                log("info", f"  Selecting category: {cat_text}")

                # Helper that verifies whether the EV form loaded after the click.
                # The EV form contains text like "Battery Number" / "Purchase Date" /
                # "Unique Identification Number" / "Electrical Bike" in the breadcrumb.
                async def is_ev_page_loaded():
                    try:
                        return await page.evaluate("""() => {
                            const txt = (document.body.innerText || '').toLowerCase();
                            // EV-specific markers
                            const evHits = [
                                'battery number',
                                'purchase date',
                                'unique identification number',
                                'unique identification number of asset',
                                'electrical bike',
                                'electric bike',
                                'transaction type'
                            ].filter(s => txt.includes(s)).length;
                            // Wrong-page markers (regular Package flow)
                            const wrongHits = [
                                'fetch vehicle data',
                                'enter registration number',
                                'standalone od policy',
                                'change in ownership in the last 12 months'
                            ].filter(s => txt.includes(s)).length;
                            // Need at least 2 EV markers AND no wrong-page markers
                            return evHits >= 2 && wrongHits === 0;
                        }""")
                    except Exception:
                        return False

                async def is_wrong_page_loaded():
                    try:
                        return await page.evaluate("""() => {
                            const txt = (document.body.innerText || '').toLowerCase();
                            return txt.includes('fetch vehicle data') ||
                                   txt.includes('enter registration number') ||
                                   txt.includes('change in ownership in the last 12 months');
                        }""")
                    except Exception:
                        return False

                # Strategies to try, in order. After each strategy that succeeds in
                # clicking, we wait for navigation and verify the EV form loaded.

                # ── DEBUG: report what we see in the DOM for Electric Bike ──
                # CRITICAL: The portal has the text "Electric Bike" in MULTIPLE places —
                # both in the visible center modal AND in a hidden left side-menu
                # (the 'nysa' nav drawer). Hidden elements report negative X coords.
                # We MUST filter to only ON-SCREEN visible elements.
                debug_info = await page.evaluate("""(catText) => {
                    const lower = catText.toLowerCase().trim();
                    const vw = window.innerWidth;
                    const vh = window.innerHeight;
                    const matches = [];
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if (t === lower) {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) continue;
                            // ON-SCREEN check — must be within viewport bounds
                            const onScreen = r.left >= 0 && r.top >= 0 &&
                                             r.right <= vw && r.bottom <= vh;
                            // walk up to find clickable ancestor
                            let click = null;
                            let p = el;
                            for (let i = 0; i < 8; i++) {
                                if (!p) break;
                                const handler = p.getAttribute('ng-reflect-router-link') ||
                                                p.getAttribute('routerlink') ||
                                                p.getAttribute('href') ||
                                                p.onclick;
                                if (handler || p.tagName === 'A' || p.tagName === 'BUTTON') {
                                    click = {
                                        tag: p.tagName, cls: p.className,
                                        href: p.getAttribute('href') || '',
                                        routerLink: p.getAttribute('routerlink') ||
                                                    p.getAttribute('ng-reflect-router-link') || '',
                                        rect: p.getBoundingClientRect()
                                    };
                                    break;
                                }
                                p = p.parentElement;
                            }
                            matches.push({
                                tag: el.tagName,
                                cls: el.className,
                                rect: r,
                                onScreen: onScreen,
                                clickAncestor: click
                            });
                        }
                    }
                    return {viewport: {w: vw, h: vh}, matches: matches};
                }""", cat_text)

                vw = debug_info['viewport']['w']
                vh = debug_info['viewport']['h']
                matches = debug_info['matches']
                log("info", f"  DEBUG: viewport {vw}x{vh}, found {len(matches)} 'Electric Bike' element(s)")
                visible_count = 0
                for i, m in enumerate(matches):
                    r = m['rect']
                    ca = m.get('clickAncestor')
                    onScreen = m.get('onScreen')
                    marker = "✓" if onScreen else "✗ OFF-SCREEN"
                    if onScreen: visible_count += 1
                    log("info", f"    [{i}] {marker} tag={m['tag']} cls={(m.get('cls') or '')[:40]} "
                                f"@({r['x']:.0f},{r['y']:.0f}) {r['width']:.0f}x{r['height']:.0f}")
                    if ca:
                        log("info", f"        -> clickable: {ca['tag']} cls={(ca.get('cls') or '')[:40]} "
                                    f"href='{ca.get('href','')[:30]}' routerLink='{ca.get('routerLink','')[:30]}'")
                if visible_count == 0:
                    log("warn", "  ⚠ No ON-SCREEN 'Electric Bike' element found! "
                                "Check if the modal is actually visible.")

                async def click_strategy_clickable_ancestor():
                    """Find Electric Bike text (ON-SCREEN ONLY), walk UP to nearest <a>,
                    <button>, or element with click handler, then click via native .click()."""
                    js_clicked = await page.evaluate("""(catText) => {
                        const lower = catText.toLowerCase().trim();
                        const vw = window.innerWidth;
                        const vh = window.innerHeight;
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                            if (t !== lower) continue;
                            const r = el.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) continue;
                            // MUST be on-screen
                            if (r.left < 0 || r.top < 0 || r.right > vw || r.bottom > vh) continue;
                            if (r.width > 300 || r.height > 200) continue;

                            let p = el;
                            for (let i = 0; i < 8; i++) {
                                if (!p) break;
                                const isClickable =
                                    p.tagName === 'A' || p.tagName === 'BUTTON' ||
                                    p.getAttribute('routerlink') ||
                                    p.getAttribute('ng-reflect-router-link') ||
                                    p.onclick ||
                                    (p.getAttribute('class') || '').toLowerCase()
                                      .match(/\\b(card|tile|item|product|btn|nysa)\\b/);
                                if (isClickable) {
                                    p.click();
                                    return {ok: true, tag: p.tagName,
                                            cls: p.className,
                                            x: p.getBoundingClientRect().left + p.getBoundingClientRect().width/2,
                                            y: p.getBoundingClientRect().top + p.getBoundingClientRect().height/2};
                                }
                                p = p.parentElement;
                            }
                        }
                        return {ok: false};
                    }""", cat_text)
                    if js_clicked.get('ok'):
                        log("info", f"  → native .click() on {js_clicked['tag']} "
                                    f"cls={(js_clicked.get('cls') or '')[:40]} "
                                    f"@({js_clicked.get('x',0):.0f},{js_clicked.get('y',0):.0f})")
                        return True
                    return False

                # Strategies to try, in order. After each strategy that succeeds in
                # clicking, we wait for navigation and verify the EV form loaded.
                async def click_strategy_text_locator():
                    """Click via Playwright text locator on ON-SCREEN 'Electric Bike' text."""
                    try:
                        # Iterate ALL matches and pick the on-screen one
                        all_loc = page.get_by_text(cat_text, exact=True)
                        n = await all_loc.count()
                        for i in range(n):
                            loc = all_loc.nth(i)
                            try:
                                box = await loc.bounding_box()
                                if not box:
                                    continue
                                vp = page.viewport_size or {"width": 1366, "height": 768}
                                if (box['x'] < 0 or box['y'] < 0 or
                                    box['x'] + box['width'] > vp['width'] or
                                    box['y'] + box['height'] > vp['height']):
                                    continue
                                await loc.scroll_into_view_if_needed()
                                await page.wait_for_timeout(300)
                                box = await loc.bounding_box()
                                if box:
                                    await page.mouse.click(box['x'] + box['width']/2,
                                                           box['y'] + box['height']/2)
                                    return True
                            except Exception:
                                continue
                    except Exception:
                        pass
                    return False

                async def click_strategy_card_with_image():
                    """Find a card containing both an image and the text 'Electric Bike'
                    (ON-SCREEN ONLY), then click the IMAGE/ICON area."""
                    coords = await page.evaluate("""(catText) => {
                        const lower = catText.toLowerCase().trim();
                        const vw = window.innerWidth;
                        const vh = window.innerHeight;

                        const candidates = [];
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                            if (t === lower || t.replace(/\\s+/g,' ') === lower) {
                                const r = el.getBoundingClientRect();
                                if (r.width === 0 || r.height === 0) continue;
                                if (r.left < 0 || r.top < 0 || r.right > vw || r.bottom > vh) continue;
                                if (r.width > 400 || r.height > 200) continue;
                                candidates.push(el);
                            }
                        }

                        for (const labelEl of candidates) {
                            let p = labelEl.parentElement;
                            for (let i = 0; i < 6; i++) {
                                if (!p) break;
                                const img = p.querySelector('img');
                                const r = p.getBoundingClientRect();
                                if (img && r.width > 60 && r.width < 220
                                    && r.height > 60 && r.height < 220
                                    && r.left >= 0 && r.top >= 0
                                    && r.right <= vw && r.bottom <= vh) {
                                    img.scrollIntoView({block:'center'});
                                    const ir = img.getBoundingClientRect();
                                    return {
                                        x: ir.left + ir.width/2,
                                        y: ir.top + ir.height/2,
                                        cardX: r.left + r.width/2,
                                        cardY: r.top + r.height/2,
                                        cardW: r.width,
                                        cardH: r.height
                                    };
                                }
                                p = p.parentElement;
                            }
                        }
                        return null;
                    }""", cat_text)
                    if not coords:
                        return False
                    log("info", f"  EV card found: card@({coords['cardX']:.0f},{coords['cardY']:.0f})"
                                f" size={coords['cardW']:.0f}x{coords['cardH']:.0f}")
                    await page.mouse.click(coords['x'], coords['y'])
                    return True

                async def click_strategy_xpath():
                    """XPath on exact text node, then click the parent card."""
                    try:
                        coords = await page.evaluate("""(catText) => {
                            const xp = `//*[normalize-space(text())='${catText}']`;
                            const r = document.evaluate(xp, document, null,
                                XPathResult.FIRST_ORDERED_NODE_TYPE, null);
                            const node = r.singleNodeValue;
                            if (!node) return null;
                            // Walk up to card container
                            let target = node;
                            let p = node.parentElement;
                            for (let i = 0; i < 5; i++) {
                                if (!p) break;
                                const rr = p.getBoundingClientRect();
                                if (p.querySelector('img') && rr.width < 220 && rr.width > 60) {
                                    target = p; break;
                                }
                                p = p.parentElement;
                            }
                            target.scrollIntoView({block:'center'});
                            const tr = target.getBoundingClientRect();
                            return {x: tr.left + tr.width/2, y: tr.top + tr.height/2};
                        }""", cat_text)
                        if coords:
                            await page.mouse.click(coords['x'], coords['y'])
                            return True
                    except Exception:
                        pass
                    return False

                async def click_strategy_card_top():
                    """Some Angular cards register click on their top portion only.
                    Click the upper-third of the card (where the icon is)."""
                    coords = await page.evaluate("""(catText) => {
                        const lower = catText.toLowerCase().trim();
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText||'').trim().toLowerCase();
                            if (t === lower) {
                                const r = el.getBoundingClientRect();
                                if (r.width === 0 || r.height === 0) continue;
                                if (r.width > 400) continue;
                                // Walk up to card
                                let p = el.parentElement;
                                for (let i = 0; i < 5; i++) {
                                    if (!p) break;
                                    const pr = p.getBoundingClientRect();
                                    if (p.querySelector('img') && pr.width > 60 && pr.width < 220) {
                                        p.scrollIntoView({block:'center'});
                                        const fr = p.getBoundingClientRect();
                                        // Click upper-third
                                        return {x: fr.left + fr.width/2,
                                                y: fr.top + fr.height/3};
                                    }
                                    p = p.parentElement;
                                }
                            }
                        }
                        return null;
                    }""", cat_text)
                    if coords:
                        await page.mouse.click(coords['x'], coords['y'])
                        return True
                    return False

                strategies = [
                    ("clickable-ancestor", click_strategy_clickable_ancestor),
                    ("card+image",  click_strategy_card_with_image),
                    ("text-locator", click_strategy_text_locator),
                    ("xpath",       click_strategy_xpath),
                    ("card-top",    click_strategy_card_top),
                ]

                cat_clicked = False
                for label, strategy in strategies:
                    if cat_clicked:
                        break
                    log("info", f"  Trying click strategy: {label}")
                    success = False
                    try:
                        success = await strategy()
                    except Exception as e:
                        log("info", f"    strategy '{label}' raised: {e}")

                    if not success:
                        log("info", f"    strategy '{label}' did not click")
                        continue

                    log("ok", f"  -> Click attempted via '{label}', verifying page...")

                    # Wait a bit for page to load/navigate
                    await page.wait_for_timeout(2000)
                    try:
                        await page.wait_for_load_state("networkidle", timeout=8000)
                    except Exception:
                        pass

                    # Verify EV form loaded
                    for _v in range(8):
                        if await is_ev_page_loaded():
                            cat_clicked = True
                            log("ok", f"  ✓ EV vehicle details form detected (strategy: {label})")
                            break
                        if await is_wrong_page_loaded():
                            log("warn", f"  ✗ Wrong page detected (regular Package flow). "
                                        f"Going back to retry with another strategy.")
                            # Navigate back to motor home so we can retry
                            try:
                                await page.go_back()
                                await page.wait_for_timeout(1500)
                                # Re-open category modal — try clicking Motor again
                                await smart_click(page, ["Motor", "Motor Insurance"], "Motor")
                                await page.wait_for_timeout(1500)
                                # Re-open the category slider button
                                slider_clicked = await page.evaluate("""() => {
                                    for (const el of document.querySelectorAll('*')) {
                                        const t = (el.innerText||'').trim().toLowerCase();
                                        if (t.includes('four wheeler') || t.includes('two wheeler')) {
                                            const r = el.getBoundingClientRect();
                                            if (r.width > 0) return true;
                                        }
                                    }
                                    return false;
                                }""")
                                if not slider_clicked:
                                    # Click the orange ">>" slider
                                    coords = await page.evaluate("""() => {
                                        const sliders = document.querySelectorAll(
                                            '[class*="slider"], [class*="toggle"], [class*="arrow"]');
                                        for (const s of sliders) {
                                            const r = s.getBoundingClientRect();
                                            if (r.width > 0 && r.left < 100) {
                                                return {x: r.left + r.width/2, y: r.top + r.height/2};
                                            }
                                        }
                                        return null;
                                    }""")
                                    if coords:
                                        await page.mouse.click(coords['x'], coords['y'])
                                        await page.wait_for_timeout(1500)
                            except Exception as e:
                                log("warn", f"    couldn't navigate back: {e}")
                            break  # break out of verify loop, try next strategy
                        await page.wait_for_timeout(700)

                if not cat_clicked:
                    log("warn", f"  Could not navigate to EV form automatically — please click "
                                f"'Electric Bike' (NOT Two Wheeler) in the modal manually")
                    await pause(f"Click '{cat_text}' card in the modal so the EV vehicle "
                                f"details form opens, then press ENTER")
                    # Verify after manual click
                    await page.wait_for_timeout(1500)
                    if await is_ev_page_loaded():
                        cat_clicked = True
                        log("ok", "  ✓ EV form detected after manual click")
                    else:
                        log("warn", "  ✗ EV form still not detected — proceeding anyway")

                # NOTE: Electric Bike has NO Package/TP badge — skip the type click
                log("info", "  EV mode: no Package/TP badge to select for Electric Bike")

                await page.wait_for_timeout(2000)
                steps.append("motor")


                # ── STEP 1d: EV Vehicle Details form ───────────────────────
                # After clicking "Electric Bike", the EV-specific vehicle
                # details page loads directly (no SAOD, no Policy Type tab,
                # no Registration Number / Fetch). The form has these fields:
                #   - Transaction Type      (dropdown)
                #   - Battery Number        (text input)
                #   - Purchase Date         (date input DD/MM/YYYY)
                #   - Customer Type         (radio: Individual / Corporate)
                #   - Manufacturer          (autocomplete text)
                #   - Model                 (dropdown)
                #   - Year of Manufacture   (dropdown)
                #   - Policy Start Date     (date)
                #   - Policy End Date       (date)
                #   - Unique Identification Number of Asset  (text)
                #   - Invoice Number        (text)
                #   - Customer State        (text / autocomplete)
                steps.append("new")
                steps.append("fetch_details")

            else:
                # Skipping STEP 1: already on EV details page after Another Policy
                log("step", "Skipping STEP 1 (resuming from EV details page)")
                steps.append("motor"); steps.append("new"); steps.append("fetch_details")
                try:
                    await page.wait_for_load_state("networkidle", timeout=10000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)


            # ══════════════════════════════════════════════════════
            # STEP 4 — Fill EV Vehicle Details
            # ══════════════════════════════════════════════════════
            log("step", "STEP 4 — Filling EV Vehicle Details")

            # Wait for the EV Vehicle Details form to be ready
            log("info", "  Waiting for EV vehicle details form to load...")
            for _vd_wait in range(40):
                try:
                    vd_ready = await page.evaluate("""() => {
                        const txt = (document.body.innerText||'').toLowerCase();
                        return txt.includes('battery number') ||
                               txt.includes('purchase date') ||
                               txt.includes('unique identification number');
                    }""")
                    if vd_ready:
                        log("ok", f"  EV Vehicle Details form ready (attempt {_vd_wait+1})")
                        break
                except Exception:
                    pass
                await page.wait_for_timeout(500)
            await page.wait_for_timeout(1500)

            # ── Helper: find an input/textarea by its visible label text ──
            async def find_input_by_label(label_text, max_dist=180):
                """Locate the input/textarea closest to a given label.
                Returns dict with x, y, width, height, formCtrl, placeholder.
                """
                return await page.evaluate(r"""([label, maxDist]) => {
                    const lower = label.toLowerCase();
                    // Find element whose direct/short text matches label
                    let labelEl = null;
                    let bestArea = Infinity;
                    for (const el of document.querySelectorAll('label, span, div, p, mat-label')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase()
                            .replace(/\s*\*\s*$/, '').trim();
                        if (t === lower) {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) continue;
                            // Skip giant containers
                            if (r.width > 600 && r.height > 80) continue;
                            const area = r.width * r.height;
                            if (area < bestArea) { bestArea = area; labelEl = el; }
                        }
                    }
                    // Fallback: starts-with match
                    if (!labelEl) {
                        for (const el of document.querySelectorAll('label, span, div, p, mat-label')) {
                            const t = (el.innerText || el.textContent || '').trim().toLowerCase()
                                .replace(/\s*\*\s*$/, '').trim();
                            if (t.startsWith(lower) && t.length < lower.length + 8) {
                                const r = el.getBoundingClientRect();
                                if (r.width === 0 || r.height === 0) continue;
                                if (r.width > 600 && r.height > 80) continue;
                                const area = r.width * r.height;
                                if (area < bestArea) { bestArea = area; labelEl = el; }
                            }
                        }
                    }
                    if (!labelEl) return null;
                    labelEl.scrollIntoView({block:'center'});
                    const lr = labelEl.getBoundingClientRect();
                    // Search for the closest input/textarea/mat-select BELOW the label
                    let best = null, bestDist = Infinity, bestType = '';
                    const candidates = document.querySelectorAll(
                        'input, textarea, mat-select, select, [role="combobox"]'
                    );
                    for (const inp of candidates) {
                        const r = inp.getBoundingClientRect();
                        if (r.width === 0 || r.height === 0) continue;
                        const t = (inp.type || '').toLowerCase();
                        if (['hidden', 'submit', 'button'].includes(t)) continue;
                        // Vertical distance: input must be at or below the label
                        const dy = r.top - lr.bottom;
                        if (dy < -10 || dy > maxDist) continue;
                        // Horizontal alignment: input center should be near label center
                        const dx = Math.abs((r.left + r.width/2) - (lr.left + lr.width/2));
                        if (dx > 280) continue;
                        const dist = Math.abs(dy) * 2 + dx;
                        if (dist < bestDist) {
                            bestDist = dist; best = inp;
                            bestType = inp.tagName.toLowerCase() + ':' + (inp.type || '');
                        }
                    }
                    if (!best) return null;
                    best.scrollIntoView({block:'center'});
                    const br = best.getBoundingClientRect();
                    return {
                        x: br.left + br.width/2,
                        y: br.top + br.height/2,
                        width: br.width,
                        height: br.height,
                        tag: bestType,
                        formCtrl: best.getAttribute('formcontrolname') || '',
                        placeholder: best.placeholder || '',
                        currentValue: best.value || ''
                    };
                }""", [label_text, max_dist])

            # ── Helper: type plain text into the input under a label ──
            async def fill_text_by_label(label_text, value):
                if not value:
                    return False
                value = str(value)
                info = await find_input_by_label(label_text)
                if not info:
                    log("warn", f"  [{label_text}] input not found")
                    return False
                log("info", f"  [{label_text}]: {value}")
                await page.mouse.click(info['x'], info['y'])
                await page.wait_for_timeout(200)
                await page.keyboard.press("Control+a")
                await page.keyboard.press("Delete")
                await page.wait_for_timeout(150)
                await page.keyboard.type(value, delay=40)
                await page.wait_for_timeout(300)
                # Click outside to close any picker / blur
                await page.keyboard.press("Tab")
                await page.wait_for_timeout(300)
                log("ok", f"  -> {label_text}: {value}")
                return True

            # ── Helper: type a date into a DD/MM/YYYY input ──
            async def fill_date_by_label(label_text, value):
                if not value:
                    return False
                value = str(value).strip()
                # Normalize to DD/MM/YYYY
                m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", value)
                if m:
                    dd, mm, yy = m.groups()
                    value = f"{int(dd):02d}/{int(mm):02d}/{yy}"
                info = await find_input_by_label(label_text)
                if not info:
                    log("warn", f"  [{label_text}] date input not found")
                    return False
                log("info", f"  [{label_text}]: {value}")
                await page.mouse.click(info['x'], info['y'])
                await page.wait_for_timeout(200)
                await page.keyboard.press("Control+a")
                await page.keyboard.press("Delete")
                await page.wait_for_timeout(150)
                # Some Angular date pickers ignore typing — try typed first, then JS set
                await page.keyboard.type(value, delay=60)
                await page.wait_for_timeout(300)
                # Verify
                v_now = await page.evaluate("""([x,y]) => {
                    const inps = document.querySelectorAll('input');
                    for (const inp of inps) {
                        const r = inp.getBoundingClientRect();
                        if (Math.abs(r.left+r.width/2-x)<5 && Math.abs(r.top+r.height/2-y)<5) {
                            return inp.value;
                        }
                    }
                    return '';
                }""", [info['x'], info['y']])
                if (v_now or '').strip() != value:
                    # Try JS-set fallback
                    await page.evaluate("""([x,y,v]) => {
                        for (const inp of document.querySelectorAll('input')) {
                            const r = inp.getBoundingClientRect();
                            if (Math.abs(r.left+r.width/2-x)<5 && Math.abs(r.top+r.height/2-y)<5) {
                                inp.focus();
                                const setter = Object.getOwnPropertyDescriptor(
                                    window.HTMLInputElement.prototype, 'value').set;
                                setter.call(inp, v);
                                inp.dispatchEvent(new Event('input', {bubbles:true}));
                                inp.dispatchEvent(new Event('change', {bubbles:true}));
                                inp.dispatchEvent(new Event('blur', {bubbles:true}));
                                return;
                            }
                        }
                    }""", [info['x'], info['y'], value])
                    await page.wait_for_timeout(300)
                # Close any open date popup
                await page.keyboard.press("Escape")
                await page.wait_for_timeout(200)
                log("ok", f"  -> {label_text}: {value}")
                return True

            # ── Helper: select a dropdown option by label & visible value ──
            async def pick_dropdown_by_label(label_text, value):
                if not value:
                    return False
                value = str(value).strip()

                # ── Pre-check 1: is the displayed value already correct? ──
                # The portal often shows the dropdown's selected text in a span/div
                # near the label. Read that text FIRST before clicking — if it matches,
                # don't open the dropdown at all (avoids the dropdown getting stuck open).
                displayed = await page.evaluate("""([lbl]) => {
                    const lower = lbl.toLowerCase().trim();
                    // Find the label
                    let labelEl = null;
                    let bestArea = Infinity;
                    for (const el of document.querySelectorAll('label, span, div, p, mat-label')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase()
                            .replace(/\\s*\\*\\s*$/, '').trim();
                        if (t === lower || t === lower + ' *') {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) continue;
                            if (r.width > 600 && r.height > 80) continue;
                            const area = r.width * r.height;
                            if (area < bestArea) { bestArea = area; labelEl = el; }
                        }
                    }
                    if (!labelEl) return null;
                    const lr = labelEl.getBoundingClientRect();
                    // Look for a sibling/nearby element that holds the displayed value
                    // (typically a span inside the dropdown trigger, just below the label)
                    const candidates = document.querySelectorAll(
                        'mat-select-trigger, .mat-select-value-text, .mat-select-value, ' +
                        '.ng-select-container .ng-value, [class*="select-value"], ' +
                        '.dropdown-toggle, span, div'
                    );
                    let best = null, bestDist = Infinity;
                    for (const c of candidates) {
                        const r = c.getBoundingClientRect();
                        if (r.width === 0 || r.height === 0) continue;
                        // Must be below or right of label and close to it
                        const dy = r.top - lr.bottom;
                        if (dy < -10 || dy > 80) continue;
                        const dx = Math.abs((r.left + r.width/2) - (lr.left + lr.width/2));
                        if (dx > 220) continue;
                        const t = (c.innerText || c.textContent || '').trim();
                        if (!t || t.length > 60) continue;
                        // Skip the label itself
                        if (t.toLowerCase() === lower) continue;
                        const dist = Math.abs(dy) + dx;
                        if (dist < bestDist) {
                            bestDist = dist;
                            best = t;
                        }
                    }
                    return best;
                }""", [label_text])

                if displayed and displayed.strip().lower() == value.lower():
                    log("ok", f"  [{label_text}]: already '{value}' — skipping")
                    return True

                info = await find_input_by_label(label_text)
                if not info:
                    log("warn", f"  [{label_text}] dropdown not found")
                    return False

                log("info", f"  [{label_text}]: {value} (current displayed: '{displayed or info.get('currentValue','')}')")

                # ── Open the dropdown ──
                await page.mouse.click(info['x'], info['y'])
                await page.wait_for_timeout(800)

                async def close_dropdown():
                    """Close the dropdown by Escape + clicking neutral area."""
                    try:
                        await page.keyboard.press("Escape")
                        await page.wait_for_timeout(200)
                        # Click somewhere harmless (top-left blank area)
                        await page.mouse.click(50, 50)
                        await page.wait_for_timeout(300)
                    except Exception:
                        pass

                # ── Click the matching option, searching ONLY inside the overlay panel ──
                # Angular Material renders mat-option inside <body> at .cdk-overlay-pane.
                # We search there to avoid matching the page's regular text.
                result = await page.evaluate("""(val) => {
                    const v = val.toLowerCase().trim();
                    // Look for option panels (in priority order)
                    const panelSelectors = [
                        '.cdk-overlay-pane .mat-mdc-select-panel',
                        '.cdk-overlay-pane .mat-select-panel',
                        '.cdk-overlay-pane [role="listbox"]',
                        '.cdk-overlay-pane',
                        '.mat-select-panel',
                        '.mat-mdc-select-panel',
                        '[role="listbox"]'
                    ];
                    let panel = null;
                    for (const sel of panelSelectors) {
                        const found = document.querySelectorAll(sel);
                        for (const p of found) {
                            const r = p.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0) {
                                panel = p; break;
                            }
                        }
                        if (panel) break;
                    }
                    if (!panel) {
                        return {ok: false, reason: 'no-panel-visible'};
                    }
                    const opts = panel.querySelectorAll(
                        'mat-option, [role="option"], .mat-option, .mat-mdc-option, ' +
                        'li, .dropdown-item, span'
                    );
                    if (opts.length === 0) {
                        return {ok: false, reason: 'no-options', count: 0};
                    }
                    // Exact match
                    for (const o of opts) {
                        const t = (o.innerText || o.textContent || '').trim().toLowerCase();
                        if (t === v) {
                            o.scrollIntoView({block:'center'});
                            o.click();
                            return {ok: true, mode: 'EXACT', text: t};
                        }
                    }
                    // Partial match (option starts with value, or value starts with option)
                    for (const o of opts) {
                        const t = (o.innerText || o.textContent || '').trim().toLowerCase();
                        if (!t) continue;
                        if (t.startsWith(v) || v.startsWith(t) || t.includes(v)) {
                            o.scrollIntoView({block:'center'});
                            o.click();
                            return {ok: true, mode: 'PARTIAL', text: t};
                        }
                    }
                    // Collect option texts for debug
                    const all = Array.from(opts).map(o => (o.innerText||'').trim())
                                 .filter(t => t).slice(0, 10);
                    return {ok: false, reason: 'no-match', options: all};
                }""", value)

                if result.get('ok'):
                    log("ok", f"  -> {label_text}: {value} ({result['mode']}: '{result['text']}')")
                    await close_dropdown()
                    return True

                # ── Fallback: native <select> at the same coordinates ──
                native_ok = await page.evaluate("""([x,y,val]) => {
                    for (const sel of document.querySelectorAll('select')) {
                        const r = sel.getBoundingClientRect();
                        if (Math.abs(r.left+r.width/2-x)<10 && Math.abs(r.top+r.height/2-y)<10) {
                            for (const opt of sel.options) {
                                if (opt.text.trim().toLowerCase() === val.toLowerCase()) {
                                    sel.value = opt.value;
                                    sel.dispatchEvent(new Event('change',{bubbles:true}));
                                    return true;
                                }
                            }
                        }
                    }
                    return false;
                }""", [info['x'], info['y'], value])
                if native_ok:
                    log("ok", f"  -> {label_text}: {value} (<select>)")
                    await close_dropdown()
                    return True

                # ── Couldn't find — log details and close dropdown ──
                if result.get('reason') == 'no-panel-visible':
                    log("warn", f"  -> {label_text}: dropdown panel not visible after click")
                elif result.get('reason') == 'no-match':
                    opts = result.get('options', [])
                    log("warn", f"  -> {label_text}: '{value}' not found. "
                                f"Available: {opts}")
                else:
                    log("warn", f"  -> {label_text}: '{value}' — {result.get('reason','unknown')}")

                await close_dropdown()
                return False

            # ── Helper: type into autocomplete and pick matching option ──
            async def fill_autocomplete_by_label(label_text, value):
                if not value:
                    return False
                value = str(value).strip()
                info = await find_input_by_label(label_text)
                if not info:
                    log("warn", f"  [{label_text}] autocomplete not found")
                    return False
                log("info", f"  [{label_text}]: {value}")
                await page.mouse.click(info['x'], info['y'])
                await page.wait_for_timeout(200)
                await page.keyboard.press("Control+a")
                await page.keyboard.press("Delete")
                await page.wait_for_timeout(150)
                # Type only first chunk to trigger autocomplete suggestions
                search_term = value.split("-")[0].split("(")[0].strip()
                await page.keyboard.type(search_term, delay=80)
                await page.wait_for_timeout(1200)
                # Try clicking exact then partial match
                full_val = value.upper()
                for _scroll in range(8):
                    result = await page.evaluate("""(v) => {
                        const val = v.trim().toUpperCase();
                        const opts = document.querySelectorAll(
                            'mat-option, [role="option"], .cdk-overlay-pane li, .ng-option');
                        if (opts.length === 0) return 'NO_OPTIONS';
                        for (const o of opts) {
                            const t = (o.innerText||o.textContent||'').trim().toUpperCase();
                            if (t === val) { o.scrollIntoView({block:'center'}); o.click();
                                              return 'EXACT||'+o.innerText.trim(); }
                        }
                        for (const o of opts) {
                            const t = (o.innerText||o.textContent||'').trim().toUpperCase();
                            if (t.includes(val) || val.includes(t)) {
                                o.scrollIntoView({block:'center'}); o.click();
                                return 'PARTIAL||'+o.innerText.trim();
                            }
                        }
                        return 'NO_MATCH||' + opts.length;
                    }""", full_val)
                    if result.startswith("EXACT") or result.startswith("PARTIAL"):
                        log("ok", f"  -> {label_text}: {result.split('||')[1]}")
                        await page.wait_for_timeout(400)
                        return True
                    if result == "NO_OPTIONS":
                        if _scroll == 0:
                            await page.wait_for_timeout(800); continue
                        break
                    # Scroll and retry
                    await page.evaluate("""() => {
                        const panel = document.querySelector(
                            '.mat-autocomplete-panel, .cdk-overlay-pane, [role="listbox"]');
                        if (panel) panel.scrollBy({top: 200});
                    }""")
                    await page.wait_for_timeout(700)
                # Last resort: keep typed text and Tab away
                await page.keyboard.press("Tab")
                await page.wait_for_timeout(300)
                log("warn", f"  [{label_text}]: '{value}' typed but no dropdown match")
                return False

            # ── Helper: click a radio button labelled `label_text` ──
            async def click_radio_by_label(label_text):
                ok = await page.evaluate("""(lbl) => {
                    const lower = lbl.toLowerCase().trim();
                    // Strategy 1: input[type=radio] whose label/parent text matches
                    for (const r of document.querySelectorAll('input[type="radio"]')) {
                        const lab = r.closest('label') ||
                                    document.querySelector('label[for="' + r.id + '"]') ||
                                    r.parentElement;
                        const t = (lab?.innerText || lab?.textContent || '').trim().toLowerCase();
                        if (t === lower || t.startsWith(lower)) {
                            r.scrollIntoView({block:'center'}); r.click();
                            return true;
                        }
                    }
                    // Strategy 2: mat-radio-button — match by inner text
                    for (const mr of document.querySelectorAll('mat-radio-button, [role="radio"]')) {
                        const t = (mr.innerText || mr.textContent || '').trim().toLowerCase();
                        if (t === lower || t.startsWith(lower)) {
                            mr.scrollIntoView({block:'center'}); mr.click();
                            return true;
                        }
                    }
                    // Strategy 3: small element whose text == label, click it
                    let bestEl = null, bestArea = Infinity;
                    for (const el of document.querySelectorAll('label, span, div')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if (t === lower) {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) continue;
                            const area = r.width * r.height;
                            if (area < bestArea) { bestArea = area; bestEl = el; }
                        }
                    }
                    if (bestEl) {
                        bestEl.scrollIntoView({block:'center'});
                        bestEl.click();
                        return true;
                    }
                    return false;
                }""", label_text)
                return ok

            # ────────────────────────────────────────────────────────
            # 1. Transaction Type (dropdown — usually pre-set to "New")
            # ────────────────────────────────────────────────────────
            txn_type = D.get("transaction_type", "New").strip() or "New"
            await pick_dropdown_by_label("Transaction Type", txn_type)
            await page.wait_for_timeout(400)

            # ────────────────────────────────────────────────────────
            # 2. Battery Number (text)
            # ────────────────────────────────────────────────────────
            await fill_text_by_label("Battery Number", D.get("battery_number", ""))
            await page.wait_for_timeout(400)

            # ────────────────────────────────────────────────────────
            # 3. Purchase Date (date DD/MM/YYYY)
            # ────────────────────────────────────────────────────────
            await fill_date_by_label("Purchase Date", D.get("purchase_date", ""))
            await page.wait_for_timeout(400)

            # ────────────────────────────────────────────────────────
            # 4. Customer Type (radio: Individual / Corporate)
            # ────────────────────────────────────────────────────────
            cust_type = D.get("customer_type", "Individual").strip()
            if cust_type:
                target = "Corporate" if cust_type.lower().startswith("corp") else "Individual"
                log("info", f"  [Customer Type]: {target}")
                if await click_radio_by_label(target):
                    log("ok", f"  -> Customer Type: {target}")
                else:
                    log("warn", f"  Customer Type radio '{target}' not found")
                await page.wait_for_timeout(500)

            # ────────────────────────────────────────────────────────
            # 5. Manufacturer (autocomplete)
            # ────────────────────────────────────────────────────────
            await fill_autocomplete_by_label("Manufacturer", D.get("manufacturer", ""))
            await page.wait_for_timeout(800)

            # ────────────────────────────────────────────────────────
            # 6. Model (dropdown — populated after manufacturer)
            # ────────────────────────────────────────────────────────
            await pick_dropdown_by_label("Model", D.get("model", ""))
            await page.wait_for_timeout(800)

            # ────────────────────────────────────────────────────────
            # 7. Year of Manufacture (dropdown)
            # ────────────────────────────────────────────────────────
            yom = D.get("year_of_manufacture", "")
            if yom:
                await pick_dropdown_by_label("Year of Manufacture", str(yom))
                await page.wait_for_timeout(400)

            # ────────────────────────────────────────────────────────
            # 8. Policy Start Date (date)
            # ────────────────────────────────────────────────────────
            await fill_date_by_label("Policy Start Date", D.get("policy_start_date", ""))
            await page.wait_for_timeout(400)

            # ────────────────────────────────────────────────────────
            # 9. Policy End Date (date — may auto-fill)
            # ────────────────────────────────────────────────────────
            await fill_date_by_label("Policy End Date", D.get("policy_end_date", ""))
            await page.wait_for_timeout(400)

            # ────────────────────────────────────────────────────────
            # 10. Unique Identification Number of Asset (text)
            # ────────────────────────────────────────────────────────
            await fill_text_by_label("Unique Identification Number of Asset",
                                     D.get("uin_asset", ""))
            await page.wait_for_timeout(400)
            # Some forms use shorter label
            if not D.get("uin_asset"):
                pass  # nothing to do
            else:
                # Try fallback if the long label didn't match
                pass

            # ────────────────────────────────────────────────────────
            # 11. Invoice Number (text)
            # ────────────────────────────────────────────────────────
            await fill_text_by_label("Invoice Number", D.get("invoice_number", ""))
            await page.wait_for_timeout(400)

            # ────────────────────────────────────────────────────────
            # 12. Customer State (autocomplete / text)
            # ────────────────────────────────────────────────────────
            cust_state = D.get("customer_state", "") or D.get("state", "")
            if cust_state:
                # Try autocomplete first, then fall back to plain text fill
                if not await fill_autocomplete_by_label("Customer State", cust_state):
                    await fill_text_by_label("Customer State", cust_state)
                await page.wait_for_timeout(400)

            # Optional: Watt field (some EV variants ask for wattage)
            watt = D.get("watt", "")
            if watt:
                if not await fill_text_by_label("Watt", watt):
                    await fill_text_by_label("Wattage", watt)

            steps.append("fill_details")
            log("ok", "══ SECTION DONE: EV Vehicle Details filled ══")
            await page.wait_for_timeout(1500)


            # ── STEP 5: Click GET QUOTE ──────────────────────────────
            log("step", "STEP 5 — GET QUOTE")
            # Wait for page to fully settle after all field fills
            log("info", "Waiting for page to settle before GET QUOTE...")
            try:
                await page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(1000)

            async def click_by_text(texts):
                """Click any visible element matching text — works for any tag."""
                if isinstance(texts, str): texts = [texts]
                for text in texts:
                    for tag in ["button", "input", "a", "span", "div"]:
                        try:
                            loc = page.locator(f"{tag}:has-text('{text}')").first
                            await loc.wait_for(state="visible", timeout=2000)
                            await loc.scroll_into_view_if_needed()
                            await loc.click()
                            log("ok", f"  -> Clicked '{text}' via <{tag}>")
                            return True
                        except Exception:
                            continue
                    # JS scan — find smallest matching element and mouse.click it
                    try:
                        coords = await page.evaluate(f"""() => {{
                            let best = null, bestArea = Infinity;
                            for (const el of document.querySelectorAll('*')) {{
                                const rect = el.getBoundingClientRect();
                                if (rect.width === 0 || rect.height === 0) continue;
                                const txt = (el.innerText || el.value || '').trim();
                                if (txt === '{text}') {{
                                    const area = rect.width * rect.height;
                                    if (area < bestArea) {{ bestArea = area; best = el; }}
                                }}
                            }}
                            if (best) {{
                                best.scrollIntoView({{block:'center'}});
                                const r = best.getBoundingClientRect();
                                return {{x: r.left+r.width/2, y: r.top+r.height/2}};
                            }}
                            return null;
                        }}""")
                        if coords:
                            await page.wait_for_timeout(300)
                            await page.mouse.click(coords['x'], coords['y'])
                            log("ok", f"  -> Clicked '{text}' via JS mouse.click at ({coords['x']:.0f},{coords['y']:.0f})")
                            return True
                    except Exception:
                        pass
                return False

            if not await click_by_text(["Get Quote", "GET QUOTE", "Get Quotation"]):
                log("warn", "GET QUOTE not found — check browser")
                await pause("Click 'GET QUOTE' manually, then press ENTER")
            try:
                await page.wait_for_load_state("networkidle", timeout=30000)
            except Exception:
                pass
            await page.wait_for_timeout(5000)
            steps.append("get_quote")

            # ── IDV fill — after Get Quote (portal shows IDV input on the quote/customize page) ──
            if D.get("idv"):
                idv_val = str(D["idv"]).strip().split(".")[0]
                log("info", f"Setting IDV (post-Get Quote): {idv_val}")

                # Wait for IDV input to appear (up to 20 x 500ms = 10s)
                idv_found = False
                for _attempt in range(20):
                    idv_found = await page.evaluate("""() => {
                        let inp = document.getElementById('sirangeinput');
                        if (!inp) inp = document.querySelector('input[name="amountInput"]');
                        if (!inp) {
                            const slider = document.querySelector(
                                'input[type="range"], input.custom-range');
                            if (!slider) return false;
                            const all = Array.from(document.querySelectorAll('input'));
                            const idx = all.indexOf(slider);
                            for (let i = idx - 1; i >= 0; i--) {
                                const c = all[i];
                                if (['range','checkbox','radio','hidden','submit','button']
                                    .includes(c.type)) continue;
                                if (c.getBoundingClientRect().width === 0) continue;
                                inp = c; break;
                            }
                        }
                        if (!inp) return false;
                        inp.scrollIntoView({block:'center'});
                        return true;
                    }""")
                    if idv_found:
                        log("fill", f"IDV input found (attempt {_attempt+1})")
                        break
                    await page.wait_for_timeout(500)

                # Wait for default value to populate before overwriting
                if idv_found:
                    for _dv in range(20):
                        idv_current = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) {
                                const slider = document.querySelector(
                                    'input[type="range"], input.custom-range');
                                if (!slider) return '';
                                const all = Array.from(document.querySelectorAll('input'));
                                const idx = all.indexOf(slider);
                                for (let i = idx - 1; i >= 0; i--) {
                                    const c = all[i];
                                    if (['range','checkbox','radio','hidden','submit','button']
                                        .includes(c.type)) continue;
                                    if (c.getBoundingClientRect().width === 0) continue;
                                    return c.value;
                                }
                            }
                            return inp ? inp.value : '';
                        }""") or ""
                        if idv_current.strip() and idv_current.strip() not in ("0", ""):
                            log("fill", f"IDV default populated: {idv_current.strip()}")
                            break
                        await page.wait_for_timeout(500)

                if not idv_found:
                    log("warn", "IDV input not found on quote page — skipping IDV fill")
                else:
                    await page.wait_for_timeout(400)
                    # Single fill attempt — the guard loop below re-applies only if Angular resets it
                    idv_coords = await page.evaluate("""() => {
                        let inp = document.getElementById('sirangeinput');
                        if (!inp) inp = document.querySelector('input[name="amountInput"]');
                        if (!inp) {
                            const slider = document.querySelector(
                                'input[type="range"], input.custom-range');
                            if (!slider) return null;
                            const all = Array.from(document.querySelectorAll('input'));
                            const idx = all.indexOf(slider);
                            for (let i = idx - 1; i >= 0; i--) {
                                const c = all[i];
                                if (['range','checkbox','radio','hidden','submit','button']
                                    .includes(c.type)) continue;
                                if (c.getBoundingClientRect().width === 0) continue;
                                inp = c; break;
                            }
                        }
                        if (!inp) return null;
                        inp.scrollIntoView({block:'center'});
                        const r = inp.getBoundingClientRect();
                        return {x: r.left + r.width/2, y: r.top + r.height/2};
                    }""")
                    if idv_coords:
                        await page.mouse.click(idv_coords['x'], idv_coords['y'], click_count=3)
                        await page.wait_for_timeout(150)
                        await page.keyboard.press("Control+a")
                        await page.keyboard.press("Delete")
                        await page.wait_for_timeout(80)
                        await page.keyboard.type(idv_val, delay=80)
                        await page.wait_for_timeout(400)
                        # Commit via Angular native input setter
                        await page.evaluate("""(val) => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) {
                                const slider = document.querySelector(
                                    'input[type="range"], input.custom-range');
                                if (!slider) return;
                                const all = Array.from(document.querySelectorAll('input'));
                                const idx = all.indexOf(slider);
                                for (let i = idx - 1; i >= 0; i--) {
                                    const c = all[i];
                                    if (['range','checkbox','radio','hidden','submit','button']
                                        .includes(c.type)) continue;
                                    if (c.getBoundingClientRect().width === 0) continue;
                                    inp = c; break;
                                }
                            }
                            if (!inp) return;
                            const setter = Object.getOwnPropertyDescriptor(
                                window.HTMLInputElement.prototype, 'value').set;
                            setter.call(inp, ''); setter.call(inp, val);
                            inp.dispatchEvent(new Event('input', {bubbles: true}));
                            inp.dispatchEvent(new Event('change', {bubbles: true}));
                        }""", idv_val)
                        await page.wait_for_timeout(600)
                        actual = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            return inp ? inp.value : '';
                        }""") or ""
                        if actual.strip() == idv_val:
                            log("ok", f"IDV confirmed: {actual.strip()}")
                        else:
                            log("warn", f"IDV got '{actual.strip()}', expected '{idv_val}' — guard will re-apply if Angular resets")

                    # IDV Guard: re-apply up to 10x if Angular resets it
                    for _guard in range(10):
                        current_idv = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            return inp ? inp.value : '';
                        }""") or ""
                        if str(current_idv).strip() == idv_val:
                            await page.wait_for_timeout(600)
                            recheck = await page.evaluate("""() => {
                                let inp = document.getElementById('sirangeinput');
                                if (!inp) inp = document.querySelector('input[name="amountInput"]');
                                return inp ? inp.value : '';
                            }""") or ""
                            if str(recheck).strip() == idv_val:
                                log("ok", f"IDV stable: {recheck}")
                                break
                        log("warn", f"IDV reverted to '{current_idv}' — re-applying {idv_val}")
                        idv_gc = await page.evaluate("""() => {
                            let inp = document.getElementById('sirangeinput');
                            if (!inp) inp = document.querySelector('input[name="amountInput"]');
                            if (!inp) return null;
                            inp.scrollIntoView({block:'center'});
                            const r = inp.getBoundingClientRect();
                            return {x: r.left+r.width/2, y: r.top+r.height/2};
                        }""")
                        if idv_gc:
                            await page.mouse.click(idv_gc['x'], idv_gc['y'], click_count=3)
                            await page.wait_for_timeout(150)
                            await page.keyboard.press("Delete")
                            await page.keyboard.type(idv_val, delay=80)
                            await page.wait_for_timeout(500)
                        else:
                            break

                    # Click "Update" button after IDV is set
                    log("info", "Clicking Update button after IDV...")
                    update_clicked = False
                    for _upd in range(5):
                        try:
                            upd = page.get_by_text("Update", exact=True).first
                            await upd.wait_for(state="visible", timeout=2000)
                            await upd.click()
                            log("ok", "  -> Update clicked")
                            update_clicked = True
                            break
                        except Exception:
                            pass
                        # JS fallback
                        try:
                            coords = await page.evaluate("""() => {
                                for (const el of document.querySelectorAll('button, input[type="submit"], a')) {
                                    const t = (el.innerText||el.value||'').trim();
                                    if (t === 'Update' || t === 'UPDATE') {
                                        el.scrollIntoView({block:'center'});
                                        const r = el.getBoundingClientRect();
                                        if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                    }
                                }
                                return null;
                            }""")
                            if coords:
                                await page.mouse.click(coords['x'], coords['y'])
                                log("ok", "  -> Update clicked (JS)")
                                update_clicked = True
                                break
                        except Exception:
                            pass
                        await page.wait_for_timeout(500)
                    if not update_clicked:
                        log("warn", "  Update button not found — click manually")
                    try:
                        await page.wait_for_load_state("networkidle", timeout=15000)
                    except Exception:
                        pass
                    await page.wait_for_timeout(2000)
            # ── STEP 6 / 7 / 7c / 8 — SKIPPED for Electric Bike category ──
            # After IDV update the portal stays on the Plan Selection page.
            # We skip Add/Remove Cover, Add-ons, Recalculate and Submit
            # Additional Covers entirely and jump straight to Additional Discounts.
            log("step", "STEP 6/7/7c/8 — SKIPPED (electric bike: going directly to Additional Discounts)")
            steps.append("customize_opened")
            steps.append("addons_selected")
            steps.append("recalculated")
            steps.append("submit_coverages")

            # ── STEP 9: Additional Discounts (if enabled in Excel) ────
            if D.get("additional_discount"):
                log("step", "STEP 9 — Additional Discounts (Loading)")

                # Wait for Plan Selection page to load (it has "Buy Now" buttons)
                log("info", "  Waiting for Plan Selection page to load...")
                for _ps in range(40):
                    has_plan = await page.evaluate("""() => {
                        const txt = document.body.innerText.toLowerCase();
                        return txt.includes('buy now') || txt.includes('plan selection')
                            || txt.includes('additional discounts');
                    }""")
                    if has_plan:
                        log("ok", "  -> Plan Selection page loaded")
                        break
                    await page.wait_for_timeout(500)
                else:
                    log("warn", "  Plan Selection page timed out — continuing anyway")

                await page.wait_for_timeout(2000)

                # Scroll to bottom to find "+Add" button in Additional Discounts section
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await page.wait_for_timeout(1000)

                # Click "+Add" button in Additional Discounts section
                add_clicked = False

                # Strategy 1: Find the "Additional Discounts" section, then click "+ Add" near it
                try:
                    # The button shows as "+ Add" — try multiple text patterns
                    for btn_text in ["+ Add", "+Add", "Add"]:
                        loc = page.get_by_text(btn_text, exact=True)
                        cnt = await loc.count()
                        if cnt > 0:
                            # Click the last match (likely the one at the bottom near Additional Discounts)
                            el = loc.last
                            await el.scroll_into_view_if_needed()
                            await page.wait_for_timeout(300)
                            await el.click()
                            log("ok", f"  -> '{btn_text}' button clicked")
                            add_clicked = True
                            break
                except Exception as e:
                    log("info", f"  get_by_text failed: {e}")

                # Strategy 2: Use Playwright locator with role
                if not add_clicked:
                    try:
                        btn_loc = page.get_by_role("button", name="Add")
                        if await btn_loc.count() > 0:
                            await btn_loc.last.scroll_into_view_if_needed()
                            await page.wait_for_timeout(300)
                            await btn_loc.last.click()
                            log("ok", "  -> 'Add' button clicked via role")
                            add_clicked = True
                    except Exception:
                        pass

                # Strategy 3: JS — find element containing "Add" near "Additional Discounts"
                if not add_clicked:
                    coords = await page.evaluate("""() => {
                        // First find "Additional Discounts" section y-position
                        let adY = 0;
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText || '').trim();
                            if (t.includes('Additional Discounts') && t.length < 100) {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) { adY = r.top; break; }
                            }
                        }
                        // Now find a clickable element with "Add" text near that y-position
                        for (const el of document.querySelectorAll('button, a, div, span')) {
                            const t = (el.innerText || el.textContent || '').trim();
                            if (t.includes('Add') && t.length < 20) {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0 && r.height > 0) {
                                    // Must be near the Additional Discounts section (within 100px)
                                    if (adY > 0 && Math.abs(r.top - adY) > 100) continue;
                                    el.scrollIntoView({block: 'center'});
                                    const r2 = el.getBoundingClientRect();
                                    return {x: r2.left + r2.width/2, y: r2.top + r2.height/2};
                                }
                            }
                        }
                        return null;
                    }""")
                    if coords:
                        await page.wait_for_timeout(300)
                        await page.mouse.click(coords['x'], coords['y'])
                        log("ok", f"  -> '+Add' clicked via JS at ({coords['x']:.0f},{coords['y']:.0f})")
                        add_clicked = True

                if not add_clicked:
                    log("warn", "  '+Add' button not found")
                    await pause("Click '+Add' in Additional Discounts section, then press ENTER")

                await page.wait_for_timeout(1500)

                # ── Decide which checkbox to tick based on Excel dropdown ──
                # Excel "Additional Discounts Dropdown" can be:
                #   "Other Discount"  → tick "Other Discount" checkbox
                #   "Other Loading"   → tick "Other Loading"  checkbox
                # Default (blank/unknown) → fall back to "Other Loading" (legacy behaviour)
                dropdown_val = str(D.get("additional_discounts_dropdown", "") or "").strip().lower()
                if "discount" in dropdown_val and "loading" not in dropdown_val:
                    target_label = "Other Discount"
                elif "loading" in dropdown_val:
                    target_label = "Other Loading"
                else:
                    target_label = "Other Loading"  # legacy default

                # ── Tick the chosen checkbox ──
                log("info", f"  Looking for '{target_label}' checkbox...")
                ol_clicked = False

                try:
                    ol_loc = page.get_by_text(target_label, exact=False)
                    if await ol_loc.count() > 0:
                        el = ol_loc.first
                        await el.scroll_into_view_if_needed()
                        await page.wait_for_timeout(300)
                        box = await el.bounding_box()
                        if box:
                            # Click the checkbox area to the left of the text
                            click_x = box['x'] - 20
                            if click_x < 5:
                                click_x = box['x'] + 5
                            click_y = box['y'] + box['height'] / 2
                            await page.mouse.click(click_x, click_y)
                            log("ok", f"  -> '{target_label}' clicked at ({click_x:.0f},{click_y:.0f})")
                            ol_clicked = True
                            await page.wait_for_timeout(800)
                except Exception as e:
                    log("info", f"  get_by_text for {target_label} failed: {e}")

                # Fallback: try clicking the text itself
                if not ol_clicked:
                    try:
                        ol_loc = page.get_by_text(target_label, exact=False)
                        if await ol_loc.count() > 0:
                            await ol_loc.first.click()
                            log("ok", f"  -> '{target_label}' clicked (direct text)")
                            ol_clicked = True
                            await page.wait_for_timeout(800)
                    except Exception:
                        pass

                if not ol_clicked:
                    log("warn", f"  '{target_label}' not found")
                    await pause(f"Tick '{target_label}' manually, then press ENTER")

                await page.wait_for_timeout(500)

                # ── Fill the Loading% / Discount% value ──
                loading_val = D.get("loading_pct", "")
                if loading_val:
                    # Decide which placeholder/label to expect based on the
                    # checkbox we just ticked.
                    is_discount = (target_label == "Other Discount")
                    placeholder_prefix = "Enter discount" if is_discount else "Enter loading"
                    label_keyword     = "discount percent" if is_discount else "loading percent"
                    field_keyword     = "discount" if is_discount else "loading"

                    log("info", f"  Entering {target_label} % value: {loading_val}")
                    filled = False

                    # Strategy 1: Find input by placeholder "Enter discount/loading %"
                    try:
                        inp = page.get_by_placeholder(placeholder_prefix, exact=False)
                        if await inp.count() > 0:
                            await inp.first.scroll_into_view_if_needed()
                            await page.wait_for_timeout(300)
                            await inp.first.click()
                            await inp.first.fill(str(loading_val))
                            log("ok", f"  -> '{loading_val}' entered via placeholder match")
                            filled = True
                    except Exception as e:
                        log("info", f"  Placeholder search failed: {e}")

                    # Strategy 2: Find input inside the expanded section via JS
                    if not filled:
                        try:
                            js_filled = await page.evaluate("""(args) => {
                                const val = args.val;
                                const fieldKw = args.fieldKw;
                                const labelKw = args.labelKw;
                                // Find input with placeholder containing the field keyword
                                for (const inp of document.querySelectorAll('input')) {
                                    const ph = (inp.placeholder || '').toLowerCase();
                                    if (ph.includes(fieldKw)) {
                                        const r = inp.getBoundingClientRect();
                                        if (r.width > 0 && r.height > 0) {
                                            inp.focus();
                                            inp.value = '';
                                            inp.dispatchEvent(new Event('input', {bubbles:true}));
                                            inp.value = val;
                                            inp.dispatchEvent(new Event('input', {bubbles:true}));
                                            inp.dispatchEvent(new Event('change', {bubbles:true}));
                                            return 'placeholder';
                                        }
                                    }
                                }
                                // Find "<labelKw>" label, then get the next input
                                for (const el of document.querySelectorAll('*')) {
                                    const t = (el.innerText || '').trim().toLowerCase();
                                    if (t.includes(labelKw) && t.length < 30) {
                                        const r = el.getBoundingClientRect();
                                        if (r.width === 0) continue;
                                        // Look for input below this label
                                        const inputs = document.querySelectorAll('input');
                                        let closest = null, closestDist = Infinity;
                                        for (const inp of inputs) {
                                            const ir = inp.getBoundingClientRect();
                                            if (ir.width === 0 || ir.height === 0) continue;
                                            const dist = Math.abs(ir.top - r.bottom);
                                            if (dist < closestDist && ir.top >= r.top - 10) {
                                                closestDist = dist;
                                                closest = inp;
                                            }
                                        }
                                        if (closest && closestDist < 80) {
                                            closest.focus();
                                            closest.value = '';
                                            closest.dispatchEvent(new Event('input', {bubbles:true}));
                                            closest.value = val;
                                            closest.dispatchEvent(new Event('input', {bubbles:true}));
                                            closest.dispatchEvent(new Event('change', {bubbles:true}));
                                            return 'label';
                                        }
                                    }
                                }
                                return null;
                            }""", {"val": str(loading_val), "fieldKw": field_keyword, "labelKw": label_keyword})
                            if js_filled:
                                log("ok", f"  -> '{loading_val}' entered via JS ({js_filled})")
                                filled = True
                        except Exception as e:
                            log("info", f"  JS fill failed: {e}")

                    if not filled:
                        log("warn", f"  Could not enter {target_label} % value")
                        await pause(f"Enter '{loading_val}' in the {target_label} % field, then press ENTER")

                await page.wait_for_timeout(800)

                # ── Click "Apply & Continue" button ──
                log("info", "  Clicking 'Apply & Continue'...")
                ac_clicked = False

                # Strategy 1: Exact text match for the button
                try:
                    ac_loc = page.get_by_text("Apply & Continue", exact=True)
                    if await ac_loc.count() > 0:
                        await ac_loc.first.scroll_into_view_if_needed()
                        await page.wait_for_timeout(300)
                        await ac_loc.first.click()
                        log("ok", "  -> 'Apply & Continue' clicked (exact match)")
                        ac_clicked = True
                except Exception:
                    pass

                # Strategy 2: Playwright get_by_role button
                if not ac_clicked:
                    try:
                        btn = page.get_by_role("button", name="Apply & Continue")
                        if await btn.count() > 0:
                            await btn.first.scroll_into_view_if_needed()
                            await page.wait_for_timeout(300)
                            await btn.first.click()
                            log("ok", "  -> 'Apply & Continue' clicked (role button)")
                            ac_clicked = True
                    except Exception:
                        pass

                # Strategy 3: JS — find button/clickable with exact "Apply & Continue" text
                if not ac_clicked:
                    coords = await page.evaluate("""() => {
                        for (const el of document.querySelectorAll('button, a, div, span')) {
                            const t = (el.innerText || el.textContent || '').trim();
                            if (t === 'Apply & Continue' || t === 'Apply &amp; Continue'
                                || t === 'Apply and Continue') {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0 && r.height > 0) {
                                    el.scrollIntoView({block: 'center'});
                                    const r2 = el.getBoundingClientRect();
                                    return {x: r2.left + r2.width/2, y: r2.top + r2.height/2};
                                }
                            }
                        }
                        return null;
                    }""")
                    if coords:
                        await page.wait_for_timeout(300)
                        await page.mouse.click(coords['x'], coords['y'])
                        log("ok", f"  -> 'Apply & Continue' clicked via JS at ({coords['x']:.0f},{coords['y']:.0f})")
                        ac_clicked = True

                if not ac_clicked:
                    log("warn", "  'Apply & Continue' not found")
                    await pause("Click 'Apply & Continue' manually, then press ENTER")

                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)
                steps.append("additional_discount")
            else:
                log("info", "STEP 9 — Additional Discounts: SKIPPED (not enabled in Excel)")

            # ── STEP 10: Click "Buy Now" on Basic Plan card ──────────
            log("step", "STEP 10 — Click Buy Now on Basic Plan")

            # Wait for Plan Selection page to be ready
            await page.wait_for_timeout(2000)
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(500)
            await page.evaluate("window.scrollTo(0, 0)")
            await page.wait_for_timeout(500)

            buy_clicked = False

            # Strategy 1: Find "Buy Now" button on Basic Plan (first/left card)
            try:
                buy_loc = page.get_by_text("Buy Now", exact=True)
                if await buy_loc.count() > 0:
                    # Click the FIRST "Buy Now" (Basic Plan is the left card)
                    await buy_loc.first.scroll_into_view_if_needed()
                    await page.wait_for_timeout(300)
                    await buy_loc.first.click()
                    log("ok", "  -> 'Buy Now' clicked (Basic Plan)")
                    buy_clicked = True
            except Exception as e:
                log("info", f"  get_by_text failed: {e}")

            # Strategy 2: Playwright role
            if not buy_clicked:
                try:
                    btn = page.get_by_role("button", name="Buy Now")
                    if await btn.count() > 0:
                        await btn.first.scroll_into_view_if_needed()
                        await page.wait_for_timeout(300)
                        await btn.first.click()
                        log("ok", "  -> 'Buy Now' clicked via role")
                        buy_clicked = True
                except Exception:
                    pass

            # Strategy 3: JS
            if not buy_clicked:
                coords = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('button, a, div, span')) {
                        const t = (el.innerText || el.textContent || '').trim();
                        if (t === 'Buy Now') {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0) {
                                el.scrollIntoView({block: 'center'});
                                const r2 = el.getBoundingClientRect();
                                return {x: r2.left + r2.width/2, y: r2.top + r2.height/2};
                            }
                        }
                    }
                    return null;
                }""")
                if coords:
                    await page.mouse.click(coords['x'], coords['y'])
                    log("ok", f"  -> 'Buy Now' clicked via JS at ({coords['x']:.0f},{coords['y']:.0f})")
                    buy_clicked = True

            if not buy_clicked:
                log("warn", "  'Buy Now' button not found")
                await pause("Click 'Buy Now' on the Basic Plan card, then press ENTER")

            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)
            steps.append("buy_now")

            # ── STEP 11: Proposal Form (Engine/Chassis + KYC popup) ──
            log("step", "STEP 11 — Proposal Form")
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            # Wait until Motor Number / Battery Number OR Chassis fields appear in DOM.
            # EV proposal page uses different labels than regular motor: "Motor Number /
            # Battery Number" and "Chasis Number" (with a typo).
            log("info", "Waiting for Motor Number / Battery Number / Chassis fields to appear...")
            for _ef in range(60):
                fields_ready = await page.evaluate("""() => {
                    const txt = document.body.innerText.toLowerCase();
                    // Regular motor labels
                    const hasEngine  = txt.includes('engine number');
                    const hasChassis = txt.includes('chassis number') || txt.includes('chasis number');
                    // EV-specific labels
                    const hasMotorBat = txt.includes('motor number') || txt.includes('battery number');
                    const inputs = document.querySelectorAll('input');
                    return (hasEngine || hasChassis || hasMotorBat) && inputs.length > 0;
                }""")
                if fields_ready:
                    log("info", f"  Motor/Battery/Chassis fields ready (attempt {_ef+1})")
                    break
                await page.wait_for_timeout(500)

            async def mouse_fill(find_js, value, field_name):
                await page.evaluate(f"() => {{ const el={find_js}; if(el) el.scrollIntoView({{block:'center'}}); }}")
                await page.wait_for_timeout(200)
                c = await page.evaluate(f"() => {{ const el={find_js}; if(!el) return null; const r=el.getBoundingClientRect(); return r.width>0?{{x:r.left+r.width/2,y:r.top+r.height/2}}:null; }}")
                if c:
                    await page.mouse.click(c['x'], c['y'])
                    await page.wait_for_timeout(120)
                    await page.keyboard.press("Control+a")
                    await page.keyboard.press("Delete")
                    await page.wait_for_timeout(80)
                    await page.keyboard.type(str(value), delay=40)
                    await page.wait_for_timeout(200)
                    log("fill", f"  -> {field_name}: {value}")
                    return True
                log("warn", f"  {field_name} not found")
                return False

            async def click_last_button(text_upper, label):
                log("info", f"Clicking {label}")
                await page.evaluate(f"""() => {{
                    const all = Array.from(document.querySelectorAll('button,input[type="submit"]'))
                        .filter(el => (el.innerText||el.value||'').trim().toUpperCase() === '{text_upper}');
                    if (all.length) all[all.length-1].scrollIntoView({{block:'center'}});
                }}""")
                await page.wait_for_timeout(400)
                c = await page.evaluate(f"""() => {{
                    const all = Array.from(document.querySelectorAll('button,input[type="submit"]'))
                        .filter(el => (el.innerText||el.value||'').trim().toUpperCase() === '{text_upper}')
                        .filter(el => {{ const r=el.getBoundingClientRect(); return r.width>0&&r.height>0; }});
                    if (!all.length) return null;
                    for (const el of [...all].reverse()) {{
                        let node=el.parentElement;
                        for (let i=0;i<12;i++) {{
                            if(!node) break;
                            const s=window.getComputedStyle(node);
                            if (s.position==='fixed'||s.position==='absolute'||
                                ['modal','popup','overlay','dialog','cdk-overlay'].some(c=>node.classList.contains(c))) {{
                                const r=el.getBoundingClientRect();
                                return {{x:r.left+r.width/2,y:r.top+r.height/2}};
                            }}
                            node=node.parentElement;
                        }}
                    }}
                    const el=all[all.length-1];
                    const r=el.getBoundingClientRect();
                    return {{x:r.left+r.width/2,y:r.top+r.height/2}};
                }}""")
                if c:
                    await page.mouse.click(c['x'], c['y'])
                    log("ok", f"  -> {label} clicked at ({c['x']:.0f},{c['y']:.0f})")
                    return True
                log("warn", f"  {label} not found")
                return False

            # ── 11a: Motor Number / Battery Number (EV) = Engine Number field ──
            # On the EV Proposal page the first field is labelled
            # "Motor Number / Battery Number" — we fill it with battery_number.
            battery_val_11 = D.get("battery_number", "").strip()
            if battery_val_11:
                motor_bat_js = """(() => {
                    // Strategy 1: placeholder contains 'motor' or 'battery'
                    const byPh = Array.from(document.querySelectorAll('input')).find(i => {
                        const ph = (i.placeholder||'').toLowerCase();
                        return ph.includes('motor') || ph.includes('battery');
                    });
                    if (byPh) return byPh;
                    // Strategy 2: label text contains 'motor number' or 'battery number'
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if ((t.includes('motor number') || t.includes('battery number')) && t.length < 60) {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0) continue;
                            let best = null, bestDist = Infinity;
                            for (const inp of document.querySelectorAll('input')) {
                                const ir = inp.getBoundingClientRect();
                                if (ir.width === 0 || ir.height === 0) continue;
                                const dist = Math.abs(ir.left - r.left) + Math.abs(ir.top - r.bottom);
                                if (dist < bestDist) { bestDist = dist; best = inp; }
                            }
                            if (best && bestDist < 180) return best;
                        }
                    }
                    // Strategy 3: fallback — regular "engine number" label
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if (t.includes('engine number') && t.length < 30) {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0) continue;
                            let best = null, bestDist = Infinity;
                            for (const inp of document.querySelectorAll('input')) {
                                const ir = inp.getBoundingClientRect();
                                if (ir.width === 0 || ir.height === 0) continue;
                                const dist = Math.abs(ir.left - r.left) + Math.abs(ir.top - r.bottom);
                                if (dist < bestDist) { bestDist = dist; best = inp; }
                            }
                            if (best && bestDist < 150) return best;
                        }
                    }
                    return null;
                })()"""
                await mouse_fill(motor_bat_js, battery_val_11, "Motor Number / Battery Number")
            await page.wait_for_timeout(300)

            # ── 11b: Chasis Number = Unique Identification Number of Asset ──
            chassis_val_11 = D.get("uin_asset", "").strip()
            if chassis_val_11:
                chassis_js = """(() => {
                    // Only consider EMPTY inputs (skip already-filled Motor/Battery field)
                    const emptyInputs = Array.from(document.querySelectorAll('input')).filter(i => {
                        const r = i.getBoundingClientRect();
                        return r.width > 0 && r.height > 0 && (i.value || '').trim() === '';
                    });
                    // Strategy 1: placeholder on empty input
                    const byPh = emptyInputs.find(i => {
                        const ph = (i.placeholder||'').toLowerCase();
                        return ph.includes('chassis') || ph.includes('chasis');
                    });
                    if (byPh) return byPh;
                    // Strategy 2: label proximity — only empty inputs
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if ((t.includes('chassis') || t.includes('chasis')) && t.length < 40) {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0) continue;
                            let best = null, bestDist = Infinity;
                            for (const inp of emptyInputs) {
                                const ir = inp.getBoundingClientRect();
                                const dist = Math.abs(ir.left - r.left) + Math.abs(ir.top - r.bottom);
                                if (dist < bestDist) { bestDist = dist; best = inp; }
                            }
                            if (best && bestDist < 180) return best;
                        }
                    }
                    return null;
                })()"""
                await mouse_fill(chassis_js, chassis_val_11, "Chasis Number")
            await page.wait_for_timeout(300)

            # ── 11b2: Manufacturer (re-fill on Proposal page) ────────
            mfr_val_11 = D.get("manufacturer", "").strip()
            if mfr_val_11:
                log("info", f"  Re-filling Manufacturer on Proposal page: {mfr_val_11}")
                mfr_js = """(() => {
                    // By placeholder
                    const byPh = Array.from(document.querySelectorAll('input')).find(i =>
                        (i.placeholder||'').toLowerCase().includes('manufacturer'));
                    if (byPh) return byPh;
                    // By label proximity
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if (t === 'manufacturer' || t === 'manufacturer *') {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0) continue;
                            let best = null, bestDist = Infinity;
                            for (const inp of document.querySelectorAll('input')) {
                                const ir = inp.getBoundingClientRect();
                                if (ir.width === 0 || ir.height === 0) continue;
                                const dist = Math.abs(ir.left - r.left) + Math.abs(ir.top - r.bottom);
                                if (dist < bestDist) { bestDist = dist; best = inp; }
                            }
                            if (best && bestDist < 180) return best;
                        }
                    }
                    return null;
                })()"""
                await mouse_fill(mfr_js, mfr_val_11, "Manufacturer (Proposal page)")
            await page.wait_for_timeout(300)

            # ── 11b3: Model (re-fill on Proposal page) ───────────────
            model_val_11 = D.get("model", "").strip()
            if model_val_11:
                log("info", f"  Re-filling Model on Proposal page: {model_val_11}")
                model_js = """(() => {
                    // By placeholder
                    const byPh = Array.from(document.querySelectorAll('input')).find(i =>
                        (i.placeholder||'').toLowerCase().includes('model'));
                    if (byPh) return byPh;
                    // By label proximity
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if (t === 'model' || t === 'model *') {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0) continue;
                            let best = null, bestDist = Infinity;
                            for (const inp of document.querySelectorAll('input')) {
                                const ir = inp.getBoundingClientRect();
                                if (ir.width === 0 || ir.height === 0) continue;
                                const dist = Math.abs(ir.left - r.left) + Math.abs(ir.top - r.bottom);
                                if (dist < bestDist) { bestDist = dist; best = inp; }
                            }
                            if (best && bestDist < 180) return best;
                        }
                    }
                    return null;
                })()"""
                await mouse_fill(model_js, model_val_11, "Model (Proposal page)")
            await page.wait_for_timeout(300)

            # ── 11b4: Watt (after Manufacturer and Model) ────────────────
            watt_val_11 = D.get("watt", "").strip()
            if watt_val_11:
                watt_js = """(() => {
                    // Strategy 1: placeholder contains 'watt'
                    const byPh = Array.from(document.querySelectorAll('input')).find(i =>
                        (i.placeholder||'').toLowerCase().includes('watt'));
                    if (byPh) return byPh;
                    // Strategy 2: label text proximity
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                        if ((t === 'watt' || t === 'watt *' || t === 'wattage' || t === 'wattage *') && t.length < 20) {
                            const r = el.getBoundingClientRect();
                            if (r.width === 0) continue;
                            let best = null, bestDist = Infinity;
                            for (const inp of document.querySelectorAll('input')) {
                                const ir = inp.getBoundingClientRect();
                                if (ir.width === 0 || ir.height === 0) continue;
                                const dist = Math.abs(ir.left - r.left) + Math.abs(ir.top - r.bottom);
                                if (dist < bestDist) { bestDist = dist; best = inp; }
                            }
                            if (best && bestDist < 180) return best;
                        }
                    }
                    return null;
                })()"""
                await mouse_fill(watt_js, watt_val_11, "Watt")
            await page.wait_for_timeout(300)

            # Mark engine/chassis cells green in Excel (Req 2 — prevent duplicate entry)
            mark_engine_chassis_green(D)

            # ── 11c: Check if KYC already done (loop case) ──────────
            pin_val = D.get("pin_code","").strip()   # Applicant Details pincode — always available
            gst_val = D.get("gst_number","").strip()       # defined here so always available
            kyc_already_done = await page.evaluate("""() =>
                document.body.innerText.toLowerCase().includes('your kyc is done')
            """)

            if kyc_already_done:
                log("ok", "  -> KYC already done — skipping KYC popup steps")
                steps.append("kyc")
                log("ok", "══ SECTION DONE: KYC ══")
            else:
                # ── 11c: Complete KYC radio ───────────────────────────────
                log("info", "Clicking Complete KYC radio")
                await page.evaluate("""() => {
                    for (const r of document.querySelectorAll('input[type="radio"]')) {
                        const lbl=r.closest('label')||document.querySelector('label[for="'+r.id+'"]')||r.parentElement;
                        if ((lbl?.innerText||lbl?.textContent||'').toLowerCase().includes('complete kyc'))
                            { r.scrollIntoView({block:'center'}); return; }
                    }
                }""")
                await page.wait_for_timeout(300)
                kyc_c = await page.evaluate("""() => {
                    for (const r of document.querySelectorAll('input[type="radio"]')) {
                        const lbl=r.closest('label')||document.querySelector('label[for="'+r.id+'"]')||r.parentElement;
                        if ((lbl?.innerText||lbl?.textContent||'').toLowerCase().includes('complete kyc')) {
                            const rect=r.getBoundingClientRect();
                            if (rect.width>0) return {x:rect.left+rect.width/2,y:rect.top+rect.height/2};
                        }
                    }
                    for (const el of document.querySelectorAll('*')) {
                        if ((el.innerText||el.textContent||'').trim().toLowerCase()==='complete kyc') {
                            const rect=el.getBoundingClientRect();
                            if (rect.width>0) return {x:rect.left+rect.width/2,y:rect.top+rect.height/2};
                        }
                    }
                    return null;
                }""")
                if kyc_c:
                    await page.mouse.click(kyc_c['x'], kyc_c['y'])
                    log("ok", "  -> Complete KYC radio clicked")
                log("info", "Waiting for KYC popup...")
                for _ in range(20):
                    if await page.evaluate("""() =>
                        document.body.innerText.toLowerCase().includes('select entity type')||
                        document.body.innerText.toLowerCase().includes('customer onboarding')
                    """):
                        log("ok", "  -> KYC popup opened"); break
                    await page.wait_for_timeout(500)
                await page.wait_for_timeout(800)

                # ── 11d: Entity Type = Corporate ─────────────────────────
                log("info", "Selecting Entity Type: Corporate")
                for sel_el in await page.query_selector_all('select'):
                    try:
                        near = await page.evaluate("""(s)=>{ let n=s.parentElement; for(let i=0;i<8;i++){if(!n)break;if((n.innerText||'').toLowerCase().includes('entity type'))return true;n=n.parentElement;} return false;}""", sel_el)
                        if not near: continue
                        await sel_el.scroll_into_view_if_needed()
                        await sel_el.select_option(label="Corporate")
                        await page.wait_for_timeout(1200)
                        log("ok", "  -> Entity Type: Corporate"); break
                    except Exception as e:
                        log("warn", f"  Entity Type error: {e}")

                # ── 11e: Click GST document-type button ───────────────────
                # The portal renders document-type choices (PAN/CKYC/CIN/GST) as
                # styled button-like elements, NOT standard radio <input>s.
                # Strategy: find the element whose ONLY visible text is exactly "GST"
                # and whose bounding box is small (it's a button badge, not a container).
                # Retry up to 10 times in case the popup is still animating in.
                log("info", "Clicking GST document type button")
                gst_clicked = False
                for _gst_try in range(10):
                    gst_coords = await page.evaluate("""() => {
                        // Walk every element; pick the one whose trimmed innerText === 'GST'
                        // with the smallest bounding area (the button, not a container).
                        let best = null, bestArea = Infinity;
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText || el.textContent || '').trim();
                            if (t !== 'GST') continue;
                            const r = el.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) continue;
                            const area = r.width * r.height;
                            if (area < bestArea) { bestArea = area; best = el; }
                        }
                        if (!best) return null;
                        best.scrollIntoView({block: 'center'});
                        const r = best.getBoundingClientRect();
                        return {x: r.left + r.width / 2, y: r.top + r.height / 2};
                    }""")
                    if gst_coords:
                        await page.mouse.click(gst_coords['x'], gst_coords['y'])
                        log("ok", f"  -> GST button clicked at ({gst_coords['x']:.0f},{gst_coords['y']:.0f})")
                        gst_clicked = True
                        break
                    await page.wait_for_timeout(400)
                if not gst_clicked:
                    log("warn", "  GST button not found")

                # Wait until the label below the doc-type buttons changes to
                # something GST-related ("GSTIN Number", "GST", "GSTIN") —
                # meaning the portal has re-rendered the input fields for GST.
                log("info", "  Waiting for GSTIN field to appear...")
                for _ in range(25):
                    gstin_label_visible = await page.evaluate("""() => {
                        const body = (document.body.innerText || '').toLowerCase();
                        return body.includes('gstin') || body.includes('gst number');
                    }""")
                    if gstin_label_visible:
                        log("ok", "  -> GSTIN field visible")
                        break
                    await page.wait_for_timeout(400)
                await page.wait_for_timeout(500)

                # ── 11f: Fill GSTIN ───────────────────────────────────────
                # ── 11g: Fill PIN Code ────────────────────────────────────
                # After switching to GST tab the popup shows:
                #   Field label "GSTIN Number" → text input
                #   Field label "Pin Code"     → text input
                # Fill by finding the input closest to each label.

                async def fill_kyc_field_by_label(label_texts, value, field_name):
                    """Find visible input nearest to a label containing any of label_texts, fill it."""
                    coords = await page.evaluate("""(labels) => {
                        const lowerLabels = labels.map(l => l.toLowerCase());
                        // Find the label element
                        let labelEl = null;
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText || el.textContent || '').trim().toLowerCase();
                            if (lowerLabels.some(l => t === l || t.startsWith(l)) && t.length < 40) {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0 && r.height > 0) { labelEl = el; break; }
                            }
                        }
                        if (!labelEl) return null;
                        const lr = labelEl.getBoundingClientRect();
                        // Find the closest visible, editable input below/near this label
                        let best = null, bestDist = Infinity;
                        for (const inp of document.querySelectorAll('input')) {
                            if (inp.readOnly || inp.disabled) continue;
                            const t = (inp.type || '').toLowerCase();
                            if (['checkbox','radio','hidden','submit','button'].includes(t)) continue;
                            const r = inp.getBoundingClientRect();
                            if (r.width === 0 || r.height === 0) continue;
                            // Prefer inputs below the label (positive dy), penalise ones above
                            const dy = r.top - lr.bottom;
                            const dx = Math.abs((r.left + r.width/2) - (lr.left + lr.width/2));
                            const dist = dx + Math.max(dy, 0) * 0.5 + Math.max(-dy, 0) * 3;
                            if (dist < bestDist) { bestDist = dist; best = inp; }
                        }
                        if (!best || bestDist > 300) return null;
                        best.scrollIntoView({block: 'center'});
                        const r = best.getBoundingClientRect();
                        return {x: r.left + r.width / 2, y: r.top + r.height / 2};
                    }""", label_texts)
                    if not coords:
                        log("warn", f"  {field_name} input not found by label")
                        return False
                    await page.mouse.click(coords['x'], coords['y'])
                    await page.wait_for_timeout(150)
                    await page.keyboard.press("Control+a")
                    await page.keyboard.press("Delete")
                    await page.wait_for_timeout(80)
                    await page.keyboard.type(str(value), delay=60)
                    await page.wait_for_timeout(300)
                    # Commit via Angular native setter
                    await page.evaluate("""({cx, cy, val}) => {
                        const el = document.elementFromPoint(cx, cy);
                        if (!el) return;
                        const inp = el.tagName === 'INPUT' ? el : el.closest('input');
                        if (!inp) return;
                        const setter = Object.getOwnPropertyDescriptor(
                            window.HTMLInputElement.prototype, 'value').set;
                        setter.call(inp, val);
                        inp.dispatchEvent(new Event('input',  {bubbles: true}));
                        inp.dispatchEvent(new Event('change', {bubbles: true}));
                        inp.dispatchEvent(new Event('blur',   {bubbles: true}));
                    }""", {"cx": coords['x'], "cy": coords['y'], "val": str(value)})
                    log("ok", f"  -> {field_name}: {value}")
                    return True

                gst_val = D.get("gst_number", "").strip()
                kyc_pin_val = D.get("kyc_pincode", D.get("pin_code", "")).strip()

                if gst_val:
                    filled = await fill_kyc_field_by_label(
                        ["gstin number", "gstin", "gst number", "gst no"],
                        gst_val, "GSTIN (KYC popup)")
                    if not filled:
                        # Positional fallback: first visible text input in popup
                        await mouse_fill(
                            "Array.from(document.querySelectorAll('input')).filter(i=>{ const r=i.getBoundingClientRect(); return r.width>0&&r.height>0&&!i.readOnly&&!i.disabled&&!['checkbox','radio','hidden','submit','button'].includes((i.type||'').toLowerCase()); })[0]",
                            gst_val, "GSTIN (KYC popup) [positional]")

                if kyc_pin_val:
                    filled = await fill_kyc_field_by_label(
                        ["pin code", "pincode", "pin"],
                        kyc_pin_val, "PIN Code (KYC)")
                    if not filled:
                        # Positional fallback: second visible text input in popup
                        await mouse_fill(
                            "Array.from(document.querySelectorAll('input')).filter(i=>{ const r=i.getBoundingClientRect(); return r.width>0&&r.height>0&&!i.readOnly&&!i.disabled&&!['checkbox','radio','hidden','submit','button'].includes((i.type||'').toLowerCase()); })[1]",
                            kyc_pin_val, "PIN Code (KYC) [positional]")

                # ── 11h: Consent checkbox ─────────────────────────────────
                # Ensure the consent checkbox is ticked. It may already be pre-checked.
                log("info", "Ensuring consent checkbox is ticked")
                await page.wait_for_timeout(300)
                cb = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        let node = el.parentElement;
                        for (let i = 0; i < 8; i++) {
                            if (!node) break;
                            const t = (node.innerText || node.textContent || '').toLowerCase();
                            if (t.includes('consent') || t.includes('hereby')) {
                                el.scrollIntoView({block: 'center'});
                                const r = el.getBoundingClientRect();
                                return r.width > 0
                                    ? {x: r.left+r.width/2, y: r.top+r.height/2, checked: el.checked}
                                    : null;
                            }
                            node = node.parentElement;
                        }
                    }
                    return null;
                }""")
                if cb:
                    if not cb.get('checked', False):
                        await page.mouse.click(cb['x'], cb['y'])
                        await page.wait_for_timeout(400)
                        log("ok", f"  -> Consent checkbox ticked at ({cb['x']:.0f},{cb['y']:.0f})")
                    else:
                        log("ok", f"  -> Consent checkbox already ticked at ({cb['x']:.0f},{cb['y']:.0f})")
                else:
                    log("warn", "  Consent checkbox not found")

                await page.wait_for_timeout(300)
                await click_last_button("SUBMIT", "SUBMIT")
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)
                steps.append("kyc")
                log("ok", "══ SECTION DONE: KYC ══")

            # ── STEP 12: Declaration → PROCEED ────────────────────────
            # Skip PROCEED if KYC already done — Applicant Details already visible
            if kyc_already_done:
                log("step", "STEP 12 — Declaration PROCEED (skipped — KYC already done)")
                steps.append("proceed")
            else:
                log("step", "STEP 12 — Declaration PROCEED")
                log("info", "Waiting for page to fully load before PROCEED...")
                try:
                    await page.wait_for_load_state("networkidle", timeout=20000)
                except Exception:
                    pass
                decl_cb = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        if (el.checked) continue;
                        let node=el.parentElement;
                        for (let i=0;i<6;i++) {
                            if (!node) break;
                            const t=(node.innerText||node.textContent||'').toLowerCase();
                            if (t.includes('hereby declare')||t.includes('premium')||t.includes('declaration')) {
                                el.scrollIntoView({block:'center'});
                                const r=el.getBoundingClientRect();
                                return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2}:null;
                            }
                            node=node.parentElement;
                        }
                    }
                    return null;
                }""")
                if decl_cb:
                    await page.mouse.click(decl_cb['x'], decl_cb['y'])
                    await page.wait_for_timeout(400)
                    log("ok", "  -> Declaration checkbox ticked")
                proceed_clicked = False
                for _proc in range(25):
                    try:
                        for txt in ["PROCEED", "Proceed", "proceed"]:
                            loc = page.get_by_role("button", name=txt)
                            if await loc.count() > 0:
                                await loc.last.scroll_into_view_if_needed()
                                await page.wait_for_timeout(300)
                                await loc.last.click(force=True)
                                log("ok", f"  -> PROCEED clicked (locator)")
                                proceed_clicked = True
                                break
                        if proceed_clicked:
                            break
                    except Exception:
                        pass
                    coords = await page.evaluate("""() => {
                        const all = Array.from(document.querySelectorAll('button, input[type="submit"]'))
                            .filter(el => (el.innerText||el.value||'').trim().toLowerCase().includes('proceed'));
                        if (!all.length) return null;
                        const el = all[all.length - 1];
                        el.scrollIntoView({block:'center'});
                        const r = el.getBoundingClientRect();
                        return r.width > 0 ? {x: r.left+r.width/2, y: r.top+r.height/2} : null;
                    }""")
                    if coords:
                        await page.wait_for_timeout(300)
                        await page.mouse.click(coords['x'], coords['y'])
                        log("ok", f"  -> PROCEED clicked at ({coords['x']:.0f},{coords['y']:.0f})")
                        proceed_clicked = True
                        break
                    log("info", f"  PROCEED not found yet (attempt {_proc+1}/25), waiting...")
                    await page.wait_for_timeout(2000)
                if not proceed_clicked:
                    log("warn", "  PROCEED not found after all retries")
                try:
                    await page.wait_for_load_state("networkidle", timeout=20000)
                except Exception:
                    pass
                await page.wait_for_timeout(2000)
                steps.append("proceed")

            # ── STEP 12.5: Applicant Details Form ─────────────────────
            log("step", "STEP 12.5 — Applicant Details → Create Proposal")
            for _ in range(20):
                if await page.evaluate("""() => document.body.innerText.includes('Applicant Details')||document.body.innerText.includes('Customer Details')||document.body.innerText.includes('Create Proposal')"""):
                    log("ok", "  -> Applicant Details loaded"); break
                await page.wait_for_timeout(500)
            await page.wait_for_timeout(600)

            # ── DIAGNOSTIC: dump all visible inputs in Customer Details ──
            # This tells us the real attributes (placeholder/formControlName/id/name)
            # so we know exactly how to target each field.
            input_map = await page.evaluate("""() => {
                // Find Customer Details section
                let section = null;
                for (const el of document.querySelectorAll('*')) {
                    const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                    if ((t === 'customer details' || t === 'kyc details') &&
                         el.querySelectorAll('input').length >= 2) {
                        section = el; break;
                    }
                }
                const root = section || document;
                return Array.from(root.querySelectorAll('input,textarea,mat-select,select'))
                    .filter(el => {
                        const r = el.getBoundingClientRect();
                        return r.width > 0 && r.height > 0;
                    })
                    .map((el, idx) => ({
                        idx,
                        tag:         el.tagName,
                        type:        el.type || '',
                        placeholder: el.placeholder || '',
                        fc:          el.getAttribute('formcontrolname') || '',
                        id:          el.id || '',
                        name:        el.name || '',
                        ariaLabel:   el.getAttribute('aria-label') || '',
                        value:       el.value || '',
                        x:           Math.round(el.getBoundingClientRect().left + el.getBoundingClientRect().width/2),
                        y:           Math.round(el.getBoundingClientRect().top  + el.getBoundingClientRect().height/2),
                    }));
            }""")
            log("info", f"  Customer Details inputs found: {len(input_map)}")
            for im in input_map:
                log("info", f"  [{im['idx']}] {im['tag']} type={im['type']!r:8} ph={im['placeholder']!r:20} fc={im['fc']!r:20} id={im['id']!r:20} name={im['name']!r:15} aria={im['ariaLabel']!r:20} val={im['value']!r}")

            # ── Helper: find input by any known attribute, then by position ──
            async def find_and_fill(field_name, value, *,
                                    ph_includes=(), ph_exact=(),
                                    fc_includes=(), id_includes=(),
                                    aria_includes=(), name_includes=(),
                                    position=None):
                """
                Find the right input using every available attribute, then fill it.
                Falls back to positional index within Customer Details as last resort.
                Always uses: click → Ctrl+A → Delete → type → JS native setter.
                """
                js = """([phInc, phExact, fcInc, idInc, ariaInc, nameInc, pos]) => {
                    let section = null;
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                        if ((t === 'customer details' || t === 'kyc details') &&
                             el.querySelectorAll('input').length >= 2) {
                            section = el; break;
                        }
                    }
                    const root = section || document;
                    const visibleInputs = Array.from(
                        root.querySelectorAll('input,textarea')
                    ).filter(el => {
                        if (el.readOnly || el.disabled) return false;
                        const t = (el.type||'').toLowerCase();
                        if (['checkbox','radio','hidden','submit','button'].includes(t)) return false;
                        const r = el.getBoundingClientRect();
                        return r.width > 0 && r.height > 0;
                    });

                    // Try attribute-based match first
                    for (const inp of visibleInputs) {
                        const ph   = (inp.placeholder||'').toLowerCase();
                        const fc   = (inp.getAttribute('formcontrolname')||'').toLowerCase();
                        const id   = (inp.id||'').toLowerCase();
                        const aria = (inp.getAttribute('aria-label')||'').toLowerCase();
                        const nm   = (inp.name||'').toLowerCase();
                        const matched =
                            phExact.some(v => ph === v.toLowerCase()) ||
                            phInc.some(v   => ph.includes(v.toLowerCase())) ||
                            fcInc.some(v   => fc.includes(v.toLowerCase())) ||
                            idInc.some(v   => id.includes(v.toLowerCase())) ||
                            ariaInc.some(v => aria.includes(v.toLowerCase())) ||
                            nameInc.some(v => nm.includes(v.toLowerCase()));
                        if (matched) {
                            inp.scrollIntoView({block:'center'});
                            const r = inp.getBoundingClientRect();
                            return {x: r.left+r.width/2, y: r.top+r.height/2, how:'attr'};
                        }
                    }
                    // Positional fallback
                    if (pos !== null && pos < visibleInputs.length) {
                        const inp = visibleInputs[pos];
                        inp.scrollIntoView({block:'center'});
                        const r = inp.getBoundingClientRect();
                        return {x: r.left+r.width/2, y: r.top+r.height/2, how:'pos:'+pos};
                    }
                    return null;
                }"""
                coords = await page.evaluate(js, [
                    list(ph_includes), list(ph_exact),
                    list(fc_includes), list(id_includes),
                    list(aria_includes), list(name_includes),
                    position
                ])
                if not coords:
                    log("warn", f"  {field_name} not found")
                    return False
                log("info", f"  Filling {field_name} via {coords['how']} at ({coords['x']},{coords['y']})")
                await page.mouse.click(coords['x'], coords['y'])
                await page.wait_for_timeout(150)
                await page.keyboard.press("Control+A")
                await page.wait_for_timeout(60)
                await page.keyboard.press("Delete")
                await page.wait_for_timeout(60)
                await page.keyboard.type(str(value), delay=40)
                await page.wait_for_timeout(250)
                # Angular native setter
                await page.evaluate("""({cx, cy, val}) => {
                    const el = document.elementFromPoint(cx, cy);
                    if (!el) return;
                    const inp = (el.tagName==='INPUT'||el.tagName==='TEXTAREA') ? el
                        : (el.closest('input') || el.closest('textarea'));
                    if (!inp) return;
                    const proto = inp.tagName==='TEXTAREA'
                        ? window.HTMLTextAreaElement.prototype
                        : window.HTMLInputElement.prototype;
                    const setter = Object.getOwnPropertyDescriptor(proto, 'value').set;
                    setter.call(inp, val);
                    inp.dispatchEvent(new Event('input',  {bubbles:true}));
                    inp.dispatchEvent(new Event('change', {bubbles:true}));
                    inp.dispatchEvent(new Event('blur',   {bubbles:true}));
                }""", {"cx": coords['x'], "cy": coords['y'], "val": str(value)})
                log("ok", f"  -> {field_name}: {value}")
                return True

            # ── Title dropdown ─────────────────────────────────────────
            # Diagnostic: MAT-SELECT fc='title' id='mat-select-6'
            # Must click the mat-select trigger to open the overlay, then click M/s mat-option.
            log("info", "Selecting Title: M/s")
            title_done = False

            # Find the mat-select trigger by formcontrolname='title'
            title_trigger = await page.evaluate("""() => {
                // Direct: mat-select[formcontrolname='title']
                const ms = document.querySelector('mat-select[formcontrolname="title"]');
                if (ms) {
                    ms.scrollIntoView({block:'center'});
                    const r = ms.getBoundingClientRect();
                    return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                }
                // Fallback: mat-form-field whose label = 'title'
                for (const ff of document.querySelectorAll('mat-form-field')) {
                    const lbl = ff.querySelector('mat-label,label');
                    if (!lbl) continue;
                    const t = (lbl.innerText||lbl.textContent||'').trim().toLowerCase();
                    if (t === 'title' || t === 'title *') {
                        const ms2 = ff.querySelector('mat-select');
                        if (ms2) {
                            ms2.scrollIntoView({block:'center'});
                            const r = ms2.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                    }
                }
                return null;
            }""")
            if title_trigger:
                await page.mouse.click(title_trigger['x'], title_trigger['y'])
                await page.wait_for_timeout(500)
                # Wait for mat-option overlay to render
                for _tw in range(10):
                    has_opts = await page.evaluate("() => document.querySelectorAll('mat-option').length > 0")
                    if has_opts: break
                    await page.wait_for_timeout(150)
                # Click the M/s option
                clicked = await page.evaluate("""() => {
                    for (const opt of document.querySelectorAll('mat-option')) {
                        const t = (opt.innerText||opt.textContent||'').trim();
                        if (t === 'M/s' || t.toLowerCase() === 'm/s') {
                            opt.click(); return t;
                        }
                    }
                    // If M/s not listed, click first available option
                    for (const opt of document.querySelectorAll('mat-option')) {
                        const t = (opt.innerText||opt.textContent||'').trim();
                        const r = opt.getBoundingClientRect();
                        if (t && r.width>0 && r.height>0) { opt.click(); return '(first) ' + t; }
                    }
                    return null;
                }""")
                if clicked:
                    log("ok", f"  -> Title: {clicked}")
                    title_done = True
                else:
                    await page.keyboard.press("Escape")
                    log("warn", "  Title overlay opened but no mat-option found")
            else:
                log("warn", "  Title mat-select not found — skipping")
            await page.wait_for_timeout(400)

            # ── Insured Name ──────────────────────────────────────────
            # Motor form: position 0 in Customer Details (after Title select)
            # fc candidates: 'insuredName','insured','name','applicantName'
            insured_name = D.get("insured_name","").strip()
            if insured_name:
                await find_and_fill("Insured Name", insured_name,
                    ph_includes=("insured name","insured","applicant name"),
                    fc_includes=("insuredname","insured","applicantname","name"),
                    id_includes=("insured","applicant"),
                    aria_includes=("insured name","insured"),
                    position=0)
            await page.wait_for_timeout(300)

            # ── DOB ───────────────────────────────────────────────────
            # Find by formcontrolname='dob' first — the only reliable attribute on this form.
            # Type DD/MM/YYYY with slashes so no digits bleed into the next field.
            # Pattern: click → Escape → click → Escape → Ctrl+A → Delete → type
            # Commit via Angular native setter + blur — do NOT press Tab.
            dob_str = D.get("dob","").strip()
            if dob_str:
                log("info", f"Filling DOB: {dob_str}")
                try:
                    parts = dob_str.replace("-","/").split("/")
                    dob_formatted = f"{int(parts[0]):02d}/{int(parts[1]):02d}/{int(parts[2])}"
                    log("info", f"  Typing: {dob_formatted}")

                    dob_coords = await page.evaluate("""() => {
                        // Strategy 1: exact formcontrolname='dob'
                        let dobInp = document.querySelector("input[formcontrolname='dob']");
                        // Strategy 2: broader attribute match
                        if (!dobInp) {
                            const inputs = Array.from(document.querySelectorAll('input')).filter(el => {
                                if (el.readOnly||el.disabled) return false;
                                const t=(el.type||'').toLowerCase();
                                if (['checkbox','radio','hidden','submit','button'].includes(t)) return false;
                                const r=el.getBoundingClientRect();
                                return r.width>0&&r.height>0;
                            });
                            dobInp = inputs.find(i => {
                                const ph   = (i.placeholder||'').toLowerCase();
                                const fc   = (i.getAttribute('formcontrolname')||'').toLowerCase();
                                const id   = (i.id||'').toLowerCase();
                                const aria = (i.getAttribute('aria-label')||'').toLowerCase();
                                return ph.includes('dob')||ph.includes('date')||ph.includes('birth')
                                    || fc.includes('dob')||fc.includes('date')||fc.includes('birth')
                                    || id.includes('dob')||id.includes('date')||id.includes('birth')
                                    || aria.includes('dob')||aria.includes('date')||aria.includes('birth');
                            }) || inputs[1];  // positional fallback: index 1
                        }
                        if (!dobInp) return null;
                        dobInp.scrollIntoView({block:'center'});
                        const r = dobInp.getBoundingClientRect();
                        return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                    }""")
                    if not dob_coords:
                        raise ValueError("DOB input not found")

                    # STAR-file Escape pattern: click → Escape → click → Escape → clear → type
                    await page.mouse.click(dob_coords['x'], dob_coords['y'])
                    await page.wait_for_timeout(300)
                    await page.keyboard.press("Escape")
                    await page.wait_for_timeout(200)
                    await page.mouse.click(dob_coords['x'], dob_coords['y'])
                    await page.wait_for_timeout(200)
                    await page.keyboard.press("Escape")
                    await page.wait_for_timeout(200)
                    await page.keyboard.press("Control+A")
                    await page.wait_for_timeout(60)
                    await page.keyboard.press("Delete")
                    await page.wait_for_timeout(60)
                    # Type with slashes DD/MM/YYYY — prevents digit bleeding to next field
                    for ch in dob_formatted:
                        await page.keyboard.type(ch, delay=80)
                        await page.wait_for_timeout(50)
                    await page.wait_for_timeout(300)
                    # Commit via Angular native setter + blur — do NOT press Tab
                    await page.evaluate("""({cx, cy, val}) => {
                        const el = document.elementFromPoint(cx,cy);
                        const inp = el&&(el.tagName==='INPUT'?el:el.closest('input'));
                        if (!inp) return;
                        const setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
                        setter.call(inp, val);
                        inp.dispatchEvent(new Event('input',  {bubbles:true}));
                        inp.dispatchEvent(new Event('change', {bubbles:true}));
                        inp.dispatchEvent(new Event('blur',   {bubbles:true}));
                    }""", {"cx": dob_coords['x'], "cy": dob_coords['y'], "val": dob_formatted})
                    await page.wait_for_timeout(200)
                    actual_dob = await page.evaluate("""({cx,cy}) => {
                        const el=document.elementFromPoint(cx,cy);
                        const inp=el&&(el.tagName==='INPUT'?el:el.closest('input'));
                        return inp?inp.value:null;
                    }""", {"cx": dob_coords['x'], "cy": dob_coords['y']})
                    log("ok" if actual_dob else "warn",
                        f"  -> DOB set: {actual_dob}" if actual_dob else "  DOB field empty after typing")
                except Exception as e:
                    log("warn", f"  DOB error: {e}")

            # ── Address Line ──────────────────────────────────────────
            # Motor form: position 2
            addr_val = D.get("address_line","").strip()
            if addr_val:
                await find_and_fill("Address Line", addr_val,
                    ph_includes=("address",),
                    fc_includes=("address","addressline","addr"),
                    id_includes=("address","addr"),
                    aria_includes=("address",),
                    name_includes=("address",),
                    position=2)
            await page.wait_for_timeout(300)

            # ── Mobile ────────────────────────────────────────────────
            # Motor form: position 3
            mob_val = D.get("mobile_number","").strip()
            if mob_val:
                await find_and_fill("Mobile", mob_val,
                    ph_includes=("mobile","phone"),
                    fc_includes=("mobile","phone","mobileNo","contactNo","contact"),
                    id_includes=("mobile","phone"),
                    aria_includes=("mobile","phone"),
                    name_includes=("mobile","phone"),
                    position=3)
            await page.wait_for_timeout(300)

            # ── Email ─────────────────────────────────────────────────
            # Motor form: position 4
            email_val = D.get("email","").strip()
            if email_val:
                await find_and_fill("Email", email_val,
                    ph_includes=("email",),
                    fc_includes=("email","emailId","emailAddress"),
                    id_includes=("email",),
                    aria_includes=("email",),
                    name_includes=("email",),
                    position=4)
            await page.wait_for_timeout(300)

            # ── Pincode — retry until State field populates ──────────
            # Motor form: position 5
            pin_val = D.get("pin_code","").strip()
            if pin_val:
                for _pin_try in range(10):
                    filled = await find_and_fill("Pincode", pin_val,
                        ph_includes=("pincode","pin code","pin"),
                        fc_includes=("pincode","pin","zipcode","zip","postalcode"),
                        id_includes=("pin","zip"),
                        aria_includes=("pincode","pin code"),
                        name_includes=("pin","zip"),
                        position=5)
                    await page.wait_for_timeout(1000)
                    state_populated = await page.evaluate("""() => {
                        for (const inp of document.querySelectorAll('input[readonly],input[disabled]')) {
                            if ((inp.value||'').trim().length > 1) return true;
                        }
                        const body = (document.body.innerText||'').toLowerCase();
                        return ['rajasthan','maharashtra','gujarat','karnataka','delhi',
                                'west bengal','uttar pradesh','tamil nadu','punjab',
                                'haryana','telangana','kerala'].some(s => body.includes(s));
                    }""")
                    if state_populated:
                        log("ok", f"  State populated after pincode (attempt {_pin_try+1})")
                        break
                    log("warn", f"  State not populated — retyping pincode (attempt {_pin_try+1}/10)...")
                    await page.wait_for_timeout(500)

            # ── City/District ─────────────────────────────────────────
            # Strategy 1: mat-select[formcontrolname="city"] — most precise selector.
            # After clicking, loop up to 15×200ms (3s) for mat-option elements to appear.
            # Strategy 2 fallback: label-proximity. Strategy 3 fallback: native <select>.
            log("info", "Selecting City/District")
            try:
                await page.wait_for_load_state("networkidle", timeout=10000)
            except Exception:
                pass
            city_done = False
            # Wait up to 12.5s for the mat-select[formcontrolname="city"] to appear
            for _city_wait in range(25):
                city_ready = await page.evaluate("""() => {
                    if (document.querySelector('mat-select[formcontrolname="city"]')) return true;
                    for (const sel of document.querySelectorAll('select')) {
                        let n = sel.parentElement;
                        for (let i=0;i<6;i++) {
                            if (!n) break;
                            const t=(n.innerText||n.textContent||'').toLowerCase();
                            if (t.includes('city')||t.includes('district')) {
                                const real = Array.from(sel.options).filter(
                                    o=>o.value&&o.text.trim()&&
                                    !['','enter city / district','select'].includes(o.text.trim().toLowerCase())
                                );
                                return real.length > 0;
                            }
                            n = n.parentElement;
                        }
                    }
                    for (const ff of document.querySelectorAll('mat-form-field')) {
                        const lbl=ff.querySelector('mat-label,label');
                        if(!lbl) continue;
                        const t=(lbl.innerText||lbl.textContent||'').trim().toLowerCase();
                        if (t.includes('city')||t.includes('district')) {
                            const ms=ff.querySelector('mat-select');
                            if(ms) return true;
                        }
                    }
                    return false;
                }""")
                if city_ready:
                    log("info", f"  City/District ready (attempt {_city_wait+1})")
                    break
                await page.wait_for_timeout(500)
            await page.wait_for_timeout(400)

            # Find the trigger coords — formcontrolname="city" first
            city_trigger = await page.evaluate("""() => {
                const ms = document.querySelector('mat-select[formcontrolname="city"]');
                if (ms) {
                    ms.scrollIntoView({block:'center'});
                    const r = ms.getBoundingClientRect();
                    return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2, via:'fc'} : null;
                }
                // Fallback: label-proximity via mat-form-field
                for (const ff of document.querySelectorAll('mat-form-field')) {
                    const lbl=ff.querySelector('mat-label,label');
                    if(!lbl) continue;
                    const t=(lbl.innerText||lbl.textContent||'').trim().toLowerCase();
                    if (t.includes('city')||t.includes('district')) {
                        const ms2=ff.querySelector('mat-select');
                        if(ms2){ms2.scrollIntoView({block:'center'});const r=ms2.getBoundingClientRect();return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2,via:'lbl'}:null;}
                    }
                }
                for (const lbl of document.querySelectorAll('label,span,div')) {
                    const t=(lbl.innerText||lbl.textContent||'').toLowerCase();
                    if((t.includes('city')||t.includes('district'))&&t.length<30){
                        let node=lbl.parentElement;
                        for(let i=0;i<5;i++){
                            if(!node) break;
                            const trigger=node.querySelector('select,[role="combobox"],[role="listbox"],[class*="dropdown"],[class*="select"],[class*="ng-select"]');
                            if(trigger){trigger.scrollIntoView({block:'center'});const r=trigger.getBoundingClientRect();return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2,via:'prox'}:null;}
                            node=node.parentElement;
                        }
                    }
                }
                for (const el of document.querySelectorAll('*')) {
                    const t=(el.innerText||'').trim();
                    if(t==='Enter City / District'){const r=el.getBoundingClientRect();if(r.width>0&&r.height>0)return{x:r.left+r.width/2,y:r.top+r.height/2,via:'ph'};}
                }
                return null;
            }""")
            if city_trigger:
                await page.mouse.click(city_trigger['x'], city_trigger['y'])
                log("info", f"  -> City dropdown trigger clicked (via={city_trigger.get('via','')})")
                # Loop up to 15×200ms = 3s waiting for mat-option elements to appear in DOM
                city_opt = None
                for _opt_wait in range(15):
                    await page.wait_for_timeout(200)
                    city_opt = await page.evaluate("""() => {
                        for (const el of document.querySelectorAll('mat-option,li,[role="option"],div.option,div.item')) {
                            const t=(el.innerText||el.textContent||'').trim();
                            if(t&&t!=='Enter City / District'&&t.length>1&&t.length<40){
                                const r=el.getBoundingClientRect();
                                if(r.width>0&&r.height>0)return{x:r.left+r.width/2,y:r.top+r.height/2,txt:t};
                            }
                        }
                        return null;
                    }""")
                    if city_opt:
                        log("info", f"  -> City options appeared after {(_opt_wait+1)*200}ms")
                        break
                if city_opt:
                    await page.mouse.click(city_opt['x'], city_opt['y'])
                    await page.wait_for_timeout(400)
                    log("ok", f"  -> City selected: {city_opt['txt']}")
                    city_done = True
                else:
                    log("warn", "  City options did not appear after 3s — trying native select fallback")

            if not city_done:
                for sel_el in await page.query_selector_all('select'):
                    try:
                        near = await page.evaluate("""(s)=>{let n=s.parentElement;for(let i=0;i<5;i++){if(!n)break;const t=(n.innerText||n.textContent||'').toLowerCase();if(t.includes('city')||t.includes('district'))return true;n=n.parentElement;}return false;}""", sel_el)
                        if not near: continue
                        opts = await page.evaluate("(s)=>Array.from(s.options).map(o=>({val:o.value,txt:o.text.trim()}))", sel_el)
                        best = next((o for o in opts if o['val'] and o['txt'] and o['txt'].upper() not in ('','ENTER CITY / DISTRICT','SELECT')), None)
                        if best:
                            await sel_el.select_option(value=best['val'])
                            await page.wait_for_timeout(400)
                            log("ok", f"  -> City (select): {best['txt']}")
                            city_done = True; break
                    except Exception as e:
                        log("warn", f"  City select error: {e}")

            # ── Add GST Details checkbox ──────────────────────────────
            if gst_val:
                log("info", "Ticking Add GST Details checkbox")
                await page.evaluate("window.scrollBy(0, 400)")
                await page.wait_for_timeout(400)
                gst_cb = await page.evaluate("""() => {
                    // Match the checkbox whose OWN label/sibling says "Add GST Details"
                    // — NOT the whole container which includes all three checkbox labels.
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        // Check: <label> wrapping this checkbox
                        const wrap = el.closest('label');
                        if (wrap) {
                            const t = (wrap.innerText||wrap.textContent||'').toLowerCase();
                            if (t.includes('add gst') || (t.includes('gst') && t.includes('detail'))) {
                                el.scrollIntoView({block:'center'});
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                            continue;
                        }
                        // Check: <label for=id> associated with this checkbox
                        if (el.id) {
                            const lbl = document.querySelector('label[for="'+el.id+'"]');
                            if (lbl) {
                                const t = (lbl.innerText||lbl.textContent||'').toLowerCase();
                                if (t.includes('add gst') || (t.includes('gst') && t.includes('detail'))) {
                                    el.scrollIntoView({block:'center'});
                                    const r = el.getBoundingClientRect();
                                    if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                }
                            }
                        }
                        // Check: immediately adjacent sibling text node / span
                        let sib = el.nextElementSibling || el.parentElement;
                        if (sib) {
                            const t = (sib.innerText||sib.textContent||'').toLowerCase().trim();
                            if (t.length < 40 && (t.includes('add gst') || (t.includes('gst') && t.includes('detail')))) {
                                el.scrollIntoView({block:'center'});
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                        }
                    }
                    return null;
                }""")
                if gst_cb:
                    await page.mouse.click(gst_cb['x'], gst_cb['y'])
                    await page.wait_for_timeout(800)
                    log("ok", "  -> Add GST Details checkbox ticked")
                else:
                    log("warn", "  Add GST Details checkbox not found")

                # ── GSTIN field ───────────────────────────────────────
                # Wait up to 3s for the GSTIN input to appear after checkbox tick
                gstin_coords = None
                for _gst_wait in range(15):
                    await page.wait_for_timeout(200)
                    gstin_coords = await page.evaluate("""() => {
                        // 1. formcontrolname contains 'gstin' or 'gst'
                        const byFc = Array.from(document.querySelectorAll('input')).find(i => {
                            const fc = (i.getAttribute('formcontrolname')||'').toLowerCase();
                            return fc.includes('gstin') || fc === 'gstno' || fc === 'gstnumber';
                        });
                        if (byFc && !byFc.readOnly && !byFc.disabled) {
                            byFc.scrollIntoView({block:'center'});
                            const r = byFc.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // 2. placeholder includes gstin
                        const byPh = Array.from(document.querySelectorAll('input')).find(i =>
                            (i.placeholder||'').toLowerCase().includes('gstin') ||
                            (i.placeholder||'').toLowerCase().includes('enter gstin')
                        );
                        if (byPh && !byPh.readOnly && !byPh.disabled) {
                            byPh.scrollIntoView({block:'center'});
                            const r = byPh.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // 3. nearest input to a label that says "GSTIN Number" or "GSTIN"
                        for (const el of document.querySelectorAll('label,mat-label,span,div')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if (t.includes('gstin') || t === 'gst number') {
                                const r = el.getBoundingClientRect();
                                if (r.width === 0) continue;
                                let best = null, bestDist = Infinity;
                                for (const inp of document.querySelectorAll('input')) {
                                    if (inp.readOnly || inp.disabled) continue;
                                    const ir = inp.getBoundingClientRect();
                                    if (ir.width === 0 || ir.height === 0) continue;
                                    const dist = Math.abs(ir.left - r.left) + Math.abs(ir.top - r.bottom);
                                    if (dist < bestDist) { bestDist = dist; best = inp; }
                                }
                                if (best && bestDist < 200) {
                                    best.scrollIntoView({block:'center'});
                                    const br = best.getBoundingClientRect();
                                    return {x:br.left+br.width/2, y:br.top+br.height/2};
                                }
                            }
                        }
                        return null;
                    }""")
                    if gstin_coords:
                        log("info", f"  GSTIN field appeared after {(_gst_wait+1)*200}ms")
                        break
                if gstin_coords:
                    # Fill with confirm loop: up to 3 attempts in case Angular re-renders
                    gstin_confirmed = False
                    for _gst_fill in range(3):
                        await page.mouse.click(gstin_coords['x'], gstin_coords['y'])
                        await page.wait_for_timeout(150)
                        await page.keyboard.press("Control+A")
                        await page.wait_for_timeout(60)
                        await page.keyboard.press("Delete")
                        await page.wait_for_timeout(60)
                        await page.keyboard.type(gst_val, delay=40)
                        await page.wait_for_timeout(200)
                        await page.evaluate("""({cx,cy,val}) => {
                            const el = document.elementFromPoint(cx,cy);
                            const inp = el&&(el.tagName==='INPUT'?el:el.closest('input'));
                            if (!inp) return;
                            const s = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
                            s.call(inp, val);
                            inp.dispatchEvent(new Event('input',  {bubbles:true}));
                            inp.dispatchEvent(new Event('change', {bubbles:true}));
                            inp.dispatchEvent(new Event('blur',   {bubbles:true}));
                        }""", {"cx": gstin_coords['x'], "cy": gstin_coords['y'], "val": gst_val})
                        # Poll up to 10×200ms to confirm field holds the value
                        for _chk in range(10):
                            await page.wait_for_timeout(200)
                            held = await page.evaluate("""({cx,cy,val}) => {
                                const el = document.elementFromPoint(cx,cy);
                                const inp = el&&(el.tagName==='INPUT'?el:el.closest('input'));
                                return inp ? inp.value.trim().toUpperCase() === val.trim().toUpperCase() : false;
                            }""", {"cx": gstin_coords['x'], "cy": gstin_coords['y'], "val": gst_val})
                            if held:
                                gstin_confirmed = True
                                break
                        if gstin_confirmed:
                            log("ok", f"  -> GSTIN filled and confirmed: {gst_val} (attempt {_gst_fill+1})")
                            break
                        log("warn", f"  GSTIN value lost — retrying (attempt {_gst_fill+1}/3)...")
                    if not gstin_confirmed:
                        log("warn", "  GSTIN could not be confirmed after 3 attempts")
                    # Let Angular fully settle before the next section touches the form
                    await page.wait_for_timeout(1200)
                else:
                    log("warn", "  GSTIN input not found — skipping")

            # ── Loan / Lease / Hypothecation Details ─────────────────
            loan_val         = str(D.get("loan_hypothecation","")).strip().lower()
            financier_name   = str(D.get("financier_name",   "")).strip()
            financier_branch = str(D.get("financier_branch", "")).strip()

            if loan_val == "yes":
                log("info", "Ticking Loan / Lease / Hypothecation Details checkbox")
                await page.evaluate("window.scrollBy(0, 300)")
                await page.wait_for_timeout(400)
                loan_cb = await page.evaluate("""() => {
                    // Match the checkbox whose OWN label/sibling mentions loan/hypothecation
                    // — NOT the whole container which includes all three checkbox labels.
                    for (const el of document.querySelectorAll('input[type="checkbox"]')) {
                        const wrap = el.closest('label');
                        if (wrap) {
                            const t = (wrap.innerText||wrap.textContent||'').toLowerCase();
                            if (t.includes('loan') || t.includes('hypothecation') || t.includes('lease')) {
                                el.scrollIntoView({block:'center'});
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                            continue;
                        }
                        if (el.id) {
                            const lbl = document.querySelector('label[for="'+el.id+'"]');
                            if (lbl) {
                                const t = (lbl.innerText||lbl.textContent||'').toLowerCase();
                                if (t.includes('loan') || t.includes('hypothecation') || t.includes('lease')) {
                                    el.scrollIntoView({block:'center'});
                                    const r = el.getBoundingClientRect();
                                    if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                }
                            }
                        }
                        let sib = el.nextElementSibling || el.parentElement;
                        if (sib) {
                            const t = (sib.innerText||sib.textContent||'').toLowerCase().trim();
                            if (t.length < 60 && (t.includes('loan') || t.includes('hypothecation') || t.includes('lease'))) {
                                el.scrollIntoView({block:'center'});
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                        }
                    }
                    return null;
                }""")
                if loan_cb:
                    await page.mouse.click(loan_cb['x'], loan_cb['y'])
                    await page.wait_for_timeout(800)
                    log("ok", "  -> Loan/Hypothecation checkbox ticked")
                else:
                    log("warn", "  Loan/Hypothecation checkbox not found")

                # Wait for the Loan section to expand, then click Loan/Hypothecation radio
                # The portal uses mat-radio-button. Retry up to 8× waiting for it to appear.
                await page.wait_for_timeout(600)
                loan_radio = None
                for _lr in range(8):
                    loan_radio = await page.evaluate("""() => {
                        // Strategy 1: mat-radio-button whose text includes loan/hypothecation
                        for (const rb of document.querySelectorAll('mat-radio-button')) {
                            const t = (rb.innerText||rb.textContent||'').trim().toLowerCase();
                            if (t.includes('loan') || t.includes('hypothecation')) {
                                rb.scrollIntoView({block:'center'});
                                const r = rb.getBoundingClientRect();
                                if (r.width > 0) return {x:r.left+r.width/2, y:r.top+r.height/2};
                            }
                        }
                        // Strategy 2: input[type=radio] near loan/hypothecation label
                        for (const r of document.querySelectorAll('input[type="radio"]')) {
                            const lbl = r.closest('label') ||
                                        document.querySelector('label[for="'+r.id+'"]') ||
                                        r.parentElement;
                            const t = (lbl?.innerText||lbl?.textContent||'').toLowerCase();
                            if (t.includes('loan') || t.includes('hypothecation')) {
                                r.scrollIntoView({block:'center'});
                                const rect = r.getBoundingClientRect();
                                if (rect.width > 0) return {x:rect.left+rect.width/2, y:rect.top+rect.height/2};
                            }
                        }
                        // Strategy 3: any element whose exact text is 'Loan/Hypothecation'
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText||el.textContent||'').trim();
                            if (t === 'Loan/Hypothecation' || t === 'Loan / Hypothecation') {
                                const r = el.getBoundingClientRect();
                                if (r.width > 0) return {x:r.left+r.width/2, y:r.top+r.height/2};
                            }
                        }
                        return null;
                    }""")
                    if loan_radio: break
                    await page.wait_for_timeout(400)
                if loan_radio:
                    await page.mouse.click(loan_radio['x'], loan_radio['y'])
                    await page.wait_for_timeout(600)
                    log("ok", "  -> Loan/Hypothecation radio selected")
                else:
                    log("warn", "  Loan/Hypothecation radio not found")

                # Wait for Financier Name/Branch fields to appear
                await page.wait_for_timeout(500)

                # Fill Financier Name — find by formcontrolname OR label proximity
                if financier_name:
                    fn_coords = await page.evaluate("""() => {
                        // formcontrolname
                        const byFc = Array.from(document.querySelectorAll('input')).find(i => {
                            const fc = (i.getAttribute('formcontrolname')||'').toLowerCase();
                            return fc.includes('financier') && (fc.includes('name') || fc === 'financiername' || fc === 'financier');
                        });
                        if (byFc && !byFc.readOnly && !byFc.disabled) {
                            byFc.scrollIntoView({block:'center'});
                            const r = byFc.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // placeholder
                        const byPh = Array.from(document.querySelectorAll('input')).find(i =>
                            (i.placeholder||'').toLowerCase().includes('financier name') ||
                            (i.placeholder||'').toLowerCase().includes('financier')
                        );
                        if (byPh && !byPh.readOnly && !byPh.disabled) {
                            byPh.scrollIntoView({block:'center'});
                            const r = byPh.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // label proximity
                        for (const el of document.querySelectorAll('label,mat-label,span,div')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if (t === 'financier name' || t === 'financier name *') {
                                const r = el.getBoundingClientRect();
                                if (r.width === 0) continue;
                                let best = null, bd = Infinity;
                                for (const inp of document.querySelectorAll('input')) {
                                    if (inp.readOnly||inp.disabled) continue;
                                    const ir = inp.getBoundingClientRect();
                                    if (!ir.width||!ir.height) continue;
                                    const d = Math.abs(ir.left-r.left)+Math.abs(ir.top-r.bottom);
                                    if (d < bd) { bd=d; best=inp; }
                                }
                                if (best && bd < 200) { best.scrollIntoView({block:'center'}); const br=best.getBoundingClientRect(); return {x:br.left+br.width/2,y:br.top+br.height/2}; }
                            }
                        }
                        return null;
                    }""")
                    if fn_coords:
                        await page.mouse.click(fn_coords['x'], fn_coords['y'])
                        await page.wait_for_timeout(120)
                        await page.keyboard.press("Control+A")
                        await page.keyboard.press("Delete")
                        await page.wait_for_timeout(60)
                        await page.keyboard.type(financier_name, delay=40)
                        await page.wait_for_timeout(200)
                        await page.evaluate("""({cx,cy,val})=>{const el=document.elementFromPoint(cx,cy);const inp=el&&(el.tagName==='INPUT'?el:el.closest('input'));if(!inp)return;const s=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;s.call(inp,val);inp.dispatchEvent(new Event('input',{bubbles:true}));inp.dispatchEvent(new Event('change',{bubbles:true}));inp.dispatchEvent(new Event('blur',{bubbles:true}));}""",
                                            {"cx": fn_coords['x'], "cy": fn_coords['y'], "val": financier_name})
                        log("ok", f"  -> Financier Name: {financier_name}")
                    else:
                        log("warn", "  Financier Name not found")

                # Fill Financier Branch — find by formcontrolname OR label proximity
                if financier_branch:
                    fb_coords = await page.evaluate("""() => {
                        // formcontrolname
                        const byFc = Array.from(document.querySelectorAll('input')).find(i => {
                            const fc = (i.getAttribute('formcontrolname')||'').toLowerCase();
                            return fc.includes('branch');
                        });
                        if (byFc && !byFc.readOnly && !byFc.disabled) {
                            byFc.scrollIntoView({block:'center'});
                            const r = byFc.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // placeholder
                        const byPh = Array.from(document.querySelectorAll('input')).find(i =>
                            (i.placeholder||'').toLowerCase().includes('branch')
                        );
                        if (byPh && !byPh.readOnly && !byPh.disabled) {
                            byPh.scrollIntoView({block:'center'});
                            const r = byPh.getBoundingClientRect();
                            return r.width>0 ? {x:r.left+r.width/2, y:r.top+r.height/2} : null;
                        }
                        // label proximity
                        for (const el of document.querySelectorAll('label,mat-label,span,div')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if (t === 'financier branch' || t === 'financier branch *') {
                                const r = el.getBoundingClientRect();
                                if (r.width === 0) continue;
                                let best = null, bd = Infinity;
                                for (const inp of document.querySelectorAll('input')) {
                                    if (inp.readOnly||inp.disabled) continue;
                                    const ir = inp.getBoundingClientRect();
                                    if (!ir.width||!ir.height) continue;
                                    const d = Math.abs(ir.left-r.left)+Math.abs(ir.top-r.bottom);
                                    if (d < bd) { bd=d; best=inp; }
                                }
                                if (best && bd < 200) { best.scrollIntoView({block:'center'}); const br=best.getBoundingClientRect(); return {x:br.left+br.width/2,y:br.top+br.height/2}; }
                            }
                        }
                        return null;
                    }""")
                    if fb_coords:
                        await page.mouse.click(fb_coords['x'], fb_coords['y'])
                        await page.wait_for_timeout(120)
                        await page.keyboard.press("Control+A")
                        await page.keyboard.press("Delete")
                        await page.wait_for_timeout(60)
                        await page.keyboard.type(financier_branch, delay=40)
                        await page.wait_for_timeout(200)
                        await page.evaluate("""({cx,cy,val})=>{const el=document.elementFromPoint(cx,cy);const inp=el&&(el.tagName==='INPUT'?el:el.closest('input'));if(!inp)return;const s=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;s.call(inp,val);inp.dispatchEvent(new Event('input',{bubbles:true}));inp.dispatchEvent(new Event('change',{bubbles:true}));inp.dispatchEvent(new Event('blur',{bubbles:true}));}""",
                                            {"cx": fb_coords['x'], "cy": fb_coords['y'], "val": financier_branch})
                        log("ok", f"  -> Financier Branch: {financier_branch}")
                    else:
                        log("warn", "  Financier Branch not found")

            # ── Create Proposal ───────────────────────────────────────
            log("info", "Clicking Create Proposal")
            await page.wait_for_timeout(400)
            cp_c = await page.evaluate("""() => {
                for (const el of document.querySelectorAll('button,input[type="submit"]')) {
                    if ((el.innerText||el.value||'').trim().toLowerCase().includes('create proposal')) {
                        el.scrollIntoView({block:'center'});
                        const r=el.getBoundingClientRect();
                        return r.width>0?{x:r.left+r.width/2,y:r.top+r.height/2}:null;
                    }
                }
                return null;
            }""")
            if cp_c:
                await page.mouse.click(cp_c['x'], cp_c['y'])
                log("ok", f"  -> Create Proposal clicked at ({cp_c['x']:.0f},{cp_c['y']:.0f})")
            else:
                log("warn", "  Create Proposal not found")
                await pause("Click 'Create Proposal' manually, then press ENTER")
            try:
                await page.wait_for_load_state("networkidle", timeout=20000)
            except Exception:
                pass
            await page.wait_for_timeout(2000)
            steps.append("proposal_created")
            log("ok", "══ SECTION DONE: Applicant Details / Proposal Created ══")

            # ── STEP 14: Payment ─────────────────────────────────────
            log("step", "STEP 14 — Payment")

            # Scroll to bottom so Pay Now button is in viewport
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(1000)

            # ── STEP 13: Proposal Number ──────────────────────────────
            # Read proposal number — retry up to 10x (page may take a few seconds to load)
            log("step", "STEP 13 — Saving Proposal Number")

            proposal_no = ""
            for _pn_try in range(25):
                proposal_no = await page.evaluate(r"""() => {
                    // Portal shows "Proposal Number: 4500823477" as a single element text.
                    // Strategy 1: find element containing "proposal number" and extract value after colon.
                    for (const el of document.querySelectorAll('*')) {
                        if ((el.children || []).length > 8) continue;
                        const t = (el.innerText || el.textContent || '').trim();
                        if (t.toLowerCase().includes('proposal number') ||
                            t.toLowerCase().includes('proposal no')) {
                            const m = t.match(/proposal\s*(?:number|no\.?)\s*[:\-]?\s*([A-Z0-9]{6,20})/i);
                            if (m && m[1]) return m[1].trim();
                        }
                    }
                    // Strategy 2: full body text scan
                    const allText = document.body.innerText || '';
                    const m = allText.match(/Proposal\s*(?:Number|No\.?)\s*[:\-]?\s*([A-Z0-9]{6,20})/i);
                    if (m && m[1]) return m[1].trim();
                    return null;
                }""") or ""
                if proposal_no:
                    break
                log("info", f"  Proposal number not visible yet, waiting... (attempt {_pn_try+1}/25)")
                await page.wait_for_timeout(1000)

            if proposal_no:
                log("ok", f"Proposal Number: {proposal_no}")
            else:
                log("warn", "Could not auto-read proposal number")
                proposal_no = input("      Type the Proposal Number from browser, press ENTER: ").strip()

            with open(LOG_PATH, "a") as f:
                f.write(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
                        f"{D['manufacturer_model']}  |  {D.get('engine_number','')}  |  {proposal_no}\n")
            log("ok", "Saved to proposal_log.txt")
            save_proposal_number(D, proposal_no)
            steps.append("proposal_saved")

            # Now click Pay Now — scroll to bottom, get fresh coords after scroll settles
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(800)
            pay_now_c = await page.evaluate("""() => {
                // Find Pay Now — scroll it into view, wait, then return fresh coords
                for (const el of document.querySelectorAll('*')) {
                    const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                    if (t !== 'pay now') continue;
                    el.scrollIntoView({block:'center'});
                    return true;
                }
                return false;
            }""")
            await page.wait_for_timeout(600)  # wait for scroll to settle
            pay_now_coords = await page.evaluate("""() => {
                for (const el of document.querySelectorAll('*')) {
                    const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                    if (t !== 'pay now') continue;
                    const r = el.getBoundingClientRect();
                    if (r.width > 0 && r.height > 0) {
                        return {x: r.left+r.width/2, y: r.top+r.height/2};
                    }
                }
                return null;
            }""")
            if pay_now_coords:
                await page.mouse.click(pay_now_coords['x'], pay_now_coords['y'])
                log("ok", "  -> Pay Now clicked")
                # Wait for payment page to fully load before attempting Advance PID
                try:
                    await page.wait_for_load_state("networkidle", timeout=15000)
                except Exception:
                    pass
                await page.wait_for_timeout(5000)

                # ── STEP 13 retry: read proposal number from payment page ──
                # The Premium Details card on the payment page shows proposal number
                # as bold plain text — try to capture it here if STEP 13 failed earlier
                if not proposal_no:
                    log("info", "  Retrying proposal number read from payment page...")
                    proposal_no_retry = await page.evaluate("""() => {
                        const isPN = s => /^[0-9]{8,15}$/.test((s||'').trim());
                        for (const el of document.querySelectorAll('*')) {
                            const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                            if (t !== 'proposal number' && t !== 'proposal no') continue;
                            let card = el.parentElement;
                            for (let i = 0; i < 5; i++) {
                                if (!card) break;
                                const walker = document.createTreeWalker(
                                    card, NodeFilter.SHOW_TEXT);
                                let node;
                                while ((node = walker.nextNode())) {
                                    const v = node.nodeValue.trim();
                                    if (isPN(v)) return v;
                                }
                                for (const child of card.querySelectorAll('*')) {
                                    if (child === el) continue;
                                    const v = (child.innerText||'').trim();
                                    if (isPN(v)) return v;
                                }
                                card = card.parentElement;
                            }
                        }
                        return null;
                    }""") or ""
                    if proposal_no_retry:
                        proposal_no = proposal_no_retry
                        log("ok", f"  Proposal Number (payment page): {proposal_no}")
                        with open(LOG_PATH, "a") as f:
                            f.write(f"{datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
                                    f"{D['manufacturer_model']}  |  {D.get('engine_number','')}  |  {proposal_no}\n")
                        save_proposal_number(D, proposal_no)
                    else:
                        log("warn", "  Proposal number still not found on payment page")

            else:
                log("warn", "  Pay Now button not found")

            if D["auto_pay"]:
                # Wait for payment page to fully load before searching for Advance PID
                log("info", "Waiting for payment page to load before Advance PID...")
                try:
                    await page.wait_for_load_state("networkidle", timeout=25000)
                except Exception:
                    pass
                # Auto-complete payment — click Advance PID
                # Try on current page AND any newly opened page/tab
                adv_clicked = False

                async def try_click_advance_pid(pg):
                    # Strategy 1: Playwright locator by text
                    try:
                        loc = pg.get_by_text("Advance PID", exact=True)
                        if await loc.count() > 0:
                            await loc.first.scroll_into_view_if_needed()
                            await pg.wait_for_timeout(300)
                            await loc.first.click(force=True)
                            return True
                    except Exception:
                        pass
                    # Strategy 2: XPath — find element whose text is exactly "Advance PID"
                    try:
                        loc2 = pg.locator("xpath=//*[normalize-space(text())='Advance PID']")
                        if await loc2.count() > 0:
                            await loc2.first.scroll_into_view_if_needed()
                            await pg.wait_for_timeout(300)
                            await loc2.first.click(force=True)
                            return True
                    except Exception:
                        pass
                    # Strategy 3: JS — click parent of text node "Advance PID"
                    try:
                        done = await pg.evaluate("""() => {
                            const walker = document.createTreeWalker(
                                document.body, NodeFilter.SHOW_TEXT);
                            let node;
                            while ((node = walker.nextNode())) {
                                if (node.nodeValue.trim().toLowerCase() === 'advance pid') {
                                    const el = node.parentElement;
                                    el.scrollIntoView({block:'center'});
                                    el.click();
                                    return true;
                                }
                            }
                            return false;
                        }""")
                        if done:
                            return True
                    except Exception:
                        pass
                    return False

                # Poll until "Advance PID" text appears on the page —
                # the payment page can take several minutes to fully load.
                # Wait up to 10 minutes (120 × 5s) before giving up.
                log("info", "  Waiting for Advance PID to appear on page...")
                for _adv_wait in range(120):
                    adv_visible = await page.evaluate("""() => {
                        const walker = document.createTreeWalker(
                            document.body, NodeFilter.SHOW_TEXT);
                        let node;
                        while ((node = walker.nextNode())) {
                            if (node.nodeValue.trim().toLowerCase() === 'advance pid')
                                return true;
                        }
                        return false;
                    }""")
                    if adv_visible:
                        log("info", f"  Advance PID found (attempt {_adv_wait+1})")
                        break
                    # Also check other open tabs
                    try:
                        for pg in page.context.pages:
                            if pg == page:
                                continue
                            adv_visible = await pg.evaluate("""() => {
                                const walker = document.createTreeWalker(
                                    document.body, NodeFilter.SHOW_TEXT);
                                let node;
                                while ((node = walker.nextNode())) {
                                    if (node.nodeValue.trim().toLowerCase() === 'advance pid')
                                        return true;
                                }
                                return false;
                            }""")
                            if adv_visible:
                                break
                    except Exception:
                        pass
                    if adv_visible:
                        break
                    if _adv_wait % 6 == 5:   # log every 30s so user knows it's working
                        log("info", f"  Still waiting for Advance PID... ({(_adv_wait+1)*5}s)")
                    await page.wait_for_timeout(5000)

                # Now click it
                for _adv_try in range(5):
                    adv_clicked = await try_click_advance_pid(page)
                    if adv_clicked:
                        break
                    try:
                        for pg in page.context.pages:
                            if pg == page:
                                continue
                            adv_clicked = await try_click_advance_pid(pg)
                            if adv_clicked:
                                break
                    except Exception:
                        pass
                    if adv_clicked:
                        break
                    await page.wait_for_timeout(1000)

                if adv_clicked:
                    log("ok", "  -> Advance PID clicked")
                    # Wait for Advance PID page to fully load
                    try:
                        await page.wait_for_load_state("networkidle", timeout=15000)
                    except Exception:
                        pass
                    await page.wait_for_timeout(3000)

                    # ── Tick the PID Number checkbox matching Excel pid_number ──
                    pid_val = str(D.get("pid_number", "")).strip()
                    if pid_val:
                        log("info", f"  Selecting PID Number: {pid_val}")
                        pid_ticked = False

                        # ── DEBUG: dump all rows and cells to find PID structure ──
                        debug_rows = await page.evaluate("""() => {
                            const out = [];
                            document.querySelectorAll('tr').forEach((row, ri) => {
                                const cells = Array.from(row.querySelectorAll('td,th'))
                                    .map(c => (c.innerText||'').trim().slice(0,40));
                                const hasCb = !!row.querySelector('input[type="checkbox"]');
                                if (cells.length > 0)
                                    out.push({ri, cells, hasCb});
                            });
                            return out.slice(0,20);
                        }""")
                        for dr in debug_rows:
                            log("info", f"  [ROW {dr['ri']}] cb={dr['hasCb']} cells={dr['cells']}")

                        # ── Also dump all inputs on page ──
                        debug_inputs = await page.evaluate("""() => {
                            return Array.from(document.querySelectorAll('input')).slice(0,10).map(i => ({
                                type: i.type, ph: i.placeholder, val: (i.value||'').slice(0,30),
                                cls: (i.className||'').slice(0,40)
                            }));
                        }""")
                        for di in debug_inputs:
                            log("info", f"  [INP] type={di['type']} ph='{di['ph']}' val='{di['val']}' cls='{di['cls']}'")

                        async def tick_pid_on_page(pg):
                            """Type PID in search box to filter table, then click its checkbox."""
                            # Step 1: type PID number into the Search box to filter the table
                            try:
                                search_inp = await pg.query_selector('input[placeholder*="Search" i], input[type="search"]')
                                if search_inp:
                                    await search_inp.triple_click()
                                    await search_inp.type(pid_val, delay=80)
                                    await pg.wait_for_timeout(1500)
                                    log("info", f"  Typed '{pid_val}' into search box")
                                else:
                                    log("info", "  No search box found")
                            except Exception as ex:
                                log("info", f"  Search box error: {ex}")

                            for _s in range(5):
                                # Step 2: scroll checkbox into view
                                scrolled = await pg.evaluate("""(pidVal) => {
                                    const rows = document.querySelectorAll('tr');
                                    for (const row of rows) {
                                        for (const cell of row.querySelectorAll('td, th')) {
                                            const t = (cell.innerText||'').trim();
                                            if (t === pidVal) {
                                                const cb = row.querySelector('input[type="checkbox"]');
                                                if (cb) { cb.scrollIntoView({block:'center'}); return 'cb-found'; }
                                                return 'row-found-no-cb';
                                            }
                                        }
                                    }
                                    return 'not-found';
                                }""", pid_val)
                                log("info", f"  PID scroll attempt {_s+1}: {scrolled}")
                                if scrolled == 'not-found':
                                    await pg.wait_for_timeout(1000)
                                    continue
                                if scrolled == 'row-found-no-cb':
                                    # No checkbox in tr — try clicking the row itself
                                    coords = await pg.evaluate("""(pidVal) => {
                                        const rows = document.querySelectorAll('tr');
                                        for (const row of rows) {
                                            for (const cell of row.querySelectorAll('td,th')) {
                                                if ((cell.innerText||'').trim() === pidVal) {
                                                    const r = row.getBoundingClientRect();
                                                    if (r.width > 0) return {x: r.left+20, y: r.top+r.height/2};
                                                }
                                            }
                                        }
                                        return null;
                                    }""", pid_val)
                                    if coords:
                                        await pg.wait_for_timeout(400)
                                        await pg.mouse.click(coords['x'], coords['y'])
                                        return True
                                # Step 3: re-read fresh coords after scroll settles
                                await pg.wait_for_timeout(500)
                                coords = await pg.evaluate("""(pidVal) => {
                                    const rows = document.querySelectorAll('tr');
                                    for (const row of rows) {
                                        for (const cell of row.querySelectorAll('td, th')) {
                                            if ((cell.innerText||'').trim() === pidVal) {
                                                const cb = row.querySelector('input[type="checkbox"]');
                                                if (cb) {
                                                    const cr = cb.getBoundingClientRect();
                                                    if (cr.width > 0)
                                                        return {x: cr.left+cr.width/2, y: cr.top+cr.height/2};
                                                }
                                            }
                                        }
                                    }
                                    return null;
                                }""", pid_val)
                                if coords:
                                    await pg.mouse.click(coords['x'], coords['y'])
                                    await pg.wait_for_timeout(300)
                                    return True
                                await pg.wait_for_timeout(1000)
                            return False

                        # Try on current page first
                        pid_ticked = await tick_pid_on_page(page)
                        # If not found, try all other open tabs
                        if not pid_ticked:
                            try:
                                for pg in page.context.pages:
                                    if pg == page:
                                        continue
                                    log("info", f"  Trying PID on tab: {pg.url}")
                                    pid_ticked = await tick_pid_on_page(pg)
                                    if pid_ticked:
                                        break
                            except Exception:
                                pass

                        if pid_ticked:
                            log("ok", f"  -> PID {pid_val} checkbox ticked")
                        else:
                            log("warn", f"  PID {pid_val} not found — tick manually")
                    else:
                        log("info", "  No PID Number in Excel — skipping PID checkbox")

                    await page.wait_for_timeout(1000)

                    # ── Click Pay button after PID checkbox ticked ──
                    log("info", "Clicking Pay button...")
                    pay_clicked = False
                    for _pay in range(5):
                        pay_coords = await page.evaluate("""() => {
                            for (const el of document.querySelectorAll('button, input[type="submit"], a')) {
                                const t = (el.innerText||el.textContent||el.value||'').trim().toLowerCase();
                                if (t === 'pay') {
                                    el.scrollIntoView({block:'center'});
                                    const r = el.getBoundingClientRect();
                                    if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                }
                            }
                            return null;
                        }""")
                        if pay_coords:
                            await page.wait_for_timeout(400)
                            # Re-read coords fresh after scroll
                            pay_coords2 = await page.evaluate("""() => {
                                for (const el of document.querySelectorAll('button, input[type="submit"], a')) {
                                    const t = (el.innerText||el.textContent||el.value||'').trim().toLowerCase();
                                    if (t === 'pay') {
                                        const r = el.getBoundingClientRect();
                                        if (r.width > 0) return {x: r.left+r.width/2, y: r.top+r.height/2};
                                    }
                                }
                                return null;
                            }""")
                            if pay_coords2:
                                await page.mouse.click(pay_coords2['x'], pay_coords2['y'])
                                log("ok", "  -> Pay clicked")
                                pay_clicked = True
                                break
                        await page.wait_for_timeout(1000)
                    if not pay_clicked:
                        log("warn", "  Pay button not found — click manually")
                        await pause("Click 'Pay' manually, then press ENTER")

                    log("ok", "Payment initiated!")
                    steps.append("payment_done")
                else:
                    log("warn", "Advance PID button not found — complete payment manually")
                    await pause("Complete payment manually in the browser, then press ENTER")
                    steps.append("payment_manual")
            else:
                log("info", "auto_pay = FALSE — complete payment manually")
                await pause("Complete payment manually in the browser, then press ENTER")
                steps.append("payment_manual")

            save_result(D, "SUCCESS", proposal_no, steps)

            print(f"""
+--------------------------------------------------------------+
|  ALL DONE!                                                   |
|  Vehicle  : {D['manufacturer_model']:<51}|
|  Proposal : {proposal_no:<51}|
|  Saved    : proposal_log.txt + OnlyEV.xlsx (Results tab)    |
+--------------------------------------------------------------+""")

            # ── Capture Policy Number from success page ──────────────
            log("info", "Waiting for success page and Policy Number...")
            try:
                await page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)

            # Retry up to 60x × 2s = 2 minutes — success page can take time.
            # Uses 4 strategies so no layout variation is missed.
            policy_no = ""
            for _pol in range(60):
                policy_no = await page.evaluate("""() => {
                    // Policy number pattern: 4digits/9digits/2digits/3digits
                    function isPolicyNo(v) {
                        return v && v.length > 6 &&
                               /[0-9]/.test(v) &&
                               v.includes('/') &&
                               !/policy number/i.test(v);
                    }

                    // ── Strategy 1: find label "Policy Number", read next sibling ──
                    let labelEl = null, labelSize = Infinity;
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim();
                        if (t.toLowerCase() === 'policy number') {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0) {
                                const sz = r.width * r.height;
                                if (sz < labelSize) { labelEl = el; labelSize = sz; }
                            }
                        }
                    }
                    if (labelEl) {
                        // Check next siblings
                        let sib = labelEl.nextElementSibling;
                        for (let i = 0; i < 5; i++) {
                            if (!sib) break;
                            const v = (sib.innerText||sib.textContent||'').trim();
                            if (isPolicyNo(v)) return v;
                            sib = sib.nextElementSibling;
                        }
                        // Walk up parent chain, collect all text nodes, find value after label
                        let row = labelEl.parentElement;
                        for (let i = 0; i < 6; i++) {
                            if (!row) break;
                            const texts = [];
                            const walker = document.createTreeWalker(row, NodeFilter.SHOW_TEXT);
                            let node;
                            while ((node = walker.nextNode())) {
                                const v = node.nodeValue.trim();
                                if (v) texts.push(v);
                            }
                            const idx = texts.findIndex(t => t.toLowerCase() === 'policy number');
                            if (idx !== -1) {
                                for (let j = idx+1; j < Math.min(idx+5, texts.length); j++) {
                                    if (isPolicyNo(texts[j])) return texts[j];
                                }
                            }
                            row = row.parentElement;
                        }
                    }

                    // ── Strategy 2: scan ALL text nodes for policy number pattern ──
                    const walker2 = document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
                    let node2;
                    while ((node2 = walker2.nextNode())) {
                        const v = node2.nodeValue.trim();
                        // Match format like 3005/433241549/00/000
                        if (v.match(/^[0-9]{4}[/][0-9]{9}[/][0-9]{2}[/][0-9]{3}$/)) return v;
                    }

                    // ── Strategy 3: scan visible elements for policy-like value ──
                    for (const el of document.querySelectorAll('span,p,div,td,h1,h2,h3,h4,strong,b')) {
                        const v = (el.innerText||'').trim();
                        if (v.match(/^[0-9]{4}[/][0-9]+[/][0-9]{2}[/][0-9]{3}$/)) return v;
                    }

                    return null;
                }""") or ""
                if policy_no:
                    log("ok", f"  Policy Number found (attempt {_pol+1}): {policy_no}")
                    break
                if _pol % 5 == 4:
                    log("info", f"  Still waiting for Policy Number... ({(_pol+1)*2}s)")
                await page.wait_for_timeout(2000)

            if policy_no:
                log("ok", f"  Policy Number: {policy_no}")
                save_policy_number(D, policy_no)
            else:
                log("warn", "  Policy Number not found — saving empty")

            # ── Click "Another Policy" to go back for next row ──────

            another_clicked = False
            for _ap in range(15):
                # Strategy 1: Playwright get_by_text (most reliable)
                try:
                    loc = page.get_by_text("Another Policy", exact=True)
                    if await loc.count() > 0:
                        await loc.first.scroll_into_view_if_needed()
                        await page.wait_for_timeout(400)
                        await loc.first.click(force=True)
                        log("ok", "  -> Another Policy clicked (get_by_text)")
                        another_clicked = True
                        break
                except Exception:
                    pass
                # Strategy 2: XPath exact text
                try:
                    loc2 = page.locator("xpath=//*[normalize-space(text())='Another Policy']")
                    if await loc2.count() > 0:
                        await loc2.first.scroll_into_view_if_needed()
                        await page.wait_for_timeout(400)
                        await loc2.first.click(force=True)
                        log("ok", "  -> Another Policy clicked (xpath)")
                        another_clicked = True
                        break
                except Exception:
                    pass
                # Strategy 3: JS scan all elements including div/span
                coords = await page.evaluate("""() => {
                    for (const el of document.querySelectorAll('*')) {
                        const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                        if (t === 'another policy') {
                            const r = el.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0) {
                                el.scrollIntoView({block:'center'});
                                return {x: r.left+r.width/2, y: r.top+r.height/2};
                            }
                        }
                    }
                    return null;
                }""")
                if coords:
                    await page.wait_for_timeout(400)
                    await page.mouse.click(coords['x'], coords['y'])
                    log("ok", "  -> Another Policy clicked (JS scan)")
                    another_clicked = True
                    break
                await page.wait_for_timeout(1000)
            if not another_clicked:
                log("warn", "  Another Policy button not found — waiting extra 30s and retrying once")
                await page.wait_for_timeout(30000)
                try:
                    loc = page.get_by_text("Another Policy")
                    if await loc.count() > 0:
                        await loc.first.click(force=True)
                        log("ok", "  -> Another Policy clicked (late retry)")
                        another_clicked = True
                except Exception:
                    pass
                if not another_clicked:
                    log("warn", "  Another Policy not found after all retries — skipping")

            # Wait for the details page to reload
            try:
                await page.wait_for_load_state("networkidle", timeout=15000)
            except Exception:
                pass
            await page.wait_for_timeout(3000)
            return D["_excel_row"] + 1  # signal to process next row

        except Exception as e:
            log("warn", f"Error: {e}")
            save_result(D, "FAILED", proposal_no, steps, str(e))
            await pause("Check the browser. Press ENTER to continue to next row.")
            return D.get("_excel_row", 4) + 1  # skip failed row, try next


# ═══════════════════════════════════════════════════════════════════
#  ENTRY POINT
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # ── Read first row to preview and confirm before starting ────
    D_first = read_details(start_row=4)

    print("\n  First record from OnlyEV.xlsx:")
    print(f"     Battery No   : {D_first.get('battery_number','')}")
    print(f"     Mobile       : {D_first.get('mobile_number','')}")
    print(f"     Txn Type     : {D_first.get('transaction_type','')}")
    print(f"     Customer Type: {D_first.get('customer_type','')}")
    print(f"     Manufacturer : {D_first.get('manufacturer','')}")
    print(f"     Model        : {D_first.get('model','')}")
    print(f"     Watt         : {D_first.get('watt','')}")
    print(f"     Year of Mfg  : {D_first.get('year_of_manufacture','')}")
    print(f"     Policy Start : {D_first.get('policy_start_date','')}")
    print(f"     Policy End   : {D_first.get('policy_end_date','')}")
    print(f"     UIN of Asset : {D_first.get('uin_asset','')}")
    print(f"     Invoice No   : {D_first.get('invoice_number','')}")
    print(f"     Customer Stt : {D_first.get('customer_state','')}")
    print(f"     IDV (Rs)     : {D_first.get('idv','')}")
    print(f"     Mandatory Cv : {'YES' if D_first.get('mandatory_covers') else 'NO'}")
    print(f"     Addl Disc    : {'YES (' + str(D_first.get('additional_discounts_dropdown','') or 'Other Loading') + ') @ ' + str(D_first.get('loading_pct','')) + '%' if D_first.get('additional_discount') else 'NO'}")
    print(f"     GST          : {D_first.get('gst_number','') or '(not provided)'}")
    print(f"     KYC Pincode  : {D_first.get('kyc_pincode','')}")
    print(f"     Insured Name : {D_first.get('insured_name','')}")
    print(f"     DOB          : {D_first.get('dob','')}")
    print(f"     Auto Pay     : {D_first.get('auto_pay')}")

    print("""
  ──────────────────────────────────────────────────────────────
  BEFORE pressing ENTER, make sure:

  1. Chrome is open with remote debugging. Start it with Win+R:
     "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe"
     --remote-debugging-port=9222 --user-data-dir="C:\\chrome-nysa"

  2. You are logged in to nysa.icicilombard.com (home page visible)
  ──────────────────────────────────────────────────────────────
    """)

    input("  Press ENTER to connect to your Chrome and start...")

    # ── Loop through all data rows ───────────────────────────────
    next_row = 4
    row_count = 0
    while True:
        try:
            D = read_details(start_row=next_row)
        except SystemExit:
            print("\n  No more data rows found. All done!")
            break

        row_count += 1

        # ── Skip if Proposal Number already filled ───────────────
        existing_proposal = str(D.get("proposal_number", "") or "").strip()
        if existing_proposal:
            print(f"\n[SKIP] Row {D['_excel_row']} already has Proposal Number: {existing_proposal} — skipping")
            next_row = D["_excel_row"] + 1
            continue

        # ── Skip if Engine/Chassis already marked green (full flow done) ──
        if is_engine_chassis_green(D):
            print(f"\n[SKIP] Row {D['_excel_row']} engine/chassis already green (policy issued) — skipping")
            next_row = D["_excel_row"] + 1
            continue

        print(f"\n{'='*60}")
        print(f"  Processing row {D['_excel_row']} (record #{row_count})")
        print(f"  Battery: {D.get('battery_number','')}  |  Model: {D.get('manufacturer','')[:20]} — {D.get('model','')[:20]}")
        print(f"{'='*60}")

        next_row = asyncio.run(run(D, skip_to_step4=(row_count > 1)))

        if next_row is None:
            # run() didn't return a next row — stop
            print("\n  Processing stopped.")
            break

        # Check if there are more rows
        try:
            read_details(start_row=next_row)  # just check — will raise SystemExit if no data
        except SystemExit:
            print(f"\n  All {row_count} record(s) processed successfully!")
            break

    print("\n  Script finished.")