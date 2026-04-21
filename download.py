"""
ICICI NYSA — Policy PDF Downloader
Reads policy.xlsx, searches each Policy Number on Nysa Policies tab,
downloads PDF to DOC folder, updates Status column in Excel.

USAGE:
  STEP 1: Start Chrome with remote debugging (run ONCE in Win+R):
    "C:\Program Files\Google\Chrome\Application\chrome.exe"
    --remote-debugging-port=9222 --user-data-dir="C:\chrome-nysa"
  STEP 2: Log in to nysa.icicilombard.com
  STEP 3: Run: python download.py
"""

import asyncio
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright

# ── Paths ──────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
EXCEL_PATH  = BASE_DIR / "policy.xlsx"
DOC_DIR     = BASE_DIR / "DOC"
CHROME_PORT = 9222


# ── Helpers ────────────────────────────────────────────────────
def log(kind, msg):
    icons = {"step":"[STEP]","ok":"  [OK]  ","warn":"  [WARN]","info":"  [INFO]"}
    print(f"{icons.get(kind,'[?]')} {msg}")


async def pause(msg="Done? Press ENTER to continue"):
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, lambda: input(f"\n  {msg}\n"))


# ── Connection — identical to icici_nysa_LATEST_STAR.py ────────
async def connect_to_chrome(playwright):
    """Attach to Chrome started with --remote-debugging-port=9222."""
    try:
        browser = await playwright.chromium.connect_over_cdp(
            f"http://localhost:{CHROME_PORT}"
        )
    except Exception as e:
        print(f"""
  ERROR: Cannot connect to Chrome on port {CHROME_PORT}

  Make sure you started Chrome with this command (Win+R, paste it):
  "C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\\chrome-nysa"

  Then login to nysa.icicilombard.com and run this script again.
  Details: {e}
        """)
        input("Press ENTER to exit...")
        raise SystemExit(1)

    # Find the NYSA tab
    nysa_page = None
    all_pages = [p for ctx in browser.contexts for p in ctx.pages]
    for p in all_pages:
        if "icicilombard.com" in p.url:
            nysa_page = p
            break
    if not nysa_page:
        if all_pages:
            nysa_page = all_pages[-1]
            log("warn", f"NYSA tab not found — using: {nysa_page.url}")
            await pause("Navigate to nysa.icicilombard.com, then press ENTER")
        else:
            print("  No open tabs found in Chrome.")
            input("Press ENTER to exit...")
            raise SystemExit(1)

    await nysa_page.bring_to_front()
    log("ok", f"Connected! Tab URL: {nysa_page.url}")
    return browser, nysa_page


# ── Excel helpers ──────────────────────────────────────────────
def read_policies():
    if not EXCEL_PATH.exists():
        print(f"\n  Cannot find: {EXCEL_PATH}")
        print("  Creating template policy.xlsx...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for i, h in enumerate(["Policy Number","Registration Number","Engine Number","Chassis Number","Proposal Number","Policy Start Date","Policy End Date","Status"], 1):
            ws.cell(1, i, h)
        wb.save(EXCEL_PATH)
        print(f"  Created: {EXCEL_PATH}  — fill in Policy Numbers and run again.")
        input("\nPress ENTER to exit...")
        raise SystemExit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    headers = {}
    for col in range(1, 13):
        v = ws.cell(1, col).value
        if v:
            headers[str(v).strip().lower()] = col

    policies = []
    for row in range(2, ws.max_row + 1):
        pn = ws.cell(row, headers.get("policy number", 1)).value
        if not pn or not str(pn).strip():
            continue
        status = str(ws.cell(row, headers.get("status", 7)).value or "").strip().lower()
        if status == "downloaded":
            continue
        policies.append({
            "row":             row,
            "policy_number":   str(pn).strip(),
            "status_col":      headers.get("status", 8),
        })
    return policies


def update_row(row, status_col, status, registration="", engine="", chassis="", proposal="", policy_start="", policy_end=""):
    """Write status + scraped data back to Excel row.
    Uses a temp-file-then-replace strategy to prevent corruption:
    always writes to policy.xlsx.tmp first, then renames over the original.
    If the script crashes mid-write, the original file is untouched.
    """
    import shutil
    tmp_path = EXCEL_PATH.with_suffix(".xlsx.tmp")
    for _attempt in range(3):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active
            # Find header columns
            headers = {}
            for col in range(1, 13):
                v = ws.cell(1, col).value
                if v:
                    headers[str(v).strip().lower()] = col
            ws.cell(row=row, column=status_col, value=status)
            # Always write all fields — prevents previous row values carrying over
            ws.cell(row=row, column=headers.get("registration number", 2), value=registration or "")
            ws.cell(row=row, column=headers.get("engine number",       3), value=engine       or "")
            ws.cell(row=row, column=headers.get("chassis number",      4), value=chassis      or "")
            ws.cell(row=row, column=headers.get("proposal number",     5), value=proposal     or "")
            ws.cell(row=row, column=headers.get("policy start date",   6), value=policy_start or "")
            ws.cell(row=row, column=headers.get("policy end date",     7), value=policy_end   or "")
            # Save to temp file first — if this crashes the original is safe
            wb.save(str(tmp_path))
            # Only replace original once temp is fully written
            shutil.move(str(tmp_path), str(EXCEL_PATH))
            log("ok", f"Excel updated -> row {row}: status='{status}'"
                + (f" reg={registration}"        if registration else "")
                + (f" engine={engine}"           if engine       else "")
                + (f" chassis={chassis}"         if chassis      else "")
                + (f" proposal={proposal}"       if proposal     else "")
                + (f" start={policy_start}"      if policy_start else "")
                + (f" end={policy_end}"          if policy_end   else ""))
            return
        except PermissionError:
            log("warn", "  Excel is open — close policy.xlsx and press ENTER")
            input("  Press ENTER after closing Excel...")
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception as e:
            log("warn", f"Could not update Excel: {e}")
            if tmp_path.exists():
                tmp_path.unlink()
            return


# ── Download one policy PDF ────────────────────────────────────
async def download_policy_pdf(page, policy, is_first=True):
    policy_no = policy["policy_number"]
    log("step", f"Processing: {policy_no}")

    # 1. First policy only — navigate to portal, click Motor, click Nysa Policies.
    #    From 2nd policy onwards the browser is already on Nysa Policies page,
    #    so skip directly to the Policy Number radio + search box.
    if is_first:
        if "nysa.icicilombard.com" not in page.url:
            log("info", "  Navigating to NYSA portal...")
            await page.goto("https://nysa.icicilombard.com/#/dashboard", wait_until="domcontentloaded")
            await page.wait_for_timeout(800)

        # 1a. Click Motor — retry until "Nysa Policies" tab actually appears.
        #     The dashboard card text contains "Motor" as a child node, so we
        #     target the SMALLEST visible element whose exact text is "Motor"
        #     (the card heading), not a parent that contains extra text.
        #     After each click we check for "Nysa Policies" — if it doesn't
        #     appear within 3s we click Motor again (click may not have landed).
        log("info", "Clicking Motor and waiting for Nysa Policies tab...")
        nysa_tab_visible = False
        # Helper: run a page.evaluate but swallow "Execution context was destroyed"
        # errors. After the Motor click the SPA navigates, which tears down the JS
        # context; the next evaluate on this page must not bubble up an exception
        # — we just treat it as "not ready yet" and let the loop retry.
        async def safe_eval(js, *args):
            try:
                if args:
                    return await page.evaluate(js, *args)
                return await page.evaluate(js)
            except Exception as _e:
                msg = str(_e).lower()
                if ("execution context was destroyed" in msg
                    or "navigation" in msg
                    or "target closed" in msg
                    or "frame was detached" in msg):
                    return None
                raise

        for _m in range(30):
            # Check if Nysa Policies already visible — if yes, no need to click
            nysa_tab_visible = await safe_eval("""() => {
                for (const el of document.querySelectorAll('*')) {
                    if ((el.innerText||el.textContent||'').trim() === 'Nysa Policies') {
                        return el.getBoundingClientRect().width > 0;
                    }
                }
                return false;
            }""")
            if nysa_tab_visible:
                log("ok", f"  -> Nysa Policies tab visible (attempt {_m+1})")
                break
            # If safe_eval returned None (context destroyed), wait for the page
            # to settle and retry the outer loop.
            if nysa_tab_visible is None:
                try:
                    await page.wait_for_load_state("domcontentloaded", timeout=5000)
                except Exception:
                    pass
                await page.wait_for_timeout(400)
                continue

            # Find the smallest element with exact text "Motor" — avoids
            # matching parent cards that contain extra child text.
            # Scroll it into view first so mouse.click lands correctly.
            mc = await safe_eval("""() => {
                let best = null, bestArea = Infinity;
                for (const el of document.querySelectorAll('*')) {
                    const t = (el.innerText||el.textContent||'').trim();
                    if (t !== 'Motor' && t !== 'Motor Insurance') continue;
                    const r = el.getBoundingClientRect();
                    if (r.width === 0 || r.height === 0) continue;
                    const area = r.width * r.height;
                    if (area < bestArea) { bestArea = area; best = el; }
                }
                if (!best) return null;
                best.scrollIntoView({block:'center', inline:'center'});
                const r = best.getBoundingClientRect();
                return {x: r.left+r.width/2, y: r.top+r.height/2};
            }""")
            if mc:
                await page.wait_for_timeout(300)   # let scroll settle
                try:
                    await page.mouse.click(mc['x'], mc['y'])
                except Exception as _ce:
                    log("info", f"  Motor click raised ({str(_ce)[:60]}) — retrying")
                    await page.wait_for_timeout(500)
                    continue
                log("info", f"  Motor clicked (attempt {_m+1}), waiting for Nysa Policies tab...")
                # Wait up to 3s for Nysa Policies tab to appear after this click.
                # Use safe_eval — the click usually triggers an SPA navigation which
                # destroys the JS context briefly.
                for _w in range(6):
                    await page.wait_for_timeout(500)
                    appeared = await safe_eval("""() => {
                        for (const el of document.querySelectorAll('*')) {
                            if ((el.innerText||el.textContent||'').trim() === 'Nysa Policies') {
                                return el.getBoundingClientRect().width > 0;
                            }
                        }
                        return false;
                    }""")
                    if appeared:
                        nysa_tab_visible = True
                        break
                if nysa_tab_visible:
                    log("ok", f"  -> Motor click confirmed — Nysa Policies tab appeared")
                    break
                log("info", "  Nysa Policies tab not appeared yet — re-clicking Motor...")
            else:
                await page.wait_for_timeout(400)
        if not nysa_tab_visible:
            log("warn", "  Nysa Policies tab never appeared — continuing anyway")

        # 1b. Click Nysa Policies tab
        log("info", "Clicking Nysa Policies tab...")
        for _n in range(15):
            try:
                loc = page.get_by_text("Nysa Policies", exact=True)
                if await loc.count() > 0:
                    await loc.first.click()
                    log("ok", "  -> Nysa Policies clicked")
                    break
            except Exception:
                pass
            c = await page.evaluate("""() => {
                for (const el of document.querySelectorAll('*')) {
                    if ((el.innerText||el.textContent||'').trim() === 'Nysa Policies') {
                        const r = el.getBoundingClientRect();
                        if (r.width > 0) return {x:r.left+r.width/2, y:r.top+r.height/2};
                    }
                } return null;
            }""")
            if c:
                await page.mouse.click(c['x'], c['y'])
                log("ok", "  -> Nysa Policies clicked (coords)")
                break
            await page.wait_for_timeout(500)

        try:
            await page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass

        # Wait for the Nysa Policies search form to be FULLY rendered before
        # moving on. Simply waiting on networkidle isn't enough — the SPA can
        # report networkidle while the search form/radio buttons are still
        # being painted. We poll until both the "Policy Number" radio option
        # AND the GET POLICIES button are visible. This is what makes the 1st
        # policy behave the same as the 2nd, 3rd, etc.
        log("info", "Waiting for Nysa Policies search form to load...")
        form_ready = False
        for _fr in range(40):   # up to 20s
            form_ready = await page.evaluate("""() => {
                // Policy Number radio label visible?
                let has_radio = false;
                for (const el of document.querySelectorAll('label,span')) {
                    const t = (el.innerText||el.textContent||'').trim().toLowerCase();
                    if (t === 'policy number' || t.includes('policy number')) {
                        const r = el.getBoundingClientRect();
                        if (r.width > 0 && r.height > 0) { has_radio = true; break; }
                    }
                }
                // GET POLICIES button visible?
                let has_button = false;
                for (const el of document.querySelectorAll('button,input[type="submit"]')) {
                    const t = (el.innerText||el.value||'').trim().toUpperCase();
                    if (t.includes('GET POLICIES') || t.includes('GET POLICY')) {
                        const r = el.getBoundingClientRect();
                        if (r.width > 0 && r.height > 0) { has_button = true; break; }
                    }
                }
                return has_radio && has_button;
            }""")
            if form_ready:
                log("ok", f"  -> Nysa Policies form ready (attempt {_fr+1})")
                break
            await page.wait_for_timeout(500)
        if not form_ready:
            log("warn", "  Search form never reported ready — continuing anyway")
        # Small settle wait after form is ready — lets event handlers attach
        await page.wait_for_timeout(800)
    else:
        log("info", "  Skipping Motor/Nysa Policies click (already on page)")

    # 2. Click Policy Number radio
    log("info", "Clicking Policy Number radio...")
    for _r in range(10):
        rc = await page.evaluate("""() => {
            for (const r of document.querySelectorAll('input[type="radio"]')) {
                const lbl = r.closest('label') ||
                            document.querySelector('label[for="'+r.id+'"]') ||
                            r.parentElement;
                if ((lbl?.innerText||lbl?.textContent||'').trim().toLowerCase().includes('policy number')) {
                    r.scrollIntoView({block:'center'});
                    const rect = r.getBoundingClientRect();
                    if (rect.width > 0) return {x:rect.left+rect.width/2, y:rect.top+rect.height/2};
                }
            }
            for (const el of document.querySelectorAll('label,span')) {
                if ((el.innerText||'').trim().toLowerCase() === 'policy number') {
                    const rect = el.getBoundingClientRect();
                    if (rect.width > 0) return {x:rect.left+rect.width/2, y:rect.top+rect.height/2};
                }
            }
            return null;
        }""")
        if rc:
            await page.mouse.click(rc['x'], rc['y'])
            log("ok", "  -> Policy Number radio clicked")
            break
        await page.wait_for_timeout(200)

    # 3. Type policy number in search box
    log("info", f"Typing policy number: {policy_no}")
    for _s in range(10):
        inp = await page.query_selector(
            'input[placeholder*="Policy Number" i],'
            'input[placeholder*="Enter Policy" i],'
            'input[placeholder*="Search" i]'
        )
        if inp:
            await inp.scroll_into_view_if_needed()
            await inp.click()
            await page.keyboard.press("Control+a")
            await page.keyboard.press("Delete")
            await page.keyboard.type(policy_no, delay=50)
            log("ok", f"  -> Typed: {policy_no}")
            break
        await page.wait_for_timeout(100)

    # 4. Click GET POLICIES
    log("info", "Clicking GET POLICIES...")
    for _g in range(10):
        try:
            loc = page.get_by_text("GET POLICIES", exact=True)
            if await loc.count() > 0:
                await loc.first.click(force=True)
                log("ok", "  -> GET POLICIES clicked")
                break
        except Exception:
            pass
        gc = await page.evaluate("""() => {
            for (const el of document.querySelectorAll('button,input[type="submit"]')) {
                const t = (el.innerText||el.value||'').trim().toUpperCase();
                if (t.includes('GET POLICIES') || t.includes('GET POLICY')) {
                    el.scrollIntoView({block:'center'});
                    const r = el.getBoundingClientRect();
                    if (r.width > 0) return {x:r.left+r.width/2, y:r.top+r.height/2};
                }
            } return null;
        }""")
        if gc:
            await page.mouse.click(gc['x'], gc['y'])
            log("ok", "  -> GET POLICIES clicked (coords)")
            break
        await page.wait_for_timeout(200)

    # Wait for the results table to show data for THIS policy number.
    # On 2nd+ policies the previous search result is still visible — we must
    # wait until the table contains the CURRENT policy number before scraping.
    try:
        await page.wait_for_load_state("networkidle", timeout=20000)
    except Exception:
        pass
    policy_no_clean = policy_no.strip().upper()
    results_found = False
    for _res in range(60):
        results_ready = await page.evaluate("""(pno) => {
            for (const tbl of document.querySelectorAll('table')) {
                const headers = Array.from(tbl.querySelectorAll('th'));
                const hasResultCols = headers.some(th => {
                    const t = (th.innerText||th.textContent||'').toLowerCase();
                    return t.includes('engine') || t.includes('chassis');
                });
                if (!hasResultCols) continue;
                // Must have a data row containing the current policy number
                const rows = Array.from(tbl.querySelectorAll('tr'));
                for (let i = 1; i < rows.length; i++) {
                    const tds = Array.from(rows[i].querySelectorAll('td'));
                    if (tds.some(td =>
                        (td.textContent||'').trim().toUpperCase().includes(pno)
                    )) return true;
                }
            }
            return false;
        }""", policy_no_clean)
        if results_ready:
            log("info", f"  Results for {policy_no} loaded (attempt {_res+1})")
            results_found = True
            break
        await page.wait_for_timeout(500)

    if not results_found:
        log("warn", f"  No results found for {policy_no} after 30s — skipping")
        update_row(policy["row"], policy["status_col"], "Not Found")
        return False

    # Extra wait: the row text appeared, but the Policy PDF column cell/icon
    # inside the row may still be mounting. Poll until the row has at least
    # as many <td> cells as the header has <th> cells — this confirms the row
    # is structurally complete (not a half-rendered placeholder). Also wait
    # for networkidle to let any per-row XHRs finish.
    log("info", "  Waiting for results row to be fully rendered...")
    for _rr in range(30):   # up to 15s
        row_ready = await page.evaluate("""(pno) => {
            for (const tbl of document.querySelectorAll('table')) {
                const ths = tbl.querySelectorAll('th');
                if (ths.length === 0) continue;
                const hasResultCols = Array.from(ths).some(th => {
                    const t = (th.innerText||th.textContent||'').toLowerCase();
                    return t.includes('engine') || t.includes('chassis');
                });
                if (!hasResultCols) continue;
                const rows = Array.from(tbl.querySelectorAll('tbody tr, tr'));
                for (const tr of rows) {
                    const tds = tr.querySelectorAll('td');
                    if (tds.length === 0) continue;
                    // Check this row contains the current policy number
                    const rowHasPolicy = Array.from(tds).some(td =>
                        (td.textContent||'').trim().toUpperCase().includes(pno)
                    );
                    if (!rowHasPolicy) continue;
                    // Row must have at least as many TDs as header has THs
                    // (meaning all columns are rendered, not just the first few)
                    if (tds.length >= ths.length - 1) return true;
                }
            }
            return false;
        }""", policy_no_clean)
        if row_ready:
            log("ok", f"  -> Row fully rendered (attempt {_rr+1})")
            break
        await page.wait_for_timeout(500)

    # Network quiet-down after row render — important so that any per-row
    # resource fetches (PDF icon images, tooltip JS, etc.) finish before we
    # try to click the Policy PDF icon.
    try:
        await page.wait_for_load_state("networkidle", timeout=10000)
    except Exception:
        pass
    await page.wait_for_timeout(500)

    # 4b. Scrape Engine / Chassis / Proposal from the results table.
    #
    # Actual column order (from UI screenshot):
    #   Registration Number | Chassis Number | Engine Number |
    #   Policy/Covernote number | Customer Name | Policy Start Date |
    #   Policy End Date | Proposal Number | View Proposal | Policy Status |
    #   Policy PDF | Send PDF
    #
    # We target ONLY the table whose headers include "engine" or "chassis"
    # — this avoids accidentally reading the filter/search form table.

    scraped = await page.evaluate("""() => {
        const result = {registration:'', engine:'', chassis:'', proposal:'', policy_start:'', policy_end:'', policy_in_row:'', debug:''};

        function mapHeaders(cells) {
            const colMap = {};
            cells.forEach((cell, i) => {
                const t = (cell.innerText || cell.textContent || '')
                            .replace(/\s+/g, ' ').trim().toLowerCase();
                if (t.includes('registration'))
                    colMap.registration = i;
                if (t.includes('engine'))
                    colMap.engine   = i;
                if (t.includes('chassis'))
                    colMap.chassis  = i;
                if (t.includes('proposal') && !t.includes('view'))
                    colMap.proposal = i;
                if (t.includes('policy') && (t.includes('covernote') || t.includes('number')))
                    colMap.policy_no = i;
                if (t.includes('policy') && t.includes('start') && t.includes('date'))
                    colMap.policy_start = i;
                if (t.includes('policy') && t.includes('end') && t.includes('date'))
                    colMap.policy_end = i;
            });
            return colMap;
        }

        for (const tbl of document.querySelectorAll('table')) {
            const allTh = Array.from(tbl.querySelectorAll('th'));
            if (allTh.length === 0) continue;
            const colMap = mapHeaders(allTh);
            result.debug = 'th=' + allTh.length + ' colMap=' + JSON.stringify(colMap);
            if (!('engine' in colMap) && !('chassis' in colMap)) continue;
            for (const tr of Array.from(tbl.querySelectorAll('tr'))) {
                const tds = Array.from(tr.querySelectorAll('td'));
                if (tds.length === 0) continue;
                if ('registration'  in colMap && tds[colMap.registration])
                    result.registration  = (tds[colMap.registration].textContent  || '').trim();
                if ('engine'       in colMap && tds[colMap.engine])
                    result.engine        = (tds[colMap.engine].textContent       || '').trim();
                if ('chassis'      in colMap && tds[colMap.chassis])
                    result.chassis       = (tds[colMap.chassis].textContent      || '').trim();
                if ('proposal'     in colMap && tds[colMap.proposal])
                    result.proposal      = (tds[colMap.proposal].textContent     || '').trim();
                if ('policy_no'    in colMap && tds[colMap.policy_no])
                    result.policy_in_row = (tds[colMap.policy_no].textContent    || '').trim();
                if ('policy_start' in colMap && tds[colMap.policy_start])
                    result.policy_start  = (tds[colMap.policy_start].textContent || '').trim();
                if ('policy_end'   in colMap && tds[colMap.policy_end])
                    result.policy_end    = (tds[colMap.policy_end].textContent   || '').trim();
                break;
            }
            if (result.engine || result.chassis || result.proposal) return result;
        }
        return result;
    }""")
    log("info", f"  Scrape debug: {scraped.get('debug','')}")
    log("info", f"  Scraped — reg:'{scraped.get('registration','')}' engine:'{scraped.get('engine','')}' chassis:'{scraped.get('chassis','')}' proposal:'{scraped.get('proposal','')}' start:'{scraped.get('policy_start','')}' end:'{scraped.get('policy_end','')}' policy_in_row:'{scraped.get('policy_in_row','')}'")

    # CRITICAL: verify the scraped row belongs to THIS policy number.
    # If the page was slow and still shows the previous policy's results,
    # the policy number in the data row will NOT match the current one.
    # In that case write empty strings — never carry over another policy's data.
    policy_in_row = scraped.get("policy_in_row", "").strip().upper().replace(" ", "")
    policy_no_norm = policy_no.strip().upper().replace(" ", "")
    data_is_correct = (policy_in_row == policy_no_norm) or (policy_no_norm in policy_in_row)

    if not data_is_correct and (scraped.get("engine") or scraped.get("chassis")):
        log("warn", f"  STALE DATA DETECTED — table shows '{policy_in_row}' but current is '{policy_no_norm}' — saving empty to prevent carry-over")
        scraped["registration"] = ""
        scraped["engine"]       = ""
        scraped["chassis"]      = ""
        scraped["proposal"]     = ""
        scraped["policy_start"] = ""
        scraped["policy_end"]   = ""

    # Save to Excel immediately — status "In Progress" marks it started
    update_row(
        policy["row"], policy["status_col"], "In Progress",
        registration=scraped.get("registration", ""),
        engine=scraped.get("engine", ""),
        chassis=scraped.get("chassis", ""),
        proposal=scraped.get("proposal", ""),
        policy_start=scraped.get("policy_start", ""),
        policy_end=scraped.get("policy_end", "")
    )

    # 5. Scroll table right to reveal Policy PDF column
    log("info", "Scrolling table right...")
    for _sc in range(15):
        visible = await page.evaluate("""() => {
            return (document.body.innerText||'').toLowerCase().includes('policy pdf');
        }""")
        if visible:
            log("info", "  Policy PDF column visible")
            break
        await page.evaluate("""() => {
            document.querySelectorAll('*').forEach(el => {
                if (el.scrollWidth > el.clientWidth + 10) {
                    const s = window.getComputedStyle(el);
                    if (s.overflowX === 'auto' || s.overflowX === 'scroll')
                        el.scrollLeft += 300;
                }
            });
        }""")
        await page.wait_for_timeout(300)

    # 6. Click Policy PDF icon using page.mouse.click on exact screen coordinates.
    #    JS .click() on a CDP-connected Chrome does not trigger the browser's
    #    native file download — only a real mouse event does.
    log("info", "Clicking Policy PDF download icon...")
    safe_name = policy_no.replace("/", "_").replace("\\", "_")
    pdf_path = DOC_DIR / f"{safe_name}.pdf"

    download_done = False
    import shutil
    try:
        # Find "Policy PDF" column index within the RESULTS table only.
        # query_selector_all("th") picks up th from ALL tables on the page
        # (filter forms etc.) giving wrong index like 27.
        # We scope to the results table — the one whose headers include "engine".
        pdf_col_idx = -1
        results_table = None
        for tbl in await page.query_selector_all("table"):
            ths = await tbl.query_selector_all("th")
            for th in ths:
                raw = await th.inner_text()
                if "engine" in " ".join(raw.split()).lower():
                    results_table = tbl
                    break
            if results_table:
                break

        if results_table:
            ths_in_table = await results_table.query_selector_all("th")
            for i, th in enumerate(ths_in_table):
                raw = await th.inner_text()
                txt = " ".join(raw.split()).lower()
                if "policy" in txt and "pdf" in txt and "send" not in txt:
                    pdf_col_idx = i
                    break

        log("info", f"  Policy PDF column index: {pdf_col_idx}")

        if pdf_col_idx == -1:
            log("warn", "  Policy PDF column not found — cannot click icon")
        else:
            # Scroll table right by clicking the ► right-arrow button at the
            # end of the horizontal scrollbar (visible in screen recording at
            # far right of the scrollbar track). Clicking it repeatedly scrolls
            # the table right until Policy PDF column enters the viewport.
            # This avoids drag coordinate issues with scrollbar thumb position.

            # First scroll the page down so the table scrollbar is in viewport
            await page.evaluate(
                "() => { document.querySelectorAll('*').forEach(el => {"
                "  const s = window.getComputedStyle(el);"
                "  if (s.overflowX==='auto'||s.overflowX==='scroll')"
                "    el.scrollLeft = el.scrollWidth;"
                "}); }"
            )
            await page.wait_for_timeout(500)
            log("info", "  Scrolled table container fully right")

            # Also find the right-arrow (►) button of the scrollbar and click it
            arrow_info = await page.evaluate("""() => {
                // The ► arrow is usually an element at the far right of the
                // scrollbar track — find it by looking for the rightmost visible
                // small clickable element near the scrollbar area
                const scrollers = Array.from(document.querySelectorAll('*')).filter(el => {
                    const s = window.getComputedStyle(el);
                    return (s.overflowX === 'auto' || s.overflowX === 'scroll')
                           && el.scrollWidth > el.clientWidth + 10
                           && el.clientWidth > 100;
                });
                if (!scrollers.length) return null;
                const scroller = scrollers.reduce((a,b) => a.clientWidth > b.clientWidth ? a : b);
                const r = scroller.getBoundingClientRect();
                // Clamp bottom to viewport so it's always on screen
                const visBottom = Math.min(r.bottom, window.innerHeight - 5);
                return {
                    left:  r.left,
                    right: Math.min(r.right, window.innerWidth - 5),
                    y:     visBottom - 6
                };
            }""")

            # scrollLeft = scrollWidth above already scrolls fully right.
            # Do NOT click the right-arrow — it triggers spurious downloads.
            await page.wait_for_timeout(500)
            log("info", "  Table scrolled right via scrollLeft")

            # Find "Policy PDF" th in Python, scroll it into view,
            # then use bounding_box() for exact screen coordinates.
            # No viewport width check — scrollIntoView ensures it is on screen.
            pdf_th_el = None
            for th in await page.query_selector_all("th"):
                txt = " ".join((await th.inner_text()).split()).lower()
                if txt == "policy pdf":
                    pdf_th_el = th
                    break

            # Before reading coordinates, wait for the PDF cell in the target
            # row to actually contain something clickable (icon/button/link).
            # An empty <td></td> will give valid coordinates but the click will
            # land on blank space and nothing downloads. Poll for up to 10s.
            if pdf_th_el:
                for _pic in range(20):
                    cell_has_icon = await pdf_th_el.evaluate("""(th) => {
                        const table = th.closest('table');
                        if (!table) return false;
                        // Column index of Policy PDF
                        const ths = Array.from(table.querySelectorAll('th'));
                        const colIdx = ths.indexOf(th);
                        if (colIdx < 0) return false;
                        // First data row's cell at that column
                        const firstRow = table.querySelector('tbody tr') || table.querySelector('tr:nth-child(2)');
                        if (!firstRow) return false;
                        const tds = firstRow.querySelectorAll('td');
                        if (colIdx >= tds.length) return false;
                        const cell = tds[colIdx];
                        // Cell is 'ready' if it has an img, svg, button, or a
                        // non-whitespace child of nonzero size.
                        const clickable = cell.querySelector('img, svg, button, a, i[class], [role="button"]');
                        if (clickable) {
                            const r = clickable.getBoundingClientRect();
                            if (r.width > 0 && r.height > 0) return true;
                        }
                        return false;
                    }""")
                    if cell_has_icon:
                        log("info", f"  Policy PDF icon present in row (attempt {_pic+1})")
                        break
                    await page.wait_for_timeout(500)
                else:
                    log("warn", "  Policy PDF icon never appeared in row — clicking anyway")

            icon_coords = None
            if pdf_th_el:
                await pdf_th_el.scroll_into_view_if_needed()
                await page.wait_for_timeout(300)
                th_box = await pdf_th_el.bounding_box()
                if th_box:
                    center_x = th_box["x"] + th_box["width"] / 2
                    log("info", f"  Policy PDF th at x={th_box['x']:.0f} w={th_box['width']:.0f} -> center_x={center_x:.0f}")
                    # Get the first data td from the SAME table as the Policy PDF th
                    # by walking up to the table element, then finding its first td
                    first_td_in_table = await pdf_th_el.evaluate_handle(
                        "(th) => th.closest('table').querySelector('tbody tr td, tr td')"
                    )
                    td_box = await first_td_in_table.bounding_box() if first_td_in_table else None
                    if td_box and td_box["height"] > 0:
                        icon_coords = {
                            "x": center_x,
                            "y": td_box["y"] + td_box["height"] / 2
                        }
                    else:
                        # Fallback: use th bottom + 40px (row is just below header)
                        icon_coords = {
                            "x": center_x,
                            "y": th_box["y"] + th_box["height"] + 40
                        }

            if not icon_coords:
                log("warn", "  Could not get Policy PDF icon coordinates after scrolling")
            else:
                await page.wait_for_timeout(300)
                log("info", f"  Policy PDF icon at ({icon_coords['x']:.0f},{icon_coords['y']:.0f})")
                async with page.expect_download(timeout=60000) as dl_info:
                    await page.mouse.click(icon_coords["x"], icon_coords["y"])
                    log("ok", "  -> Policy PDF icon clicked (mouse)")

                dl = await dl_info.value
                failure = await dl.failure()
                if failure:
                    log("warn", f"  Download failed: {failure}")
                else:
                    # Strategy A: wait for the CDP-named UUID file to appear in DOC_DIR.
                    # Strategy B: if that never appears (common on the 1st download
                    #            before setDownloadBehavior has fully propagated),
                    #            fall back to Playwright's dl.save_as() — BUT do so
                    #            only after giving CDP its chance first.
                    found_file = None
                    for _fw in range(60):   # up to 30s
                        # Look for any new file in DOC_DIR (UUID name, no extension)
                        for f in DOC_DIR.iterdir():
                            if f.is_file() and f.suffix.lower() != ".pdf" and f.stat().st_size > 1024:
                                found_file = f
                                break
                        if found_file:
                            break
                        # Also check if save_as already wrote it as .pdf
                        if pdf_path.exists() and pdf_path.stat().st_size > 1024:
                            found_file = pdf_path
                            break
                        await page.wait_for_timeout(500)

                    if not found_file:
                        # Strategy B: ask Playwright for the download path.
                        # dl.path() blocks until the download finishes, so we get
                        # the real file even if CDP setDownloadBehavior was late.
                        log("info", "  No UUID file found — asking Playwright for download path")
                        try:
                            pw_path = await dl.path()
                            if pw_path:
                                import shutil as _sh
                                pw_p = Path(pw_path)
                                # Wait for Playwright to finish writing
                                for _pw_wait in range(20):
                                    if pw_p.exists() and pw_p.stat().st_size > 1024:
                                        break
                                    await page.wait_for_timeout(500)
                                if pw_p.exists() and pw_p.stat().st_size > 1024:
                                    _sh.copy(str(pw_p), str(pdf_path))
                                    log("ok", f"  Copied from Playwright path: {pw_p.name} ({pw_p.stat().st_size:,} bytes)")
                                    found_file = pdf_path
                                else:
                                    log("info", f"  Playwright path exists={pw_p.exists()} size={pw_p.stat().st_size if pw_p.exists() else 0}")
                        except Exception as _pe:
                            log("info", f"  dl.path() failed: {_pe}")

                        # Strategy C (last resort): save_as to target path
                        if not found_file:
                            log("info", "  Trying dl.save_as() as last resort")
                            try:
                                await dl.save_as(str(pdf_path))
                            except Exception as _se:
                                log("info", f"  save_as failed: {_se}")
                            # Wait longer for save_as — on slow first downloads it
                            # can take 5-10 seconds to actually land on disk.
                            for _sa_wait in range(20):   # up to 10s
                                if pdf_path.exists() and pdf_path.stat().st_size > 1024:
                                    break
                                await page.wait_for_timeout(500)
                            if pdf_path.exists() and pdf_path.stat().st_size > 1024:
                                found_file = pdf_path

                    if found_file and found_file != pdf_path:
                        found_file.rename(pdf_path)
                        log("ok", f"  Renamed {found_file.name} -> {pdf_path.name}")

                    if pdf_path.exists() and pdf_path.stat().st_size > 1024:
                        size = pdf_path.stat().st_size
                        log("ok", f"  -> Saved: {pdf_path.name} ({size:,} bytes)")
                        download_done = True
                    else:
                        size = pdf_path.stat().st_size if pdf_path.exists() else 0
                        log("warn", f"  File too small or missing ({size} bytes)")
                        if pdf_path.exists(): pdf_path.unlink()

        # Remove any remaining leftover temp files (UUID names, no extension)
        for tmp in DOC_DIR.iterdir():
            if tmp.is_file() and tmp.suffix.lower() != ".pdf":
                try:
                    tmp.unlink()
                    log("info", f"  Removed temp file: {tmp.name}")
                except Exception:
                    pass
    except Exception as e:
        log("warn", f"  Download error: {e}")

    status = "Downloaded" if download_done else "Failed"
    update_row(
        policy["row"], policy["status_col"], status,
        registration=scraped.get("registration", ""),
        engine=scraped.get("engine", ""),
        chassis=scraped.get("chassis", ""),
        proposal=scraped.get("proposal", ""),
        policy_start=scraped.get("policy_start", ""),
        policy_end=scraped.get("policy_end", "")
    )
    return download_done


# ── Main ───────────────────────────────────────────────────────
async def main():
    DOC_DIR.mkdir(exist_ok=True)
    log("info", f"DOC folder: {DOC_DIR}")

    policies = read_policies()
    if not policies:
        print("\n  No unprocessed policies found in policy.xlsx")
        print("  (Rows with Status='Downloaded' are skipped)")
        return

    print(f"\n  Found {len(policies)} policy/policies to download:")
    for p in policies:
        print(f"    Row {p['row']}: {p['policy_number']}")

    print("""
  ──────────────────────────────────────────────────────────────
  Before pressing ENTER make sure:
  1. Chrome is running with --remote-debugging-port=9222
  2. You are logged in to nysa.icicilombard.com
  ──────────────────────────────────────────────────────────────
    """)
    input("  Press ENTER to start downloading...")

    async with async_playwright() as p:
        browser, page = await connect_to_chrome(p)

        # Set download path via CDP
        # IMPORTANT: do this BEFORE any page interaction so Chrome has the setting
        # registered by the time the first download click happens. We also attach
        # the CDP session to the browser-level target (not just the page) so that
        # new tabs/popups triggered by a PDF click also honour the download path.
        try:
            cdp = await page.context.new_cdp_session(page)
            await cdp.send("Browser.setDownloadBehavior", {
                "behavior": "allowAndName",
                "downloadPath": str(DOC_DIR.resolve()),
                "eventsEnabled": True
            })
            # Give CDP a moment to register the setting. Without this small wait
            # the first download click often completes BEFORE Chrome has applied
            # the download-path override, causing the file to either land in the
            # user's default Downloads folder or get reported as failed. Second
            # and later downloads always work because the setting is active by then.
            await page.wait_for_timeout(1500)
            log("ok", f"Download path: {DOC_DIR.resolve()}")
        except Exception as e:
            log("warn", f"Could not set download path: {e}")

        ok, fail = 0, 0
        for i, policy in enumerate(policies, 1):
            print(f"\n{'='*60}")
            print(f"  Policy {i}/{len(policies)}: {policy['policy_number']}")
            print(f"{'='*60}")
            try:
                result = await download_policy_pdf(page, policy, is_first=(i == 1))
                if result: ok += 1
                else: fail += 1
            except Exception as e:
                log("warn", f"Error: {e}")
                update_row(policy["row"], policy["status_col"], f"Error: {str(e)[:50]}")
                fail += 1
            await page.wait_for_timeout(300)

        print(f"""
+--------------------------------------------------------------+
|  DOWNLOAD COMPLETE                                           |
|  Success : {ok:<51}|
|  Failed  : {fail:<51}|
|  PDFs in : {str(DOC_DIR):<51}|
+--------------------------------------------------------------+""")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n  Stopped by user.")
    except SystemExit:
        input("\n  Press ENTER to close...")
    except Exception as e:
        print(f"\n  Unexpected error: {e}")
        import traceback; traceback.print_exc()
    finally:
        input("\n  Press ENTER to close...")